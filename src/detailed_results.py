from __future__ import annotations

from datetime import date, datetime, time
from pathlib import Path
import re
from typing import Any

try:
    from .results_model import RESULTS_MODEL_FILENAME, read_result_explorer_model, model_index, model_sheet
except Exception:  # pragma: no cover - import fallback for direct execution
    from src.results_model import RESULTS_MODEL_FILENAME, read_result_explorer_model, model_index, model_sheet

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_to_tuple


CHART_DASHBOARD_DATA_SHEET = '_Chart Dashboard Data'
DETAILED_RESULTS_MAX_FULL_CELLS = 12000
DETAILED_RESULTS_SAFE_MAX_ROWS = 260
DETAILED_RESULTS_SAFE_MAX_COLS = 48
CHART_MAX_POINTS = 45
CHART_MAX_SERIES = 10
CHART_MAX_SLICES = 12



def _clean_sheet_title(sheet_name: str) -> str:
    """Return a UI-facing page name without spreadsheet ordering prefixes.

    Sheet tabs use hierarchical ordinal prefixes (e.g. "1A. Executive
    Summary", "1H. Spending Summary"), not just plain numbers ("22.
    Glossary") - the prefix can include trailing letters.
    """
    text = str(sheet_name or "").strip()
    text = re.sub(r"^\s*\d+[A-Za-z]*\.\s*", "", text)
    return text or str(sheet_name or "Results")


def _is_chart_dashboard_sheet(sheet_name: str) -> bool:
    low = str(sheet_name or "").strip().lower()
    if "chart" in low and "dashboard" in low:
        return True
    # Refactored workbook tab name.  Excel sheet names are flat, so the
    # section/letter prefix is part of the visible tab name.
    return bool(re.match(r"^\d+[a-z]\.\s*charts$", low) or low == "charts")


def _json_value(value: Any) -> Any:
    """Convert workbook cell values into JSON-safe, readable result values."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat(timespec="seconds")
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    return str(value)


def _cell_value_kind(value: Any, number_format: str = "") -> str:
    """Classify workbook cells so the UI can apply readable formatting.

    Excel stores many dollar cells as plain numbers plus a number format. The
    detailed-results explorer preserves the raw value, but also tags common
    currency/percentage/date cells so the browser can render compact,
    human-readable values such as $250K instead of 250000.
    """
    fmt = str(number_format or "").lower()
    if value is None:
        return "blank"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, (datetime, date, time)):
        return "date"
    if isinstance(value, (int, float)):
        if "%" in fmt:
            return "percent"
        currency_tokens = ("$", "[$", "accounting", "usd", "eur", "gbp", "¥", "€", "£")
        if any(token in fmt for token in currency_tokens):
            return "currency"
        return "number"
    return "text"


def _display_value(value: Any, number_format: str = "") -> str:
    if value is None:
        return ""
    kind = _cell_value_kind(value, number_format)
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat(timespec="seconds")
    if isinstance(value, (int, float)):
        n = float(value)
        if kind == "currency":
            k = round(abs(n) / 1000)
            return ("-" if n < 0 else "") + "$" + f"{k:,}" + "K"
        if kind == "percent":
            pct = n * 100 if abs(n) <= 1 else n
            return f"{pct:.0f}%"
        if isinstance(value, float):
            # Keep result data compact without hiding precision when it matters.
            if value.is_integer():
                return str(int(value))
            return f"{value:.6f}".rstrip("0").rstrip(".")
    return str(value)


def _trim(values: list[Any]) -> list[Any]:
    out = list(values)
    while out and (out[-1] is None or str(out[-1]).strip() == ""):
        out.pop()
    return out


def _is_blank_row(values: list[Any]) -> bool:
    return not any(str(v).strip() for v in values if v is not None)


def _looks_like_section_title(values: list[Any]) -> bool:
    nonblank = [str(v).strip() for v in values if v is not None and str(v).strip()]
    if not nonblank:
        return False
    if len(nonblank) == 1:
        text = nonblank[0]
        # Workbook section titles are often merged across columns, all-caps, or
        # descriptive standalone rows before a table begins.
        return bool(text) and (len(text) <= 140)
    return len(nonblank) <= 2 and sum(len(x) for x in nonblank) <= 120


def _section_title(sheet_name: str, start_row: int, rows: list[dict[str, Any]], index: int) -> str:
    if rows:
        first_cells = rows[0].get("cells") or []
        nonblank = [str(c.get("display", "")).strip() for c in first_cells if str(c.get("display", "")).strip()]
        if nonblank and _looks_like_section_title([c for c in nonblank]):
            title = " — ".join(nonblank[:2]).strip()
            if title:
                return title
    # Avoid exposing spreadsheet row ranges or mechanical section numbers in the UI.
    # If a sheet has no natural in-sheet title, use the user-facing page title.
    return _clean_sheet_title(sheet_name)


def _sheet_category(sheet_name: str) -> str:
    name = str(sheet_name or "")
    low = name.lower()
    if any(x in low for x in ["monte carlo", "survivor", "stress", "scenario", "life insurance", "ltc", "long-term care", "market-luck", "divorce", "qdro"]):
        return "Stress Tests"
    if any(x in low for x in ["roth conversion", "social security", "state residency", "s-corp", "llc", "entity", "charitable", "asset location", "asset allocation", "withdrawal strategy", "optimizer", "planning lever", "estate"]):
        return "Strategy"
    if any(x in low for x in ["system configuration", "diagnostic", "schema", "reference data", "security master"]):
        return "System Configuration"
    return "Reports"


def _categories_from_sheets(sheets: list[dict[str, Any]]) -> list[dict[str, Any]]:
    categories: list[dict[str, Any]] = []
    by_category: dict[str, list[dict[str, Any]]] = {}
    for sheet in sheets:
        by_category.setdefault(sheet["category"], []).append({
            "name": sheet["name"],
            "display_name": sheet.get("display_name") or _clean_sheet_title(sheet["name"]),
            "row_count": sheet.get("row_count", 0),
            "section_count": sheet.get("chart_count", (len(sheet.get("sections") or []) if "sections" in sheet else sheet.get("section_count", 0))),
            "chart_count": sheet.get("chart_count", 0),
            "kind": sheet.get("kind", "table"),
        })
    category_order = [
        "Reports",
        "Strategy",
        "Stress Tests",
        "System Configuration",
        "Other workbook detail",
    ]
    for name in category_order:
        if name in by_category:
            categories.append({"name": name, "sheets": by_category[name]})
    for name in sorted(k for k in by_category if k not in category_order):
        categories.append({"name": name, "sheets": by_category[name]})
    return categories


def _sheet_to_sections(ws: Any, *, max_rows: int | None = None, max_cols: int | None = None, preview_note: str = "") -> dict[str, Any]:
    workbook_max_row = int(ws.max_row or 0)
    workbook_max_col = int(ws.max_column or 0)
    max_row, max_col, auto_bounded, auto_note = _safe_sheet_limits(ws, max_rows, max_cols)
    truncated = bool(auto_bounded or (max_rows and workbook_max_row > max_rows) or (max_cols and workbook_max_col > max_cols))
    preview_note = preview_note or auto_note
    raw_rows: list[dict[str, Any] | None] = []
    used_max_col = 0

    for r in range(1, max_row + 1):
        values = _trim([ws.cell(r, c).value for c in range(1, max_col + 1)])
        if _is_blank_row(values):
            raw_rows.append(None)
            continue
        used_max_col = max(used_max_col, len(values))
        cells = []
        for c, value in enumerate(values, start=1):
            cell = ws.cell(r, c)
            number_format = getattr(cell, "number_format", "") or ""
            cells.append({
                "column": get_column_letter(c),
                "value": _json_value(value),
                "display": _display_value(value, number_format),
                "kind": _cell_value_kind(value, number_format),
                "number_format": str(number_format),
            })
        raw_rows.append({"number": r, "cells": cells})

    sections: list[dict[str, Any]] = []
    current: list[dict[str, Any]] = []
    current_start = 1
    fallback_title = _clean_sheet_title(ws.title)

    def flush() -> None:
        nonlocal current, current_start
        if not current:
            return
        title = _section_title(ws.title, current_start, current, len(sections) + 1)
        # A section with no natural title of its own (blank-row spacing before
        # a table or total row, not a real subsection break) falls back to the
        # generic sheet name. If the previous section already carries that same
        # sheet name in its own (real) title, merge into it instead of stacking
        # a redundant "Sheet Name" heading with no distinguishing content -
        # e.g. Spending Summary's blank-row-separated metrics/table/total blocks
        # showed as three headings, two of them bare duplicates of the sheet name.
        if (
            title == fallback_title
            and sections
            and fallback_title.lower() in sections[-1]["title"].lower()
        ):
            prev = sections[-1]
            prev["rows"].extend(current)
            prev["row_count"] = len(prev["rows"])
            prev["end_row"] = current[-1]["number"]
            current = []
            return
        sections.append({
            "title": title,
            "start_row": current_start,
            "end_row": current[-1]["number"],
            "row_count": len(current),
            "rows": current,
        })
        current = []

    previous_blank = True
    for entry in raw_rows:
        if entry is None:
            # A single merged/standalone title row commonly precedes a table with
            # one visual spacer row. Keep it attached to the table so the UI
            # presents the logical section, not a title-only fragment.
            if current and not (len(current) == 1 and _looks_like_section_title([c.get("display", "") for c in current[0].get("cells", [])])):
                flush()
            previous_blank = True
            continue
        if previous_blank and not current:
            current_start = entry["number"]
        current.append(entry)
        previous_blank = False
    flush()

    non_empty_rows = sum(section["row_count"] for section in sections)
    return {
        "name": ws.title,
        "display_name": _clean_sheet_title(ws.title),
        "category": _sheet_category(ws.title),
        "row_count": non_empty_rows,
        "workbook_row_count": workbook_max_row,
        "column_count": used_max_col,
        "workbook_column_count": workbook_max_col,
        "columns": [get_column_letter(c) for c in range(1, used_max_col + 1)],
        "section_count": len(sections),
        "sections": sections,
        "preview": bool(preview_note or truncated),
        "truncated": truncated,
        "preview_note": preview_note or ("Showing a fast preview of this large workbook sheet." if truncated else ""),
    }


def _visible_worksheets(wb: Any) -> list[Any]:
    """Return all visible workbook sheets (every tab the user can see in Excel)."""
    return [
        ws for ws in wb.worksheets
        if str(getattr(ws, "sheet_state", "visible") or "visible") == "visible"
    ]


def _number(value: Any) -> float:
    try:
        if value is None or value == "":
            return 0.0
        return float(value)
    except Exception:
        return 0.0



def _is_asset_allocation_sheet(sheet_name: str) -> bool:
    low = str(sheet_name or "").lower()
    return "asset allocation" in low


def _safe_sheet_limits(ws: Any, max_rows: int | None, max_cols: int | None) -> tuple[int, int, bool, str]:
    """Choose UI-safe workbook bounds for result explorer table pages.

    The explorer is a browser UI, not Excel. Some workbook sheets have a large
    styled/used range or many helper columns. Rendering every cell can make the
    request appear stuck at 92%, so large pages are bounded and clearly marked as
    a UI preview while the downloadable workbook remains the full source.
    """
    workbook_max_row = int(ws.max_row or 0)
    workbook_max_col = int(ws.max_column or 0)
    explicit = max_rows is not None or max_cols is not None
    row_limit = min(workbook_max_row, max_rows) if max_rows else workbook_max_row
    col_limit = min(workbook_max_col, max_cols) if max_cols else workbook_max_col
    large_used_range = (row_limit * max(1, col_limit)) > DETAILED_RESULTS_MAX_FULL_CELLS
    visually_dense = _is_asset_allocation_sheet(getattr(ws, "title", "")) and workbook_max_col > DETAILED_RESULTS_SAFE_MAX_COLS
    bounded = False
    note = ""
    if not explicit and (large_used_range or visually_dense):
        row_limit = min(workbook_max_row, DETAILED_RESULTS_SAFE_MAX_ROWS)
        col_limit = min(workbook_max_col, DETAILED_RESULTS_SAFE_MAX_COLS)
        bounded = workbook_max_row > row_limit or workbook_max_col > col_limit
        note = (
            "This result page has a large workbook range, so the UI shows a fast, "
            "readable preview instead of trying to render every workbook cell. "
            "Download the workbook for the full Excel sheet."
        )
    return row_limit, col_limit, bounded, note


def _compact_series(years: list[Any], series: list[dict[str, Any]]) -> tuple[list[Any], list[dict[str, Any]], bool]:
    """Limit native UI chart payloads so dashboard rendering stays responsive."""
    truncated = False
    clean_series = [s for s in series if any(abs(_number(v)) > 1e-9 for v in (s.get("values") or []))]
    if len(clean_series) > CHART_MAX_SERIES:
        clean_series = clean_series[:CHART_MAX_SERIES]
        truncated = True
    if len(years) > CHART_MAX_POINTS:
        step = max(1, int(round(len(years) / CHART_MAX_POINTS)))
        idxs = list(range(0, len(years), step))
        if idxs[-1] != len(years) - 1:
            idxs.append(len(years) - 1)
        years = [years[i] for i in idxs]
        clean_series = [{**s, "values": [(s.get("values") or [])[i] if i < len(s.get("values") or []) else 0 for i in idxs]} for s in clean_series]
        truncated = True
    return years, clean_series, truncated


def _compact_pie(slices: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], bool]:
    total = sum(max(0.0, _number(s.get("value"))) for s in slices)
    min_material_value = max(1.0, total * 0.000001) if total > 0 else 1.0
    slices = [s for s in slices if _number(s.get("value")) > min_material_value]
    slices.sort(key=lambda s: _number(s.get("value")), reverse=True)
    if len(slices) <= CHART_MAX_SLICES:
        return slices, False
    keep = slices[: CHART_MAX_SLICES - 1]
    other = sum(_number(s.get("value")) for s in slices[CHART_MAX_SLICES - 1 :])
    if other > 0:
        keep.append({"label": "Other", "value": other})
    return keep, True

def _year_rows(ws: Any, year_col: int = 1, start_row: int = 5) -> list[tuple[int, Any]]:
    rows: list[tuple[int, Any]] = []
    for r in range(start_row, min(int(ws.max_row or 0), start_row + 400) + 1):
        value = ws.cell(r, year_col).value
        if value in (None, ""):
            # The hidden chart-data sheet has multiple tables. Stop at the
            # first blank after the main annual projection block.
            if rows:
                break
            continue
        try:
            year = int(value)
        except Exception:
            break
        rows.append((r, year))
    return rows


def _series_from_columns(ws: Any, *, year_col: int, value_cols: list[int], header_row: int = 4, start_row: int = 5) -> tuple[list[Any], list[dict[str, Any]]]:
    rows = _year_rows(ws, year_col=year_col, start_row=start_row)
    years = [yr for _, yr in rows]
    series: list[dict[str, Any]] = []
    for col in value_cols:
        label = str(ws.cell(header_row, col).value or f"Series {col}").strip()
        if not label:
            continue
        series.append({
            "label": label,
            "values": [_number(ws.cell(r, col).value) for r, _ in rows],
        })
    return years, series


def _pie_from_chart_data(ws: Any, *, data_col: int) -> list[dict[str, Any]]:
    start = 0
    for r in range(1, int(ws.max_row or 0) + 1):
        if str(ws.cell(r, 1).value or "").strip().lower() == "asset class":
            start = r + 1
            break
    if not start:
        return []
    data: list[dict[str, Any]] = []
    for r in range(start, int(ws.max_row or 0) + 1):
        name = str(ws.cell(r, 1).value or "").strip()
        if not name:
            break
        data.append({"label": name, "value": _number(ws.cell(r, data_col).value)})
    return data




def _safe_chart_title(chart: Any, fallback: str = "Chart") -> str:
    """Extract a readable title from an openpyxl chart object."""
    try:
        title = getattr(chart, "title", None)
        if title is None:
            return fallback
        tx = getattr(title, "tx", None)
        rich = getattr(tx, "rich", None)
        parts: list[str] = []
        for para in getattr(rich, "p", []) or []:
            for run in getattr(para, "r", []) or []:
                text = getattr(run, "t", "")
                if text:
                    parts.append(str(text))
        text = "".join(parts).strip()
        return text or fallback
    except Exception:
        return fallback


def _chart_series_label(wb: Any, series: Any, value_ref: str | None, fallback: str) -> str:
    """Extract a readable series label from an openpyxl series."""
    try:
        tx = getattr(series, "tx", None)
        v = getattr(tx, "v", None)
        if v not in (None, ""):
            return str(v).strip() or fallback
        str_ref = getattr(tx, "strRef", None)
        f = getattr(str_ref, "f", None)
        values = _values_from_ref(wb, f)
        if values:
            label = str(values[0]).strip()
            if label:
                return label
    except Exception:
        pass
    # Older workbook charts sometimes store the label in the cell immediately
    # above the value range rather than in a SeriesLabel object.
    if value_ref:
        try:
            sheet, bounds = range_to_tuple(value_ref)
            min_col, min_row, _max_col, _max_row = bounds
            if sheet in wb.sheetnames and min_row > 1:
                label = str(wb[sheet].cell(min_row - 1, min_col).value or "").strip()
                if label:
                    return label
        except Exception:
            pass
    return fallback


def _values_from_ref(wb: Any, ref: str | None) -> list[Any]:
    """Read values from an Excel range formula such as 'Sheet'!$A$1:$A$5."""
    if not ref:
        return []
    try:
        sheet, bounds = range_to_tuple(str(ref))
    except Exception:
        return []
    if sheet not in wb.sheetnames:
        return []
    min_col, min_row, max_col, max_row = bounds
    ws = wb[sheet]
    values: list[Any] = []
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            values.append(_json_value(ws.cell(r, c).value))
    return values


def _series_refs(series: Any) -> tuple[str | None, str | None]:
    """Return category and value range formulas for an openpyxl chart series."""
    cat_ref = None
    val_ref = None
    try:
        cat = getattr(series, "cat", None)
        cat_ref = getattr(getattr(cat, "numRef", None), "f", None) or getattr(getattr(cat, "strRef", None), "f", None)
    except Exception:
        cat_ref = None
    try:
        val = getattr(series, "val", None)
        val_ref = getattr(getattr(val, "numRef", None), "f", None)
    except Exception:
        val_ref = None
    return cat_ref, val_ref


def _embedded_charts_from_excel_objects(wb: Any, chart_ws: Any) -> tuple[list[dict[str, Any]], bool]:
    """Rebuild native UI chart specs directly from embedded Excel chart objects.

    This is the compatibility path for workbooks that were built before the
    hidden `_Chart Dashboard Data` sheet existed, or for any workbook where the
    hidden helper ranges are missing. It reads the chart formulas and follows
    them back to their source cells so the browser can still render charts.
    """
    charts: list[dict[str, Any]] = []
    compacted = False
    for idx, chart in enumerate(list(getattr(chart_ws, "_charts", []) or []), start=1):
        series_objs = list(getattr(chart, "series", []) or [])
        if not series_objs:
            continue
        title = _safe_chart_title(chart, f"Chart {idx}")
        chart_class = chart.__class__.__name__.lower()
        first_cat_ref, first_val_ref = _series_refs(series_objs[0])
        # Pie charts use one value series plus one category range.
        if "pie" in chart_class:
            labels = _values_from_ref(wb, first_cat_ref)
            vals = [_number(v) for v in _values_from_ref(wb, first_val_ref)]
            slices = [
                {"label": str(label or f"Slice {i + 1}"), "value": vals[i] if i < len(vals) else 0}
                for i, label in enumerate(labels)
            ]
            slices, was_compacted = _compact_pie(slices)
            compacted = compacted or was_compacted
            if slices:
                charts.append({"type": "pie", "title": title, "unit": "currency", "slices": slices})
            continue

        x = _values_from_ref(wb, first_cat_ref)
        rendered_series: list[dict[str, Any]] = []
        for sidx, ser in enumerate(series_objs, start=1):
            _cat_ref, val_ref = _series_refs(ser)
            values = [_number(v) for v in _values_from_ref(wb, val_ref)]
            if not values:
                continue
            label = _chart_series_label(wb, ser, val_ref, f"Series {sidx}")
            rendered_series.append({"label": label, "values": values})
        if not x and rendered_series:
            x = list(range(1, max(len(s.get("values") or []) for s in rendered_series) + 1))
        chart_type = "line" if "line" in chart_class else "stacked_bar"
        x, rendered_series, was_compacted = _compact_series(x, rendered_series)
        compacted = compacted or was_compacted
        if x and rendered_series:
            charts.append({"type": chart_type, "title": title, "unit": "currency", "x": x, "series": rendered_series})
    return charts, compacted


def _chart_dashboard_visible_table_charts(chart_ws: Any) -> tuple[list[dict[str, Any]], bool]:
    """Compatibility path for old visible chart helper tables."""
    charts: list[dict[str, Any]] = []
    compacted = False

    def add_xy(title: str, chart_type: str, years: list[Any], series: list[dict[str, Any]]) -> None:
        nonlocal compacted
        years2, series2, was_compacted = _compact_series(years, series)
        compacted = compacted or was_compacted
        if years2 and series2:
            charts.append({"type": chart_type, "title": title, "unit": "currency", "x": years2, "series": series2})

    years, nw_series = _series_from_columns(chart_ws, year_col=1, value_cols=list(range(2, 9)))
    add_xy("Net Worth by Component", "stacked_bar", years, nw_series)
    years, inc_series = _series_from_columns(chart_ws, year_col=11, value_cols=list(range(12, 27)))
    add_xy("Cash Flow — Income & Portfolio Draws", "stacked_bar", years, inc_series)
    years, exp_series = _series_from_columns(chart_ws, year_col=29, value_cols=list(range(30, 39)))
    add_xy("Cash Flow — Spending & Taxes", "stacked_bar", years, exp_series)
    years, mc_series = _series_from_columns(chart_ws, year_col=43, value_cols=list(range(44, 49)))
    add_xy("Net Worth Percentile Bands — Monte Carlo", "line", years, mc_series)
    before, was_compacted = _compact_pie(_pie_from_chart_data(chart_ws, data_col=2))
    compacted = compacted or was_compacted
    if before:
        charts.append({"type": "pie", "title": "Current Portfolio Allocation", "unit": "currency", "slices": before})
    after, was_compacted = _compact_pie(_pie_from_chart_data(chart_ws, data_col=3))
    compacted = compacted or was_compacted
    if after:
        charts.append({"type": "pie", "title": "Target Portfolio Allocation", "unit": "currency", "slices": after})
    return charts, compacted



def _norm_header(value: Any) -> str:
    return " ".join(str(value or "").replace("Σ", "sum").lower().split())


def _find_visible_sheet(wb: Any, *terms: str) -> Any | None:
    wanted = [str(t or "").lower() for t in terms if str(t or "").strip()]
    for ws in _visible_worksheets(wb):
        low = str(getattr(ws, "title", "") or "").lower()
        if all(term in low for term in wanted):
            return ws
    return None


def _header_map(ws: Any, header_row: int) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for col in range(1, int(ws.max_column or 0) + 1):
        text = str(ws.cell(header_row, col).value or "").strip()
        if text:
            mapping[_norm_header(text)] = col
    return mapping


def _first_matching_col(mapping: dict[str, int], label: str) -> int | None:
    target = _norm_header(label)
    if target in mapping:
        return mapping[target]
    for key, col in mapping.items():
        if target and (target in key or key in target):
            return col
    return None


def _years_from_sheet(ws: Any, year_col: int, start_row: int) -> list[tuple[int, Any]]:
    rows: list[tuple[int, Any]] = []
    for r in range(start_row, min(int(ws.max_row or 0), start_row + 500) + 1):
        value = ws.cell(r, year_col).value
        if value in (None, ""):
            if rows:
                break
            continue
        try:
            year = int(value)
        except Exception:
            if rows:
                break
            continue
        # Avoid accidentally reading ages or other identifiers as future years.
        if year < 1900 or year > 2300:
            if rows:
                break
            continue
        rows.append((r, year))
    return rows


def _series_from_named_headers(
    ws: Any,
    *,
    year_col: int,
    header_row: int,
    start_row: int,
    labels: list[str],
) -> tuple[list[Any], list[dict[str, Any]]]:
    rows = _years_from_sheet(ws, year_col=year_col, start_row=start_row)
    years = [yr for _, yr in rows]
    mapping = _header_map(ws, header_row)
    series: list[dict[str, Any]] = []
    used_cols: set[int] = set()
    for label in labels:
        col = _first_matching_col(mapping, label)
        if not col or col in used_cols:
            continue
        vals = [_number(ws.cell(r, col).value) for r, _ in rows]
        if any(abs(v) > 1e-9 for v in vals):
            series.append({"label": str(ws.cell(header_row, col).value or label), "values": vals})
            used_cols.add(col)
    return years, series


def _chart_dashboard_projection_fallback(wb: Any) -> tuple[list[dict[str, Any]], bool]:
    """Build native chart specs from ordinary result sheets when chart objects
    or hidden chart-source ranges are not available.

    This avoids a dead-end Chart Dashboard page for workbooks created before the
    hidden helper sheet existed or for workbook variants where Excel chart source
    formulas are unavailable to openpyxl. The charts are intentionally based on
    the same user-facing projection sheets exposed elsewhere in Results Explorer.
    """
    charts: list[dict[str, Any]] = []
    compacted = False

    def add_xy(title: str, chart_type: str, years: list[Any], series: list[dict[str, Any]]) -> None:
        nonlocal compacted
        years2, series2, was_compacted = _compact_series(years, series)
        compacted = compacted or was_compacted
        if years2 and series2:
            charts.append({"type": chart_type, "title": title, "unit": "currency", "x": years2, "series": series2})

    nw = _find_visible_sheet(wb, "net worth", "projection")
    if nw is not None:
        years, series = _series_from_named_headers(
            nw,
            year_col=1,
            header_row=2,
            start_row=3,
            labels=["Σ Ann", "Σ PreTax", "Σ Roth", "Σ Trust", "HSA", "Home Equity", "Σ Other", "TOTAL NW"],
        )
        # Keep the total out of the stacked bars when component series exist.
        if len(series) > 1:
            series = [s for s in series if "total" not in str(s.get("label", "")).lower()]
        add_xy("Net Worth by Component", "stacked_bar", years, series)

    cash = _find_visible_sheet(wb, "cash flow", "projection")
    if cash is not None:
        income_labels = [
            "Earned", "Matthew SS", "Patricia SS", "Pension", "W Single Ann", "W Joint Ann",
            "H Single Ann", "H Joint Ann", "Note P+I", "RMD Dist", "Σ Draws",
        ]
        years, series = _series_from_named_headers(cash, year_col=1, header_row=2, start_row=3, labels=income_labels)
        add_xy("Cash Flow — Income & Portfolio Draws", "stacked_bar", years, series)

        spending_labels = ["Spend Base", "Rec Extra", "Lump", "Mortgage + RE Tax", "Rent", "Fed Tax", "State Tax", "NIIT"]
        years, series = _series_from_named_headers(cash, year_col=1, header_row=2, start_row=3, labels=spending_labels)
        add_xy("Cash Flow — Spending & Taxes", "stacked_bar", years, series)

        balance_labels = ["NW Check"]
        years, series = _series_from_named_headers(cash, year_col=1, header_row=2, start_row=3, labels=balance_labels)
        add_xy("Net Worth Trend", "line", years, series)

    return charts, compacted

def _chart_dashboard_sheet(wb: Any, chart_ws: Any) -> dict[str, Any]:
    """Return a chart-only representation of the workbook chart dashboard.

    Excel charts are not rendered by openpyxl, so the app builds native UI chart
    cards from the same hidden source ranges the workbook charts use. This keeps
    the explorer focused on the visuals and avoids showing chart helper tables.
    """
    charts: list[dict[str, Any]] = []
    compacted = False

    def add_xy_chart(title: str, chart_type: str, years: list[Any], series: list[dict[str, Any]]) -> None:
        nonlocal compacted
        years2, series2, was_compacted = _compact_series(years, series)
        compacted = compacted or was_compacted
        if years2 and series2:
            charts.append({
                "type": chart_type,
                "title": title,
                "unit": "currency",
                "x": years2,
                "series": series2,
            })

    source = "hidden chart-source data"
    if CHART_DASHBOARD_DATA_SHEET in wb.sheetnames:
        ws = wb[CHART_DASHBOARD_DATA_SHEET]
        years, nw_series = _series_from_columns(ws, year_col=1, value_cols=list(range(2, 9)))
        add_xy_chart("Net Worth by Component", "stacked_bar", years, nw_series)
        years, inc_series = _series_from_columns(ws, year_col=11, value_cols=list(range(12, 27)))
        add_xy_chart("Cash Flow — Income & Portfolio Draws", "stacked_bar", years, inc_series)
        years, exp_series = _series_from_columns(ws, year_col=29, value_cols=list(range(30, 39)))
        add_xy_chart("Cash Flow — Spending & Taxes", "stacked_bar", years, exp_series)
        years, mc_series = _series_from_columns(ws, year_col=43, value_cols=list(range(44, 49)))
        add_xy_chart("Net Worth Percentile Bands — Monte Carlo", "line", years, mc_series)
        before, was_compacted = _compact_pie(_pie_from_chart_data(ws, data_col=2))
        compacted = compacted or was_compacted
        if before:
            charts.append({"type": "pie", "title": "Current Portfolio Allocation", "unit": "currency", "slices": before})
        after, was_compacted = _compact_pie(_pie_from_chart_data(ws, data_col=3))
        compacted = compacted or was_compacted
        if after:
            charts.append({"type": "pie", "title": "Target Portfolio Allocation", "unit": "currency", "slices": after})

    if not charts:
        embedded, was_compacted = _embedded_charts_from_excel_objects(wb, chart_ws)
        if embedded:
            charts = embedded
            compacted = compacted or was_compacted
            source = "embedded Excel chart source ranges"

    if not charts:
        visible, was_compacted = _chart_dashboard_visible_table_charts(chart_ws)
        if visible:
            charts = visible
            compacted = compacted or was_compacted
            source = "legacy visible chart-source tables"

    if not charts:
        derived, was_compacted = _chart_dashboard_projection_fallback(wb)
        if derived:
            charts = derived
            compacted = compacted or was_compacted
            source = "projection result sheets"

    note = f"Chart Dashboard is shown as native UI charts rebuilt from {source}. Chart source tables are not displayed in the explorer."
    if compacted:
        note += " Some long chart series are sampled/limited for browser responsiveness; download the workbook for the full Excel chart detail."
    if not charts:
        note = "No browser-native Chart Dashboard charts were found in this workbook. Open the detailed projection result pages or download the workbook for any embedded Excel-only visuals."
    return {
        "name": chart_ws.title,
        "category": _sheet_category(chart_ws.title),
        "kind": "chart_dashboard",
        "row_count": len(charts),
        "column_count": 0,
        "section_count": len(charts),
        "chart_count": len(charts),
        "charts": charts,
        "chart_note": note,
        "preview": False,
        "truncated": False,
    }


def _workbook_not_found(path: Path) -> dict[str, Any]:
    return {"success": False, "error": "No workbook found; build the workbook first.", "sheets": [], "categories": []}


def _excel_workbook_detailed_index(workbook_path: str | Path) -> dict[str, Any]:
    """Return a lightweight workbook index for fast navigation.

    This intentionally does not scan every populated cell. The UI loads this
    first so the Detailed Results screen can render quickly, then fetches the
    selected sheet on demand.
    """
    path = Path(workbook_path)
    if not path.exists():
        return _workbook_not_found(path)
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        sheets: list[dict[str, Any]] = []
        for ws in _visible_worksheets(wb):
            is_chart_dashboard = _is_chart_dashboard_sheet(ws.title)
            sheets.append({
                "name": ws.title,
                "category": _sheet_category(ws.title),
                "row_count": 6 if is_chart_dashboard else int(ws.max_row or 0),
                "column_count": 0 if is_chart_dashboard else int(ws.max_column or 0),
                "section_count": 6 if is_chart_dashboard else 0,
                "chart_count": 6 if is_chart_dashboard else 0,
                "kind": "chart_dashboard" if is_chart_dashboard else "table",
                "loaded": False,
                "preview": False,
                "preview_note": "Chart-only view; source ranges are hidden from the explorer." if is_chart_dashboard else "",
            })
        return {
            "success": True,
            "mode": "index",
            "workbook": path.name,
            "sheet_count": len(sheets),
            "sheets": sheets,
            "categories": _categories_from_sheets(sheets),
        }
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _excel_workbook_detailed_sheet(workbook_path: str | Path, sheet_name: str) -> dict[str, Any]:
    """Return full row/section data for one workbook sheet."""
    path = Path(workbook_path)
    if not path.exists():
        return _workbook_not_found(path)
    target = str(sheet_name or "")
    chart_target = _is_chart_dashboard_sheet(target)
    wb = load_workbook(path, data_only=True, read_only=not chart_target)
    try:
        if not target:
            visible = _visible_worksheets(wb)
            ws = visible[0] if visible else None
        else:
            ws = wb[target] if target in wb.sheetnames and str(getattr(wb[target], "sheet_state", "visible") or "visible") == "visible" else None
        if ws is None:
            return {"success": False, "error": f"Workbook sheet not found: {target}", "sheet": None}
        if _is_chart_dashboard_sheet(ws.title):
            sheet = _chart_dashboard_sheet(wb, ws)
        else:
            sheet = _sheet_to_sections(ws)
        return {"success": True, "mode": "sheet", "workbook": path.name, "sheet": sheet}
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _excel_workbook_detailed_results(workbook_path: str | Path) -> dict[str, Any]:
    """Return workbook data as UI-friendly sheets and collapsible sections.

    The workbook remains the source of truth. This helper intentionally exposes
    every non-blank row and cell, but groups rows by natural blank-row-separated
    workbook sections so the UI can use progressive disclosure instead of a flat
    sheet dump.
    """
    path = Path(workbook_path)
    if not path.exists():
        return _workbook_not_found(path)

    # Use normal workbook mode here so embedded chart objects are available
    # when callers ask for the full detailed-results payload. The index/single
    # sheet endpoints remain the preferred faster UI path.
    wb = load_workbook(path, data_only=True, read_only=False)
    try:
        sheets = [
            _chart_dashboard_sheet(wb, ws) if _is_chart_dashboard_sheet(ws.title) else _sheet_to_sections(ws)
            for ws in _visible_worksheets(wb)
        ]
        return {
            "success": True,
            "workbook": path.name,
            "sheet_count": len(sheets),
            "sheets": sheets,
            "categories": _categories_from_sheets(sheets),
        }
    finally:
        try:
            wb.close()
        except Exception:
            pass



def _result_model_for_workbook(workbook_path: str | Path) -> dict[str, Any] | None:
    """Load the v10 semantic result model sidecar for a workbook, if present."""
    path = Path(workbook_path)
    return read_result_explorer_model(path.with_name(RESULTS_MODEL_FILENAME))


def _merge_model_index_with_excel(model_payload: dict[str, Any], workbook_path: str | Path) -> dict[str, Any]:
    """Use the semantic model first, then append Excel-only legacy pages.

    This keeps the 8.4 architecture fast and UI-native for modeled pages while
    retaining full backward-compatible access to workbook-only sheets until all
    pages have been migrated into the semantic model.
    """
    out = model_index(model_payload)
    try:
        excel = _excel_workbook_detailed_index(workbook_path)
    except Exception:
        excel = None
    if not (excel and excel.get("success")):
        return out
    existing = {str(s.get("name")) for s in out.get("sheets") or []}
    added = []
    for s in excel.get("sheets") or []:
        if str(s.get("name")) not in existing:
            x = dict(s)
            x["source"] = "excel_parser_fallback"
            added.append(x)
    if added:
        out["sheets"] = (out.get("sheets") or []) + added
        out["sheet_count"] = len(out["sheets"])
        out["categories"] = _categories_from_sheets(out["sheets"])
        out["fallback_sheet_count"] = len(added)
    return out


def workbook_detailed_index(workbook_path: str | Path) -> dict[str, Any]:
    """Return Results Explorer navigation, preferring the v10 semantic model."""
    model_payload = _result_model_for_workbook(workbook_path)
    if model_payload:
        return _merge_model_index_with_excel(model_payload, workbook_path)
    return _excel_workbook_detailed_index(workbook_path)


def workbook_detailed_sheet(workbook_path: str | Path, sheet_name: str) -> dict[str, Any]:
    """Return a result page, preferring semantic-model data for new builds."""
    model_payload = _result_model_for_workbook(workbook_path)
    if model_payload:
        modeled = model_sheet(model_payload, sheet_name)
        if modeled:
            return modeled
    return _excel_workbook_detailed_sheet(workbook_path, sheet_name)


def workbook_detailed_results(workbook_path: str | Path) -> dict[str, Any]:
    """Return all result pages, preferring the semantic model and falling back to Excel."""
    model_payload = _result_model_for_workbook(workbook_path)
    if model_payload:
        return _merge_model_index_with_excel(model_payload, workbook_path)
    return _excel_workbook_detailed_results(workbook_path)
