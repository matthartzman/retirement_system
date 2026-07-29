"""User-managed workbook column-width formatting.

The Settings → Workbook Formatting UI lets a user override individual column
widths in the generated Excel workbook. This module is the single source of
truth for that feature, shared by:

  * the HTTP route layer (build the sheet -> table -> column tree for the UI,
    read/write the saved overrides), and
  * the workbook build (apply saved overrides after the reference-template
    layout pass so user edits always win).

Design notes
------------
Excel column width is a per-column-letter, per-sheet property. Two stacked
tables on one sheet therefore share a column's width, so overrides are keyed by
(sheet title, column letter). "Tables" are a purely organizational grouping for
the UI: a sheet's wide matrix layouts (Net Worth, Cash Flow, ...) place a merged
section banner across the top of each column group, and we surface each such
group as a table. Sheets whose header is a single full-width banner (most report
sheets) have exactly one table, and the UI collapses that layer away.

Column titles are read from the sheet's header band (the sub-header cell for a
column), falling back to the column letter when a column has no header text.
"""
from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any, Optional

from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter, column_index_from_string

from ..workspace_context import workspace_input_dir

# #209/#210/#212/#228: sheet letters (1A, 2J, 3F, ...) are computed fresh every
# build from which sheets survive module gating (see workbook_common.
# SHEET_LETTER_ORDER / compute_final_sheet_renames), so the same feature can
# land on a different letter build to build. Overrides/alignments/template
# layout are keyed by each sheet's STABLE name instead, and the build passes
# its live {stable: final_title} map so those lookups always hit.
#
# This UI route, though, only has an already-built .xlsx on disk -- there's no
# live build-time map to consult. Recover the stable identity from the DISPLAY
# TITLE portion of the sheet's current final name (the part after "1A. "),
# which is static and shared with workbook_common.SHEET_DISPLAY_TITLES.
# Duplicated here (not imported) to keep this lightweight settings route from
# pulling in workbook_common's heavy engine-module imports.
_DISPLAY_TITLE_TO_STABLE = {
    'Executive Summary': '1. Executive Summary',
    'Net Worth': '5. Net Worth Projection',
    'Cash Flow': '6. Cash Flow Projection',
    'Balance Sheet': '3. Balance Sheet',
    'Charts': '8. Charts Dashboard',
    'Lifetime Taxes': '7. Lifetime Tax',
    # #221: Core Spending removed -- merged into Spending Summary.
    'Spending Summary': '29. Spending Summary',
    'Roth Conversion': '11. Roth Conversion',
    'Asset Allocation': '4. Asset Allocation',
    'State Residency': '13. State Residency',
    'Social Security': '10. Social Security',
    'S-Corp vs LLC': 'S-Corp vs LLC',
    'Charitable Giving': '12. Charitable Giving',
    'Estate & Legacy Planning': '14. Estate Plan',
    'Planning Levers': '27. Planning Levers',
    'Tax-Loss Harvesting': '12B. Tax-Loss Harvesting',
    'Education Funding': '30. Education Funding',
    'Equity Compensation': '35. Equity Compensation',
    'Special-Needs Planning': '36. Special-Needs Planning',
    'Business Succession': '34. Business Succession',
    'Gain Harvesting': '12C. Gain Harvesting',
    'Monte Carlo': '15. Market-Luck Stress Test',
    'Survivor': '18. Survivor Stress Test',
    'LTC + Life Insurance': '19. Life Insurance',
    'Existing Life Insurance': '31. Existing Life Insurance',
    'Disability Income': '32. Disability Income',
    'P&C Umbrella': '33. P&C Umbrella',
    'Plan Data': 'Plan Data',
    'Assumptions': '2. Assumptions',
    'Account Reconciliation': '25. Account Reconciliation',
    'Quality Control': '21. Quality Control',
    'RMD Audit': '20. RMD Audit',
    'Methodology': '23. Methodology',
    'Glossary': '22. Glossary',
}

_FINAL_TITLE_PREFIX_RE = re.compile(r'^[1-4][A-Z]\.\s*')


def stable_name_for_sheet_title(title: str) -> str:
    """Resolve a workbook sheet's current final title back to its stable
    (build-time) name, for keying overrides/alignments/template-layout data
    that must survive letters shifting build to build. Falls back to the
    title itself for anything not in the map (section dividers, unrecognized
    titles) -- those are already stable/unrenamed."""
    display = _FINAL_TITLE_PREFIX_RE.sub('', title or '', count=1)
    return _DISPLAY_TITLE_TO_STABLE.get(display, title)

OVERRIDES_FILENAME = "workbook_format_overrides.json"
# Horizontal-alignment overrides are stored separately from width overrides so
# the existing width file's flat {sheet: {col: width}} shape never needs a
# migration.
ALIGNMENTS_FILENAME = "workbook_format_alignments.json"
ALIGN_VALUES = ("left", "center", "right")

# Excel's built-in default column width (characters) when a column has no
# explicit width set. Used so the UI shows a concrete editable number.
DEFAULT_COL_WIDTH = 8.43

# Sensible clamp so a stray value can't produce an unusable sheet.
MIN_WIDTH = 1.0
MAX_WIDTH = 255.0

# Header band search depth. Report sheets use a title banner in row 1 and
# column headers in row 2; wide matrix sheets add a sub-header row too.
_HEADER_BAND_ROWS = 4


def overrides_path(input_dir: Optional[Path] = None) -> Path:
    base = Path(input_dir) if input_dir is not None else workspace_input_dir()
    return base / OVERRIDES_FILENAME


def load_overrides(input_dir: Optional[Path] = None) -> dict[str, dict[str, float]]:
    """Return {sheet_title: {column_letter: width}}; empty when none saved."""
    path = overrides_path(input_dir)
    try:
        with open(path, "r", encoding="utf-8") as f:
            raw = json.load(f)
    except (FileNotFoundError, ValueError, OSError):
        return {}
    return _sanitize_overrides(raw)


def save_overrides(data: dict, input_dir: Optional[Path] = None) -> dict[str, dict[str, float]]:
    """Validate, clamp, and persist overrides, REPLACING the whole file.

    Only safe when `data` is known to be the complete, current set of every
    sheet's saved overrides (e.g. a full-file import/restore). The live
    Settings -> Workbook Formatting UI must never call this directly -- use
    `merge_overrides` instead, or an edit to one sheet silently deletes every
    other sheet's previously-saved widths. See merge_overrides docstring.
    """
    clean = _sanitize_overrides(data)
    path = overrides_path(input_dir)
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(clean, f, indent=1, sort_keys=True)
    return clean


def merge_overrides(patch: dict, input_dir: Optional[Path] = None) -> dict[str, dict[str, float]]:
    """Merge a partial {sheet: {column_letter: width}} patch into the
    persisted overrides file, returning the complete resulting mapping.

    This is the only writer the Settings -> Workbook Formatting UI (and its
    `/api/workbook-format` POST route) should ever call. Column-width edits
    happen one sheet/column at a time; a naive "replace the whole file with
    whatever the client just sent" (the old `save_overrides` behavior) would
    silently wipe out every other sheet's saved widths the moment the
    client's in-memory copy is even slightly stale (a race, a second tab, a
    reload mid-edit) -- which is exactly the "changes keep reverting" bug this
    function exists to prevent.

    A sheet/column absent from `patch` is left exactly as persisted. A
    column present in `patch` with a non-positive or non-numeric width
    deletes just that column's override (this is how the UI's per-column
    "Reset" control signals "go back to automatic width").
    """
    current = load_overrides(input_dir)
    if isinstance(patch, dict):
        for sheet, cols in patch.items():
            if not isinstance(sheet, str) or not sheet.strip() or not isinstance(cols, dict):
                continue
            for letter, width in cols.items():
                norm = _normalize_letter(letter)
                if norm is None:
                    continue
                try:
                    w = float(width)
                except (TypeError, ValueError):
                    w = None
                if w is None or w <= 0:
                    current.get(sheet, {}).pop(norm, None)
                    if sheet in current and not current[sheet]:
                        current.pop(sheet, None)
                else:
                    current.setdefault(sheet, {})[norm] = round(max(MIN_WIDTH, min(MAX_WIDTH, w)), 2)
    path = overrides_path(input_dir)
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(current, f, indent=1, sort_keys=True)
    return current


def _sanitize_overrides(raw: Any) -> dict[str, dict[str, float]]:
    out: dict[str, dict[str, float]] = {}
    if not isinstance(raw, dict):
        return out
    for sheet, cols in raw.items():
        if not isinstance(sheet, str) or not sheet.strip() or not isinstance(cols, dict):
            continue
        sheet_out: dict[str, float] = {}
        for letter, width in cols.items():
            norm = _normalize_letter(letter)
            if norm is None:
                continue
            try:
                w = float(width)
            except (TypeError, ValueError):
                continue
            if w <= 0:
                continue
            sheet_out[norm] = round(max(MIN_WIDTH, min(MAX_WIDTH, w)), 2)
        if sheet_out:
            out[sheet] = sheet_out
    return out


def _normalize_letter(letter: Any) -> Optional[str]:
    if not isinstance(letter, str):
        return None
    letter = letter.strip().upper()
    if not letter.isalpha():
        return None
    try:
        column_index_from_string(letter)
    except (ValueError, KeyError):
        return None
    return letter


def alignments_path(input_dir: Optional[Path] = None) -> Path:
    base = Path(input_dir) if input_dir is not None else workspace_input_dir()
    return base / ALIGNMENTS_FILENAME


def load_alignments(input_dir: Optional[Path] = None) -> dict[str, dict[str, str]]:
    """Return {sheet_title: {column_letter: 'left'|'center'|'right'}}; empty when none saved."""
    path = alignments_path(input_dir)
    try:
        with open(path, "r", encoding="utf-8") as f:
            raw = json.load(f)
    except (FileNotFoundError, ValueError, OSError):
        return {}
    return _sanitize_alignments(raw)


def save_alignments(data: dict, input_dir: Optional[Path] = None) -> dict[str, dict[str, str]]:
    """Validate and persist horizontal-alignment overrides, REPLACING the
    whole file. See save_overrides's docstring -- the same "must be the
    complete current state or it silently deletes other sheets" caveat
    applies here. The live UI must use `merge_alignments` instead.
    """
    clean = _sanitize_alignments(data)
    path = alignments_path(input_dir)
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(clean, f, indent=1, sort_keys=True)
    return clean


def merge_alignments(patch: dict, input_dir: Optional[Path] = None) -> dict[str, dict[str, str]]:
    """Merge a partial {sheet: {column_letter: 'left'|'center'|'right'}} patch
    into the persisted alignments file, returning the complete resulting
    mapping. See merge_overrides's docstring for why this (not
    save_alignments) is the writer the UI must use.

    A sheet/column absent from `patch` is left exactly as persisted. A
    column present in `patch` with an unrecognized/empty value deletes just
    that column's alignment override (reset to automatic).
    """
    current = load_alignments(input_dir)
    if isinstance(patch, dict):
        for sheet, cols in patch.items():
            if not isinstance(sheet, str) or not sheet.strip() or not isinstance(cols, dict):
                continue
            for letter, align in cols.items():
                norm_letter = _normalize_letter(letter)
                if norm_letter is None:
                    continue
                norm_align = _normalize_align(align)
                if norm_align is None:
                    current.get(sheet, {}).pop(norm_letter, None)
                    if sheet in current and not current[sheet]:
                        current.pop(sheet, None)
                else:
                    current.setdefault(sheet, {})[norm_letter] = norm_align
    path = alignments_path(input_dir)
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(current, f, indent=1, sort_keys=True)
    return current


def _normalize_align(value: Any) -> Optional[str]:
    if not isinstance(value, str):
        return None
    v = value.strip().lower()
    if v in ("l", "left"):
        return "left"
    if v in ("c", "center", "centre"):
        return "center"
    if v in ("r", "right"):
        return "right"
    return None


def _sanitize_alignments(raw: Any) -> dict[str, dict[str, str]]:
    out: dict[str, dict[str, str]] = {}
    if not isinstance(raw, dict):
        return out
    for sheet, cols in raw.items():
        if not isinstance(sheet, str) or not sheet.strip() or not isinstance(cols, dict):
            continue
        sheet_out: dict[str, str] = {}
        for letter, align in cols.items():
            norm_letter = _normalize_letter(letter)
            norm_align = _normalize_align(align)
            if norm_letter is None or norm_align is None:
                continue
            sheet_out[norm_letter] = norm_align
        if sheet_out:
            out[sheet] = sheet_out
    return out


# ─────────────────────────────────────────────────────────────────────────────
# Structure detection: sheet -> table(s) -> columns
# ─────────────────────────────────────────────────────────────────────────────
def _merged_anchor_cell(ws, row: int, col: int):
    """#243: resolve (row, col) to the top-left anchor cell of its merged range.

    openpyxl only stores a value/style on a merged range's top-left cell --
    every other cell in the range reports value=None and is a read-only
    MergedCell. Callers that check `cell.value` or set `.alignment` on
    whatever column a user's override happens to name must resolve through
    this first, or the check/write silently no-ops for every column of a
    merged row except the one that happens to be the anchor.
    """
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            return ws.cell(mr.min_row, mr.min_col)
    return ws.cell(row, col)


def _effective_width(ws, letter: str) -> float:
    dim = ws.column_dimensions.get(letter)
    if dim is not None and dim.width is not None:
        return round(float(dim.width), 2)
    default = getattr(ws.sheet_format, "defaultColWidth", None)
    return round(float(default), 2) if default else DEFAULT_COL_WIDTH


def _effective_align(ws, letter: str, band_end: int) -> str:
    """First data row's horizontal alignment for this column, else 'left'."""
    col = column_index_from_string(letter)
    max_row = ws.max_row or 1
    for r in range(band_end + 1, max_row + 1):
        cell = _merged_anchor_cell(ws, r, col)
        if cell.value in (None, ""):
            continue
        h = cell.alignment.horizontal if cell.alignment else None
        return h if h in ALIGN_VALUES else "left"
    return "left"


def _header_band_end(ws) -> int:
    """Row index (1-based) of the last header row before the data body.

    Report sheets put a merged title/section banner in row 1 and the real
    column headers in the row beneath it, so a row is treated as part of the
    header band when it is (a) styled like a header, (b) itself a merged banner
    row, or (c) the row directly under a banner. The band stops at the first
    ordinary data row.
    """
    max_row = ws.max_row or 1
    limit = min(max_row, _HEADER_BAND_ROWS)
    banner_rows = {
        mr.min_row
        for mr in ws.merged_cells.ranges
        if mr.min_row <= limit and (mr.max_col - mr.min_col) >= 1
    }
    band = 1
    for r in range(1, limit + 1):
        row_cells = [ws.cell(r, c) for c in range(1, (ws.max_column or 1) + 1)]
        nonblank = [c for c in row_cells if c.value not in (None, "")]
        styled = sum(
            1
            for c in nonblank
            if (c.font and c.font.bold) or (c.fill and c.fill.fill_type)
        )
        header_like = bool(nonblank) and styled >= max(1, len(nonblank) * 0.5)
        if header_like or r in banner_rows or (r - 1) in banner_rows:
            band = max(band, r)
        elif r > band:
            break
    return band


def _column_title(ws, letter: str, band_end: int, group_banner_rows: set[int]) -> Optional[str]:
    """Deepest non-empty header cell for a column, skipping group-banner rows."""
    col = column_index_from_string(letter)
    for r in range(band_end, 0, -1):
        if r in group_banner_rows:
            continue
        val = ws.cell(r, col).value
        if isinstance(val, str) and val.strip():
            return " ".join(val.split())
        if isinstance(val, (int, float)):
            return str(val)
    return None


def _row1_groups(ws) -> list[tuple[int, int, str]]:
    """Multi-column merged banners in row 1: (min_col, max_col, text).

    Only wide matrix sheets, which stack two or more side-by-side merged
    section banners across the top (Net Worth, Cash Flow, ...), are treated as
    multi-table. A lone banner is just the sheet title, so a single group is
    reported as no groups (the sheet is one table).
    """
    groups = []
    for mr in ws.merged_cells.ranges:
        if mr.min_row == 1 and (mr.max_col - mr.min_col) >= 1:
            text = ws.cell(1, mr.min_col).value
            groups.append((mr.min_col, mr.max_col, " ".join(str(text or "").split())))
    groups.sort()
    if len(groups) < 2:
        return []
    return groups


def _columns_with_content(ws, band_end: int) -> list[int]:
    """Columns that have either an explicit width or any header-band text."""
    cols = []
    for c in range(1, (ws.max_column or 1) + 1):
        letter = get_column_letter(c)
        dim = ws.column_dimensions.get(letter)
        has_width = dim is not None and dim.width is not None
        has_header = any(
            ws.cell(r, c).value not in (None, "") for r in range(1, band_end + 1)
        )
        if has_width or has_header:
            cols.append(c)
    return cols


def build_sheet_tree(ws, sheet_overrides: dict[str, float], sheet_aligns: Optional[dict[str, str]] = None) -> Optional[dict]:
    """Build one sheet's {sheet, single_table, tables:[...]} node, or None."""
    sheet_aligns = sheet_aligns or {}
    band_end = _header_band_end(ws)
    groups = _row1_groups(ws)
    content_cols = _columns_with_content(ws, band_end)
    if not content_cols:
        return None
    group_banner_rows = {1} if groups else set()

    def _col_node(col_idx: int) -> dict:
        letter = get_column_letter(col_idx)
        title = _column_title(ws, letter, band_end, group_banner_rows)
        return {
            "col": letter,
            "title": title or letter,
            "width": _effective_width(ws, letter),
            "overridden": letter in sheet_overrides,
            "align": sheet_aligns.get(letter) or _effective_align(ws, letter, band_end),
            "align_overridden": letter in sheet_aligns,
        }

    tables: list[dict] = []
    if groups:
        # Walk content columns left-to-right; each column belongs to the row-1
        # group (merged banner) covering it, or forms its own single-column
        # table when no banner covers it (e.g. a lone TOTAL column).
        assigned: set[int] = set()
        for (c0, c1, name) in groups:
            cols = [c for c in content_cols if c0 <= c <= c1]
            if not cols:
                continue
            assigned.update(cols)
            tables.append({"name": name or None, "columns": [_col_node(c) for c in cols]})
        for c in content_cols:
            if c in assigned:
                continue
            header = ws.cell(1, c).value
            name = " ".join(str(header).split()) if isinstance(header, str) and header.strip() else None
            tables.append({"name": name, "columns": [_col_node(c)], "_orphan_col": c})
        # Preserve left-to-right sheet order across grouped + orphan tables.
        tables.sort(key=lambda t: column_index_from_string(t["columns"][0]["col"]))
        for t in tables:
            t.pop("_orphan_col", None)

    single_table = len(tables) <= 1
    if single_table:
        tables = [{"name": None, "columns": [_col_node(c) for c in content_cols]}]

    # #209/#210/#212/#228: "sheet" is the STABLE key -- this is what gets sent
    # back on save, so it must match the key overrides/alignments are stored
    # under. "display" is the sheet's current final title, for the UI to show
    # the user; it can differ letter-to-letter across builds.
    return {
        "sheet": stable_name_for_sheet_title(ws.title),
        "display": ws.title,
        "single_table": single_table,
        "tables": tables,
    }


def build_format_tree(workbook_path: str | Path, overrides: Optional[dict] = None, alignments: Optional[dict] = None) -> dict:
    """Introspect a built workbook into the UI tree.

    Returns {available: bool, sheets: [sheet_node, ...]}. `available` is False
    when the workbook file does not exist yet (no build has run).
    """
    import openpyxl

    overrides = _sanitize_overrides(overrides or {})
    alignments = _sanitize_alignments(alignments or {})
    path = Path(workbook_path)
    if not path.exists():
        return {"available": False, "sheets": []}

    wb = openpyxl.load_workbook(path, read_only=False)
    sheets = []
    for ws in wb.worksheets:
        if getattr(ws, "sheet_state", "visible") != "visible":
            continue
        stable = stable_name_for_sheet_title(ws.title)
        node = build_sheet_tree(ws, overrides.get(stable, {}), alignments.get(stable, {}))
        if node is not None:
            sheets.append(node)
    return {"available": True, "sheets": sheets}


def apply_overrides(wb, input_dir: Optional[Path] = None, sheet_renames: Optional[dict] = None) -> None:
    """Apply saved column-width overrides to a workbook in place.

    #209/#210/#212/#228: overrides are keyed by each sheet's STABLE (build-
    time) name so a saved override keeps applying to the same sheet even when
    module gating shifts its letter. `sheet_renames` is this build's live
    {stable_name: final_title} map (workbook_common.FINAL_SHEET_RENAMES);
    omitting it falls back to treating the JSON key as the sheet's current
    title, for any caller still on the old final-name keying.
    """
    overrides = load_overrides(input_dir)
    if not overrides:
        return
    sheet_renames = sheet_renames or {}
    by_title = {ws.title: ws for ws in wb.worksheets}
    for sheet, cols in overrides.items():
        ws = by_title.get(sheet_renames.get(sheet, sheet))
        if ws is None:
            continue
        for letter, width in cols.items():
            ws.column_dimensions[letter].width = width


def apply_alignments(wb, input_dir: Optional[Path] = None, sheet_renames: Optional[dict] = None) -> None:
    """Apply saved horizontal-alignment overrides to a workbook's data rows in place.

    Header-band rows are left untouched -- their alignment is deliberate
    (e.g. centered, bold titles), so only rows below the header are re-aligned.

    #243: resolves each (row, column) through _merged_anchor_cell first, since
    a merged row's content and style live only on its top-left anchor cell --
    every other cell in the range is empty, so applying alignment (or even
    detecting the row has content) against a non-anchor column previously
    silently no-op'd for merged rows.

    Keyed by stable sheet name; see apply_overrides for what `sheet_renames` is.
    """
    alignments = load_alignments(input_dir)
    if not alignments:
        return
    sheet_renames = sheet_renames or {}
    by_title = {ws.title: ws for ws in wb.worksheets}
    for sheet, cols in alignments.items():
        ws = by_title.get(sheet_renames.get(sheet, sheet))
        if ws is None:
            continue
        band_end = _header_band_end(ws)
        max_row = ws.max_row or 1
        for letter, align in cols.items():
            col_idx = column_index_from_string(letter)
            for r in range(band_end + 1, max_row + 1):
                cell = _merged_anchor_cell(ws, r, col_idx)
                if cell.value in (None, ""):
                    continue
                old = cell.alignment or Alignment()
                cell.alignment = Alignment(
                    horizontal=align,
                    vertical=old.vertical,
                    text_rotation=old.text_rotation,
                    wrap_text=old.wrap_text,
                    shrink_to_fit=old.shrink_to_fit,
                    indent=old.indent,
                )
