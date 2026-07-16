
"""
Retirement, Tax & Estate Plan Builder — v5 Institutional
Workbook/report orchestration layer. Financial engines live in extracted modules.
"""

import csv, math, random, datetime, sys, traceback
from copy import copy
from collections import defaultdict
import os

from .. import taxes as _td  # consolidated from tax_data
from .. import core as _ar  # consolidated from account_registry
from .. import core as _aa  # consolidated from account_access
from .. import optimization as _ao  # consolidated from allocation_optimizer
from ..core import *  # shared projection/tax/annuity primitives  # consolidated from engine_core
from ..market_data import PRICE_CACHE, fetch_price, price_source, pricing_diagnostics, pricing_source_summary, write_pricing_diagnostics  # consolidated from market_data_providers

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              numbers as xl_numbers)
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.shapes import GraphicalProperties

# ─────────────────────────────────────────────────────────────────────────────
# 1.  DATA INPUT
# ─────────────────────────────────────────────────────────────────────────────
from ..data_io import load_csv, parse_client, build_plan_from_json  # consolidated from data_parser
from ..config_backend import load_active_config  # Version 7 active CSV/JSON/YAML/SQLite config loader
from ..workspace_context import workspace_output_dir, workspace_input_dir, sanitize_id
from ..report_compute import prepare_config_from_sectioned_data, run_projection_artifacts

# ─────────────────────────────────────────────────────────────────────────────
# 2.  LIVE PRICE FETCHING
# ─────────────────────────────────────────────────────────────────────────────
# Implemented in market_data_providers.py and imported above.

# ─────────────────────────────────────────────────────────────────────────────
# 3.  STYLE HELPERS
# ─────────────────────────────────────────────────────────────────────────────

NAVY   = '1F3864'
BLUE   = '2E75B6'
ORANGE = 'C55A11'
GREEN  = '375623'
GRAY   = 'F2F2F2'
LGRAY  = 'D9D9D9'
DGRAY  = '595959'
GOLD   = 'FFC000'
RED    = 'C00000'
WHITE  = 'FFFFFF'
YELLOW_INPUT = 'FFFF00'
BLUE_TEXT    = '0000FF'

THIN = Side(style='thin')
MED  = Side(style='medium')

def thin_border():
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def med_border():
    return Border(left=MED, right=MED, top=MED, bottom=MED)

def fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def hdr_font(bold=True, color=WHITE, size=11):
    return Font(name='Arial', bold=bold, color=color, size=size)

def body_font(bold=False, color='000000', size=10):
    return Font(name='Arial', bold=bold, color=color, size=size)

def input_style(ws, cell):
    cell.fill = PatternFill('solid', fgColor=YELLOW_INPUT)
    cell.font = Font(name='Arial', color=BLUE_TEXT, size=10)
    cell.border = thin_border()

FMT_DOLLAR   = '$#,##0;($#,##0);"-"'
FMT_DOLLAR_K = '$#,##0;($#,##0);"-"'
# Cash Flow / Net Worth sheets: amounts strictly between -$1 and $1 are
# rounding noise, not real dollars, so they display as "-" like an exact zero
# instead of "$0"/"($0)". Excel custom formats allow two bracketed
# conditions, so this bands the outer sections to |value| >= 1.
FMT_DOLLAR_ZERO_BAND = '[>=1]$#,##0;[<=-1]($#,##0);"-"'
FMT_PCT      = '0.0%'
FMT_YEAR     = '0'
FMT_INT      = '#,##0'

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def write_hdr(ws, row, col, text, bg=NAVY, fg=WHITE, bold=True, span=1, size=11):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = fill(bg)
    c.font = Font(name='Arial', bold=bold, color=fg, size=size)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = thin_border()
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+span-1)
    return c

def write_cell(ws, row, col, value, fmt=None, bold=False, bg=None, fg='000000',
               align='left', border=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name='Arial', bold=bold, color=fg, size=10)
    c.alignment = Alignment(horizontal=align, vertical='center')
    if border:
        c.border = thin_border()
    if bg:
        c.fill = fill(bg)
    if fmt:
        c.number_format = fmt
    return c

def section_title(ws, row, text, span=8, bg=None):
    ws.row_dimensions[row].height = 22
    c = ws.cell(row=row, column=1, value=text)
    c.fill = fill(bg or NAVY)
    c.font = Font(name='Arial', bold=True, color=WHITE, size=13)
    c.alignment = Alignment(horizontal='left', vertical='center')
    if span > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)

# ─────────────────────────────────────────────────────────────────────────────
# 4.  DATA PARSING
# ─────────────────────────────────────────────────────────────────────────────
# Implemented in data_parser.py and imported above.

# ─────────────────────────────────────────────────────────────────────────────
# 5.  TAX / ANNUITY HELPERS
# ─────────────────────────────────────────────────────────────────────────────
# Engine primitives are imported from engine_core.py. Workbook/reporting code must
# not carry separate financial implementations.

# ─────────────────────────────────────────────────────────────────────────────
# 6.  PROJECTION / MONTE CARLO DELEGATION
# ─────────────────────────────────────────────────────────────────────────────
# `project`/`monte_carlo` are not called directly in this module, but several
# sibling modules (e.g. sheets_stress.py) rely on getting them via their own
# `from .workbook_common import *` wildcard import — do not remove as "unused".
from ..planning_engines import project  # consolidated from projection_engine
from ..planning_engines import monte_carlo  # consolidated from monte_carlo_engine
from ..planning_engines import optimize_roth_conversion_strategy

# ─────────────────────────────────────────────────────────────────────────────
# 9.  WORKBOOK BUILDER
# ─────────────────────────────────────────────────────────────────────────────

SECTION_COLOR = {
    '1': '2E75B6',  # Reports — blue
    '2': '7030A0',  # Optimizers — purple
    '3': 'C00000',  # Risk & Stress Tests — red
    '4': 'B45F06',  # System — amber
    'H': '4472C4',  # Hidden/helper sheets — medium blue
}

# Final workbook tab layout. Excel exposes one flat tab strip, so visible
# section-summary tabs are numbered sequentially and their child tabs use
# section-letter labels (1A, 1B, ...). The source builders still create the
# legacy numbered sheets first; workbook_builder.apply_final_workbook_structure
# then merges/relabels/reorders them into this presentation layout.
WORKBOOK_SECTION_LAYOUT = [
    {
        'section': '1. Reports',
        'code': '1',
        'description': 'Read-only plan reports and advisor-review outputs.',
        'sheets': [
            '1A. Executive Summary', '1B. Net Worth', '1C. Cash Flow',
            '1D. Balance Sheet', '1E. Charts', '1F. Lifetime Taxes',
            '1G. Core Spending', '1H. Spending Summary',
        ],
    },
    {
        'section': '2. Optimizers',
        'code': '2',
        'description': 'Decision-support modules and optimization outputs.',
        'sheets': [
            '2A. Roth Conversion', '2B. Asset Allocation', '2C. State Residency',
            '2D. Social Security', '2E. S-Corp vs LLC', '2F. Charitable Giving',
            '2G. Estate & Legacy Planning', '2H. Planning Levers', '2I. Tax-Loss Harvesting',
        ],
    },
    {
        'section': '3. Risk & Stress Tests',
        'code': '3',
        'description': 'Monte Carlo, survivor, and protection stress tests.',
        'sheets': ['3A. Monte Carlo', '3B. Survivor', '3C. LTC + Life Insurance'],
    },
    {
        'section': '4. System',
        'code': '4',
        'description': 'Plan data snapshot, assumptions, reconciliation, quality control, RMD audit, methodology, and glossary.',
        'sheets': [
            '4A. Plan Data', '4B. Assumptions', '4C. Account Reconciliation',
            '4D. Quality Control', '4E. RMD Audit', '4F. Methodology', '4G. Glossary',
        ],
    },
]

# Legacy build-time sheet set used by the existing sheet builders. These names
# are intentionally kept stable so the computation/build code can remain
# low-risk while the user-facing workbook is reorganized at the end.
V5_LAYOUT = [
    ('1. Executive Summary', '1'),
    ('2. Assumptions', '4'),
    ('3. Balance Sheet', '1'),
    ('4. Asset Allocation', '2'),
    ('5. Net Worth Projection', '1'),
    ('6. Cash Flow Projection', '1'),
    ('7. Lifetime Tax', '1'),
    ('8. Charts Dashboard', '1'),
    ('9. Retirement Strategy', '1'),
    ('10. Social Security', '2'),
    ('11. Roth Conversion', '2'),
    ('12. Charitable Giving', '2'),
    ('12B. Tax-Loss Harvesting', '2'),
    ('13. State Residency', '2'),
    ('14. Estate Plan', '2'),
    ('15. Market-Luck Stress Test', '3'),
    ('16. Scenario Analysis', 'H'),
    ('17. LTC Stress Test', '3'),
    ('18. Survivor Stress Test', '3'),
    ('19. Life Insurance', '3'),
    ('20. RMD Audit', '4'),
    ('21. Quality Control', '4'),
    ('22. Glossary', '4'),
    ('23. Methodology', '4'),
    ('24. Asset Location', '2'),
    ('25. Account Reconciliation', '4'),
    ('26. Workbook Warnings', 'H'),
    ('27. Planning Levers', '2'),
    ('28. Core Spending', '1'),
    ('29. Spending Summary', '1'),
]



QC_CHECKS = []   # [(sheet_name, check, status, detail)]

def qc(sheet, check, passed, detail=''):
    QC_CHECKS.append((sheet, check, 'PASS' if passed else 'FAIL', detail))

# ─────────────────────────────────────────────────────────────────────────────
# Sheet builders
# ─────────────────────────────────────────────────────────────────────────────

def auto_fit_columns(ws, min_width=8, max_width=50, skip_rows=2):
    """Auto-fit column widths based on cell content, skipping title rows.
    Respects min/max bounds. Uses openpyxl cell values only (no rendering)."""
    col_widths = {}
    for row in ws.iter_rows(min_row=skip_rows + 1):
        for cell in row:
            if cell.value is None:
                continue
            col = cell.column
            try:
                val_str = str(cell.value)
                # Format numbers nicely for width estimation
                if isinstance(cell.value, float):
                    if cell.number_format and '%' in cell.number_format:
                        val_str = f'{cell.value:.1%}'
                    elif cell.number_format and '$' in cell.number_format:
                        val_str = f'${cell.value:,.0f}'
                    else:
                        val_str = f'{cell.value:.2f}'
                length = max(len(val_str), 6)
            except Exception:
                length = 8
            col_widths[col] = max(col_widths.get(col, min_width), length)
    # Also check header row
    for cell in ws[skip_rows]:
        if cell.value:
            col_widths[cell.column] = max(col_widths.get(cell.column, min_width),
                                          len(str(cell.value)) + 1)
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = min(max(width + 2, min_width), max_width)



# Export private helper names to sheet modules using star imports.
__all__ = [name for name in globals() if not name.startswith("__")]


# ─────────────────────────────────────────────────────────────────────────────
# Workbook-wide readability pass
# ─────────────────────────────────────────────────────────────────────────────
def _display_len(value):
    """Approximate rendered character length for data-driven sizing."""
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return 10
    s = str(value)
    if s.startswith('='):
        return min(len(s), 10)
    return max(len(line) for line in s.splitlines() or [''])


def _is_textual(value):
    if value is None:
        return False
    return isinstance(value, str) and not value.startswith('=')


def _is_heading_row(ws, row_idx):
    """Identify rows that should not determine column width.

    Table headers and merged section titles are excluded, but label/value rows
    are kept as data even when bold/fill styling is applied.
    """
    if row_idx <= 2:
        return True
    cells = [ws.cell(row=row_idx, column=c) for c in range(1, ws.max_column + 1)]
    nonblank = [c for c in cells if c.value not in (None, '')]
    if not nonblank:
        return False
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row_idx <= mr.max_row and mr.min_col == 1 and (mr.max_col - mr.min_col + 1) >= 3:
            return True
    if len(nonblank) >= 3:
        bold = sum(1 for c in nonblank if c.font and c.font.bold)
        filled = sum(1 for c in nonblank if c.fill and c.fill.fill_type)
        if bold >= max(2, len(nonblank) * 0.5) or filled >= max(2, len(nonblank) * 0.5):
            return True
    return False


def optimize_workbook_layout(wb, target_total_width=118):
    """Apply workbook-wide sizing using data rows, not headings.

    Caps are intentionally expressed in Excel character units approximating the
    requested pixels: dollars <= 71px (~9.4 chars), integer columns <= 40px
    (~5 chars), label-style text columns <= ~36 chars, and narrative/notes
    columns (a column whose longest cell exceeds narrative_threshold) <= ~46
    chars. Headings are wrapped within those caps and do not widen columns.

    Text-cap calibration: derived from a reference workbook
    (groups_accounts.xlsx) whose author hand-set column widths. Short/coded
    columns there were sized to ~1.0x their longest value (e.g. a 34-char
    category column got a 33.1-char width) rather than the ~28-char cap this
    algorithm used previously — so text_cap is raised to 36 to stop clipping
    ordinary label columns. Free-text notes columns in that same file held up
    to 141 characters but were still only given ~27-46 char widths (~0.33x
    ratio, not a full fit) — so genuinely long narrative columns get their own
    higher-but-still-capped narrative_cap (46) instead of sharing text_cap,
    trading a bit of width for far fewer wrapped lines per row.
    """
    import math as _math
    from copy import copy as _copy
    from openpyxl.styles import Alignment as _Alignment

    dollar_cap = (71 - 5) / 7
    text_cap = 36.0
    narrative_cap = 46.0
    narrative_threshold = 60  # column max content length beyond which it's treated as free-text notes, not a label
    int_cap = (40 - 5) / 7
    pct_cap = 7.5
    num_cap = 12.0
    min_text = 10.0
    min_dollar = (71 - 5) / 7  # 71 px minimum for dollar columns
    min_int = 4.4
    min_num = 7.0

    def _header_text(ws, col):
        vals = []
        for r in range(1, min(ws.max_row, 12) + 1):
            val = ws.cell(r, col).value
            if isinstance(val, str) and val.strip():
                vals.append(val.lower())
        return ' '.join(vals)

    def _cell_len(cell):
        val = cell.value
        if val is None:
            return 0
        if isinstance(val, (int, float)):
            fmt = (cell.number_format or '').lower()
            if '$' in fmt or 'accounting' in fmt:
                return 10
            if '%' in fmt:
                return 6
            if float(val).is_integer():
                return len(str(int(abs(val)))) + (1 if val < 0 else 0)
            return 9
        return _display_len(val)

    def _classify(ws, col, cells):
        vals = [c.value for c in cells if c.value not in (None, '')]
        header = _header_text(ws, col)
        if not vals:
            return 'text'
        if any(isinstance(v, str) and not v.startswith('=') for v in vals):
            return 'text'
        if all(isinstance(v, (int, float)) for v in vals):
            nf = ' '.join((c.number_format or '') for c in cells).lower()
            if '$' in nf or 'accounting' in nf or any(k in header for k in ('$','amount','value','balance','tax','cash','income','spend','expense','asset','liability','net worth','rmd','ira','roth','trust','portfolio','withdrawal','premium','deduction','contribution')):
                return 'dollar'
            if '%' in nf or any(k in header for k in ('percent','%','rate','probability','success')):
                return 'percent'
            if all(float(v).is_integer() for v in vals):
                if any(k in header for k in ('year','age','item','#','count','row')) or max(abs(float(v)) for v in vals) < 10000:
                    return 'integer'
            return 'number'
        return 'text'

    for ws in wb.worksheets:
        if getattr(ws, 'sheet_state', 'visible') != 'visible':
            continue
        max_row = ws.max_row or 1
        max_col = ws.max_column or 1
        heading_rows = {r for r in range(1, max_row + 1) if _is_heading_row(ws, r)}
        specs = {}
        for col in range(1, max_col + 1):
            cells = [ws.cell(r, col) for r in range(1, min(max_row, 400) + 1)
                     if r not in heading_rows and ws.cell(r, col).value not in (None, '')]
            kind = _classify(ws, col, cells)
            lengths = sorted([_cell_len(c) for c in cells] or [0])
            p85 = lengths[int((len(lengths) - 1) * 0.85)] if lengths else 0
            max_len = max(lengths or [0])
            if kind == 'dollar':
                width = max(min_dollar, min(dollar_cap, p85 + 1)); wrap = False
            elif kind == 'integer':
                width = max(min_int, min(int_cap, p85 + 1)); wrap = False
            elif kind == 'percent':
                width = max(5.5, min(pct_cap, p85 + 1)); wrap = False
            elif kind == 'text':
                # Genuinely long free-text (notes/descriptions) gets a wider
                # cap than short labels/category names, since squeezing 100+
                # character narrative into a label-width column just produces
                # excessively tall wrapped rows.
                cap = narrative_cap if max_len > narrative_threshold else text_cap
                width = max(min_text, min(cap, (p85 + 2) if p85 else 14)); wrap = True
            else:
                width = max(min_num, min(num_cap, p85 + 1)); wrap = False
            ws.column_dimensions[get_column_letter(col)].width = round(width, 1)
            specs[col] = (kind, width, wrap, max_len)

        if specs and max(specs) <= 18:
            total_width = sum(width for _kind, width, _wrap, _max_len in specs.values())
            if total_width > target_total_width:
                mins = {
                    'text': 8.0,
                    'dollar': min_dollar,
                    'integer': min_int,
                    'percent': 5.5,
                    'number': 6.0,
                }
                shrinkable = sum(max(0.0, width - mins.get(kind, 6.0)) for kind, width, _wrap, _max_len in specs.values())
                if shrinkable > 0:
                    factor = min(1.0, (total_width - target_total_width) / shrinkable)
                    for col, (kind, width, wrap, max_len) in list(specs.items()):
                        floor = mins.get(kind, 6.0)
                        new_width = width - max(0.0, width - floor) * factor
                        ws.column_dimensions[get_column_letter(col)].width = round(new_width, 1)
                        specs[col] = (kind, new_width, wrap, max_len)

        for row_idx in range(1, max_row + 1):
            max_lines = 1
            any_wrap = False
            for col, (kind, width, wrap, _) in specs.items():
                cell = ws.cell(row_idx, col)
                if cell.value in (None, ''):
                    continue
                if kind == 'text' or _is_heading_row(ws, row_idx):
                    # Standard vertical alignment across every sheet: 'center'.
                    # Wrapped multi-line text cells still center vertically
                    # rather than pinning to the top, so a wrapped notes cell
                    # lines up visually with its single-line neighbors in the
                    # same row instead of looking anchored high.
                    old = cell.alignment or _Alignment()
                    cell.alignment = _Alignment(
                        horizontal=old.horizontal,
                        vertical='center',
                        text_rotation=old.text_rotation,
                        wrap_text=True,
                        shrink_to_fit=old.shrink_to_fit,
                        indent=old.indent,
                    )
                    any_wrap = True
                    effective_width = width
                    for mr in ws.merged_cells.ranges:
                        if mr.min_row <= row_idx <= mr.max_row and mr.min_col <= col <= mr.max_col:
                            effective_width = sum(specs.get(c, ('text', width, True, 0))[1] for c in range(mr.min_col, mr.max_col + 1))
                            break
                    text = str(cell.value)
                    lines = 0
                    for line in text.splitlines() or ['']:
                        lines += max(1, _math.ceil(len(line) / max(effective_width, 1)))
                    max_lines = max(max_lines, min(lines, 6))
                else:
                    # Numeric column kinds: dollars/percents/generic numbers are
                    # right-aligned, integers/years/ages centered; all vertically
                    # centered. Only fill alignment that is currently unset so any
                    # deliberate per-cell alignment from sheet builders is preserved.
                    old = cell.alignment or _Alignment()
                    if old.horizontal is None:
                        cell.alignment = _Alignment(
                            horizontal='center' if kind == 'integer' else 'right',
                            vertical='center',
                            text_rotation=old.text_rotation,
                            wrap_text=False,
                            shrink_to_fit=old.shrink_to_fit,
                            indent=old.indent,
                        )
            if any_wrap:
                ws.row_dimensions[row_idx].height = min(90, max(15, max_lines * (18 if row_idx <= 2 else 15)))

        if ws.title in {'2B. Asset Allocation', '4. Asset Allocation'}:
            for row_idx in (211, 285):
                cell = ws.cell(row=row_idx, column=1)
                old = cell.alignment or _Alignment()
                cell.alignment = _Alignment(
                    horizontal=old.horizontal,
                    vertical=old.vertical or 'top',
                    text_rotation=old.text_rotation,
                    wrap_text=True,
                    shrink_to_fit=old.shrink_to_fit,
                    indent=old.indent,
                )
            # Height is no longer forced here: minimize_row_heights() sizes
            # every wrapped row (including these) from actual content once
            # final column widths are known.

    return wb


def apply_numeric_centering(wb):
    """Percentages and 2-4 digit plain numbers (ages, years, counts, ranks...)
    are always centered horizontally, overriding whatever alignment the
    individual sheet builder set. Dollar-formatted cells are left untouched
    (they stay right-aligned) since the rule only covers %/plain numbers.
    """
    from openpyxl.styles import Alignment as _Alignment
    for ws in wb.worksheets:
        if getattr(ws, 'sheet_state', 'visible') != 'visible':
            continue
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if val is None:
                    continue
                nf = (cell.number_format or '').strip()
                if not nf or nf.lower() == 'general':
                    continue
                is_percent = '%' in nf
                is_plain_int = (not is_percent and '$' not in nf
                                 and 'accounting' not in nf.lower()
                                 and set(nf) <= set('0#,;()- '))
                center = False
                if is_percent:
                    center = True
                elif is_plain_int and isinstance(val, (int, float)):
                    digits = len(str(int(abs(val))))
                    if 2 <= digits <= 4:
                        center = True
                if not center:
                    continue
                old = cell.alignment or _Alignment()
                cell.alignment = _Alignment(
                    horizontal='center',
                    vertical=old.vertical or 'center',
                    text_rotation=old.text_rotation,
                    wrap_text=old.wrap_text,
                    shrink_to_fit=old.shrink_to_fit,
                    indent=old.indent,
                )
    return wb


# ─────────────────────────────────────────────────────────────────────────────
# Exact column widths / row heights from the reference formatting workbook
# ─────────────────────────────────────────────────────────────────────────────
# Captured once from "template for column widths and height.xlsx" (hand-tuned
# by the user) into a JSON sidecar keyed by final user-facing sheet title, so
# the build doesn't depend on that external file's path at run time.
import json as _json
from pathlib import Path as _Path

_TEMPLATE_LAYOUT_PATH = _Path(__file__).with_name('_template_layout_data.json')


def _load_template_layout():
    try:
        with open(_TEMPLATE_LAYOUT_PATH, 'r', encoding='utf-8') as f:
            return _json.load(f)
    except FileNotFoundError:
        return {}


TEMPLATE_LAYOUT = _load_template_layout()


def apply_template_layout(wb):
    """Pin exact column widths and row heights captured from the reference
    formatting workbook, keyed by final sheet title.

    Runs after optimize_workbook_layout so its heuristic sizing is the
    fallback for any column/row the template doesn't specify (e.g. columns
    added by data that wasn't present when the template was captured). The
    row heights pinned here are themselves a starting point, not final: they
    get recomputed by minimize_row_heights() once every column-width pass
    (including user overrides from Settings -> Workbook Formatting) has run,
    since a pinned height only matches the template's own column widths.
    """
    for ws in wb.worksheets:
        spec = TEMPLATE_LAYOUT.get(ws.title)
        if not spec:
            continue
        for letter, width in spec.get('cols', {}).items():
            ws.column_dimensions[letter].width = width
        for row_str, height in spec.get('rows', {}).items():
            ws.row_dimensions[int(row_str)].height = height
    return wb


def minimize_row_heights(wb):
    """Shrink (or grow) every content row to the minimum height that fully
    displays its text at the sheet's FINAL column widths.

    Must run last, after every column-width pass (heuristic, template,
    user overrides), since how many lines a wrapped cell needs depends on
    the width it will actually render at. Handles text merged across
    multiple columns (the wrap width is the merged range's combined column
    width) and, for robustness, text merged across multiple rows (the
    needed height is spread across the spanned rows). Rows with no cell
    content are left untouched so chart-anchor spacer rows aren't disturbed.
    """
    import math as _math
    from collections import defaultdict

    EXCEL_MAX_ROW_HEIGHT = 409.0  # points; Excel's own hard ceiling
    LINE_PAD = 4.0  # points of padding per line, above the raw font size
    DEFAULT_WIDTH = 8.43  # Excel's default column width when none is set

    for ws in wb.worksheets:
        if getattr(ws, 'sheet_state', 'visible') != 'visible':
            continue
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        if not max_row or not max_col:
            continue

        col_width = {}
        sheet_default = getattr(ws.sheet_format, 'defaultColWidth', None) or DEFAULT_WIDTH
        for c in range(1, max_col + 1):
            dim = ws.column_dimensions.get(get_column_letter(c))
            col_width[c] = float(dim.width) if (dim and dim.width) else float(sheet_default)

        merge_span = {(mr.min_row, mr.min_col): mr for mr in ws.merged_cells.ranges}

        row_needed = defaultdict(float)
        for row_idx in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                cell = ws.cell(row_idx, c)
                if cell.value in (None, ''):
                    continue
                font_size = float(cell.font.size) if (cell.font and cell.font.size) else 10.0
                line_pt = font_size + LINE_PAD
                mr = merge_span.get((row_idx, c))
                col_span = range(mr.min_col, mr.max_col + 1) if mr else range(c, c + 1)
                row_span = range(mr.min_row, mr.max_row + 1) if mr else range(row_idx, row_idx + 1)
                wrap = bool(cell.alignment and cell.alignment.wrap_text)
                # Numbers/dates render as one unbroken token (no spaces to
                # wrap at), so Excel never splits them across lines even with
                # wrap_text on -- only word-wrappable strings actually grow a
                # row. Measuring a raw float's repr would otherwise wildly
                # overstate the height a numeric cell needs.
                if wrap and isinstance(cell.value, str):
                    eff_width = max(sum(col_width.get(cc, DEFAULT_WIDTH) for cc in col_span), 1.0)
                    lines = 0
                    for line in (cell.value.splitlines() or ['']):
                        lines += max(1, _math.ceil(len(line) / eff_width))
                    lines = max(1, lines)
                else:
                    lines = 1
                per_row_pts = (lines * line_pt) / max(1, len(row_span))
                for r in row_span:
                    if per_row_pts > row_needed[r]:
                        row_needed[r] = per_row_pts

        for row_idx, pts in row_needed.items():
            ws.row_dimensions[row_idx].height = round(min(EXCEL_MAX_ROW_HEIGHT, pts), 1)

    return wb


# Refresh exports after workbook-wide helpers are defined.
__all__ = [name for name in globals() if not name.startswith("__")]
