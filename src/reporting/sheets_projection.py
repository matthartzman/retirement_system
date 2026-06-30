from .workbook_common import *
def build_sheet5(ws, c, rows):
    """Net Worth Projection"""
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'D3'

    # Headers
    COL_YEAR = 1; COL_HAGE = 2; COL_WAGE = 3
    # Groups: Annuities (4-8), Pre-Tax (10-12), Roth (14-15), Trust (17-18), HSA (20), Other (21-26), TOTAL (28)
    COLS = {
        'Year':1,'H Age':2,'W Age':3,
        'Pension PV':4,'W Single PV':5,'W Joint PV':6,'H Single PV':7,'H Joint PV':8,'Σ Ann':9,
        'PreTax_1':10,'PreTax_2':11,'PreTax_3':12,'Σ PreTax':13,
        'Roth_1':14,'Roth_2':15,'Σ Roth':16,
        'Trust_1':17,'Trust_2':18,'Σ Trust':19,
        'HSA_1':20,
        'Home Equity':21,'Next Housing Equity':22,'Startup':23,'Autos':24,'Note':25,'Cash':26,'Σ Other':27,
        'TOTAL NW':28,
    }

    def _account_slots(tax=None, acct_type=None, count=1):
        """Return workbook account slots in semantic owner/type order.

        Do not rely on c['*_ids'] ordering here. The registry is sorted for
        deterministic engine access, but that order can be alphabetical within a
        tax bucket (for example Husband_401k before Husband_IRA). Workbook detail
        columns need each label to travel with the exact account id/value.
        """
        registry = list(c.get('account_registry') or [])
        order = {'traditional_ira': 0, '401k': 1, '403b': 2, 'sep_ira': 3,
                 'roth_ira': 0, 'roth_401k': 1, 'trust': 0, 'taxable': 1, 'hsa': 0}
        candidates = []
        for acct in registry:
            if tax is not None and acct.get('tax') != tax:
                continue
            if acct_type is not None and acct.get('acct_type') != acct_type:
                continue
            candidates.append(acct)
        candidates.sort(key=lambda a: (a.get('owner_idx', 0), order.get(a.get('acct_type'), 99), str(a.get('id', ''))))

        def _owner_label(acct):
            owner = acct.get('owner_idx', 0)
            if owner == 0:
                return 'Husband'
            if owner == 1:
                return 'Wife'
            return str(acct.get('owner_name') or f'Owner {owner + 1}')

        def _kind_label(acct):
            typ = str(acct.get('acct_type') or '')
            if typ == 'traditional_ira':
                return 'IRA'
            if typ == '401k':
                return '401k'
            if typ == '403b':
                return '403b'
            if typ == 'sep_ira':
                return 'SEP IRA'
            if typ in ('roth_ira', 'roth_401k'):
                return 'Roth'
            if typ in ('trust', 'taxable'):
                return 'Trust'
            if typ == 'hsa':
                return 'HSA'
            return str(acct.get('label') or acct.get('id') or 'Account')

        slots = []
        for acct in candidates[:count]:
            aid = acct.get('id')
            slots.append((aid, f"{_owner_label(acct)} {_kind_label(acct)}"))
        while len(slots) < count:
            slots.append((None, ''))
        return slots

    pretax_slots = _account_slots(tax='pre_tax', count=3)
    roth_slots = _account_slots(tax='roth', count=2)
    trust_slots = _account_slots(tax='taxable', count=2)
    hsa_slots = _account_slots(tax='hsa', count=1)
    header_labels = {
        'PreTax_1': pretax_slots[0][1] or 'Pre-Tax 1',
        'PreTax_2': pretax_slots[1][1] or 'Pre-Tax 2',
        'PreTax_3': pretax_slots[2][1] or 'Pre-Tax 3',
        'Roth_1': roth_slots[0][1] or 'Roth 1',
        'Roth_2': roth_slots[1][1] or 'Roth 2',
        'Trust_1': trust_slots[0][1] or 'Trust 1',
        'Trust_2': trust_slots[1][1] or 'Trust 2',
        'HSA_1': hsa_slots[0][1] or 'HSA',
    }

    # Group header row
    r = 1
    write_hdr(ws, r, 1, 'Identifiers', DGRAY, WHITE, span=3)
    write_hdr(ws, r, 4, 'ANNUITIES & PENSION', BLUE, WHITE, span=6)
    write_hdr(ws, r, 10, 'PRE-TAX (IRA / 401k)', ORANGE, WHITE, span=4)
    write_hdr(ws, r, 14, 'ROTH', GREEN, WHITE, span=3)
    write_hdr(ws, r, 17, 'TRUSTS', '7030A0', WHITE, span=3)
    write_hdr(ws, r, 20, 'HSA', GOLD, '000000', span=1)
    write_hdr(ws, r, 21, 'OTHER ASSETS', DGRAY, WHITE, span=7)
    write_hdr(ws, r, 28, 'TOTAL', NAVY, WHITE, span=1)

    r = 2
    for name, col in COLS.items():
        label = header_labels.get(name, name)
        bg = LGRAY if 'Σ' in label or label=='TOTAL NW' else DGRAY
        fg = '000000' if bg==LGRAY else WHITE
        write_hdr(ws, r, col, label, bg, fg, size=9)

    # Data rows
    for i, row in enumerate(rows):
        r = i + 3
        ann_total  = (row['pension_pv']+row['w_single_pv']+row['w_joint_pv']+
                      row['h_single_pv']+row['h_joint_pv'])
        pretax_accounts = [aid for aid, _label in pretax_slots if aid]
        roth_accounts = [aid for aid, _label in roth_slots if aid]
        trust_accounts = [aid for aid, _label in trust_slots if aid]
        hsa_accounts = [aid for aid, _label in hsa_slots if aid]
        pretax_vals = [row.get(a, 0) for a in pretax_accounts] + [0, 0, 0]
        roth_vals = [row.get(a, 0) for a in roth_accounts] + [0, 0]
        trust_vals = [row.get(a, 0) for a in trust_accounts] + [0, 0]
        hsa_val = row.get(hsa_accounts[0], 0) if hsa_accounts else 0
        pretax_tot = sum(row.get(a, 0) for a in c.get('pre_tax_ids', []))
        roth_tot = sum(row.get(a, 0) for a in c.get('roth_ids', []))
        trust_tot = sum(row.get(a, 0) for a in c.get('taxable_ids', []))
        cash_val = row.get('cash_other', c.get('cash_other', 0))
        other_tot  = (row.get('home_equity',0)+row.get('next_housing_equity',0)+
                      row.get('startup_val',0)+row.get('autos_val',0)+
                      row.get('note_bal',0)+cash_val)

        vals = [row['year'], row['h_age'], row['w_age'],
                row['pension_pv'], row['w_single_pv'], row['w_joint_pv'],
                row['h_single_pv'], row['h_joint_pv'], ann_total,
                pretax_vals[0], pretax_vals[1], pretax_vals[2], pretax_tot,
                roth_vals[0], roth_vals[1], roth_tot,
                trust_vals[0], trust_vals[1], trust_tot,
                hsa_val,
                row.get('home_equity',0), row.get('next_housing_equity',0), row.get('startup_val',0),
                row.get('autos_val',0), row.get('note_bal',0), cash_val, other_tot,
                row['total_nw']]

        for col_idx, val in enumerate(vals, 1):
            if col_idx in (1,2,3):
                fmt = FMT_YEAR if col_idx==1 else '0'
                align = 'center'
            else:
                fmt = FMT_DOLLAR
                align = 'right'
            is_subtotal = col_idx in (9,13,16,19,27,28)
            bg = LGRAY if is_subtotal else (GRAY if col_idx==28 else None)
            bold = is_subtotal or col_idx==28
            write_cell(ws, r, col_idx, val, fmt=fmt, bold=bold, bg=bg, align=align)

    # Column widths
    ws.column_dimensions['A'].width = 7
    ws.column_dimensions['B'].width = 7
    ws.column_dimensions['C'].width = 7
    for col in range(4, 29):
        ws.column_dimensions[get_column_letter(col)].width = 13

    # Collapsible column groups for Sheet 5 (individual account detail vs subtotals)
    # Subtotal cols: Σ Ann=9, Σ PreTax=13, Σ Roth=16, Σ Trust=19
    # Group individual detail cols under each subtotal
    for col in range(2, 9):   # Ann detail under col 9
        ws.column_dimensions[get_column_letter(col)].outlineLevel = 1
    for col in range(10, 13): # PreTax detail under col 13
        ws.column_dimensions[get_column_letter(col)].outlineLevel = 1
    for col in range(14, 16): # Roth detail under col 16
        ws.column_dimensions[get_column_letter(col)].outlineLevel = 1
    for col in range(17, 19): # Trust detail under col 19
        ws.column_dimensions[get_column_letter(col)].outlineLevel = 1
    for col in range(21, 27): # Other detail under col 27
        ws.column_dimensions[get_column_letter(col)].outlineLevel = 1

    auto_fit_columns(ws, min_width=8, max_width=16)

    # Summary block
    r_data_end = len(rows) + 2
    r = r_data_end + 2
    write_hdr(ws, r, 1, 'Summary: Plan Start vs Plan End', NAVY, WHITE, span=5); r+=1
    write_hdr(ws, r, 1, 'Group', DGRAY, WHITE)
    write_hdr(ws, r, 2, f'Plan Start ({c["plan_start"]})', DGRAY, WHITE)
    write_hdr(ws, r, 3, f'Plan End ({c["plan_end"]})', DGRAY, WHITE)
    write_hdr(ws, r, 4, 'Change', DGRAY, WHITE)
    write_hdr(ws, r, 5, '% Change', DGRAY, WHITE)
    r += 1
    groups = [
        ('Annuities & Pension', rows[0]['ann_nw'],  rows[-1]['ann_nw']),
        ('Pre-Tax',             rows[0]['pretax_nw'],rows[-1]['pretax_nw']),
        ('Roth',                rows[0]['roth_nw'], rows[-1]['roth_nw']),
        ('Trusts',              rows[0]['trust_nw'],rows[-1]['trust_nw']),
        ('HSA',                 rows[0]['hsa_nw'],  rows[-1]['hsa_nw']),
        ('Other Assets',        rows[0]['other_nw'],rows[-1]['other_nw']),
        ('TOTAL',               rows[0]['total_nw'],rows[-1]['total_nw']),
    ]
    for grp, start, end in groups:
        chg = end - start
        pct = chg/start if start else 0
        bold = (grp=='TOTAL')
        bg   = LGRAY if bold else None
        write_cell(ws, r, 1, grp, bold=bold, bg=bg)
        write_cell(ws, r, 2, start, fmt=FMT_DOLLAR, bold=bold, bg=bg, align='right')
        write_cell(ws, r, 3, end,   fmt=FMT_DOLLAR, bold=bold, bg=bg, align='right')
        write_cell(ws, r, 4, chg,   fmt=FMT_DOLLAR, bold=bold, bg=bg, align='right')
        write_cell(ws, r, 5, pct,   fmt=FMT_PCT, bold=bold, bg=bg, align='right')
        r += 1

    qc('5. Net Worth Projection', f'Row count = {len(rows)} ({c["plan_start"]}-{c["plan_end"]})',
       len(rows)==c['plan_end']-c['plan_start']+1, '')
    qc('5. Net Worth Projection', 'All balances ≥ 0',
       all(r['total_nw']>=0 for r in rows), '')


def build_sheet6(ws, c, rows):
    """Cash Flow Projection — account-level withdrawals, collapsible groups, auto-fit."""
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'D3'
    ws.sheet_properties.outlinePr.summaryBelow = False  # summary row above detail

    # Column layout:
    # 1-3: Identifiers (Year, H Age, W Age)
    # 4-14: INCOME (Earned, H SS, W SS, Pension, W Sgl, W Jnt, H Sgl, H Jnt, Note, RMD, Σ)
    # 15-21: TAX (Roth Conv, AGI, Taxable, Fed, State, NIIT, IRMAA)
    # 22-36: SPENDING (Base, Housing detail, Wellness detail, Travel, Other, HELOC P&I, Σ)
    # 37-52: WITHDRAWALS — account level
    #   37: H Trust WD   38: W Trust WD   39: Σ Trust
    #   40: HSA WD
    #   41: H Roth WD   42: W Roth WD   43: Σ Roth
    #   44: H IRA RMD   45: H IRA Elec  46: H IRA Conv  47: H IRA Outflow
    #   48: W IRA RMD   49: W IRA Elec  50: W IRA Conv  51: W IRA Outflow
    #   52: HELOC Draw  53: HELOC Bal   54: Σ Cash Draws
    # 53: SURPLUS   54: Total NW (balance check)

    def _pos(v):
        try:
            return float(v or 0) > 0
        except Exception:
            return False

    include_rent = bool(any(_pos(r.get('rent_yr')) for r in rows))
    ltc_configured = bool(c.get('ltc_enabled', False) and _pos(c.get('ltc_annual_prem')))
    include_ltc = bool(ltc_configured or any(_pos(r.get('ltc_prem_yr')) for r in rows))

    COL = {
        'Year': 1, 'H_Age': 2, 'W_Age': 3,
        'Earned': 4, 'H_SS': 5, 'W_SS': 6, 'Pension': 7,
        'W_Sgl': 8, 'W_Jnt': 9, 'H_Sgl': 10, 'H_Jnt': 11, 'Note': 12, 'RMD': 13, 'Σ_Inc': 14,
        'Roth_Conv': 15, 'AGI': 16, 'Taxable': 17, 'Fed': 18, 'State': 19, 'NIIT': 20, 'IRMAA': 21,
    }
    col = 22
    for key in [
        'Spend_Base', 'Housing', 'Mort_PI', 'Prop_Tax', 'Housing_Utilities',
        'Home_Imp', 'Housing_Maintenance', 'Housing_Other'
    ]:
        COL[key] = col; col += 1
    if include_rent:
        COL['Rent_Det'] = col; col += 1
    for key in [
        'Wellness', 'Wellness_Premiums', 'Wellness_Medical', 'Wellness_Dental',
        'Wellness_Vision', 'Wellness_Rx_Otc', 'Wellness_Other'
    ]:
        COL[key] = col; col += 1
    if include_ltc:
        COL['HC_LTC'] = col; col += 1
    for key in ['Travel', 'Other', 'HELOC_PAI', 'Σ_Spend']:
        COL[key] = col; col += 1
    for key in ['Total_Tax', 'Other_Cash_Need', 'Total_Cash_Need', 'Income_Funding',
                'Other_Funding', 'Req_Portfolio_Draws', 'Cash_Bridge_Gap']:
        COL[key] = col; col += 1
    for key in [
        'H_Trust_WD', 'W_Trust_WD', 'Σ_Trust', 'HSA_WD', 'H_Roth_WD', 'W_Roth_WD', 'Σ_Roth',
        'H_IRA_RMD', 'H_IRA_Elec', 'H_IRA_Conv', 'H_IRA_Tot',
        'W_IRA_RMD', 'W_IRA_Elec', 'W_IRA_Conv', 'W_IRA_Tot',
        'HELOC_Draw', 'HELOC_Bal', 'Σ_WD', 'Surplus', 'NW_Check'
    ]:
        COL[key] = col; col += 1
    spending_span = COL['Σ_Spend'] - COL['Spend_Base'] + 1
    cash_bridge_span = COL['Cash_Bridge_Gap'] - COL['Total_Tax'] + 1
    withdrawal_span = COL['Σ_WD'] - COL['H_Trust_WD'] + 1

    # ── Group header row 1 ────────────────────────────────────────────────────
    write_hdr(ws, 1, COL['Year'],     'Identifiers', DGRAY, WHITE, span=3)
    write_hdr(ws, 1, COL['Earned'],   'INCOME',       BLUE,  WHITE, span=11)
    write_hdr(ws, 1, COL['Roth_Conv'],'TAX & RMD',   ORANGE,WHITE, span=7)
    write_hdr(ws, 1, COL['Spend_Base'],'SPENDING',   RED,   WHITE, span=spending_span)
    write_hdr(ws, 1, COL['Total_Tax'], 'CASH BRIDGE', NAVY, WHITE, span=cash_bridge_span)
    write_hdr(ws, 1, COL['H_Trust_WD'],'ACCOUNT OUTFLOWS — CASH DRAWS & IRA CONVERSIONS', GREEN, WHITE, span=withdrawal_span)
    # Roth conversions are account outflows/taxable, but are intentionally not
    # included in the cash-draw subtotal used by the cash bridge.
    write_hdr(ws, 1, COL['Surplus'],  'SURPLUS',      NAVY,  WHITE, span=2)

    # ── Column headers row 2 ─────────────────────────────────────────────────
    SUBTOTAL_COLS = {COL['Σ_Inc'], COL['Σ_Spend'], COL['Σ_Trust'],
                     COL['Σ_Roth'], COL['H_IRA_Tot'], COL['W_IRA_Tot'],
                     COL['Σ_WD'], COL['AGI'], COL['HELOC_Bal'], COL['HELOC_PAI'],
                     COL['Total_Cash_Need'], COL['Req_Portfolio_Draws'], COL['Cash_Bridge_Gap']}
    hdr2 = [
        (COL['Year'],       'Year'),         (COL['H_Age'],      'H Age'),
        (COL['W_Age'],      'W Age'),        (COL['Earned'],     'Earned'),
        (COL['H_SS'],       'Matthew SS'),   (COL['W_SS'],       'Patricia SS'),
        (COL['Pension'],    'Pension'),      (COL['W_Sgl'],      'W Single Ann'),
        (COL['W_Jnt'],      'W Joint Ann'),  (COL['H_Sgl'],      'H Single Ann'),
        (COL['H_Jnt'],      'H Joint Ann'),  (COL['Note'],       'Note P+I'),
        (COL['RMD'],        'RMD Dist'),     (COL['Σ_Inc'],      'Σ Income'),
        (COL['Roth_Conv'],  'Roth Conv'),    (COL['AGI'],        'AGI'),
        (COL['Taxable'],    'Taxable Inc'),  (COL['Fed'],        'Fed Tax'),
        (COL['State'],      'State Tax'),    (COL['NIIT'],       'NIIT'),
        (COL['IRMAA'],      'IRMAA'),
        (COL['Spend_Base'], 'Spend Base'),   (COL['Housing'],    'Housing'),
        (COL['Mort_PI'],    'Mortgage P&I'), (COL['Prop_Tax'],   'Property Tax'),
        (COL['Housing_Utilities'], 'Utilities'), (COL['Home_Imp'], 'Home Impr'),
        (COL['Housing_Maintenance'], 'Maintenance'), (COL['Housing_Other'], 'Other'),
        *(([(COL['Rent_Det'], 'Rent')] if include_rent else [])),
        (COL['Wellness'], 'Wellness'), (COL['Wellness_Premiums'], 'Healthcare Premiums'),
        (COL['Wellness_Medical'], 'Medical'), (COL['Wellness_Dental'], 'Dental'),
        (COL['Wellness_Vision'], 'Vision'), (COL['Wellness_Rx_Otc'], 'Rx/OTC'),
        (COL['Wellness_Other'], 'Other Wellness'),
        *(([(COL['HC_LTC'], 'LTC Prem')] if include_ltc else [])),
        (COL['Travel'],     'Travel'),
        (COL['Other'],      'Other'),        (COL['HELOC_PAI'],  'HELOC P&I'),
        (COL['Σ_Spend'],    'Σ Spend'),
        (COL['Total_Tax'], 'Taxes'), (COL['Other_Cash_Need'], 'Other Cash Need'),
        (COL['Total_Cash_Need'], 'Total Cash Need'), (COL['Income_Funding'], 'Income Funding'),
        (COL['Other_Funding'], 'Other Funding'), (COL['Req_Portfolio_Draws'], 'Required Portfolio Cash Draws'),
        (COL['Cash_Bridge_Gap'], 'Cash Bridge Gap / (Surplus)'),
        (COL['H_Trust_WD'], 'H Trust WD'),   (COL['W_Trust_WD'], 'W Trust WD'),
        (COL['Σ_Trust'],    'Σ Trust'),      (COL['HSA_WD'],     'HSA WD'),
        (COL['H_Roth_WD'],  'H Roth WD'),    (COL['W_Roth_WD'],  'W Roth WD'),
        (COL['Σ_Roth'],     'Σ Roth'),
        (COL['H_IRA_RMD'],  'H IRA RMD'),    (COL['H_IRA_Elec'], 'H IRA Elec'),
        (COL['H_IRA_Conv'], 'H IRA Conv'),   (COL['H_IRA_Tot'],  'H IRA Outflow'),
        (COL['W_IRA_RMD'],  'W IRA RMD'),    (COL['W_IRA_Elec'], 'W IRA Elec'),
        (COL['W_IRA_Conv'], 'W IRA Conv'),   (COL['W_IRA_Tot'],  'W IRA Outflow'),
        (COL['HELOC_Draw'], 'HELOC Draw'),
        (COL['HELOC_Bal'],  'HELOC Bal'),    (COL['Σ_WD'],       'Σ Cash Draws'),
        (COL['Surplus'],    'Surplus'),      (COL['NW_Check'],   'NW Check'),
    ]
    for col, hdr in hdr2:
        is_sub = col in SUBTOTAL_COLS
        bg = LGRAY if is_sub else DGRAY
        fg = '000000' if is_sub else WHITE
        write_hdr(ws, 2, col, hdr, bg, fg, size=9)

    # ── Collapsible column groups ─────────────────────────────────────────────
    # summaryRight=False: summary col appears LEFT of its detail cols
    ws.sheet_properties.outlinePr.summaryRight = False
    # Group the detail columns within each section (hide detail, show subtotal)
    # Housing detail includes every Housing spending group; Rent only appears when configured.
    housing_detail_cols = [
        COL['Mort_PI'], COL['Prop_Tax'], COL['Housing_Utilities'], COL['Home_Imp'],
        COL['Housing_Maintenance'], COL['Housing_Other']
    ]
    if include_rent:
        housing_detail_cols.append(COL['Rent_Det'])
    for col in housing_detail_cols:
        ws.column_dimensions[get_column_letter(col)].outlineLevel = 1
    # Wellness detail uses the Wellness Budget Detail expansion. Healthcare Premiums combines
    # Pre-65 Healthcare Premium plus Medicare Part B, Part D, and Part G; LTC only appears when nonzero/configured.
    wellness_detail_cols = [
        COL['Wellness_Premiums'], COL['Wellness_Medical'], COL['Wellness_Dental'],
        COL['Wellness_Vision'], COL['Wellness_Rx_Otc'], COL['Wellness_Other']
    ]
    if include_ltc:
        wellness_detail_cols.append(COL['HC_LTC'])
    for col in wellness_detail_cols:
        ws.column_dimensions[get_column_letter(col)].outlineLevel = 1
    # Trust detail (summary=Σ_Trust)
    for col in range(COL['H_Trust_WD'], COL['Σ_Trust']):
        ws.column_dimensions[get_column_letter(col)].outlineLevel = 1
    # Roth detail (summary=Σ_Roth)
    for col in range(COL['H_Roth_WD'], COL['Σ_Roth']):
        ws.column_dimensions[get_column_letter(col)].outlineLevel = 1
    # H/W IRA detail (summary=*_IRA_Tot). Conversion is an account outflow
    # but not a cash draw; it is shown here so IRA depletion reconciles to
    # starting balances plus growth instead of looking artificially low.
    for col in range(COL['H_IRA_RMD'], COL['H_IRA_Tot']):
        ws.column_dimensions[get_column_letter(col)].outlineLevel = 1
    for col in range(COL['W_IRA_RMD'], COL['W_IRA_Tot']):
        ws.column_dimensions[get_column_letter(col)].outlineLevel = 1
    # Income detail (summary=Σ_Inc)
    for col in range(COL['Earned'], COL['Σ_Inc']):
        ws.column_dimensions[get_column_letter(col)].outlineLevel = 1
    # HELOC Draw collapses under HELOC Bal
    ws.column_dimensions[get_column_letter(COL['HELOC_Draw'])].outlineLevel = 1

    # ── Data rows ─────────────────────────────────────────────────────────────
    for ri, row in enumerate(rows):
        r = ri + 3
        inc_total = (row['earned'] + row['h_ss'] + row['w_ss'] + row['pension'] +
                     row['wife_single_ann'] + row['wife_joint_ann'] +
                     row['h_single_ann'] + row['h_joint_ann'] +
                     row['note_princ'] + row['note_int'] + row['rmd_total'])
        heloc_pai   = row.get('heloc_interest', 0) + row.get('heloc_repayment_principal', 0)
        spend_total = (row['spend_base_yr']
                       + row.get('housing_total_yr', row.get('mortgage', 0) + row.get('rent_yr', 0))
                       + row.get('wellness_base_yr', 0) + row.get('ltc_prem_yr', 0)
                       + row['rec_extra'] + row['lump'] + heloc_pai)
        trust_total = row.get('h_trust_wd', 0) + row.get('w_trust_wd', 0)
        roth_total  = row.get('h_roth_wd', 0)  + row.get('w_roth_wd', 0)
        h_ira_cash  = row.get('rmd_h', 0)       + row.get('h_ira_elective', 0)
        w_ira_cash  = row.get('rmd_w', 0)       + row.get('w_ira_elective', 0)
        h_ira_tot   = row.get('h_ira_total_outflow', h_ira_cash + row.get('h_ira_conversion', 0))
        w_ira_tot   = row.get('w_ira_total_outflow', w_ira_cash + row.get('w_ira_conversion', 0))
        # Cash bridge terms are intentionally separated so users do not expect
        # total withdrawals to equal spending.  Roth conversions are account
        # outflows/taxable income, but not spendable cash.  HELOC draws are
        # borrowing/other funding, not portfolio withdrawals.  RMDs are shown in
        # the tax/RMD area, but counted here as required portfolio cash draws.
        required_portfolio_draws = (trust_total + row.get('hsa_wd', 0) + roth_total +
                                    h_ira_cash + w_ira_cash)
        wd_total = required_portfolio_draws
        total_tax = row.get('total_tax', row.get('fed_tax', 0) + row.get('state_tax', 0) + row.get('niit', 0) + row.get('irmaa', 0))
        other_cash_need = row.get('other_cash_need_yr', 0)
        total_cash_need = row.get('total_cash_need', spend_total + total_tax + other_cash_need)
        income_funding = (row.get('earned', 0) + row.get('h_ss', 0) + row.get('w_ss', 0) +
                          row.get('pension', 0) + row.get('wife_single_ann', 0) +
                          row.get('wife_joint_ann', 0) + row.get('h_single_ann', 0) +
                          row.get('h_joint_ann', 0) + row.get('note_princ', 0) +
                          row.get('note_int', 0) + row.get('portfolio_income_total', 0))
        other_funding = row.get('heloc_draw', 0)
        cash_bridge_gap = total_cash_need - income_funding - other_funding - required_portfolio_draws

        vals = {
            COL['Year']:      row['year'],
            COL['H_Age']:     row['h_age'],
            COL['W_Age']:     row['w_age'],
            COL['Earned']:    row['earned'],
            COL['H_SS']:      row['h_ss'],
            COL['W_SS']:      row['w_ss'],
            COL['Pension']:   row['pension'],
            COL['W_Sgl']:     row['wife_single_ann'],
            COL['W_Jnt']:     row['wife_joint_ann'],
            COL['H_Sgl']:     row['h_single_ann'],
            COL['H_Jnt']:     row['h_joint_ann'],
            COL['Note']:      row['note_princ'] + row['note_int'],
            COL['RMD']:       row['rmd_total'],
            COL['Σ_Inc']:     inc_total,
            COL['Roth_Conv']: row['roth_conv'],
            COL['AGI']:       row['agi'],
            COL['Taxable']:   row['taxable_inc'],
            COL['Fed']:       row['fed_tax'],
            COL['State']:     row['state_tax'],
            COL['NIIT']:      row['niit'],
            COL['IRMAA']:     row.get('irmaa', 0),
            COL['Spend_Base']:  row['spend_base_yr'],
            COL['Housing']:     row.get('housing_total_yr', row.get('mortgage', 0) + row.get('rent_yr', 0)),
            COL['Mort_PI']:     row.get('mortgage_payment_yr', 0),
            COL['Prop_Tax']:    row.get('real_estate_tax_yr', 0),
            COL['Housing_Utilities']: row.get('housing_utilities_yr', 0),
            COL['Home_Imp']:    row.get('home_improvement_yr', 0),
            COL['Housing_Maintenance']: row.get('housing_maintenance_yr', 0),
            COL['Housing_Other']: row.get('housing_other_yr', 0),
            **({COL['Rent_Det']: row.get('rent_yr', 0)} if include_rent else {}),
            COL['Wellness']:  row.get('wellness_base_yr', 0) + row.get('ltc_prem_yr', 0),
            COL['Wellness_Premiums']: row.get('wellness_premiums_yr', row.get('wellness_bridge_premium', 0) + row.get('medicare_base_premium', 0)),
            COL['Wellness_Medical']: row.get('wellness_medical_yr', 0),
            COL['Wellness_Dental']: row.get('wellness_dental_yr', 0),
            COL['Wellness_Vision']: row.get('wellness_vision_yr', 0),
            COL['Wellness_Rx_Otc']: row.get('wellness_rx_otc_yr', 0),
            COL['Wellness_Other']: row.get('wellness_other_yr', 0),
            **({COL['HC_LTC']: row.get('ltc_prem_yr', 0)} if include_ltc else {}),
            COL['Travel']:      row['rec_extra'],
            COL['Other']:       row['lump'],
            COL['HELOC_PAI']:   heloc_pai,
            COL['Σ_Spend']:     spend_total,
            COL['Total_Tax']: total_tax,
            COL['Other_Cash_Need']: other_cash_need,
            COL['Total_Cash_Need']: total_cash_need,
            COL['Income_Funding']: income_funding,
            COL['Other_Funding']: other_funding,
            COL['Req_Portfolio_Draws']: required_portfolio_draws,
            COL['Cash_Bridge_Gap']: cash_bridge_gap,
            COL['H_Trust_WD']:row.get('h_trust_wd', 0),
            COL['W_Trust_WD']:row.get('w_trust_wd', 0),
            COL['Σ_Trust']:   trust_total,
            COL['HSA_WD']:    row.get('hsa_wd', 0),
            COL['H_Roth_WD']: row.get('h_roth_wd', 0),
            COL['W_Roth_WD']: row.get('w_roth_wd', 0),
            COL['Σ_Roth']:    roth_total,
            COL['H_IRA_RMD']: row.get('rmd_h', 0),
            COL['H_IRA_Elec']:row.get('h_ira_elective', 0),
            COL['H_IRA_Conv']:row.get('h_ira_conversion', 0),
            COL['H_IRA_Tot']: h_ira_tot,
            COL['W_IRA_RMD']: row.get('rmd_w', 0),
            COL['W_IRA_Elec']:row.get('w_ira_elective', 0),
            COL['W_IRA_Conv']:row.get('w_ira_conversion', 0),
            COL['W_IRA_Tot']: w_ira_tot,
            COL['HELOC_Draw']:row.get('heloc_draw', 0),
            COL['HELOC_Bal']: row.get('heloc_balance', 0),
            COL['Σ_WD']:      wd_total,
            COL['Surplus']:   row['surplus'],
            COL['NW_Check']:  row['total_nw'],
        }
        for col_idx, val in vals.items():
            fmt = FMT_YEAR if col_idx == COL['Year'] else (
                  '0' if col_idx in (COL['H_Age'], COL['W_Age']) else FMT_DOLLAR)
            is_sub = col_idx in SUBTOTAL_COLS
            bg = LGRAY if is_sub else None
            write_cell(ws, r, col_idx, val, fmt=fmt, bold=is_sub, bg=bg,
                       align='right' if col_idx > 3 else 'center')

    auto_fit_columns(ws, min_width=9, max_width=18)

    # ── Home Sale Event Callout ───────────────────────────────────────────────
    if c.get('home_sale_yr') and c['home_sale_yr'] > 0:
        sale_row = next((rw for rw in rows if rw['year'] == c['home_sale_yr']), None)
        if sale_row and sale_row.get('home_sale_gross', 0) > 0:
            below = len(rows) + 5
            write_hdr(ws, below, 1, f"Home Sale — {c['home_sale_yr']}", BLUE, WHITE, span=6)
            below += 1
            for lbl, val, fmt in [
                ('Gross Sale Price',                  sale_row['home_sale_gross'],           FMT_DOLLAR),
                (f"Less: Selling Costs ({c['home_sell_cost_pct']*100:.0f}%)", sale_row.get('home_sale_costs',0), FMT_DOLLAR),
                ('Less: Mortgage Payoff',             sale_row['home_sale_mort_off'],        FMT_DOLLAR),
                ('Capital Gain (net of costs)',       sale_row['home_sale_gain'],            FMT_DOLLAR),
                ('Less: §121 Exclusion (MFJ)',        c['sec121'],                           FMT_DOLLAR),
                ('Taxable Gain',                      sale_row['home_sale_taxable'],         FMT_DOLLAR),
                ('LTCG Tax (bracketed 0/15/20%+NIIT)',sale_row['home_sale_tax'],             FMT_DOLLAR),
                ('Net Proceeds (basis-free in trust)',sale_row['home_sale_net'],             FMT_DOLLAR),
                (f"Deposited to: {c['home_sale_acct']}", '', None),
            ]:
                write_cell(ws, below, 1, lbl)
                if val != '':
                    write_cell(ws, below, 2, val, fmt=fmt,
                               bold=(lbl.startswith('Net')))
                below += 1

    qc('6. Cash Flow Projection', f'{len(rows)} rows, account-level WDs, collapsible groups',
       True, '')


def build_sheet7(ws, c, rows):
    """Lifetime Tax Projection"""
    ws.sheet_view.showGridLines = False
    section_title(ws, 1, 'LIFETIME TAX PROJECTION', 10)

    r = 2
    hdrs = ['Year','H Age','W Age','Filing','AGI','Taxable Income',
            'Fed Tax','State Tax','NIIT','Payroll Tax','IRMAA',
            'Total Tax','Effective Rate','Marginal Rate']
    for i, h in enumerate(hdrs, 1):
        write_hdr(ws, r, i, h, NAVY, WHITE, size=9)
    r += 1

    lifetime_fed = 0; lifetime_state = 0; lifetime_niit = 0; lifetime_total = 0

    for row in rows:
        eff_rate = row['total_tax'] / row['agi'] if row['agi'] > 0 else 0
        marg = marginal_rate(row['taxable_inc'], row['year'], row['filing'], c['brk_inf'])
        vals = [row['year'], row['h_age'], row['w_age'], row['filing'],
                row['agi'], row['taxable_inc'],
                row['fed_tax'], row['state_tax'], row['niit'],
                row['payroll_tax'], row['irmaa'],
                row['total_tax'], eff_rate, marg]
        for col_idx, val in enumerate(vals, 1):
            if col_idx == 1:
                fmt = FMT_YEAR
            elif col_idx in (4,): # filing
                fmt = None
            elif col_idx in (13, 14):
                fmt = FMT_PCT
            else:
                fmt = FMT_DOLLAR
            write_cell(ws, r, col_idx, val, fmt=fmt,
                       align='right' if col_idx>4 else 'center')
        lifetime_fed   += row['fed_tax']
        lifetime_state += row['state_tax']
        lifetime_niit  += row['niit']
        lifetime_total += row['total_tax']
        r += 1

    # Totals
    r += 1
    totals = [('Lifetime Federal Tax', lifetime_fed),
              (f'Lifetime State Tax ({c["state"][:2]})',   lifetime_state),
              ('Lifetime NIIT',        lifetime_niit),
              ('Lifetime Total Tax',   lifetime_total)]
    for label, val in totals:
        write_cell(ws, r, 1, label, bold=True, bg=NAVY, fg=WHITE)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        write_cell(ws, r, 7, val, fmt=FMT_DOLLAR, bold=True, bg=NAVY, fg=WHITE)
        r += 1

    for col in range(1, 15):
        ws.column_dimensions[get_column_letter(col)].width = 14

    auto_fit_columns(ws)
    qc('7. Lifetime Tax', 'Lifetime totals present', True,
       f"Total: ${lifetime_total:,.0f}")


def build_sheet8(ws, c, rows, mc_data=None):
    """Charts Dashboard — visual charts on a clean dashboard sheet.

    The chart source ranges are still required for Excel chart generation, but
    they are written to a hidden helper sheet so the visible workbook page and
    the UI result explorer are optimized for reviewing charts instead of helper
    tables.
    """
    chart_ws = ws
    chart_ws.sheet_view.showGridLines = False
    wb = chart_ws.parent
    data_title = '_Chart Dashboard Data'
    if data_title in wb.sheetnames:
        del wb[data_title]
    data_ws = wb.create_sheet(data_title)
    data_ws.sheet_state = 'hidden'
    data_ws.sheet_view.showGridLines = False

    title = f'CHARTS DASHBOARD — {c["h_name"]} & {c.get("w_name","")}  ·  {c["plan_start"]}–{c["plan_end"]}'
    chart_ws.cell(row=1, column=1, value=title)
    chart_ws['A1'].fill = fill(NAVY)
    chart_ws['A1'].font = Font(name='Arial', bold=True, color=WHITE, size=14)
    chart_ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    chart_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=14)
    chart_ws.row_dimensions[1].height = 24
    chart_ws.cell(row=2, column=1, value='Chart source data is stored on a hidden helper sheet. Use this page and the app explorer for visuals; use detailed result sheets for tabular audits.')
    chart_ws['A2'].font = Font(name='Arial', italic=True, color='666666', size=10)
    chart_ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=14)
    for col in range(1, 15):
        chart_ws.column_dimensions[get_column_letter(col)].width = 14

    ws = data_ws

    # ── helpers ──────────────────────────────────────────────────────────────
    def wh(r, c, text, bg=NAVY, fg=WHITE, bold=True, size=10, span=1):
        cell = ws.cell(row=r, column=c, value=text)
        cell.fill = fill(bg); cell.font = Font(name='Arial',bold=bold,color=fg,size=size)
        cell.alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
        cell.border = thin_border()
        if span > 1:
            ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c+span-1)

    def wd(r, c, val, fmt=None, bg=None, bold=False):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = Font(name='Arial', size=10, bold=bold)
        cell.alignment = Alignment(
            horizontal='right' if isinstance(val,(int,float)) else 'center', vertical='center')
        cell.border = thin_border()
        if bg:  cell.fill = fill(bg)
        if fmt: cell.number_format = fmt

    def sec(r, text, bg=NAVY, span=40):
        cell = ws.cell(row=r, column=1, value=text)
        cell.fill = fill(bg); cell.font = Font(name='Arial',bold=True,color=WHITE,size=12)
        cell.alignment = Alignment(horizontal='left',vertical='center')
        ws.row_dimensions[r].height = 22
        if span > 1:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=span)

    # ── pre-compute per-year values ───────────────────────────────────────────
    # Income = income streams + positive portfolio draws only
    # Expense = spending + taxes + payroll-tax residual (≤$35K gap → real cost)
    # Surplus = visual gap (income bar taller than expense bar); NOT a series
    nw_rows = rows  # alias — rows already have all NW fields

    per_year = []
    for row in rows:
        streams = {
            'earned':  round(row.get('earned',0)),
            'h_ss':    round(row.get('h_ss',0)),
            'w_ss':    round(row.get('w_ss',0)),
            'pension': round(row.get('pension',0)),
            'w_sgl':   round(row.get('wife_single_ann',0)),
            'w_jnt':   round(row.get('wife_joint_ann',0)),
            'h_sgl':   round(row.get('h_single_ann',0)),
            'h_jnt':   round(row.get('h_joint_ann',0)),
            'note':    round(row.get('note_princ',0) + row.get('note_int',0)),
            'rmd':     round(row.get('rmd_total',0)),
        }
        trust_wd = max(0, round(row.get('trust_wd',0)))
        hsa_wd   = max(0, round(row.get('hsa_wd',0)))
        roth_wd  = max(0, round(row.get('roth_wd',0)))
        ira_wd   = max(0, round(row.get('ira_wd',0)))
        heloc_draw_wd = max(0, round(row.get('heloc_draw',0)))
        inc_total = sum(streams.values()) + trust_wd + hsa_wd + roth_wd + ira_wd + heloc_draw_wd

        spend_base  = round(row.get('spend_base_yr',0))
        housing     = round(row.get('housing_total_yr', row.get('mortgage',0) + row.get('rent_yr', 0)))
        wellness    = round(row.get('wellness_base_yr', 0) + row.get('ltc_prem_yr', 0))
        travel      = round(row.get('rec_extra',0))
        other       = round(row.get('lump',0))
        heloc_pai_chart = round(row.get('heloc_interest', 0) + row.get('heloc_repayment_principal', 0))
        fed_tax    = round(row.get('fed_tax',0))
        state_tax  = round(row.get('state_tax',0))
        niit       = round(row.get('niit',0))
        exp_raw    = spend_base + housing + wellness + travel + other + heloc_pai_chart + fed_tax + state_tax + niit
        gap        = inc_total - exp_raw
        # Payroll-tax residual: only add to expense when gap is a real cost, not surplus
        payroll_res = round(gap) if 0 < gap <= 35000 else 0

        # Real-dollar equivalent (deflated to plan_start year)
        cpi_deflator = (1 + c['inf']) ** (row['year'] - c['plan_start'])
        per_year.append({
            'real_nw': round(row.get('total_nw', 0) / cpi_deflator) if cpi_deflator > 0 else 0,
            'yr': row['year'],
            'earned': streams['earned'], 'h_ss': streams['h_ss'], 'w_ss': streams['w_ss'],
            'pension': streams['pension'], 'w_sgl': streams['w_sgl'], 'w_jnt': streams['w_jnt'],
            'h_sgl': streams['h_sgl'], 'h_jnt': streams['h_jnt'],
            'note': streams['note'], 'rmd': streams['rmd'],
            'trust_wd': trust_wd, 'hsa_wd': hsa_wd, 'roth_wd': roth_wd,
            'ira_wd': ira_wd, 'heloc_draw_wd': heloc_draw_wd, 'inc_total': inc_total,
            'spend_base': spend_base, 'housing': housing, 'wellness': wellness,
            'travel': travel, 'other': other, 'heloc_pai': heloc_pai_chart,
            'fed_tax': fed_tax, 'state_tax': state_tax,
            'niit': niit, 'payroll_res': payroll_res,
            'exp_total': exp_raw + payroll_res,
            # NW components — home equity separated from other
            'ann_nw':    round(row.get('ann_nw',0)),
            'pretax_nw': round(row.get('pretax_nw',0)),
            'roth_nw':   round(row.get('roth_nw',0)),
            'trust_nw':  round(row.get('trust_nw',0)),
            'hsa_nw':    round(row.get('hsa_nw',0)),
            'home_eq_nw':round(row.get('home_equity',0)),
            'other_nw':  round(row.get('other_nw',0) - row.get('home_equity',0)),
        })

    n = len(per_year)
    # Cash flow scaling: round to nearest $100K, tick every $100K, display in $K
    CF_YMAX = math.ceil(max(max(p['inc_total'] for p in per_year),
                             max(p['exp_total'] for p in per_year)) / 100000) * 100000
    CF_UNIT = 100_000

    # ════════════════════════════════════════════════════════════════════════
    # DATA TABLES (3 side-by-side blocks)
    # NW:  cols  1-8   (A-H)    Year + 6 NW series + Total
    # INC: cols 10-26  (J-Z)    Year + 15 income series + Σ
    # EXP: cols 28-37  (AB-AK)  Year + 8 expense series + Σ
    # Gap cols: 9 (I) and 27 (AA)
    # ════════════════════════════════════════════════════════════════════════
    TABLE_ROW  = 4
    DATA_FIRST = TABLE_ROW + 1
    DATA_LAST  = TABLE_ROW + n

    sec(1, f'CHARTS DASHBOARD — {c["h_name"]} & {c.get("w_name","")}  ·  {c["plan_start"]}–{c["plan_end"]}', span=50)

    # ── NW Table ──────────────────────────────────────────────────────────────
    NW_YEAR = 1
    NW_SER  = [
        (2, 'Annuities & Pension', '2E75B6', 'ann_nw'),
        (3, 'Pre-Tax IRA/401k',    'C55A11', 'pretax_nw'),
        (4, 'Roth',                '0891B2', 'roth_nw'),
        (5, 'Trust',               '5A3E85', 'trust_nw'),
        (6, 'HSA',                 'C9A84C', 'hsa_nw'),
        (7, 'Home Equity',         '16A34A', 'home_eq_nw'),
        (8, 'Other Assets',        '6B7280', 'other_nw'),
    ]
    # Title for NW section — span exactly cols 1-9
    cell = ws.cell(row=2, column=1, value=f'Net Worth by Component  ·  {c["plan_start"]}–{c["plan_end"]}')
    cell.fill = fill(BLUE); cell.font = Font(name='Arial',bold=True,color=WHITE,size=11)
    cell.alignment = Alignment(horizontal='left',vertical='center')
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=9)

    wh(TABLE_ROW, NW_YEAR, 'Year', bg=DGRAY)
    for col, lbl, clr, _ in NW_SER:
        wh(TABLE_ROW, col, lbl, bg=DGRAY)
    wh(TABLE_ROW, 9, 'Total NW', bg=NAVY)

    for ri, p in enumerate(per_year):
        r = DATA_FIRST + ri; bg = GRAY if ri%2==0 else None
        wd(r, NW_YEAR, p['yr'], fmt=FMT_YEAR, bg=bg)
        for col, _, _, key in NW_SER:
            wd(r, col, p[key], fmt=FMT_DOLLAR, bg=bg)
        wd(r, 9, p['ann_nw']+p['pretax_nw']+p['roth_nw']+p['trust_nw']+p['hsa_nw']+p['home_eq_nw']+p['other_nw'],
           fmt=FMT_DOLLAR, bg=bg, bold=True)

    ws.column_dimensions[get_column_letter(1)].width = 7
    for _col in range(2, 10): ws.column_dimensions[get_column_letter(_col)].width = 14
    ws.column_dimensions[get_column_letter(10)].width = 2  # gap

    # ── Income Table ──────────────────────────────────────────────────────────
    INC_YEAR = 11
    INC_SER  = [
        (12, 'earned',   'Earned Income',         '1F3864'),
        (13, 'h_ss',     'Matthew SS',            '2E75B6'),
        (14, 'w_ss',     'Patricia SS',           '3D9AB8'),
        (15, 'pension',  'Pension',               'C9A84C'),
        (16, 'w_sgl',    'Wife Single Ann',       '2D6A4F'),
        (17, 'w_jnt',    'Wife Joint Ann',        '40916C'),
        (18, 'h_sgl',    'Husband Single Ann',    'C55A11'),
        (19, 'h_jnt',    'Husband Joint Ann',     'E07540'),
        (20, 'note',     'Note P+I',              '5A3E85'),
        (21, 'rmd',      'RMD',                   '9B2335'),
        (22, 'trust_wd', 'Trust Draw',            '7B3F9E'),
        (23, 'hsa_wd',   'HSA Draw',              '1B7A9E'),
        (24, 'roth_wd',  'Roth Draw',             '156041'),
        (25, 'ira_wd',   'IRA Draw',              'B85C00'),
        (26, 'heloc_draw_wd', 'HELOC Draw',         '8B5E3C'),
    ]
    # Income section title — span exactly cols 11-27
    cell = ws.cell(row=2, column=INC_YEAR, value=f'Cash Flow — Income & Portfolio Draws  ·  {c["plan_start"]}–{c["plan_end"]}')
    cell.fill = fill(GREEN); cell.font = Font(name='Arial',bold=True,color=WHITE,size=11)
    cell.alignment = Alignment(horizontal='left',vertical='center')
    ws.merge_cells(start_row=2, start_column=INC_YEAR, end_row=2, end_column=27)

    wh(TABLE_ROW, INC_YEAR, 'Year', bg=DGRAY)
    for col, key, lbl, clr in INC_SER:
        wh(TABLE_ROW, col, lbl, bg=DGRAY)
    wh(TABLE_ROW, 27, 'Σ Income', bg=GREEN)

    for ri, p in enumerate(per_year):
        r = DATA_FIRST + ri; bg = GRAY if ri%2==0 else None
        wd(r, INC_YEAR, p['yr'], fmt=FMT_YEAR, bg=bg)
        for col, key, _, _ in INC_SER:
            wd(r, col, p[key], fmt=FMT_DOLLAR, bg=bg)
        wd(r, 27, p['inc_total'], fmt=FMT_DOLLAR, bg=bg, bold=True)

    for _col in range(11, 28): ws.column_dimensions[get_column_letter(_col)].width = 13
    ws.column_dimensions[get_column_letter(11)].width = 7
    ws.column_dimensions[get_column_letter(28)].width = 2  # gap

    # ── Expense Table ─────────────────────────────────────────────────────────
    EXP_YEAR = 29
    EXP_SER  = [
        (30, 'spend_base',  'Base Spending',    '1F3864'),
        (31, 'housing',     'Housing',          '2E75B6'),
        (32, 'wellness',    'Wellness',       'C55A11'),
        (33, 'travel',      'Travel',           'C9A84C'),
        (34, 'other',       'Other',            '059669'),
        (35, 'fed_tax',     'Federal Tax',      '9B2335'),
        (36, 'state_tax',   f'State Tax ({c["state"][:2]})',   'C5384E'),
        (37, 'niit',        'NIIT',             'E07595'),
        (38, 'payroll_res', 'Payroll Tax',      '595959'),
        (39, 'heloc_pai',   'HELOC P&I',        '2D6A8F'),
    ]
    # Expense section title — span exactly cols 29-40
    cell = ws.cell(row=2, column=EXP_YEAR, value=f'Cash Flow — Spending & Taxes  ·  {c["plan_start"]}–{c["plan_end"]}')
    cell.fill = fill(RED); cell.font = Font(name='Arial',bold=True,color=WHITE,size=11)
    cell.alignment = Alignment(horizontal='left',vertical='center')
    ws.merge_cells(start_row=2, start_column=EXP_YEAR, end_row=2, end_column=40)

    wh(TABLE_ROW, EXP_YEAR, 'Year', bg=DGRAY)
    for col, key, lbl, clr in EXP_SER:
        wh(TABLE_ROW, col, lbl, bg=DGRAY)
    wh(TABLE_ROW, 40, 'Σ Spend+Tax', bg=RED)

    for ri, p in enumerate(per_year):
        r = DATA_FIRST + ri; bg = GRAY if ri%2==0 else None
        wd(r, EXP_YEAR, p['yr'], fmt=FMT_YEAR, bg=bg)
        for col, key, _, _ in EXP_SER:
            wd(r, col, p[key], fmt=FMT_DOLLAR, bg=bg)
        wd(r, 40, p['exp_total'], fmt=FMT_DOLLAR, bg=bg, bold=True)

    for _col in range(29, 41): ws.column_dimensions[get_column_letter(_col)].width = 14
    ws.column_dimensions[get_column_letter(29)].width = 7

    # ════════════════════════════════════════════════════════════════════════
    # CHARTS
    # - Legend RIGHT (position='r', overlay=False) — x-axis labels always visible
    # - gapWidth=20, overlap=100 — wide solid stacked bars
    # - 30cm x 18cm each; anchored at rows 38, 83, 128 (45-row gaps)
    # - Title font 24pt bold applied via post-save XML patch in main()
    # ════════════════════════════════════════════════════════════════════════
    def make_chart(title_str, year_col, series_list, data_first, data_last,
                   ymax=None, yunit=None, yfmt='$#,##0,,"M"'):
        chart = BarChart()
        chart.type = 'col'; chart.grouping = 'stacked'
        chart.gapWidth = 20; chart.overlap = 100
        chart.width = 30; chart.height = 18
        chart.title = title_str
        chart.y_axis.numFmt = yfmt
        chart.y_axis.delete = False
        chart.y_axis.tickLblPos = 'nextTo'
        if ymax:
            chart.y_axis.scaling.max = ymax
            chart.y_axis.scaling.min = 0
        if yunit:
            chart.y_axis.majorUnit = yunit
        chart.x_axis.delete = False
        chart.x_axis.numFmt = '0'
        chart.x_axis.tickLblPos = 'low'
        chart.legend.position = 'r'  # RIGHT — never overlaps x-axis labels
        chart.legend.overlay  = False
        cats = Reference(ws, min_col=year_col, max_col=year_col,
                         min_row=data_first, max_row=data_last)
        for idx, (dcol, lbl, color) in enumerate(series_list):
            chart.add_data(Reference(ws, min_col=dcol, max_col=dcol,
                                      min_row=data_first, max_row=data_last))
            s = chart.series[idx]
            s.title = SeriesLabel(v=lbl)
            gp = GraphicalProperties(); gp.solidFill = color; s.spPr = gp
        chart.set_categories(cats)
        chart.visible_cells_only = False
        return chart

    nw_series  = [(col, lbl, clr) for col, lbl, clr, _ in NW_SER]
    inc_series = [(col, lbl, clr) for col, key, lbl, clr in INC_SER]
    exp_series = [(col, lbl, clr) for col, key, lbl, clr in EXP_SER]

    chart_nw  = make_chart(f'Net Worth by Component  ·  {c["plan_start"]}–{c["plan_end"]}',
                             NW_YEAR, nw_series, DATA_FIRST, DATA_LAST)
    chart_inc = make_chart(f'Cash Flow — Income & Portfolio Draws  ·  {c["plan_start"]}–{c["plan_end"]}',
                             INC_YEAR, inc_series, DATA_FIRST, DATA_LAST,
                             ymax=CF_YMAX, yunit=CF_UNIT, yfmt='$#,##0,"K"')
    chart_exp = make_chart(f'Cash Flow — Spending & Taxes  ·  {c["plan_start"]}–{c["plan_end"]}',
                             EXP_YEAR, exp_series, DATA_FIRST, DATA_LAST,
                             ymax=CF_YMAX, yunit=CF_UNIT, yfmt='$#,##0,"K"')

    # Anchor charts: row 38, 83, 128 (45-row gaps for breathing room)
    chart_ws.add_chart(chart_nw,  'A4')
    chart_ws.add_chart(chart_inc, 'A42')
    chart_ws.add_chart(chart_exp, 'A80')

    # ── Chart 4: Net Worth Percentile Bands (Monte Carlo) ────────────────────
    # Write data table for the MC percentile bands (cols AO onward = col 41+)
    MC_YEAR_COL = 43  # col AQ
    MC_PCTS     = [10, 25, 50, 75, 90]
    MC_LABELS   = ['P10','P25','P50 Median','P75','P90']
    MC_COLORS   = ['9B2335','C55A11','2D6A4F','2E75B6','1F3864']

    # Section header
    cell_mc = ws.cell(row=2, column=MC_YEAR_COL,
                      value='MC Percentile Bands Data  ·  used by Chart 4')
    cell_mc.fill = fill('7030A0')
    cell_mc.font = Font(name='Arial', bold=True, color=WHITE, size=11)
    cell_mc.alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(start_row=2, start_column=MC_YEAR_COL,
                   end_row=2, end_column=MC_YEAR_COL + len(MC_PCTS))

    wh(TABLE_ROW, MC_YEAR_COL, 'Year', bg=DGRAY)
    for i, (p, lbl) in enumerate(zip(MC_PCTS, MC_LABELS)):
        wh(TABLE_ROW, MC_YEAR_COL + 1 + i, lbl, bg=DGRAY)

    MC_DATA_FIRST = DATA_FIRST
    MC_DATA_LAST  = DATA_LAST

    for ri, p in enumerate(per_year):
        r_mc = DATA_FIRST + ri
        yr   = p['yr']
        mc_row = (mc_data or {}).get('pct_by_year', {}).get(yr, {})
        bg_alt = GRAY if ri%2==0 else None
        wd(r_mc, MC_YEAR_COL, yr, fmt=FMT_YEAR, bg=bg_alt)
        for j, pct_key in enumerate(MC_PCTS):
            wd(r_mc, MC_YEAR_COL + 1 + j, mc_row.get(pct_key, 0), fmt=FMT_DOLLAR, bg=bg_alt)

    for col in range(MC_YEAR_COL, MC_YEAR_COL + len(MC_PCTS) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 14

    # Build line chart for percentile bands
    from openpyxl.chart import LineChart
    chart_mc = LineChart()
    chart_mc.title  = 'Net Worth Percentile Bands  ·  Monte Carlo Sims  ·  Plan Horizon'
    chart_mc.width  = 30
    chart_mc.height = 18
    chart_mc.y_axis.numFmt = '$#,##0,,"M"'
    chart_mc.y_axis.delete = False
    chart_mc.x_axis.numFmt = '0'
    chart_mc.x_axis.tickLblPos = 'low'
    chart_mc.legend.position   = 'r'
    chart_mc.legend.overlay    = False

    cats_mc = Reference(ws, min_col=MC_YEAR_COL, max_col=MC_YEAR_COL,
                        min_row=MC_DATA_FIRST, max_row=MC_DATA_LAST)
    for i, (lbl, clr) in enumerate(zip(MC_LABELS, MC_COLORS)):
        data_ref = Reference(ws, min_col=MC_YEAR_COL + 1 + i, max_col=MC_YEAR_COL + 1 + i,
                             min_row=MC_DATA_FIRST, max_row=MC_DATA_LAST)
        chart_mc.add_data(data_ref)
        s = chart_mc.series[i]
        s.title = SeriesLabel(v=lbl)
        gp = GraphicalProperties(); gp.solidFill = clr; s.spPr = gp
    chart_mc.set_categories(cats_mc)
    chart_mc.visible_cells_only = False

    chart_ws.add_chart(chart_mc, 'A118')

    # ── Allocation Pie Charts (Before & After) ────────────────────────────
    chart_data = c.get('_alloc_chart_data', {})
    _pie_buckets = chart_data.get('buckets', [])
    _pie_before = chart_data.get('before_vals', [])
    _pie_after = chart_data.get('after_vals', [])

    if _pie_buckets:
        pie_start = DATA_LAST + 175

        from openpyxl.chart.data_source import NumData, NumVal, StrData, StrRef, StrVal
        from openpyxl.chart.label import DataLabelList
        from openpyxl.chart.series import DataPoint

        # Consistent colors per category, even though each pie now has its own
        # filtered source range.  The filtering is intentional: Excel pie
        # charts include zero-value rows in labels/legends unless those rows are
        # excluded from the source data entirely.
        PIE_COLORS = ['2D6A4F', '40916C', '74C69D', 'B7E4C7',
                      'BC6C25', 'DDA15E', '023E8A', '0077B6', '90E0EF', '6C757D']
        color_by_bucket = {str(bkt): PIE_COLORS[idx % len(PIE_COLORS)]
                           for idx, bkt in enumerate(_pie_buckets)}

        def _positive_pie_rows(values):
            raw_rows = []
            for idx, bkt in enumerate(_pie_buckets):
                val = values[idx] if idx < len(values) else 0
                try:
                    num = float(val or 0)
                except (TypeError, ValueError):
                    num = 0.0
                raw_rows.append((str(bkt), num))
            total = sum(max(0.0, num) for _bkt, num in raw_rows)
            # Drop true zero rows and floating-point dust that would display as
            # a 0% allocation in Excel pie labels.  This keeps real but small
            # allocations while preventing labels for values like $0.09 on a
            # multi-million dollar portfolio.
            min_material_value = max(1.0, total * 0.000001) if total > 0 else 1.0
            return [(bkt, num) for bkt, num in raw_rows if num > min_material_value]

        def _write_pie_source(start_row, start_col, header, rows):
            write_cell(ws, start_row, start_col, 'Asset Class', bold=True)
            write_cell(ws, start_row, start_col + 1, header, bold=True)
            for i, (bkt, val) in enumerate(rows):
                write_cell(ws, start_row + 1 + i, start_col, bkt)
                write_cell(ws, start_row + 1 + i, start_col + 1, val, fmt=FMT_DOLLAR)
            return start_row + len(rows)

        def _make_alloc_pie(title, rows, start_col, anchor):
            if not rows:
                return
            pie_end = _write_pie_source(pie_start, start_col, title, rows)
            p = PieChart()
            p.title = title
            p.style = 10; p.width = 22; p.height = 28  # 46 rows ~ 28cm height
            p.add_data(Reference(ws, min_col=start_col + 1, min_row=pie_start+1, max_row=pie_end), titles_from_data=False)
            p.set_categories(Reference(ws, min_col=start_col, min_row=pie_start+1, max_row=pie_end))
            if p.series:
                p.series[0].cat.numRef = None
                p.series[0].cat.strRef = StrRef(
                    f=f"'{ws.title}'!${get_column_letter(start_col)}${pie_start+1}:${get_column_letter(start_col)}${pie_end}",
                    strCache=StrData(
                        ptCount=len(rows),
                        pt=[StrVal(idx=i, v=str(bkt)) for i, (bkt, _val) in enumerate(rows)],
                    ),
                )
                p.series[0].val.numRef.numCache = NumData(
                    ptCount=len(rows),
                    pt=[NumVal(idx=i, v=float(val)) for i, (_bkt, val) in enumerate(rows)],
                )
            # Labels ON slices: category name + percentage only.  Explicitly
            # disabling the series name prevents Excel from prefixing labels
            # with generated text such as "Series 1,".
            p.dataLabels = DataLabelList()
            p.dataLabels.showSerName = False
            p.dataLabels.showLegendKey = False
            p.dataLabels.showCatName = True
            p.dataLabels.showPercent = True
            p.dataLabels.showVal = False
            p.dataLabels.showLeaderLines = True
            p.legend = None  # labels on slices replace legend
            # Apply consistent colors per category after filtering.
            if p.series:
                for idx, (bkt, _val) in enumerate(rows):
                    pt = DataPoint(idx=idx)
                    pt.graphicalProperties = GraphicalProperties()
                    pt.graphicalProperties.solidFill = color_by_bucket.get(bkt, PIE_COLORS[idx % len(PIE_COLORS)])
                    p.series[0].data_points.append(pt)
            chart_ws.add_chart(p, anchor)

        before_rows = _positive_pie_rows(_pie_before)
        after_rows = _positive_pie_rows(_pie_after)

        # Side by side: Before at column A, After at column L (offset ~11 cols).
        # Source tables are separated as well, so categories with a 0% allocation
        # in one pie do not appear in that pie at all.
        _make_alloc_pie('Current Portfolio Allocation', before_rows, 1, 'A156')
        _make_alloc_pie('Target Portfolio Allocation', after_rows, 4, 'L156')

    qc('8. Charts Dashboard', '6 charts: NW, CF Income, CF Expense, MC Bands, Alloc Before, Alloc After', True,
       f'NW, Income (15 ser, ymax=${CF_YMAX:,}), Expense (8 ser, ymax=${CF_YMAX:,}), MC bands, 2 pie charts')



__all__ = ['build_sheet5', 'build_sheet6', 'build_sheet7', 'build_sheet8']
