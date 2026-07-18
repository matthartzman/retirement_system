"""
enterprise_pdf.py

Institutional-quality PDF report generator for the Retirement Plan System.

Ground-up design (rewritten from scratch): rather than maintaining a second,
hand-picked copy of a handful of tables, this module renders EVERY visible
worksheet of the already-built workbook (openpyxl `wb`, the same object that
gets saved to .xlsx) directly into the PDF. The workbook is the single
source of truth, so the PDF can never drift out of sync with the spreadsheet
the way a hand-maintained subset inevitably would.

Layout strategy
----------------
- Landscape letter pages, 0.32" margins on every side (outside the ~0.17"-0.25"
  non-printable border enforced by common printers), with a reserved footer
  strip below the content frame.
- Sheets with more columns than comfortably fit on one page are split into
  repeating "column bands": each band repeats the sheet's first couple of
  identifier columns (e.g. Year) plus a readable slice of the remaining
  columns, so every page stands on its own.
- Sheets with more rows than fit on one page paginate automatically via
  reportlab's native Table row-splitting, repeating the header row(s).
- Header/section-divider rows (bold + filled, as already styled in the
  source workbook) keep their real fill and font color per cell, so the
  spreadsheet's visual section grouping (INCOME/TAX/SPENDING/... bands,
  colored tab sections) carries over into the PDF. Ordinary data rows use a
  plain, minimal grid with light zebra striping for readability.
- A Table of Contents lists every sheet with the PDF page number it starts
  on, resolved via reportlab's two-pass `multiBuild`. Every page is footed
  with "Page N of TOTAL".

Requirements
------------
pip install reportlab matplotlib openpyxl
"""

import os
import tempfile
from pathlib import Path

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt

from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfgen.canvas import Canvas
from reportlab.platypus import (
    BaseDocTemplate, PageTemplate, Frame,
    Paragraph, Spacer, PageBreak, Table, TableStyle, Image,
)
from reportlab.platypus.tableofcontents import TableOfContents
from reportlab.platypus.flowables import HRFlowable


PAGE_W, PAGE_H = landscape(letter)
# Consumer and office printers enforce a non-printable border of roughly
# 0.17"-0.25" on each edge. Anything laid out inside that band displays fine on
# screen and clips when physically printed, so keep every drawn element outside
# it. The footer gets its own reserved strip below the content frame: its
# baseline sits at MARGIN (safely printable) and the content frame stops
# FOOTER_STRIP above the page edge so body content cannot collide with it.
MARGIN = 0.32 * inch
FOOTER_STRIP = 0.16 * inch
BOTTOM_MARGIN = MARGIN + FOOTER_STRIP
CONTENT_W = PAGE_W - 2 * MARGIN
CONTENT_H = PAGE_H - MARGIN - BOTTOM_MARGIN

# How many columns a single page can hold at a readable font size, and how
# many of a sheet's leading columns (typically Year, and sometimes an age or
# label column) get repeated in every column band so each page is
# self-describing without the reader flipping back to band 1.
MAX_COLS_PER_BAND = 16
ID_COLS = 2

GRID_COLOR = colors.HexColor('#D9D9D9')
DEFAULT_HEADER_BG = colors.HexColor('#1F3864')
DEFAULT_HEADER_FG = colors.white
ZEBRA = colors.HexColor('#F5F6F8')


# ─────────────────────────────────────────────────────────────────────────────
# Styles
# ─────────────────────────────────────────────────────────────────────────────

def _styles():
    ss = getSampleStyleSheet()
    ss['Title'].fontName = 'Helvetica-Bold'
    ss['Title'].fontSize = 24
    ss['Title'].leading = 30
    ss['Heading1'].fontName = 'Helvetica-Bold'
    ss['Heading1'].fontSize = 16
    ss['Heading1'].leading = 20
    ss['BodyText'].fontName = 'Helvetica'
    ss['BodyText'].fontSize = 9.5
    ss['BodyText'].leading = 13
    ss.add(ParagraphStyle(
        name='SheetTitle', fontName='Helvetica-Bold', fontSize=14, leading=17,
        spaceBefore=2, spaceAfter=2, textColor=DEFAULT_HEADER_BG,
    ))
    ss.add(ParagraphStyle(
        name='BandNote', fontName='Helvetica-Oblique', fontSize=8,
        textColor=colors.HexColor('#7F7F7F'), spaceAfter=5,
    ))
    ss.add(ParagraphStyle(
        name='TOCEntry', fontName='Helvetica', fontSize=10.5, leading=18,
        leftIndent=6,
    ))
    return ss


# ─────────────────────────────────────────────────────────────────────────────
# Cover-page chart
# ─────────────────────────────────────────────────────────────────────────────

def _chart_networth(rows, outpath):
    yrs = [r.get('year') for r in rows]
    nw = [r.get('total_nw', 0) for r in rows]
    plt.figure(figsize=(10, 3.3))
    plt.plot(yrs, nw, linewidth=2, color='#1F3864')
    plt.title('Projected Net Worth')
    plt.xlabel('Year')
    plt.ylabel('Net Worth ($)')
    plt.gca().yaxis.set_major_formatter(lambda v, _pos: f'${v/1_000_000:.1f}M' if abs(v) >= 1_000_000 else f'${v:,.0f}')
    plt.tight_layout()
    plt.savefig(outpath, dpi=170)
    plt.close()


def _money(v):
    if not isinstance(v, (int, float)):
        return str(v)
    if abs(v) >= 1_000_000:
        return f"${v/1_000_000:.2f}M"
    if abs(v) >= 1_000:
        return f"${v/1_000:.0f}K"
    return f"${v:,.0f}"


# ─────────────────────────────────────────────────────────────────────────────
# Cell formatting — read the source workbook's own number_format so the PDF
# shows the same $ / % / integer rendering as Excel, rather than raw floats.
# ─────────────────────────────────────────────────────────────────────────────

def _fmt_cell(cell):
    v = cell.value
    if v is None:
        return ''
    if isinstance(v, str):
        if v.startswith('='):
            return ''  # formula source text isn't a computed value; omit rather than confuse
        return v
    if isinstance(v, bool):
        return 'YES' if v else 'NO'
    if isinstance(v, (int, float)):
        fmt = (cell.number_format or '').lower()
        if '%' in fmt:
            return f"{v * 100:,.1f}%"
        if '$' in fmt:
            if v == 0:
                return '-'
            s = f"${abs(v):,.0f}"
            return f"({s})" if v < 0 else s
        if fmt.strip() == '0':
            return f"{int(v)}"
        if '#,##0' in fmt and '.' not in fmt:
            return f"{v:,.0f}"
        if float(v).is_integer():
            return f"{int(v):,}"
        return f"{v:,.2f}"
    return str(v)


_ALIGN_MAP = {'left': 'LEFT', 'right': 'RIGHT', 'center': 'CENTER', 'general': 'LEFT'}


def _cell_align(cell):
    h = (cell.alignment.horizontal if cell.alignment else None) or 'left'
    return _ALIGN_MAP.get(h, 'LEFT')


def _cell_colors(cell, default_bg=None, default_fg=None):
    """Best-effort read of a cell's real fill/font color; falls back to the
    given defaults for unresolved theme/indexed colors."""
    bg = default_bg
    fg = default_fg
    try:
        if cell.fill and cell.fill.patternType and cell.fill.fgColor and cell.fill.fgColor.type == 'rgb':
            rgb = cell.fill.fgColor.rgb
            if isinstance(rgb, str) and len(rgb) == 8 and rgb != '00000000':
                bg = colors.HexColor('#' + rgb[2:])
    except Exception:
        pass
    try:
        if cell.font and cell.font.color and cell.font.color.type == 'rgb':
            rgb = cell.font.color.rgb
            if isinstance(rgb, str) and len(rgb) == 8:
                fg = colors.HexColor('#' + rgb[2:])
    except Exception:
        pass
    return bg, fg


def _is_styled_row(ws, row_idx, max_col):
    """A bold + filled first cell marks a header or a mid-sheet section
    divider (e.g. a "HOW TO USE" banner) that should keep its own color
    rather than blend into the plain data-row zebra striping."""
    cell = ws.cell(row_idx, 1)
    bold = bool(cell.font and cell.font.bold)
    filled = bool(cell.fill and cell.fill.patternType)
    if bold and filled:
        return True
    # Some section rows are only bold with no explicit fill (plain white
    # section labels); still worth keeping their bold weight, but they don't
    # need special coloring.
    return False


# ─────────────────────────────────────────────────────────────────────────────
# Column banding
# ─────────────────────────────────────────────────────────────────────────────

def _column_bands(ncols, id_cols=ID_COLS, max_cols=MAX_COLS_PER_BAND):
    if ncols <= max_cols:
        return [list(range(1, ncols + 1))]
    id_cols = min(id_cols, ncols)
    data_budget = max(1, max_cols - id_cols)
    bands = []
    col = id_cols + 1
    while col <= ncols:
        end = min(ncols, col + data_budget - 1)
        bands.append(list(range(1, id_cols + 1)) + list(range(col, end + 1)))
        col = end + 1
    return bands


def _header_row_count(ws):
    """How many leading rows form this sheet's repeating header: the merged
    colored group-header band (e.g. INCOME / TAX & RMD), if present, plus the
    column-label row beneath it."""
    for mr in ws.merged_cells.ranges:
        if mr.min_row == 1 and mr.min_col == 1 and (mr.max_col - mr.min_col + 1) >= 3:
            return 2
    return 1


# ─────────────────────────────────────────────────────────────────────────────
# Table construction for one column band of one sheet
# ─────────────────────────────────────────────────────────────────────────────

def _band_table(ws, band_cols, header_rows, max_row, band_width):
    ncols = len(band_cols)
    fs = 8.0 if ncols <= 6 else (7.3 if ncols <= 11 else 6.4)
    wrap_style = ParagraphStyle('cellwrap', fontName='Helvetica', fontSize=fs, leading=fs + 2)

    src_widths = []
    for c in band_cols:
        dim = ws.column_dimensions.get(get_column_letter(c))
        src_widths.append(dim.width if dim and dim.width else 10.0)
    total_src = sum(src_widths) or 1.0
    col_widths = [(w / total_src) * band_width for w in src_widths]

    data = []
    row_indices = list(range(1, header_rows + 1)) + list(range(header_rows + 1, max_row + 1))
    for r in row_indices:
        row_vals = []
        for c in band_cols:
            cell = ws.cell(r, c)
            text = _fmt_cell(cell)
            if cell.alignment and cell.alignment.wrap_text and text:
                row_vals.append(Paragraph(text.replace('\n', '<br/>'), wrap_style))
            else:
                row_vals.append(text)
        data.append(row_vals)

    t = Table(data, colWidths=col_widths, repeatRows=header_rows)
    style = [
        ('FONTSIZE', (0, 0), (-1, -1), fs),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.25, GRID_COLOR),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
    ]

    for i, r in enumerate(row_indices):
        is_header = i < header_rows
        styled = is_header or _is_styled_row(ws, r, len(band_cols))
        if styled:
            for j, c in enumerate(band_cols):
                cell = ws.cell(r, c)
                bg, fg = _cell_colors(cell, DEFAULT_HEADER_BG if is_header else None,
                                       DEFAULT_HEADER_FG if is_header else None)
                if bg is not None:
                    style.append(('BACKGROUND', (j, i), (j, i), bg))
                if fg is not None:
                    style.append(('TEXTCOLOR', (j, i), (j, i), fg))
                style.append(('FONTNAME', (j, i), (j, i), 'Helvetica-Bold'))
                style.append(('ALIGN', (j, i), (j, i), 'CENTER' if is_header else _cell_align(cell)))
        else:
            if (i - header_rows) % 2 == 1:
                style.append(('BACKGROUND', (0, i), (-1, i), ZEBRA))
            for j, c in enumerate(band_cols):
                style.append(('ALIGN', (j, i), (j, i), _cell_align(ws.cell(r, c))))

    t.setStyle(TableStyle(style))
    return t


def _sheet_flowables(ws, styles, band_note_needed):
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1
    if max_row < 1 or max_col < 1 or ws.cell(1, 1).value is None and max_row == 1 and max_col == 1:
        return [Paragraph('(No data on this sheet.)', styles['BandNote'])]

    header_rows = min(_header_row_count(ws), max_row)
    bands = _column_bands(max_col)
    flow = []
    for bi, band_cols in enumerate(bands):
        if band_note_needed and len(bands) > 1:
            first_data_col = band_cols[ID_COLS] if len(band_cols) > ID_COLS else band_cols[0]
            flow.append(Paragraph(
                f"Columns {get_column_letter(first_data_col)}–{get_column_letter(band_cols[-1])} "
                f"of {get_column_letter(max_col)} &nbsp;(band {bi + 1} of {len(bands)})",
                styles['BandNote'],
            ))
        flow.append(_band_table(ws, band_cols, header_rows, max_row, CONTENT_W))
        if bi < len(bands) - 1:
            flow.append(Spacer(1, 0.12 * inch))
    return flow


# ─────────────────────────────────────────────────────────────────────────────
# Document template: TOC-aware, page-numbered footer
# ─────────────────────────────────────────────────────────────────────────────

class _PlanDocTemplate(BaseDocTemplate):
    def afterFlowable(self, flowable):
        if isinstance(flowable, Paragraph) and flowable.style.name == 'SheetTitle':
            self.notify('TOCEntry', (0, flowable.getPlainText(), self.page))


class _NumberedCanvas(Canvas):
    """Defers footer drawing until save(), when the true total page count is
    known, so every page can show 'Page N of TOTAL' instead of just 'Page N'."""

    def __init__(self, *args, **kwargs):
        Canvas.__init__(self, *args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        total = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self._draw_footer(total)
            Canvas.showPage(self)
        Canvas.save(self)

    def _draw_footer(self, total):
        self.saveState()
        self.setFont('Helvetica', 7)
        self.setFillColor(colors.HexColor('#595959'))
        self.drawString(MARGIN, MARGIN, 'Retirement Plan — Confidential')
        self.drawRightString(PAGE_W - MARGIN, MARGIN, f"Page {self._pageNumber} of {total}")
        self.restoreState()


# ─────────────────────────────────────────────────────────────────────────────
# Entry point
# ─────────────────────────────────────────────────────────────────────────────

def build_enterprise_pdf(wb, c, rows, mc_data, out_path='output/retirement_plan.pdf'):
    """Render the full workbook `wb` (every visible sheet) to a landscape PDF.

    `c`/`rows`/`mc_data` are used only for the cover-page KPI summary and net
    worth chart; every other page is generated directly from `wb` so the PDF
    can never diverge from the spreadsheet.
    """
    styles = _styles()

    doc = _PlanDocTemplate(
        out_path,
        pagesize=landscape(letter),
        leftMargin=MARGIN, rightMargin=MARGIN, topMargin=MARGIN, bottomMargin=BOTTOM_MARGIN,
        title='Retirement Plan', author='Retirement Plan System',
    )
    frame = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id='F1')
    doc.addPageTemplates([PageTemplate(id='main', frames=[frame])])

    story = []

    # ── Cover page ──────────────────────────────────────────────────────────
    story.append(Paragraph('Institutional Retirement Plan', styles['Title']))
    story.append(Spacer(1, 0.12 * inch))
    story.append(Paragraph(
        'This report reproduces every sheet of the retirement planning workbook — sustainability, '
        'tax strategy, portfolio structure, lifetime cash flow, estate outcomes, and Monte Carlo '
        'simulation results — across the full planning horizon, formatted for landscape printing.',
        styles['BodyText'],
    ))
    story.append(Spacer(1, 0.2 * inch))

    final_nw = rows[-1].get('total_nw', 0) if rows else 0
    lifetime_tax = sum(r.get('total_tax', 0) for r in rows) if rows else 0
    success = mc_data.get('success_rate', 0) if mc_data else 0
    ci_low = mc_data.get('success_rate_ci_low', success) if mc_data else 0
    ci_high = mc_data.get('success_rate_ci_high', success) if mc_data else 0

    # The Monte Carlo KPI is shown only when that module ran; with it off
    # mc_data is {} and this would read a misleading 0%.
    _mc_on = (
        os.environ.get('RETIREMENT_SYSTEM_FORCE_ALL_MODULES') == '1'
        or (bool((c.get('opt') or {}).get('market_luck_stress_test', True)) if c else True)
    )
    kpi_data = [
        ['Metric', 'Value'],
        ['Terminal Net Worth', _money(final_nw)],
        ['Lifetime Taxes', _money(lifetime_tax)],
    ]
    if _mc_on:
        kpi_data.append(
            ['Monte Carlo Success', f"{success * 100:.1f}% (95% CI {ci_low * 100:.1f}%–{ci_high * 100:.1f}%)"]
        )
    kpi_data.append(['Plan Horizon', f"{len(rows)} years" if rows else 'N/A'])
    kpi_table = Table(kpi_data, colWidths=[2.6 * inch, 3.2 * inch])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), DEFAULT_HEADER_BG),
        ('TEXTCOLOR', (0, 0), (-1, 0), DEFAULT_HEADER_FG),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9.5),
        ('GRID', (0, 0), (-1, -1), 0.25, GRID_COLOR),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, ZEBRA]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 0.22 * inch))

    if rows:
        tmp_fd, tmp_name = tempfile.mkstemp(suffix='.png')
        os.close(tmp_fd)
        tmp_chart = Path(tmp_name)
        _chart_networth(rows, tmp_chart)
        story.append(Image(str(tmp_chart), width=9.6 * inch, height=3.2 * inch))
    else:
        tmp_chart = None

    story.append(PageBreak())

    # ── Table of contents ───────────────────────────────────────────────────
    story.append(Paragraph('Table of Contents', styles['Heading1']))
    story.append(HRFlowable(width='100%'))
    story.append(Spacer(1, 0.1 * inch))
    toc = TableOfContents()
    toc.levelStyles = [styles['TOCEntry']]
    story.append(toc)
    story.append(PageBreak())

    # ── One section per visible worksheet ───────────────────────────────────
    for ws in wb.worksheets:
        if getattr(ws, 'sheet_state', 'visible') != 'visible':
            continue
        story.append(Paragraph(ws.title, styles['SheetTitle']))
        story.append(HRFlowable(width='100%', color=DEFAULT_HEADER_BG, thickness=1.1))
        story.append(Spacer(1, 0.06 * inch))
        story.extend(_sheet_flowables(ws, styles, band_note_needed=True))
        story.append(PageBreak())

    # Drop the trailing page break so we don't emit a blank final page.
    while story and isinstance(story[-1], PageBreak):
        story.pop()

    doc.multiBuild(story, canvasmaker=_NumberedCanvas)

    if tmp_chart is not None:
        tmp_chart.unlink(missing_ok=True)

    print(f"PDF created: {out_path}")
