"""Charts Dashboard sheet builder (Sheet 8).

Displays visual charts and dashboard:
- Net Worth by component (stacked bar chart)
- Cash Flow Income & Portfolio Draws (stacked bar chart with detail series)
- Cash Flow Spending & Taxes (stacked bar chart with detail series)
- Monte Carlo Percentile Bands (line chart for P10/P25/P50/P75/P90)
- Asset Allocation Before/After (pie charts with target rebalancing)

Chart source data is stored on a hidden helper sheet; visible workbook page
is optimized for reviewing charts instead of helper tables.
"""

import math
from .workbook_common import *
from .. import allocation_policy as _ap


def build_sheet8(ws, c, rows, mc_data=None):
    """Charts Dashboard — visual charts on a clean dashboard sheet.

    The chart source ranges are still required for Excel chart generation, but
    they are written to a hidden helper sheet so the visible workbook page and
    the UI result explorer are optimized for reviewing charts instead of helper
    tables.
    """
    chart_ws = ws
    chart_ws.sheet_view.showGridLines = False
    wb = chart_ws.parent
    data_title = '_Chart Dashboard Data'
    if data_title in wb.sheetnames:
        del wb[data_title]
    data_ws = wb.create_sheet(data_title)
    data_ws.sheet_state = 'hidden'
    data_ws.sheet_view.showGridLines = False

    title = f'CHARTS DASHBOARD — {c.get("h_nick") or c["h_name"]} & {c.get("w_nick") or c.get("w_name","")}  ·  {c["plan_start"]}–{c["plan_end"]}'
    chart_ws.cell(row=1, column=1, value=title)
    chart_ws['A1'].fill = fill(NAVY)
    chart_ws['A1'].font = Font(name='Arial', bold=True, color=WHITE, size=14)
    chart_ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    chart_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=14)
    chart_ws.row_dimensions[1].height = 24
    chart_ws.cell(row=2, column=1, value='Chart source data is stored on a hidden helper sheet. Use this page and the app explorer for visuals; use detailed result sheets for tabular audits.')
    chart_ws['A2'].font = Font(name='Arial', italic=True, color='666666', size=10)
    chart_ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=14)
    for col in range(1, 15):
        chart_ws.column_dimensions[get_column_letter(col)].width = 14

    ws = data_ws

    # ── helpers ──────────────────────────────────────────────────────────────
    def wh(r, c, text, bg=NAVY, fg=WHITE, bold=True, size=10, span=1):
        cell = ws.cell(row=r, column=c, value=text)
        cell.fill = fill(bg); cell.font = Font(name='Arial',bold=bold,color=fg,size=size)
        cell.alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
        cell.border = thin_border()
        if span > 1:
            ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c+span-1)

    def wd(r, c, val, fmt=None, bg=None, bold=False):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = Font(name='Arial', size=10, bold=bold)
        cell.alignment = Alignment(
            horizontal='right' if isinstance(val,(int,float)) else 'center', vertical='center')
        cell.border = thin_border()
        if bg:  cell.fill = fill(bg)
        if fmt: cell.number_format = fmt

    def sec(r, text, bg=NAVY, span=40):
        cell = ws.cell(row=r, column=1, value=text)
        cell.fill = fill(bg); cell.font = Font(name='Arial',bold=True,color=WHITE,size=12)
        cell.alignment = Alignment(horizontal='left',vertical='center')
        ws.row_dimensions[r].height = 22
        if span > 1:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=span)

    # ── pre-compute per-year values ───────────────────────────────────────────
    # Income = income streams + positive portfolio draws only
    # Expense = spending + taxes + payroll-tax residual (≤$35K gap → real cost)
    # Surplus = visual gap (income bar taller than expense bar); NOT a series
    nw_rows = rows  # alias — rows already have all NW fields

    per_year = []
    for row in rows:
        streams = {
            'earned':  round(row.get('earned',0)),
            'h_ss':    round(row.get('h_ss',0)),
            'w_ss':    round(row.get('w_ss',0)),
            'pension': round(row.get('pension',0)),
            'w_sgl':   round(row.get('wife_single_ann',0)),
            'w_jnt':   round(row.get('wife_joint_ann',0)),
            'h_sgl':   round(row.get('h_single_ann',0)),
            'h_jnt':   round(row.get('h_joint_ann',0)),
            'note':    round(row.get('note_princ',0) + row.get('note_int',0)),
            'rmd':     round(row.get('rmd_total',0)),
        }
        trust_wd = max(0, round(row.get('trust_wd',0)))
        hsa_wd   = max(0, round(row.get('hsa_wd',0)))
        roth_wd  = max(0, round(row.get('roth_wd',0)))
        ira_wd   = max(0, round(row.get('ira_wd',0)))
        heloc_draw_wd = max(0, round(row.get('heloc_draw',0)))
        inc_total = sum(streams.values()) + trust_wd + hsa_wd + roth_wd + ira_wd + heloc_draw_wd

        spend_base  = round(row.get('spend_base_yr',0))
        housing     = round(row.get('housing_total_yr', row.get('mortgage',0) + row.get('rent_yr', 0)))
        wellness    = round(row.get('wellness_base_yr', 0) + row.get('ltc_prem_yr', 0))
        travel      = round(row.get('rec_extra',0))
        other       = round(row.get('lump',0))
        heloc_pai_chart = round(row.get('heloc_interest', 0) + row.get('heloc_repayment_principal', 0))
        fed_tax    = round(row.get('fed_tax',0))
        state_tax  = round(row.get('state_tax',0))
        niit       = round(row.get('niit',0))
        exp_raw    = spend_base + housing + wellness + travel + other + heloc_pai_chart + fed_tax + state_tax + niit
        gap        = inc_total - exp_raw
        # Payroll-tax residual: only add to expense when gap is a real cost, not surplus
        payroll_res = round(gap) if 0 < gap <= 35000 else 0

        # Real-dollar equivalent (deflated to plan_start year)
        cpi_deflator = (1 + c['inf']) ** (row['year'] - c['plan_start'])
        per_year.append({
            'real_nw': round(row.get('total_nw', 0) / cpi_deflator) if cpi_deflator > 0 else 0,
            'yr': row['year'],
            'earned': streams['earned'], 'h_ss': streams['h_ss'], 'w_ss': streams['w_ss'],
            'pension': streams['pension'], 'w_sgl': streams['w_sgl'], 'w_jnt': streams['w_jnt'],
            'h_sgl': streams['h_sgl'], 'h_jnt': streams['h_jnt'],
            'note': streams['note'], 'rmd': streams['rmd'],
            'trust_wd': trust_wd, 'hsa_wd': hsa_wd, 'roth_wd': roth_wd,
            'ira_wd': ira_wd, 'heloc_draw_wd': heloc_draw_wd, 'inc_total': inc_total,
            'spend_base': spend_base, 'housing': housing, 'wellness': wellness,
            'travel': travel, 'other': other, 'heloc_pai': heloc_pai_chart,
            'fed_tax': fed_tax, 'state_tax': state_tax,
            'niit': niit, 'payroll_res': payroll_res,
            'exp_total': exp_raw + payroll_res,
            # NW components — home equity separated from other
            'ann_nw':    round(row.get('ann_nw',0)),
            'pretax_nw': round(row.get('pretax_nw',0)),
            'roth_nw':   round(row.get('roth_nw',0)),
            'trust_nw':  round(row.get('trust_nw',0)),
            'hsa_nw':    round(row.get('hsa_nw',0)),
            'home_eq_nw':round(row.get('home_equity',0)),
            'other_nw':  round(row.get('other_nw',0) - row.get('home_equity',0)),
        })

    n = len(per_year)
    # Cash flow scaling: round to nearest $100K, tick every $100K, display in $K
    CF_YMAX = math.ceil(max(max(p['inc_total'] for p in per_year),
                             max(p['exp_total'] for p in per_year)) / 100000) * 100000
    CF_UNIT = 100_000

    # ════════════════════════════════════════════════════════════════════════
    # DATA TABLES (3 side-by-side blocks)
    # NW:  cols  1-8   (A-H)    Year + 6 NW series + Total
    # INC: cols 10-26  (J-Z)    Year + 15 income series + Σ
    # EXP: cols 28-37  (AB-AK)  Year + 8 expense series + Σ
    # Gap cols: 9 (I) and 27 (AA)
    # ════════════════════════════════════════════════════════════════════════
    TABLE_ROW  = 4
    DATA_FIRST = TABLE_ROW + 1
    DATA_LAST  = TABLE_ROW + n

    sec(1, f'CHARTS DASHBOARD — {c.get("h_nick") or c["h_name"]} & {c.get("w_nick") or c.get("w_name","")}  ·  {c["plan_start"]}–{c["plan_end"]}', span=50)

    # ── NW Table ──────────────────────────────────────────────────────────────
    NW_YEAR = 1
    NW_SER  = [
        (2, 'Annuities & Pension', '2E75B6', 'ann_nw'),
        (3, 'Pre-Tax IRA/401k',    'C55A11', 'pretax_nw'),
        (4, 'Roth',                '0891B2', 'roth_nw'),
        (5, 'Trust',               '5A3E85', 'trust_nw'),
        (6, 'HSA',                 'C9A84C', 'hsa_nw'),
        (7, 'Home Equity',         '16A34A', 'home_eq_nw'),
        (8, 'Other Assets',        '6B7280', 'other_nw'),
    ]
    # Title for NW section — span exactly cols 1-9
    cell = ws.cell(row=2, column=1, value=f'Net Worth by Component  ·  {c["plan_start"]}–{c["plan_end"]}')
    cell.fill = fill(BLUE); cell.font = Font(name='Arial',bold=True,color=WHITE,size=11)
    cell.alignment = Alignment(horizontal='left',vertical='center')
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=9)

    wh(TABLE_ROW, NW_YEAR, 'Year', bg=DGRAY)
    for col, lbl, clr, _ in NW_SER:
        wh(TABLE_ROW, col, lbl, bg=DGRAY)
    wh(TABLE_ROW, 9, 'Total NW', bg=NAVY)

    for ri, p in enumerate(per_year):
        r = DATA_FIRST + ri; bg = GRAY if ri%2==0 else None
        wd(r, NW_YEAR, p['yr'], fmt=FMT_YEAR, bg=bg)
        for col, _, _, key in NW_SER:
            wd(r, col, p[key], fmt=FMT_DOLLAR, bg=bg)
        wd(r, 9, p['ann_nw']+p['pretax_nw']+p['roth_nw']+p['trust_nw']+p['hsa_nw']+p['home_eq_nw']+p['other_nw'],
           fmt=FMT_DOLLAR, bg=bg, bold=True)

    ws.column_dimensions[get_column_letter(1)].width = 7
    for _col in range(2, 10): ws.column_dimensions[get_column_letter(_col)].width = 14
    ws.column_dimensions[get_column_letter(10)].width = 2  # gap

    # ── Income Table ──────────────────────────────────────────────────────────
    INC_YEAR = 11
    INC_SER  = [
        (12, 'earned',   'Earned Income',         '1F3864'),
        (13, 'h_ss',     f'{c.get("h_nick") or c.get("h_name") or "Member 1"} SS',       '2E75B6'),
        (14, 'w_ss',     f'{c.get("w_nick") or c.get("w_name") or "Member 2"} SS',       '3D9AB8'),
        (15, 'pension',  'Pension',               'C9A84C'),
        (16, 'w_sgl',    f'{c.get("w_nick") or c.get("w_name") or "Member 2"} Single Ann', '2D6A4F'),
        (17, 'w_jnt',    f'{c.get("w_nick") or c.get("w_name") or "Member 2"} Joint Ann',  '40916C'),
        (18, 'h_sgl',    f'{c.get("h_nick") or c.get("h_name") or "Member 1"} Single Ann', 'C55A11'),
        (19, 'h_jnt',    f'{c.get("h_nick") or c.get("h_name") or "Member 1"} Joint Ann',  'E07540'),
        (20, 'note',     'Note P+I',              '5A3E85'),
        (21, 'rmd',      'RMD',                   '9B2335'),
        (22, 'trust_wd', 'Trust Draw',            '7B3F9E'),
        (23, 'hsa_wd',   'HSA Draw',              '1B7A9E'),
        (24, 'roth_wd',  'Roth Draw',             '156041'),
        (25, 'ira_wd',   'IRA Draw',              'B85C00'),
        (26, 'heloc_draw_wd', 'HELOC Draw',         '8B5E3C'),
    ]
    # Income section title — span exactly cols 11-27
    cell = ws.cell(row=2, column=INC_YEAR, value=f'Cash Flow — Income & Portfolio Draws  ·  {c["plan_start"]}–{c["plan_end"]}')
    cell.fill = fill(GREEN); cell.font = Font(name='Arial',bold=True,color=WHITE,size=11)
    cell.alignment = Alignment(horizontal='left',vertical='center')
    ws.merge_cells(start_row=2, start_column=INC_YEAR, end_row=2, end_column=27)

    wh(TABLE_ROW, INC_YEAR, 'Year', bg=DGRAY)
    for col, key, lbl, clr in INC_SER:
        wh(TABLE_ROW, col, lbl, bg=DGRAY)
    wh(TABLE_ROW, 27, 'Σ Income', bg=GREEN)

    for ri, p in enumerate(per_year):
        r = DATA_FIRST + ri; bg = GRAY if ri%2==0 else None
        wd(r, INC_YEAR, p['yr'], fmt=FMT_YEAR, bg=bg)
        for col, key, _, _ in INC_SER:
            wd(r, col, p[key], fmt=FMT_DOLLAR, bg=bg)
        wd(r, 27, p['inc_total'], fmt=FMT_DOLLAR, bg=bg, bold=True)

    for _col in range(11, 28): ws.column_dimensions[get_column_letter(_col)].width = 13
    ws.column_dimensions[get_column_letter(11)].width = 7
    ws.column_dimensions[get_column_letter(28)].width = 2  # gap

    # ── Expense Table ─────────────────────────────────────────────────────────
    EXP_YEAR = 29
    EXP_SER  = [
        (30, 'spend_base',  'Base Spending',    '1F3864'),
        (31, 'housing',     'Housing',          '2E75B6'),
        (32, 'wellness',    'Wellness',       'C55A11'),
        (33, 'travel',      'Travel',           'C9A84C'),
        (34, 'other',       'Other',            '059669'),
        (35, 'fed_tax',     'Federal Tax',      '9B2335'),
        (36, 'state_tax',   f'State Tax ({c["state"][:2]})',   'C5384E'),
        (37, 'niit',        'NIIT',             'E07595'),
        (38, 'payroll_res', 'Payroll Tax',      '595959'),
        (39, 'heloc_pai',   'HELOC P&I',        '2D6A8F'),
    ]
    # Expense section title — span exactly cols 29-40
    cell = ws.cell(row=2, column=EXP_YEAR, value=f'Cash Flow — Spending & Taxes  ·  {c["plan_start"]}–{c["plan_end"]}')
    cell.fill = fill(RED); cell.font = Font(name='Arial',bold=True,color=WHITE,size=11)
    cell.alignment = Alignment(horizontal='left',vertical='center')
    ws.merge_cells(start_row=2, start_column=EXP_YEAR, end_row=2, end_column=40)

    wh(TABLE_ROW, EXP_YEAR, 'Year', bg=DGRAY)
    for col, key, lbl, clr in EXP_SER:
        wh(TABLE_ROW, col, lbl, bg=DGRAY)
    wh(TABLE_ROW, 40, 'Σ Spend+Tax', bg=RED)

    for ri, p in enumerate(per_year):
        r = DATA_FIRST + ri; bg = GRAY if ri%2==0 else None
        wd(r, EXP_YEAR, p['yr'], fmt=FMT_YEAR, bg=bg)
        for col, key, _, _ in EXP_SER:
            wd(r, col, p[key], fmt=FMT_DOLLAR, bg=bg)
        wd(r, 40, p['exp_total'], fmt=FMT_DOLLAR, bg=bg, bold=True)

    for _col in range(29, 41): ws.column_dimensions[get_column_letter(_col)].width = 14
    ws.column_dimensions[get_column_letter(29)].width = 7

    # ════════════════════════════════════════════════════════════════════════
    # CHARTS
    # - Legend RIGHT (position='r', overlay=False) — x-axis labels always visible
    # - gapWidth=20, overlap=100 — wide solid stacked bars
    # - 30cm x 18cm each; anchored at rows 38, 83, 128 (45-row gaps)
    # - Title font 24pt bold applied via post-save XML patch in main()
    # ════════════════════════════════════════════════════════════════════════
    def make_chart(title_str, year_col, series_list, data_first, data_last,
                   ymax=None, yunit=None, yfmt='$#,##0,,"M"'):
        chart = BarChart()
        chart.type = 'col'; chart.grouping = 'stacked'
        chart.gapWidth = 20; chart.overlap = 100
        chart.width = 30; chart.height = 18
        chart.title = title_str
        chart.y_axis.numFmt = yfmt
        chart.y_axis.delete = False
        chart.y_axis.tickLblPos = 'nextTo'
        if ymax:
            chart.y_axis.scaling.max = ymax
            chart.y_axis.scaling.min = 0
        if yunit:
            chart.y_axis.majorUnit = yunit
        chart.x_axis.delete = False
        chart.x_axis.numFmt = '0'
        chart.x_axis.tickLblPos = 'low'
        chart.legend.position = 'r'  # RIGHT — never overlaps x-axis labels
        chart.legend.overlay  = False
        cats = Reference(ws, min_col=year_col, max_col=year_col,
                         min_row=data_first, max_row=data_last)
        for idx, (dcol, lbl, color) in enumerate(series_list):
            chart.add_data(Reference(ws, min_col=dcol, max_col=dcol,
                                      min_row=data_first, max_row=data_last))
            s = chart.series[idx]
            s.title = SeriesLabel(v=lbl)
            gp = GraphicalProperties(); gp.solidFill = color; s.spPr = gp
        chart.set_categories(cats)
        chart.visible_cells_only = False
        return chart

    nw_series  = [(col, lbl, clr) for col, lbl, clr, _ in NW_SER]
    inc_series = [(col, lbl, clr) for col, key, lbl, clr in INC_SER]
    exp_series = [(col, lbl, clr) for col, key, lbl, clr in EXP_SER]

    chart_nw  = make_chart(f'Net Worth by Component  ·  {c["plan_start"]}–{c["plan_end"]}',
                             NW_YEAR, nw_series, DATA_FIRST, DATA_LAST)
    chart_inc = make_chart(f'Cash Flow — Income & Portfolio Draws  ·  {c["plan_start"]}–{c["plan_end"]}',
                             INC_YEAR, inc_series, DATA_FIRST, DATA_LAST,
                             ymax=CF_YMAX, yunit=CF_UNIT, yfmt='$#,##0,"K"')
    chart_exp = make_chart(f'Cash Flow — Spending & Taxes  ·  {c["plan_start"]}–{c["plan_end"]}',
                             EXP_YEAR, exp_series, DATA_FIRST, DATA_LAST,
                             ymax=CF_YMAX, yunit=CF_UNIT, yfmt='$#,##0,"K"')

    # Anchor charts: row 38, 83, 128 (45-row gaps for breathing room)
    chart_ws.add_chart(chart_nw,  'A4')
    chart_ws.add_chart(chart_inc, 'A42')
    chart_ws.add_chart(chart_exp, 'A80')

    # ── Chart 4: Net Worth Percentile Bands (Monte Carlo) ────────────────────
    # Write data table for the MC percentile bands (cols AO onward = col 41+)
    MC_YEAR_COL = 43  # col AQ
    MC_PCTS     = [10, 25, 50, 75, 90]
    MC_LABELS   = ['P10','P25','P50 Median','P75','P90']
    MC_COLORS   = ['9B2335','C55A11','2D6A4F','2E75B6','1F3864']

    # Section header
    cell_mc = ws.cell(row=2, column=MC_YEAR_COL,
                      value='MC Percentile Bands Data  ·  used by Chart 4')
    cell_mc.fill = fill('7030A0')
    cell_mc.font = Font(name='Arial', bold=True, color=WHITE, size=11)
    cell_mc.alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(start_row=2, start_column=MC_YEAR_COL,
                   end_row=2, end_column=MC_YEAR_COL + len(MC_PCTS))

    wh(TABLE_ROW, MC_YEAR_COL, 'Year', bg=DGRAY)
    for i, (p, lbl) in enumerate(zip(MC_PCTS, MC_LABELS)):
        wh(TABLE_ROW, MC_YEAR_COL + 1 + i, lbl, bg=DGRAY)

    MC_DATA_FIRST = DATA_FIRST
    MC_DATA_LAST  = DATA_LAST

    for ri, p in enumerate(per_year):
        r_mc = DATA_FIRST + ri
        yr   = p['yr']
        mc_row = (mc_data or {}).get('pct_by_year', {}).get(yr, {})
        bg_alt = GRAY if ri%2==0 else None
        wd(r_mc, MC_YEAR_COL, yr, fmt=FMT_YEAR, bg=bg_alt)
        for j, pct_key in enumerate(MC_PCTS):
            wd(r_mc, MC_YEAR_COL + 1 + j, mc_row.get(pct_key, 0), fmt=FMT_DOLLAR, bg=bg_alt)

    for col in range(MC_YEAR_COL, MC_YEAR_COL + len(MC_PCTS) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 14

    # Build line chart for percentile bands
    from openpyxl.chart import LineChart
    chart_mc = LineChart()
    chart_mc.title  = 'Net Worth Percentile Bands  ·  Monte Carlo Sims  ·  Plan Horizon'
    chart_mc.width  = 30
    chart_mc.height = 18
    chart_mc.y_axis.numFmt = '$#,##0,,"M"'
    chart_mc.y_axis.delete = False
    chart_mc.x_axis.numFmt = '0'
    chart_mc.x_axis.tickLblPos = 'low'
    chart_mc.legend.position   = 'r'
    chart_mc.legend.overlay    = False

    cats_mc = Reference(ws, min_col=MC_YEAR_COL, max_col=MC_YEAR_COL,
                        min_row=MC_DATA_FIRST, max_row=MC_DATA_LAST)
    for i, (lbl, clr) in enumerate(zip(MC_LABELS, MC_COLORS)):
        data_ref = Reference(ws, min_col=MC_YEAR_COL + 1 + i, max_col=MC_YEAR_COL + 1 + i,
                             min_row=MC_DATA_FIRST, max_row=MC_DATA_LAST)
        chart_mc.add_data(data_ref)
        s = chart_mc.series[i]
        s.title = SeriesLabel(v=lbl)
        gp = GraphicalProperties(); gp.solidFill = clr; s.spPr = gp
    chart_mc.set_categories(cats_mc)
    chart_mc.visible_cells_only = False

    chart_ws.add_chart(chart_mc, 'A118')

    # ── Allocation Pie Charts (Before & After) ────────────────────────────
    chart_data = c.get('_alloc_chart_data', {})
    _pie_buckets = chart_data.get('buckets', [])
    _pie_before = chart_data.get('before_vals', [])
    _pie_after = chart_data.get('after_vals', [])

    if _pie_buckets:
        pie_start = DATA_LAST + 175

        from openpyxl.chart.data_source import NumData, NumVal, StrData, StrRef, StrVal
        from openpyxl.chart.label import DataLabelList
        from openpyxl.chart.series import DataPoint

        # Consistent colors per category, even though each pie now has its own
        # filtered source range.  The filtering is intentional: Excel pie
        # charts include zero-value rows in labels/legends unless those rows are
        # excluded from the source data entirely.
        # Distinct hues (not shades of the same color), assigned by position in
        # the full bucket list before value-filtering. 20 colors (tab10 + its
        # lighter tints) rather than 10 — the household's full bucket list
        # (asset classes + Cash + non-liquid coverage rows) commonly runs to
        # 13-16 entries, and a 10-color palette wrapped around and reused the
        # same color twice within one pie.
        PIE_COLORS = ['1F77B4', 'D62728', '2CA02C', 'FF7F0E', '9467BD',
                      '8C564B', 'E377C2', '7F7F7F', 'BCBD22', '17BECF',
                      'AEC7E8', 'FFBB78', '98DF8A', 'FF9896', 'C5B0D5',
                      'C49C94', 'F7B6D2', 'C7C7C7', 'DBDB8D', '9EDAE5']
        color_by_bucket = {str(bkt): PIE_COLORS[idx % len(PIE_COLORS)]
                           for idx, bkt in enumerate(_pie_buckets)}

        def _positive_pie_rows(values):
            raw_rows = []
            for idx, bkt in enumerate(_pie_buckets):
                val = values[idx] if idx < len(values) else 0
                try:
                    num = float(val or 0)
                except (TypeError, ValueError):
                    num = 0.0
                raw_rows.append((str(bkt), num))
            total = sum(max(0.0, num) for _bkt, num in raw_rows)
            # Drop true zero rows and floating-point dust that would display as
            # a 0% allocation in Excel pie labels.  This keeps real but small
            # allocations while preventing labels for values like $0.09 on a
            # multi-million dollar portfolio.
            min_material_value = max(1.0, total * 0.000001) if total > 0 else 1.0
            return [(bkt, num) for bkt, num in raw_rows if num > min_material_value]

        def _write_pie_source(start_row, start_col, header, rows):
            write_cell(ws, start_row, start_col, 'Asset Class', bold=True)
            write_cell(ws, start_row, start_col + 1, header, bold=True)
            for i, (bkt, val) in enumerate(rows):
                write_cell(ws, start_row + 1 + i, start_col, bkt)
                write_cell(ws, start_row + 1 + i, start_col + 1, val, fmt=FMT_DOLLAR)
            return start_row + len(rows)

        def _make_alloc_pie(title, rows, start_col, anchor):
            if not rows:
                return
            pie_end = _write_pie_source(pie_start, start_col, title, rows)
            p = PieChart()
            p.title = title
            p.style = 10; p.width = 18; p.height = 28  # 46 rows ~ 28cm height
            p.add_data(Reference(ws, min_col=start_col + 1, min_row=pie_start+1, max_row=pie_end), titles_from_data=False)
            p.set_categories(Reference(ws, min_col=start_col, min_row=pie_start+1, max_row=pie_end))
            if p.series:
                p.series[0].cat.numRef = None
                p.series[0].cat.strRef = StrRef(
                    f=f"'{ws.title}'!${get_column_letter(start_col)}${pie_start+1}:${get_column_letter(start_col)}${pie_end}",
                    strCache=StrData(
                        ptCount=len(rows),
                        pt=[StrVal(idx=i, v=str(bkt)) for i, (bkt, _val) in enumerate(rows)],
                    ),
                )
                p.series[0].val.numRef.numCache = NumData(
                    ptCount=len(rows),
                    pt=[NumVal(idx=i, v=float(val)) for i, (_bkt, val) in enumerate(rows)],
                )
            # Labels ON slices: category name + percentage only.  Explicitly
            # disabling the series name prevents Excel from prefixing labels
            # with generated text such as "Series 1,".
            p.dataLabels = DataLabelList()
            p.dataLabels.showSerName = False
            p.dataLabels.showLegendKey = False
            p.dataLabels.showCatName = True
            p.dataLabels.showPercent = True
            p.dataLabels.showVal = False
            p.dataLabels.showLeaderLines = True
            p.legend = None  # labels on slices replace legend
            # Apply consistent colors per category after filtering.
            if p.series:
                for idx, (bkt, _val) in enumerate(rows):
                    pt = DataPoint(idx=idx)
                    pt.graphicalProperties = GraphicalProperties()
                    pt.graphicalProperties.solidFill = color_by_bucket.get(bkt, PIE_COLORS[idx % len(PIE_COLORS)])
                    p.series[0].data_points.append(pt)
            chart_ws.add_chart(p, anchor)

        before_rows = _positive_pie_rows(_pie_before)
        after_rows = _positive_pie_rows(_pie_after)

        # Side by side: Before at column A, After at column P (offset ~15 cols).
        # Pie width was also trimmed from 22cm to 18cm. Slice data labels use
        # leader lines and can overhang a pie chart's own frame, so the previous
        # 22cm-wide charts 11 columns apart visually overlapped; the wider gap
        # plus narrower charts leaves clear separation even with long labels.
        # Source tables are separated as well, so categories with a 0% allocation
        # in one pie do not appear in that pie at all.
        _make_alloc_pie('Current Portfolio Allocation', before_rows, 1, 'A156')
        _make_alloc_pie('Target Portfolio Allocation', after_rows, 4, 'P156')

    # ── Efficient Frontier Scatter Chart ──────────────────────────────────
    # Native Excel scatter of the long-only mean-variance efficient frontier
    # (volatility vs expected return) with the recommended portfolio marked.
    # Source data is written to the hidden helper sheet (ws); the chart renders
    # on the visible Charts page (chart_ws), matching the other charts here.
    _ef_chart_added = False
    try:
        _ef_points = _ao.efficient_frontier(c, n_points=15)
        _ef_stats = _ao.allocation_portfolio_stats(c)
    except Exception:
        _ef_points, _ef_stats = [], None
    try:
        _cur_weights = {bkt: val for bkt, val in zip(_pie_buckets, _pie_before) if bkt in _ao.ASSET_CLASSES}
        _cur_stats = _ao.portfolio_stats_from_weights(c, _cur_weights) if _cur_weights else None
    except Exception:
        _cur_stats = None
    try:
        _max_sharpe_stats = _ao.allocation_portfolio_stats(c, force_mode=_ap.ALLOCATION_MODE_MAX_SHARPE)
    except Exception:
        _max_sharpe_stats = None
    try:
        _tangency_stats = _ao.allocation_portfolio_stats(c, force_mode=_ap.ALLOCATION_MODE_TANGENCY)
    except Exception:
        _tangency_stats = None
    if _ef_points and len(_ef_points) >= 2:
        from openpyxl.chart import ScatterChart, Series
        from openpyxl.chart.marker import Marker
        from openpyxl.drawing.line import LineProperties

        ef_src = DATA_LAST + 230
        write_cell(ws, ef_src, 1, 'Volatility', bold=True)
        write_cell(ws, ef_src, 2, 'Expected Return', bold=True)
        for _i, _p in enumerate(_ef_points):
            write_cell(ws, ef_src + 1 + _i, 1, float(_p.get('volatility', 0.0) or 0.0), fmt=FMT_PCT)
            write_cell(ws, ef_src + 1 + _i, 2, float(_p.get('return', 0.0) or 0.0), fmt=FMT_PCT)
        _ef_src_last = ef_src + len(_ef_points)

        _rec_vol = float(_ef_stats.get('volatility', 0.0) or 0.0) if _ef_stats else None
        _rec_ret = float(_ef_stats.get('expected_return', 0.0) or 0.0) if _ef_stats else None
        if _rec_vol is not None:
            write_cell(ws, ef_src, 4, 'Rec Volatility', bold=True)
            write_cell(ws, ef_src, 5, 'Rec Return', bold=True)
            write_cell(ws, ef_src + 1, 4, _rec_vol, fmt=FMT_PCT)
            write_cell(ws, ef_src + 1, 5, _rec_ret, fmt=FMT_PCT)

        _cur_vol = float(_cur_stats.get('volatility', 0.0) or 0.0) if _cur_stats else None
        _cur_ret = float(_cur_stats.get('expected_return', 0.0) or 0.0) if _cur_stats else None
        if _cur_vol is not None:
            write_cell(ws, ef_src, 7, 'Cur Volatility', bold=True)
            write_cell(ws, ef_src, 8, 'Cur Return', bold=True)
            write_cell(ws, ef_src + 1, 7, _cur_vol, fmt=FMT_PCT)
            write_cell(ws, ef_src + 1, 8, _cur_ret, fmt=FMT_PCT)

        _ms_vol = float(_max_sharpe_stats.get('volatility', 0.0) or 0.0) if _max_sharpe_stats else None
        _ms_ret = float(_max_sharpe_stats.get('expected_return', 0.0) or 0.0) if _max_sharpe_stats else None
        if _ms_vol is not None:
            write_cell(ws, ef_src, 10, 'MaxSharpe Volatility', bold=True)
            write_cell(ws, ef_src, 11, 'MaxSharpe Return', bold=True)
            write_cell(ws, ef_src + 1, 10, _ms_vol, fmt=FMT_PCT)
            write_cell(ws, ef_src + 1, 11, _ms_ret, fmt=FMT_PCT)

        _tan_vol = float(_tangency_stats.get('volatility', 0.0) or 0.0) if _tangency_stats else None
        _tan_ret = float(_tangency_stats.get('expected_return', 0.0) or 0.0) if _tangency_stats else None
        if _tan_vol is not None:
            write_cell(ws, ef_src, 13, 'Tangency Volatility', bold=True)
            write_cell(ws, ef_src, 14, 'Tangency Return', bold=True)
            write_cell(ws, ef_src + 1, 13, _tan_vol, fmt=FMT_PCT)
            write_cell(ws, ef_src + 1, 14, _tan_ret, fmt=FMT_PCT)

        _sc = ScatterChart()
        _sc.title = 'Efficient Frontier — Risk vs. Return'
        _sc.scatterStyle = 'lineMarker'
        _sc.style = 13
        _sc.width = 24
        _sc.height = 14
        _sc.x_axis.title = 'Volatility (Standard Deviation)'
        _sc.y_axis.title = 'Expected Return'
        _sc.x_axis.numFmt = '0.0%'
        _sc.y_axis.numFmt = '0.0%'
        _sc.x_axis.delete = False
        _sc.y_axis.delete = False

        _xref = Reference(ws, min_col=1, min_row=ef_src + 1, max_row=_ef_src_last)
        _yref = Reference(ws, min_col=2, min_row=ef_src + 1, max_row=_ef_src_last)
        _fs = Series(_yref, _xref, title='Efficient Frontier')
        _fs.marker = Marker(symbol='circle', size=6)
        _fs.graphicalProperties = GraphicalProperties(ln=LineProperties(solidFill='2D6A4F', w=20000))
        _sc.series.append(_fs)

        if _rec_vol is not None:
            _rxref = Reference(ws, min_col=4, min_row=ef_src + 1, max_row=ef_src + 1)
            _ryref = Reference(ws, min_col=5, min_row=ef_src + 1, max_row=ef_src + 1)
            _rs = Series(_ryref, _rxref, title='Recommended Portfolio')
            _rs.marker = Marker(symbol='diamond', size=11)
            _rs.graphicalProperties = GraphicalProperties(ln=LineProperties(noFill=True))
            _sc.series.append(_rs)

        if _cur_vol is not None:
            _cxref = Reference(ws, min_col=7, min_row=ef_src + 1, max_row=ef_src + 1)
            _cyref = Reference(ws, min_col=8, min_row=ef_src + 1, max_row=ef_src + 1)
            _cs = Series(_cyref, _cxref, title='Current Portfolio')
            _cs.marker = Marker(symbol='square', size=11)
            _cs.graphicalProperties = GraphicalProperties(ln=LineProperties(noFill=True))
            _sc.series.append(_cs)

        if _ms_vol is not None:
            _msxref = Reference(ws, min_col=10, min_row=ef_src + 1, max_row=ef_src + 1)
            _msyref = Reference(ws, min_col=11, min_row=ef_src + 1, max_row=ef_src + 1)
            _mss = Series(_msyref, _msxref, title='Max Sharpe (Risk-Budgeted)')
            _mss.marker = Marker(symbol='triangle', size=11)
            _mss.graphicalProperties = GraphicalProperties(ln=LineProperties(noFill=True))
            _sc.series.append(_mss)

        if _tan_vol is not None:
            _tanxref = Reference(ws, min_col=13, min_row=ef_src + 1, max_row=ef_src + 1)
            _tanyref = Reference(ws, min_col=14, min_row=ef_src + 1, max_row=ef_src + 1)
            _tans = Series(_tanyref, _tanxref, title='Pure Tangency')
            _tans.marker = Marker(symbol='star', size=12)
            _tans.graphicalProperties = GraphicalProperties(ln=LineProperties(noFill=True))
            _sc.series.append(_tans)

        chart_ws.add_chart(_sc, 'A216')
        _ef_chart_added = True

    _chart_count = 7 if _ef_chart_added else 6
    qc('8. Charts Dashboard',
       f'{_chart_count} charts: NW, CF Income, CF Expense, MC Bands, Alloc Before, Alloc After'
       + (', Efficient Frontier' if _ef_chart_added else ''),
       True,
       f'NW, Income (15 ser, ymax=${CF_YMAX:,}), Expense (8 ser, ymax=${CF_YMAX:,}), MC bands, 2 pie charts'
       + (', efficient frontier scatter' if _ef_chart_added else ''))
