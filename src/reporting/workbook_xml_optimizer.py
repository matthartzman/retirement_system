"""
workbook_xml_optimizer.py — optional XLSX XML post-save optimizations.

Uses lxml when available, falls back gracefully to no-op behavior. This keeps
workbook/PDF builds from failing on Windows machines where lxml wheels may not
be installed.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, Iterable, List


def _formula_values(wb: Any, formula: str) -> list[Any]:
    """Read values for an Excel chart range formula from an already-open workbook.

    Openpyxl writes chart formulas but not always chart value caches. Excel can
    recalculate those caches, but previewers and locked-down desktop viewers can
    show blank/broken charts when the cache is absent.  This helper resolves the
    formula so optimize_workbook_xml can embed numCache/strCache data directly
    into xl/charts/chart*.xml.

    ``wb`` is opened once by the caller (optimize_workbook_xml) and reused
    across every chart/formula in the workbook -- a single build can have 100+
    chart cache references, and reopening + reparsing the whole .xlsx from
    disk per formula was the dominant cost of the post-save XML patch step.
    """
    try:
        from openpyxl.utils.cell import range_to_tuple
    except ImportError:
        return []
    try:
        sheet_name, bounds = range_to_tuple(str(formula))
        min_col, min_row, max_col, max_row = bounds
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        values: list[Any] = []
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True):
            values.extend(row)
        return values
    except Exception:
        return []


def _is_number(value: Any) -> bool:
    try:
        if value is None or value == "":
            return False
        n = float(value)
        return n == n and n not in (float("inf"), float("-inf"))
    except Exception:
        return False


def _replace_chart_cache(etree: Any, root: Any, wb: Any) -> int:
    ns_chart = "http://schemas.openxmlformats.org/drawingml/2006/chart"
    c = "{" + ns_chart + "}"
    patched = 0

    def _clear(ref: Any) -> None:
        for tag in ("numCache", "strCache", "extLst"):
            for child in list(ref.findall(c + tag)):
                ref.remove(child)

    def _add_cache(ref: Any, tag: str, values: Iterable[Any], numeric: bool) -> bool:
        vals = list(values)
        if not vals:
            return False
        _clear(ref)
        cache = etree.Element(c + tag)
        pt_count = etree.SubElement(cache, c + "ptCount")
        pt_count.set("val", str(len(vals)))
        for idx, value in enumerate(vals):
            pt = etree.SubElement(cache, c + "pt")
            pt.set("idx", str(idx))
            v = etree.SubElement(pt, c + "v")
            if numeric:
                try:
                    n = float(value)
                    v.text = str(int(n)) if n.is_integer() else ("%.12g" % n)
                except Exception:
                    v.text = "0"
            else:
                v.text = "" if value is None else str(value)
        ref.append(cache)
        return True

    for ref in root.iter(c + "numRef"):
        f = ref.find(c + "f")
        formula = f.text if f is not None else ""
        vals = _formula_values(wb, formula)
        if vals and _add_cache(ref, "numCache", vals, numeric=True):
            patched += 1

    for ref in root.iter(c + "strRef"):
        f = ref.find(c + "f")
        formula = f.text if f is not None else ""
        vals = _formula_values(wb, formula)
        if vals and _add_cache(ref, "strCache", vals, numeric=False):
            patched += 1

    return patched


def optimize_workbook_xml(out_path: str) -> Dict[str, object]:
    try:
        from lxml import etree  # type: ignore
    except ImportError:
        return {"status": "skipped", "reason": "lxml_not_installed"}

    import os
    import zipfile

    path = Path(out_path)
    if not path.exists():
        return {"status": "skipped", "reason": "workbook_missing"}

    ns_chart = "http://schemas.openxmlformats.org/drawingml/2006/chart"
    ns_a = "http://schemas.openxmlformats.org/drawingml/2006/main"
    tmp_path = str(path) + ".xmlopt.tmp"
    patched = 0
    caches_patched = 0

    # Opened once and reused for every chart's numRef/strRef lookups below --
    # a workbook can have 100+ chart cache references, and reopening +
    # reparsing the whole .xlsx per formula was the dominant cost here.
    try:
        from openpyxl import load_workbook
        cache_wb = load_workbook(path, data_only=True, read_only=True)
    except Exception:
        cache_wb = None

    try:
        try:
            with zipfile.ZipFile(path, "r") as zin, zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    data = zin.read(item.filename)
                    if item.filename.startswith("xl/charts/chart") and item.filename.endswith(".xml"):
                        try:
                            root = etree.fromstring(data)
                            # Only the chart's own title (a direct child of
                            # c:chart) gets enlarged. c:title also appears nested
                            # inside c:catAx/c:valAx for axis titles -- root.iter()
                            # would match those too and inflate axis text along
                            # with the chart title, which is the bug behind the
                            # Efficient Frontier chart's oversized axis labels.
                            # Axis titles/labels are left completely alone here so
                            # they keep Excel's standard sizing.
                            chart_el = root.find(f"{{{ns_chart}}}chart")
                            main_title = (
                                chart_el.find(f"{{{ns_chart}}}title")
                                if chart_el is not None
                                else None
                            )
                            if main_title is not None:
                                # Readable and clearly larger than the (untouched,
                                # standard-sized) axis labels, without the
                                # previous 3x/66pt exaggeration.
                                for rpr in main_title.iter(f"{{{ns_a}}}defRPr"):
                                    rpr.set("sz", "2200")
                                    rpr.set("b", "1")
                                for rpr in main_title.iter(f"{{{ns_a}}}rPr"):
                                    rpr.set("sz", "2200")
                                    rpr.set("b", "1")
                            for leg_pos in root.iter(f"{{{ns_chart}}}legendPos"):
                                leg_pos.set("val", "r")
                            if cache_wb is not None:
                                caches_patched += _replace_chart_cache(etree, root, cache_wb)
                            data = etree.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=False)
                            patched += 1
                        except Exception:
                            pass
                    zout.writestr(item, data)
            # Release the read-only handle on `path` before replacing it --
            # Windows refuses to rename over a file that's still open.
            if cache_wb is not None:
                try:
                    cache_wb.close()
                except Exception:
                    pass
                cache_wb = None
            os.replace(tmp_path, path)
            return {"status": "ok", "charts_patched": patched, "chart_caches_patched": caches_patched}
        except Exception as exc:
            try:
                if Path(tmp_path).exists():
                    Path(tmp_path).unlink()
            except Exception:
                pass
            return {"status": "skipped", "reason": str(exc)}
    finally:
        if cache_wb is not None:
            try:
                cache_wb.close()
            except Exception:
                pass
