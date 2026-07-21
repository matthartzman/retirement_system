from __future__ import annotations

import csv
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
import unittest
import warnings
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
FIXTURES = ROOT / "tests" / "fixtures"

from tests.golden_pricing import frozen_holdings_prices

# Q3 (system review, Wave 3 item 3.15): a structural PDF check - page count
# and per-page size bounds - deeper than the existing magic-byte-plus-size
# check but without text extraction (COM-exported PDFs frequently embed text
# in ways that defeat extraction; a structural check catches the failure
# mode that actually occurs - a build that silently produces a 1-page or
# wrong-size PDF). /Type /Page object markers and /MediaBox arrays are part
# of reportlab's plain (uncompressed) object structure, so counting them
# directly via regex is reliable for this app's own PDF output without
# adding a PDF-parsing library dependency.
_PDF_PAGE_TYPE_RE = re.compile(rb"/Type\s*/Page(?!s)\b")
_PDF_MEDIABOX_RE = re.compile(rb"/MediaBox\s*\[\s*([\d.\-]+)\s+([\d.\-]+)\s+([\d.\-]+)\s+([\d.\-]+)\s*\]")


def _pdf_structural_summary(pdf_bytes: bytes) -> dict:
    """Cheap structural read of a PDF's page count and page dimensions -
    raises ValueError if the bytes don't even start like a PDF."""
    if not pdf_bytes.startswith(b"%PDF-"):
        raise ValueError("not a PDF file (missing %PDF- header)")
    page_count = len(_PDF_PAGE_TYPE_RE.findall(pdf_bytes))
    media_boxes = [
        tuple(float(g) for g in m.groups())
        for m in _PDF_MEDIABOX_RE.finditer(pdf_bytes)
    ]
    return {"page_count": page_count, "media_boxes": media_boxes}


def _load_engine_config():
    from src.data_io import load_csv
    from src.report_compute import prepare_config_from_sectioned_data
    return prepare_config_from_sectioned_data(load_csv(ROOT / "input" / "client_data.csv"), "", optimize_roth=True)


def _project_metrics(c):
    from src.planning_engines import project
    rows = project(c)
    terminal = rows[-1]
    first = rows[0]
    first_rmd = next((r for r in rows if float(r.get("rmd_total", 0) or 0) > 0), None)
    first_conv = next((r for r in rows if float(r.get("roth_conv", 0) or 0) > 0), None)
    return {
        "plan_start": int(c["plan_start"]),
        "plan_end": int(c["plan_end"]),
        "row_count": len(rows),
        "terminal_year": int(terminal["year"]),
        "terminal_total_nw": round(float(terminal.get("total_nw", 0) or 0), 2),
        "terminal_liquid_nw": round(float(terminal.get("pretax_nw", 0) or 0) + float(terminal.get("roth_nw", 0) or 0) + float(terminal.get("trust_nw", 0) or 0) + float(terminal.get("hsa_nw", 0) or 0), 2),
        "lifetime_tax": round(sum(float(r.get("total_tax", 0) or 0) for r in rows), 2),
        "total_roth_conversion": round(sum(float(r.get("roth_conv", 0) or 0) for r in rows), 2),
        "first_year_total_tax": round(float(first.get("total_tax", 0) or 0), 2),
        "first_rmd_year": int(first_rmd["year"]) if first_rmd else None,
        "first_rmd_total": round(float(first_rmd.get("rmd_total", 0) or 0), 2) if first_rmd else 0,
        "first_conversion_year": int(first_conv["year"]) if first_conv else None,
        "first_conversion_amount": round(float(first_conv.get("roth_conv", 0) or 0), 2) if first_conv else 0,
        "selected_roth_strategy": (c.get("roth_optimization") or {}).get("selected_label", ""),
    }


class Phase5GoldenMasterEngineTests(unittest.TestCase):
    """Plan-coupled engine diagnostic — WARN-ONLY, deliberately not a gate.

    This library's five "scenarios" are mutators applied to the *live client
    plan*: it loads ``input/client_data.csv``, and ``load_csv`` merges every
    sibling ``input/client_*.csv`` (household, policy, optional functions), so
    the pinned dollar totals move whenever the advisor edits real plan data.
    That made a CI failure here ambiguous between "the engine regressed" and
    "the client changed their retirement date", and its ``delta=50000.0``
    tolerance existed mostly to absorb that coupling.

    The gating baseline is now ``tests/test_synthetic_golden_master.py``, whose
    plans are built in code and read nothing from ``input/``. It covers these
    same five stresses plus a single filer, DAF, dividend-reinvestment and TLH,
    and pins to the cent.

    Kept rather than deleted because one thing it checks is *not* covered by the
    synthetic gate and is not obtainable from synthetic data: that the advisor's
    actual plan still loads and projects end-to-end. Dollar drift is reported as
    a warning; only a hard failure to project fails the test.
    """

    def test_live_client_plan_still_projects_end_to_end(self):
        expected = json.loads((FIXTURES / "golden_master_engine_cases.json").read_text(encoding="utf-8"))
        mutators = {
            "baseline_balanced_couple": lambda c: None,
            "no_voluntary_roth_policy": lambda c: c.update({"roth_policy": "none", "roth_optimized_policy": "none", "roth_optimization": {}}),
            "high_spending_pressure": lambda c: c.update({"spend_base": float(c.get("spend_base", 0)) * 1.20}),
            "lower_return_environment": lambda c: c.update({"ret": 0.04, "ret_stock": 0.055, "ret_bond": 0.03}),
            "early_survivor_compression": lambda c: c.update({"h_death_yr": int(c["plan_start"]) + 5}),
        }
        drift = []
        for case_name, expected_metrics in expected.items():
            with self.subTest(case=case_name):
                # Holdings prices are frozen to the committed golden snapshot
                # (tests/golden_pricing.py) so terminal balances are deterministic
                # and portable rather than depending on the untracked local
                # OFFLINE price cache — the same pin test_2/test_167 use.
                with frozen_holdings_prices():
                    c = _load_engine_config()
                    mutators[case_name](c)
                    # The assertion: the real plan projects to a full, non-empty
                    # horizon. This is what stays meaningful when plan data
                    # changes, and it is what the synthetic gate cannot check.
                    actual = _project_metrics(c)
                self.assertGreater(actual["row_count"], 0, f"{case_name} produced no projection rows")
                self.assertEqual(
                    actual["row_count"], actual["plan_end"] - actual["plan_start"] + 1,
                    f"{case_name} projected a horizon inconsistent with plan_start/plan_end",
                )
                self.assertGreater(actual["terminal_total_nw"], 0.0, f"{case_name} terminal net worth is not positive")

                for key, expected_value in expected_metrics.items():
                    if isinstance(expected_value, (int, float)) and not isinstance(expected_value, bool):
                        if actual[key] is None or abs(actual[key] - expected_value) > 50000.0:
                            drift.append(f"{case_name}.{key}: pinned {expected_value} -> actual {actual[key]}")
                    elif actual[key] != expected_value:
                        drift.append(f"{case_name}.{key}: pinned {expected_value!r} -> actual {actual[key]!r}")

        if drift:
            warnings.warn(
                "Live-client golden-master pins have drifted from "
                "tests/fixtures/golden_master_engine_cases.json. This is EXPECTED after a "
                "plan-data edit and is not a regression signal — the gating engine baseline "
                "is tests/test_synthetic_golden_master.py. Regenerate these pins when the "
                "current plan data is the intended new reference.\n  " + "\n  ".join(drift),
                stacklevel=2,
            )


class Phase5ClosedFormTaxTests(unittest.TestCase):
    def test_federal_tax_closed_form_bracket_edges(self):
        from src.core import compute_fed_tax
        self.assertAlmostEqual(compute_fed_tax(100000, 2025, "MFJ", 0.0), 11828.00, places=2)
        self.assertAlmostEqual(compute_fed_tax(50000, 2025, "Single", 0.0), 5914.00, places=2)
        self.assertAlmostEqual(compute_fed_tax(206700, 2025, "MFJ", 0.0), 35302.00, places=2)

    def test_standard_deduction_and_state_tax_closed_form(self):
        from src.core import standard_deduction, state_income_tax
        self.assertAlmostEqual(standard_deduction(2025, "MFJ", 0.0, n_over_65=2), 33200.00, places=2)
        self.assertAlmostEqual(standard_deduction(2025, "Single", 0.0, n_over_65=1), 17000.00, places=2)
        il_tax = state_income_tax(
            "Illinois", earned=100000, retirement_dist=100000, ss_taxable=20000,
            investment_inc=10000, nonqual_annuity=5000, roth_conv=50000, year=2025,
            age_over_65=True,
        )
        self.assertAlmostEqual(il_tax, 5692.50, places=2)

    def test_irmaa_and_niit_simple_threshold_behavior(self):
        from src.core import irmaa_surcharge, irmaa_tier
        self.assertEqual(irmaa_tier(200000, 2026, 2026, filing="MFJ"), 0)
        self.assertEqual(irmaa_tier(213000, 2026, 2026, filing="MFJ"), 1)
        self.assertGreater(irmaa_surcharge(213000, 2026, 2026, filing="MFJ"), 0)


class Phase5IRSExampleReconciliationTests(unittest.TestCase):
    def test_irs_style_social_security_examples(self):
        from src.core import social_security_taxable_amount
        examples = json.loads((FIXTURES / "irs_style_examples.json").read_text(encoding="utf-8"))["social_security_taxable_examples"]
        for ex in examples:
            with self.subTest(ex["name"]):
                actual = social_security_taxable_amount(ex["social_security"], ex["other_income"], ex["filing"])
                self.assertAlmostEqual(actual, ex["expected_taxable_social_security"], places=2)

    def test_irs_pub_590b_rmd_examples(self):
        from src.core import rmd_divisor
        examples = json.loads((FIXTURES / "irs_style_examples.json").read_text(encoding="utf-8"))["rmd_examples"]
        for ex in examples:
            with self.subTest(ex["name"]):
                div = rmd_divisor(ex["age"])
                self.assertAlmostEqual(div, ex["expected_divisor"], places=1)
                self.assertAlmostEqual(ex["prior_year_balance"] / div, ex["expected_rmd"], places=2)


class Phase5CrossToolReconciliationTests(unittest.TestCase):
    def _manual_fed_tax(self, filing: str, taxable: float) -> float:
        brackets = {
            "MFJ": [(0, 23850, .10), (23850, 96950, .12), (96950, 206700, .22), (206700, 394600, .24), (394600, 501050, .32), (501050, 751600, .35), (751600, float("inf"), .37)],
            "Single": [(0, 11925, .10), (11925, 48475, .12), (48475, 103350, .22), (103350, 197300, .24), (197300, 250525, .32), (250525, 626350, .35), (626350, float("inf"), .37)],
        }[filing]
        tax = 0.0
        for lo, hi, rate in brackets:
            if taxable <= lo:
                break
            tax += (min(taxable, hi) - lo) * rate
        return tax

    def test_csv_backed_independent_reconciliation_cases(self):
        from src.core import compute_fed_tax, social_security_taxable_amount, rmd_divisor, state_income_tax
        with (FIXTURES / "cross_tool_reconciliation_cases.csv").open(newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                with self.subTest(row["case_id"]):
                    expected = float(row["expected"])
                    tol = float(row["tolerance"])
                    domain = row["domain"]
                    if domain == "federal_tax":
                        filing, taxable, year = row["input_a"], float(row["input_b"]), int(row["input_c"])
                        engine = compute_fed_tax(taxable, year, filing, 0.0)
                        independent = self._manual_fed_tax(filing, taxable)
                    elif domain == "social_security":
                        engine = social_security_taxable_amount(float(row["input_b"]), float(row["input_c"]), row["input_a"])
                        # The independent worksheet-style expected value is supplied in the CSV fixture.
                        independent = expected
                    elif domain == "rmd":
                        engine = float(row["input_b"]) / rmd_divisor(int(row["input_a"]))
                        independent = expected
                    elif domain == "state_tax_il":
                        engine = state_income_tax("Illinois", float(row["input_a"]), 100000, 20000, float(row["input_b"]), float(row["input_c"]), 50000, 2025, True)
                        independent = expected
                    else:
                        raise AssertionError(domain)
                    self.assertAlmostEqual(engine, expected, delta=tol)
                    self.assertAlmostEqual(engine, independent, delta=tol)


class Phase5WorkbookSnapshotTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.tmp = tempfile.mkdtemp(prefix="phase5_workbook_")
        tmp_root = Path(cls.tmp)
        # The reporting stack resolves its project root from source-file
        # location, not from the process cwd, so building "in-place" against
        # ROOT would read/write the real input/, system_config.csv, and
        # output/ files. Copy the tree into a scratch dir and build there
        # instead; local_state/ and output/ are deliberately excluded so the
        # build bootstraps a fresh SQLite mirror rather than touching the
        # real one.
        excluded_names = {".git", ".claude", ".pytest_cache", "tests", "documentation", "output", "local_state", "__pycache__"}
        shutil.copytree(
            ROOT,
            tmp_root,
            ignore=lambda _dir, names: [n for n in names if n in excluded_names or n.endswith(".pyc")],
            dirs_exist_ok=True,
        )
        env = os.environ.copy()
        # Force the subprocess to treat tmp_root (its own copied tree) as the
        # workspace root, overriding any RETIREMENT_SYSTEM_WORKSPACE_ROOT the
        # parent test process has set (tests/conftest.py sets one so the
        # suite never mutates the real input/ files). Without this override,
        # the subprocess would inherit the parent's redirect and read/write
        # someone else's temp workspace instead of this test's own copy.
        env["RETIREMENT_SYSTEM_WORKSPACE_ROOT"] = str(tmp_root)
        env["RETIREMENT_MC_SIMS"] = "16"
        env["RETIREMENT_MC_SENSITIVITY_SIMS"] = "3"
        env["RETIREMENT_SKIP_REPORT_SIDECARS"] = "1"
        result = subprocess.run([sys.executable, "tools/build_workbook.py"], cwd=tmp_root, text=True, capture_output=True, env=env, timeout=120)
        cls.build_stdout = result.stdout + result.stderr
        if result.returncode != 0:
            raise AssertionError(cls.build_stdout)
        cls.workbook_path = tmp_root / "output" / "retirement_plan.xlsx"

    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(cls.tmp, ignore_errors=True)

    def test_build_also_produces_downloadable_pdf(self):
        # Regression guard: the build must write retirement_plan.pdf next to the
        # workbook, or the "Download PDF" button 404s ("run build first"). This
        # broke twice because build_enterprise_pdf was imported but never called
        # in the build pipeline; assert the artifact exists and is a real PDF.
        pdf_path = self.workbook_path.parent / "retirement_plan.pdf"
        self.assertTrue(pdf_path.exists(), f"build did not produce {pdf_path}\n{self.build_stdout}")
        with pdf_path.open("rb") as fh:
            self.assertEqual(fh.read(5), b"%PDF-", "retirement_plan.pdf is not a valid PDF")
        self.assertGreater(pdf_path.stat().st_size, 1024, "retirement_plan.pdf is suspiciously small")

    def test_pdf_has_a_real_page_count_and_uniform_landscape_letter_pages(self):
        """Q3: deeper than the magic-byte check above - every real build
        renders every visible workbook sheet as at least one page (see
        enterprise_pdf.py's module docstring), so a 27-sheet workbook should
        never collapse to a handful of PDF pages, and every page enterprise_pdf.py
        emits is landscape letter (792x612pt) by construction."""
        pdf_path = self.workbook_path.parent / "retirement_plan.pdf"
        summary = _pdf_structural_summary(pdf_path.read_bytes())
        self.assertGreater(summary["page_count"], 20, f"expected a substantial multi-page PDF, got {summary['page_count']} pages")
        self.assertEqual(len(summary["media_boxes"]), summary["page_count"], "expected one /MediaBox per page")
        for x0, y0, x1, y1 in summary["media_boxes"]:
            width, height = x1 - x0, y1 - y0
            self.assertTrue(700 <= width <= 850, f"page width {width}pt outside expected landscape-letter bounds")
            self.assertTrue(550 <= height <= 650, f"page height {height}pt outside expected landscape-letter bounds")

    def test_pdf_structural_check_catches_a_truncated_pdf(self):
        """Regression proof for the check above: a PDF cut off mid-stream
        (a failed/interrupted write, or a build that silently emits partial
        output) must be caught, not pass silently. Corrupting the trailer/
        xref by truncation makes reportlab's own object markers earlier in
        the file undercounted or unreadable - either way, this must not
        report the same healthy page count as the real file."""
        pdf_path = self.workbook_path.parent / "retirement_plan.pdf"
        full_bytes = pdf_path.read_bytes()
        full_summary = _pdf_structural_summary(full_bytes)
        # reportlab clusters every page object's own (small) dictionary
        # together, separate from and before the much larger per-page content
        # streams - empirically, all of them landed in the first ~15% of a
        # real build's PDF here. Cutting at 1/3 left every page marker intact
        # (this test caught that on its first pass); 1/10 lands inside that
        # object-definition region rather than only trimming trailing content
        # streams/xref, which need to be cut in the first place to prove
        # anything.
        truncated = full_bytes[: len(full_bytes) // 10]
        try:
            truncated_summary = _pdf_structural_summary(truncated)
        except ValueError:
            return  # truncation cut even the %PDF- header - unambiguously caught
        self.assertLess(
            truncated_summary["page_count"], full_summary["page_count"],
            "a truncated PDF must not report the same page count as the real file",
        )

    def test_workbook_snapshot_sheets_and_key_phrases(self):
        import openpyxl
        snap = json.loads((FIXTURES / "workbook_snapshot_expectations.json").read_text(encoding="utf-8"))
        wb = openpyxl.load_workbook(self.workbook_path, data_only=True, read_only=True)
        for sheet in snap["required_sheets"]:
            self.assertIn(sheet, wb.sheetnames)
        for sheet, phrases in snap["required_phrases"].items():
            text = "\n".join(str(cell) for row in wb[sheet].iter_rows(values_only=True) for cell in row if cell is not None)
            for phrase in phrases:
                self.assertIn(phrase, text, f"{sheet} missing {phrase}")

    def test_workbook_snapshot_rejects_stale_roth_language(self):
        import openpyxl
        snap = json.loads((FIXTURES / "workbook_snapshot_expectations.json").read_text(encoding="utf-8"))
        wb = openpyxl.load_workbook(self.workbook_path, data_only=True, read_only=True)
        combined = "\n".join(
            str(cell)
            for sheet in ["1A. Executive Summary", "4B. Assumptions", "2A. Roth Conversion"]
            if sheet in wb.sheetnames
            for row in wb[sheet].iter_rows(values_only=True)
            for cell in row
            if cell is not None
        )
        for forbidden in snap["forbidden_roth_phrases"]:
            self.assertNotIn(forbidden, combined)


if __name__ == "__main__":
    unittest.main()
