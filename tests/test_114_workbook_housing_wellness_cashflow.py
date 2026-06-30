from openpyxl import Workbook

from src.reporting.sheets_projection import build_sheet6


def _base_config(rent=0, ltc=0):
    return {
        'plan_start': 2026,
        'plan_end': 2026,
        'post_sale_rent_mo': rent,
        'ltc_annual_prem': ltc,
        'state': 'IL',
        'inf': 0.025,
        'h_name': 'Matthew',
        'w_name': 'Patricia',
    }


def _row(rent=0, ltc=0):
    return {
        'year': 2026,
        'h_age': 60,
        'w_age': 60,
        'earned': 0,
        'h_ss': 0,
        'w_ss': 0,
        'pension': 0,
        'wife_single_ann': 0,
        'wife_joint_ann': 0,
        'h_single_ann': 0,
        'h_joint_ann': 0,
        'note_princ': 0,
        'note_int': 0,
        'rmd_total': 0,
        'roth_conv': 0,
        'agi': 0,
        'taxable_inc': 0,
        'fed_tax': 0,
        'state_tax': 0,
        'niit': 0,
        'irmaa': 0,
        'spend_base_yr': 100,
        'mortgage': 150,
        'mortgage_payment_yr': 100,
        'real_estate_tax_yr': 20,
        'home_improvement_yr': 30,
        'rent_yr': rent,
        'housing_utilities_yr': 40,
        'housing_maintenance_yr': 50,
        'housing_other_yr': 60,
        'housing_total_yr': 300 + rent,
        'wellness_base_yr': 400,
        'wellness_premiums_yr': 150,
        'wellness_medical_yr': 100,
        'wellness_dental_yr': 50,
        'wellness_vision_yr': 40,
        'wellness_rx_otc_yr': 30,
        'wellness_other_yr': 30,
        'ltc_prem_yr': ltc,
        'rec_extra': 0,
        'lump': 0,
        'heloc_interest': 0,
        'heloc_repayment_principal': 0,
        'h_trust_wd': 0,
        'w_trust_wd': 0,
        'hsa_wd': 0,
        'h_roth_wd': 0,
        'w_roth_wd': 0,
        'rmd_h': 0,
        'h_ira_elective': 0,
        'rmd_w': 0,
        'w_ira_elective': 0,
        'heloc_draw': 0,
        'heloc_balance': 0,
        'surplus': 0,
        'total_nw': 0,
    }


def _headers(config, row):
    wb = Workbook()
    ws = wb.active
    build_sheet6(ws, config, [row])
    return [ws.cell(2, col).value for col in range(1, ws.max_column + 1)]


def test_cashflow_hides_rent_and_ltc_when_zero():
    headers = _headers(_base_config(rent=0, ltc=0), _row(rent=0, ltc=0))
    assert 'Rent' not in headers
    assert 'LTC Prem' not in headers


def test_cashflow_expands_all_housing_groups():
    headers = _headers(_base_config(), _row())
    for header in ['Mortgage P&I', 'Property Tax', 'Utilities', 'Home Impr', 'Maintenance', 'Other']:
        assert header in headers


def test_cashflow_combines_wellness_premiums_and_expands_detail():
    headers = _headers(_base_config(), _row())
    assert 'Healthcare Premiums' in headers
    assert 'Bridge Prem' not in headers
    assert 'Medicare' not in headers
    for header in ['Medical', 'Dental', 'Vision', 'Rx/OTC', 'Other Wellness']:
        assert header in headers


def test_cashflow_includes_rent_and_ltc_when_positive():
    headers = _headers(_base_config(rent=12000, ltc=1000), _row(rent=12000, ltc=1000))
    assert 'Rent' in headers
    assert 'LTC Prem' in headers


def test_cashflow_hides_ltc_when_policy_disabled_even_if_premium_value_remains():
    config = _base_config(rent=0, ltc=1000)
    config['ltc_enabled'] = False
    headers = _headers(config, _row(rent=0, ltc=0))
    assert 'LTC Prem' not in headers


def test_cashflow_includes_ltc_when_policy_enabled_and_premium_positive():
    config = _base_config(rent=0, ltc=1000)
    config['ltc_enabled'] = True
    headers = _headers(config, _row(rent=0, ltc=0))
    assert 'LTC Prem' in headers
