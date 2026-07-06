from openpyxl import load_workbook


def test_workbook_uses_numbered_sections_and_lettered_children(built_workbook_path):
    assert built_workbook_path.exists(), 'Generated workbook is missing'
    wb = load_workbook(built_workbook_path, read_only=False, data_only=False)
    visible = [ws.title for ws in wb.worksheets if ws.sheet_state == 'visible']
    expected = [
        '1. Reports',
        '1A. Executive Summary',
        '1B. Net Worth',
        '1C. Cash Flow',
        '1D. Balance Sheet',
        '1E. Charts',
        '1F. Lifetime Taxes',
        '1G. Core Spending',
        '1H. Spending Summary',
        '2. Optimizers',
        '2A. Roth Conversion',
        '2B. Asset Allocation',
        '2C. State Residency',
        '2D. Social Security',
        '2E. S-Corp vs LLC',
        '2F. Charitable Giving',
        '2G. Estate & Legacy Planning',
        '2H. Planning Levers',
        '3. Risk & Stress Tests',
        '3A. Monte Carlo',
        '3B. Survivor',
        '3C. LTC + Life Insurance',
        '4. System',
        '4A. Plan Data',
        '4B. Assumptions',
        '4C. Account Reconciliation',
        '4D. Quality Control',
        '4E. RMD Audit',
        '4F. Methodology',
        '4G. Glossary',
    ]
    assert visible[: len(expected)] == expected
    assert '4A. Plan Data' in visible
    assert 'Reports' not in visible
    assert 'Risk' not in visible
    assert 'Optimizers' not in visible
    assert 'System Configuration' not in visible


def test_summary_tabs_reference_child_tabs(built_workbook_path):
    wb = load_workbook(built_workbook_path, read_only=False, data_only=False)
    summary_expected = {
        '1. Reports': ['1A. Executive Summary', '1B. Net Worth', '1C. Cash Flow', '1D. Balance Sheet', '1E. Charts'],
        '2. Optimizers': ['2A. Roth Conversion', '2B. Asset Allocation', '2C. State Residency', '2D. Social Security', '2E. S-Corp vs LLC', '2G. Estate & Legacy Planning'],
        '3. Risk & Stress Tests': ['3A. Monte Carlo', '3B. Survivor', '3C. LTC + Life Insurance'],
        '4. System': ['4A. Plan Data', '4B. Assumptions', '4C. Account Reconciliation', '4D. Quality Control', '4E. RMD Audit', '4F. Methodology', '4G. Glossary'],
    }
    for sheet, children in summary_expected.items():
        ws = wb[sheet]
        values = [cell.value for row in ws.iter_rows(min_row=1, max_row=20, min_col=1, max_col=8) for cell in row]
        for child in children:
            assert child in values


def test_strategy_scorp_ltc_and_asset_location_merges_are_present(built_workbook_path):
    wb = load_workbook(built_workbook_path, read_only=False, data_only=False)
    exec_text = ' '.join(str(c.value or '') for row in wb['1A. Executive Summary'].iter_rows() for c in row)
    scorp_text = ' '.join(str(c.value or '') for row in wb['2E. S-Corp vs LLC'].iter_rows() for c in row)
    allocation_text = ' '.join(str(c.value or '') for row in wb['2B. Asset Allocation'].iter_rows() for c in row)
    ltc_text = ' '.join(str(c.value or '') for row in wb['3C. LTC + Life Insurance'].iter_rows() for c in row)
    assert 'WITHDRAWAL SEQUENCE STRATEGY' in exec_text
    assert 'S-CORP vs LLC' in scorp_text and 'LLC / Sole-Prop' in scorp_text
    assert 'ASSET-LOCATION OPTIMIZER' in allocation_text
    assert 'LONG-TERM-CARE STRESS TEST' in ltc_text
