"""Integration test for the Phase 3 workbook education section: a native
Excel scatter chart reproducing the "Probability of a Real Loss by Holding
Period" reference chart, with this household's own withdrawal-derived
holding period overlaid (see src/reporting/sheets_projection_charts.py,
build_sheet8's Holding-Period Real-Loss-Probability Chart block).

Renders build_sheet8 directly on a minimal in-memory workbook against the
repo's sample household, then inspects the resulting worksheet's chart
objects and hidden data sheet -- this exercises the real openpyxl chart API
calls (Reference/Series/ScatterChart), not just the underlying data modules
already covered by test_191.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from src.data_io import load_csv, parse_client
from src.planning_engines import project
from src.reporting.sheets_projection_charts import build_sheet8

ROOT = Path(__file__).resolve().parents[1]


def sample_config():
    data = load_csv(ROOT / "input" / "client_data.csv")
    return parse_client(data, "")


def test_build_sheet8_adds_holding_period_chart_without_raising():
    c = sample_config()
    rows = project(c)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '8. Charts Dashboard'
    build_sheet8(ws, c, rows, mc_data=None)  # must not raise

    data_ws = wb['_Holding Period Chart Data']
    found_title = False
    for row in data_ws.iter_rows():
        for cell in row:
            if cell.value == 'Holding Years':
                found_title = True
                break
        if found_title:
            break
    assert found_title, "expected the holding-period chart's source data block to be written"


def test_holding_period_data_isolated_from_shared_chart_dashboard_sheet():
    # Regression guard: src/reporting/dashboard.py's HTML-dashboard clone
    # (_extract_chart_block) scans the shared '_Chart Dashboard Data' sheet's
    # column 1 end-to-end for integer values to reconstruct the plan-years
    # axis. If the holding-period chart's data (which includes small integer
    # "holding years" like 0/3/5/...) were ever written to that shared sheet
    # instead of its own '_Holding Period Chart Data' sheet, those values
    # would be misread as extra plan years and desync the income/expense
    # series length (IndexError downstream in dashboard.py).
    c = sample_config()
    rows = project(c)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '8. Charts Dashboard'
    build_sheet8(ws, c, rows, mc_data=None)

    shared_ws = wb['_Chart Dashboard Data']
    for row in shared_ws.iter_rows():
        for cell in row:
            assert cell.value != 'Holding Years'
            assert cell.value != 'Holding-Period Bucket'


def test_build_sheet8_holding_period_chart_object_present():
    c = sample_config()
    rows = project(c)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '8. Charts Dashboard'
    build_sheet8(ws, c, rows, mc_data=None)

    chart_titles = []
    for chart in getattr(ws, '_charts', []):
        title = getattr(chart, 'title', None)
        # openpyxl chart titles are rich-text objects; str() renders them.
        chart_titles.append(str(title))
    assert any('Real Loss' in t or 'real loss' in t.lower() for t in chart_titles), chart_titles


def test_build_sheet8_bucket_table_sums_to_liquid_balance():
    c = sample_config()
    rows = project(c)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '8. Charts Dashboard'
    build_sheet8(ws, c, rows, mc_data=None)

    data_ws = wb['_Holding Period Chart Data']
    header_row = None
    for row in data_ws.iter_rows():
        values = [cell.value for cell in row]
        if 'Holding-Period Bucket' in values:
            header_row = row[0].row
            break
    assert header_row is not None

    total = 0.0
    liquid_nw = sum((c.get('balances') or {}).values())
    r = header_row + 1
    while True:
        label = data_ws.cell(r, 1).value
        if not label or not isinstance(label, str) or 'yr' not in label:
            break
        dollars = data_ws.cell(r, 2).value or 0.0
        total += float(dollars)
        r += 1
    if liquid_nw > 0:
        assert total == pytest.approx(liquid_nw, rel=1e-6)


def test_build_sheet8_then_html_dashboard_does_not_desync_series(tmp_path):
    # End-to-end regression test for the exact bug this chart block
    # triggered during development: build_sheet8 followed by
    # build_html_dashboard (the actual sequence tools/build_workbook.py
    # runs) must not raise, and must not be misled by any new content this
    # module writes to the hidden chart-data sheets.
    from src.reporting.dashboard import build_html_dashboard

    c = sample_config()
    rows = project(c)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet('8. Charts Dashboard')
    build_sheet8(ws, c, rows, mc_data=None)

    xlsx_path = tmp_path / 'workbook.xlsx'
    html_path = tmp_path / 'dashboard.html'
    wb.save(xlsx_path)

    build_html_dashboard(xlsx_path, html_path, rows, c)  # must not raise
    assert html_path.exists()


def test_holding_period_profile_degrades_gracefully_with_no_rows():
    # The new chart block's data source (holding_period_profile) must never
    # raise on degenerate input -- exercised directly here since build_sheet8
    # as a whole already assumes non-empty rows for its pre-existing earlier
    # chart sections (unrelated to this change) and is not itself
    # zero-row-safe.
    from src import holding_period as hp
    c = sample_config()
    profile = hp.holding_period_profile([], c)  # must not raise
    assert profile.get('source') in ('withdrawal_schedule', 'no_projected_withdrawals', 'no_liquid_assets')
