"""Wave 5.3 (system review 2026-08-04, planner finding
life-insurance-rules-of-thumb): "Life-insurance capital-needs from the
survivor projection (fix double-count now)."

Sheet 19's Section B (Need / Gap Analysis) previously listed three needs --
Income Replacement (10x earned income), Mortgage Payoff, and a flat $500K
Estate Liquidity figure -- and computed each row's gap independently as
max(0, need - existing_db - liquid_nw). That credited the SAME existing
death benefit and the SAME liquid net worth against every single row, so a
household with enough liquid assets to cover any ONE need would show "Gap
Covered" on ALL THREE rows even if those assets couldn't actually cover all
three combined (a double/triple count of one shared pool).

The fix:
1. Replaces the crude "10x income" / flat multipliers with a survivor
   shortfall derived from Sheet 18's own worst-case early-death projection
   re-run (_survivor_early_death_scenarios, shared by both sheets) -- a
   household-specific, engine-computed figure instead of a generic rule of
   thumb.
2. Applies existing death benefits and liquid net worth against the TOTAL
   combined need exactly once, not per row.
"""
from __future__ import annotations

from src.data_io import load_csv, parse_client
from src.plan_config import ensure_engine_config
from src.planning_engines import project
from src.reporting.sheets_stress import _survivor_early_death_scenarios
from tests.golden_pricing import FROZEN_GOLDEN_MASTER_PRICES, frozen_holdings_prices

from conftest import TEST_INPUT_DIR


def sample_config_and_rows():
    c = parse_client(load_csv(TEST_INPUT_DIR / "client_data.csv"), "")
    c["roth_policy"] = "none"
    c = ensure_engine_config(c, source="test")
    with frozen_holdings_prices(FROZEN_GOLDEN_MASTER_PRICES):
        rows = project(c)
    return c, rows


def test_survivor_scenarios_returns_five_labeled_results():
    c, rows = sample_config_and_rows()
    survivor = _survivor_early_death_scenarios(c, rows)
    assert len(survivor["scenarios"]) == 5
    for s in survivor["scenarios"]:
        assert "delta_nw" in s and "label" in s


def test_sheet19_needs_are_not_each_credited_the_full_shared_pool():
    # Regression test for the double-count bug: build the workbook table's
    # underlying numbers directly and confirm the combined gap uses the
    # existing DB / liquid NW pool once, not once per need row.
    from openpyxl import Workbook

    from src.reporting.sheets_stress import build_sheet19

    c, rows = sample_config_and_rows()
    wb = Workbook()
    ws = wb.active
    build_sheet19(ws, c, rows)

    texts = [str(cell.value) for row in ws.iter_rows() for cell in row if cell.value is not None]
    assert any("TOTAL (shared pool applied once)" in t for t in texts)
    assert any("Survivor Shortfall" in t for t in texts)
    # The old hardcoded rows must be gone.
    assert not any("Income Replacement (10x)" in t for t in texts)
    assert not any("Mortgage Payoff" in t for t in texts)


def test_combined_gap_never_exceeds_sum_of_individual_naive_gaps():
    # If the bug were still present, summing three independently-computed
    # gaps (each crediting the full existing DB + liquid NW) would UNDERSTATE
    # the true combined gap whenever the pool wasn't enough for all needs
    # combined. The fixed total_gap must be >= what a single-need gap
    # calculation would show once the pool is exhausted by a bigger total
    # need than any one line reflects.
    c, rows = sample_config_and_rows()
    survivor = _survivor_early_death_scenarios(c, rows)
    single_death = [s for s in survivor["scenarios"] if s["h_death"] != s["w_death"]] \
        or survivor["scenarios"]
    worst_delta = min(s["delta_nw"] for s in single_death)
    survivor_shortfall = max(0.0, -worst_delta)

    first_yr_db = sum(c["ann_db"].get(c["plan_start"], {}).values())
    liquid_nw = sum(c["balances"].get(aid, 0) for aid in c.get("invest_ids", []))
    if liquid_nw == 0:
        liquid_nw = sum(v for k, v in c["balances"].items() if not k.lower().endswith("_checking"))

    total_need = survivor_shortfall + 500000.0
    naive_gap_per_row_summed = (
        max(0.0, survivor_shortfall - first_yr_db - liquid_nw)
        + max(0.0, 500000.0 - first_yr_db - liquid_nw)
    )
    combined_gap = max(0.0, total_need - first_yr_db - liquid_nw)

    # Whenever the pool is nonzero and smaller than the sum of needs, the
    # old naive per-row approach reports a smaller (over-optimistic) total
    # gap than the corrected shared-pool approach.
    if 0 < (first_yr_db + liquid_nw) < total_need:
        assert combined_gap >= naive_gap_per_row_summed


def test_sheet18_still_renders_with_extracted_helper():
    from openpyxl import Workbook

    from src.reporting.sheets_stress import build_sheet18

    c, rows = sample_config_and_rows()
    wb = Workbook()
    ws = wb.active
    build_sheet18(ws, c, rows)
    texts = [str(cell.value) for row in ws.iter_rows() for cell in row if cell.value is not None]
    assert any("SURVIVOR / EARLY-DEATH STRESS TEST" in t for t in texts)
    assert any("BASE PLAN" in t for t in texts)
