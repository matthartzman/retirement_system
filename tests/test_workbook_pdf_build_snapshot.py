from __future__ import annotations

import csv
import json
import os
import shutil
import subprocess
import sys
import tempfile
import unittest
import warnings
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parents[1]

from conftest import TEST_INPUT_DIR
FIXTURES = ROOT / "tests" / "fixtures"

from tests.golden_pricing import frozen_holdings_prices


def _load_engine_config():
    from src.data_io import load_csv
    from src.report_compute import prepare_config_from_sectioned_data
    return prepare_config_from_sectioned_data(load_csv(TEST_INPUT_DIR / "client_data.csv"), "", optimize_roth=True)


def _project_metrics(c):
    from src.planning_engines import project
    rows = project(c)
    terminal = rows[-1]
    first = rows[0]
    first_rmd = next((r for r in rows if float(r.get("rmd_total", 0) or 0) > 0), None)
    first_conv = next((r for r in rows if float(r.get("roth_conv", 0) or 0) > 0), None)
    return {
        "plan_start": int(c["plan_start"]),
        "plan_end": int(c["plan_end"]),
        "row_count": len(rows),
        "terminal_year": int(terminal["year"]),
        "terminal_total_nw": round(float(terminal.get("total_nw", 0) or 0), 2),
        "terminal_liquid_nw": round(float(terminal.get("pretax_nw", 0) or 0) + float(terminal.get("roth_nw", 0) or 0) + float(terminal.get("trust_nw", 0) or 0) + float(terminal.get("hsa_nw", 0) or 0), 2),
        "lifetime_tax": round(sum(float(r.get("total_tax", 0) or 0) for r in rows), 2),
        "total_roth_conversion": round(sum(float(r.get("roth_conv", 0) or 0) for r in rows), 2),
        "first_year_total_tax": round(float(first.get("total_tax", 0) or 0), 2),
        "first_rmd_year": int(first_rmd["year"]) if first_rmd else None,
        "first_rmd_total": round(float(first_rmd.get("rmd_total", 0) or 0), 2) if first_rmd else 0,
        "first_conversion_year": int(first_conv["year"]) if first_conv else None,
        "first_conversion_amount": round(float(first_conv.get("roth_conv", 0) or 0), 2) if first_conv else 0,
        "selected_roth_strategy": (c.get("roth_optimization") or {}).get("selected_label", ""),
    }


# Phase5GoldenMasterEngineTests was deleted (system review 2026-08-04, quality
# finding `golden-master-live-plan-duplication`). Its five mutator scenarios
# duplicated five identically-named scenarios that tests/test_synthetic_golden_master.py
# already gates exactly, to the cent and mandatorily, where this was warn-only
# with a $50,000 tolerance. Its stated reason for existing -- proving the
# advisor's real plan still loads and projects -- no longer applied either:
# tests now resolve through the frozen fixture (see conftest), so it had
# stopped touching the live plan at all. tests/fixtures/golden_master_engine_cases.json
# went with it.

# Phase5ClosedFormTaxTests, Phase5IRSExampleReconciliationTests and
# Phase5CrossToolReconciliationTests moved to tests/test_core_tax_math.py
# (system review 2026-08-04, `buried-tax-math-unit-tests`): src/taxes.py is the
# highest-fan-in domain module in the codebase and its only closed-form tests
# lived in this file, whose name and other contents are about PDF structure and
# a validation-maturity roadmap item. Moved verbatim; fixtures unchanged.

@pytest.mark.slow
@pytest.mark.e2e
class Phase5WorkbookSnapshotTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.tmp = tempfile.mkdtemp(prefix="phase5_workbook_")
        tmp_root = Path(cls.tmp)
        # The reporting stack resolves its project root from source-file
        # location, not from the process cwd, so building "in-place" against
        # ROOT would read/write the real input/, system_config.csv, and
        # output/ files. Copy the tree into a scratch dir and build there
        # instead; local_state/ and output/ are deliberately excluded so the
        # build bootstraps a fresh SQLite mirror rather than touching the
        # real one.
        excluded_names = {".git", ".claude", ".pytest_cache", "tests", "documentation", "output", "local_state", "__pycache__"}
        shutil.copytree(
            ROOT,
            tmp_root,
            ignore=lambda _dir, names: [n for n in names if n in excluded_names or n.endswith(".pyc")],
            dirs_exist_ok=True,
        )
        # Overwrite the copied input/ with the frozen fixture, the same
        # committed/self-contained plan test_199 pins. The copytree above
        # copies the REAL ROOT/input as it happens to sit on disk right now --
        # which can be a blank new-plan workspace (zero starting account
        # balances, the build fails outright) or a live household with real
        # PII. Building against a deterministic fixture instead of "whatever a
        # human last saved" is what makes this an e2e/reporting-contract test
        # rather than a live-plan diagnostic (that diagnostic already exists,
        # warn-only, in test_recommendations_regression.py).
        FROZEN_DIR = ROOT / "tests" / "fixtures" / "sample_plan_frozen"
        tmp_input = tmp_root / "input"

        def _clear_readonly(func, path, _exc_info):
            # Files copied from a git checkout can carry the read-only
            # attribute on Windows, which turns a plain rmtree into
            # PermissionError: WinError 5.
            os.chmod(path, 0o666)
            func(path)

        if tmp_input.exists():
            shutil.rmtree(tmp_input, onerror=_clear_readonly)
        tmp_input.mkdir(parents=True)
        for f in sorted(FROZEN_DIR.iterdir()):
            if f.is_file():
                shutil.copy(f, tmp_input / f.name)
        env = os.environ.copy()
        # Force the subprocess to treat tmp_root (its own copied tree) as the
        # workspace root, overriding any RETIREMENT_SYSTEM_WORKSPACE_ROOT the
        # parent test process has set (tests/conftest.py sets one so the
        # suite never mutates the real input/ files). Without this override,
        # the subprocess would inherit the parent's redirect and read/write
        # someone else's temp workspace instead of this test's own copy.
        env["RETIREMENT_SYSTEM_WORKSPACE_ROOT"] = str(tmp_root)
        env["RETIREMENT_MC_SIMS"] = "16"
        env["RETIREMENT_MC_SENSITIVITY_SIMS"] = "3"
        env["RETIREMENT_SKIP_REPORT_SIDECARS"] = "1"
        result = subprocess.run([sys.executable, "tools/build_workbook.py"], cwd=tmp_root, text=True, capture_output=True, env=env, timeout=120)
        cls.build_stdout = result.stdout + result.stderr
        if result.returncode != 0:
            raise AssertionError(cls.build_stdout)
        cls.workbook_path = tmp_root / "output" / "retirement_plan.xlsx"

    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(cls.tmp, ignore_errors=True)

    def test_workbook_snapshot_sheets_and_key_phrases(self):
        import openpyxl
        snap = json.loads((FIXTURES / "workbook_snapshot_expectations.json").read_text(encoding="utf-8"))
        wb = openpyxl.load_workbook(self.workbook_path, data_only=True, read_only=True)
        for sheet in snap["required_sheets"]:
            self.assertIn(sheet, wb.sheetnames)
        for sheet, phrases in snap["required_phrases"].items():
            text = "\n".join(str(cell) for row in wb[sheet].iter_rows(values_only=True) for cell in row if cell is not None)
            for phrase in phrases:
                self.assertIn(phrase, text, f"{sheet} missing {phrase}")

    def test_monte_carlo_sheet_discloses_the_asset_location_modeling_limit(self):
        """The success rate must ship with the limit on what it can support.

        Per-account returns reach Monte Carlo as a constant per-bucket return
        offset, not as sleeve-level draws with their own volatility (see
        documentation/reports/PLANNER_SIGNOFF_2026-08-17.md finding S1). Within
        the taxable/pretax/Roth/HSA buckets every account takes the SAME annual
        market shock and differs only by that constant, so the success rate is
        structurally blind to de-risking INSIDE those buckets -- a bond tent or
        glidepath cannot move it, because the bonds fall in lockstep with
        equities.

        The cash bucket is the exception and must not be swept into the same
        claim: it grows on a short-rate path rather than the equity draw, so a
        cash reserve held in a cash account genuinely is modeled as
        low-volatility.

        Asset LOCATION is modeled and may be cited. Asset ALLOCATION de-risking
        within the market buckets may not. Without this disclosure a reader has
        no way to tell which conclusions the number supports, and the natural
        reading -- "my bond tent didn't help" -- is an artifact of the model
        rather than a finding about the plan.
        """
        import openpyxl
        from src.reporting.workbook_format_config import stable_name_for_sheet_title

        wb = openpyxl.load_workbook(self.workbook_path, data_only=True, read_only=True)
        # Resolved through the stable-name API, never a hardcoded '3A. ...':
        # section letters are recomputed per build and shift whenever an
        # optional module is toggled.
        mc_sheets = [
            name for name in wb.sheetnames
            if stable_name_for_sheet_title(name) == '15. Market-Luck Stress Test'
        ]
        if not mc_sheets:
            self.skipTest('Monte Carlo module is off in this build; no sheet to disclose on')

        text = "\n".join(
            str(cell)
            for row in wb[mc_sheets[0]].iter_rows(values_only=True)
            for cell in row if cell is not None
        )
        self.assertIn(
            'cannot credit a bond tent or de-risking glidepath held inside retirement accounts '
            'with reducing failure risk',
            text,
            f"{mc_sheets[0]} presents a success rate without disclosing that the model "
            "cannot see allocation de-risking. Restore the disclosure in "
            "sheets_stress.py's Methodology table rather than deleting this test.",
        )

    def test_hsa_schedule_discloses_its_modeling_limits(self):
        """A proposed HSA drawdown schedule must ship with its limits.

        The schedule lands on the Roth Conversion sheet because it is not an
        independent recommendation: it and the conversion plan are scored
        against one shared objective (`joint_headroom_used`/`allocate_surplus`),
        so they must be read together or not at all.

        NOTE ON WHAT THIS GUARD CAN AND CANNOT PROVE. The section is gated on
        `hsa_withdrawal_mode == 'optimize'`, and the frozen fixture this class
        builds against does not set that mode -- so on the frozen build there is
        no schedule and this test's assertions do not run. That is the correct
        outcome, not a defect: the modes that produce no schedule must not
        render a section. It does mean this test alone is NOT evidence the
        disclosure content is right. `HsaScheduleSheetSectionTests` in
        tests/test_hsa_optimizer_regression.py pins the text and the table
        directly, with the mode on and no build in the loop; that is the guard
        that can actually fail on wording. This one guards the delivery path:
        that the section reaches a real workbook at all when it does render.
        """
        import openpyxl
        from src.reporting.workbook_format_config import stable_name_for_sheet_title

        wb = openpyxl.load_workbook(self.workbook_path, data_only=True, read_only=True)
        # Resolved through the stable-name API, never a hardcoded '11. ...':
        # section letters/numbers are recomputed per build and shift whenever an
        # optional module is toggled.
        sheets = [
            name for name in wb.sheetnames
            if stable_name_for_sheet_title(name) == '11. Roth Conversion'
        ]
        if not sheets:
            self.skipTest('Roth Conversion sheet is off in this build; nowhere to disclose')

        text = "\n".join(
            str(cell)
            for row in wb[sheets[0]].iter_rows(values_only=True)
            for cell in row if cell is not None
        )
        if 'PROPOSED HSA DRAWDOWN SCHEDULE' not in text:
            self.skipTest(
                'HSA schedule optimizer is off in this build '
                "(hsa_withdrawal_mode is not 'optimize'); no schedule to disclose on"
            )
        self.assertIn('optimized on the deterministic path', text)
        self.assertIn('shares the Roth conversion objective', text)

    def test_monte_carlo_sheet_does_not_credit_the_liquidity_buffer_with_mitigating_sequence_risk(self):
        """The narrative sections must not claim what the Methodology section disclaims.

        The Interpretation section used to tell the reader that "the configured
        liquidity buffer (Trust accounts) is the primary mitigation" for
        sequence-of-returns risk. Three independent facts about this engine make
        that false, and it is client-facing advice, not description:

        1. Trust maps to the TAXABLE bucket, and within a bucket every account
           takes the same annual market shock (finding S1). A reserved dollar is
           no less exposed to an early bear market than any other taxable dollar.
        2. What the buffer actually does is set a floor under the taxable draw
           (planning_engines.liquidity_buffer_years_for_year, consumed by
           withdraw_taxable_trust). That is a withdrawal-ORDER preference: it
           redirects spending to other buckets. It never changes any dollar's
           volatility.
        3. The floor is applied in the deterministic cascade only. The
           vectorized MC scales the deterministic engine's planned bucket
           withdrawals and never re-enforces the floor against shocked balances.

        The genuine, modeled mitigation is the cash bucket: it is drawn first and
        grows on a short-rate path rather than the equity draw. The sheet may
        cite that, and must not let it be confused with the buffer -- selecting
        'Cash' as a Liquidity Buffer row's reserve_account does NOT move money
        into that bucket; only holding the reserve in a cash-type account does.
        """
        import openpyxl
        from src.reporting.workbook_format_config import stable_name_for_sheet_title

        wb = openpyxl.load_workbook(self.workbook_path, data_only=True, read_only=True)
        mc_sheets = [
            name for name in wb.sheetnames
            if stable_name_for_sheet_title(name) == '15. Market-Luck Stress Test'
        ]
        if not mc_sheets:
            self.skipTest('Monte Carlo module is off in this build; no sheet to make claims on')

        text = "\n".join(
            str(cell)
            for row in wb[mc_sheets[0]].iter_rows(values_only=True)
            for cell in row if cell is not None
        )

        # The unsupported claim, in the two forms it shipped in (section G's
        # "primary mitigation" and section E's "riding out early bear markets").
        for forbidden in (
            'is the primary mitigation',
            'without forced selling',
        ):
            self.assertNotIn(
                forbidden, text,
                f"{mc_sheets[0]} credits a mitigation the model does not simulate. "
                "Reserved taxable/Trust dollars take the same market shock as the rest "
                "of the bucket; see this test's docstring before re-wording.",
            )

        # And the correction must be present, not merely the claim absent.
        self.assertIn(
            'every reserved dollar still takes the full taxable-bucket market shock',
            text,
            f"{mc_sheets[0]} no longer tells the reader what the liquidity buffer does "
            "and does not do. Restore the correction in sheets_stress.py section G "
            "rather than deleting this test.",
        )

    def test_insurance_sheet_prints_no_client_independent_dollar_figures(self):
        """Insurance verdicts must describe THIS household, not a fixed example.

        Finding C2 was that the Executive Summary published literal dollar
        amounts gated on booleans but never varying with the plan. It was fixed
        at the location the review cited; P6 found the same class alive on the
        insurance sheet, which C2 never named:

        - "$500K face, start 2027, ~$18,500/yr" printed in the same table row
          whose Death Benefit column renders the CONFIGURED face, so a household
          with different settings got a row contradicting itself.
        - "IL estate tax > $320K" -- the exact figure C2 was raised about,
          computed nowhere, while summary_figures.credit_shelter_trust_savings
          derives the real one from the user-editable il_exempt.
        - A flat $500,000 "Estate Liquidity Buffer" need, sitting one row under
          Section B's own note boasting that these needs come from the
          household's projection rather than a generic multiple.
        - A closing recommendation naming a specific commercial product.

        The indicative premium table in Section C is deliberately still present:
        it is a market-pricing comparison, and it is now labeled as indicative
        rather than presented as a plan output.
        """
        import openpyxl
        from src.reporting.workbook_format_config import stable_name_for_sheet_title

        wb = openpyxl.load_workbook(self.workbook_path, data_only=True, read_only=True)
        sheets = [
            name for name in wb.sheetnames
            if stable_name_for_sheet_title(name) == '19. Life Insurance'
        ]
        if not sheets:
            self.skipTest('Life insurance module is off in this build; no sheet to check')

        text = "\n".join(
            str(cell)
            for row in wb[sheets[0]].iter_rows(values_only=True)
            for cell in row if cell is not None
        )

        for forbidden in (
            '$500K face, start 2027',
            '$320K',
            'Lincoln MoneyGuard',
            '★ RECOMMENDED',
            '★ OPTIMAL',
        ):
            self.assertNotIn(
                forbidden, text,
                f"{sheets[0]} prints '{forbidden}', a client-independent figure or a "
                "product name, as though it were advice derived from this plan. "
                "Derive it from the configured values or drop the figure (finding C2's class).",
            )

        # Absence is not enough -- a sheet that dropped the verdicts entirely
        # would pass the loop above. Pin that the derived text is present.
        self.assertTrue(
            'Configured:' in text or 'Not currently configured' in text,
            f"{sheets[0]} no longer states the plan's own hybrid policy configuration.",
        )
        self.assertIn(
            'Estate Liquidity Buffer (projected estate tax at the terminal estate)', text,
            f"{sheets[0]} no longer labels where the estate-liquidity need comes from.",
        )

    def test_workbook_snapshot_rejects_stale_roth_language(self):
        import openpyxl
        snap = json.loads((FIXTURES / "workbook_snapshot_expectations.json").read_text(encoding="utf-8"))
        wb = openpyxl.load_workbook(self.workbook_path, data_only=True, read_only=True)
        combined = "\n".join(
            str(cell)
            for sheet in ["1A. Executive Summary", "4B. Assumptions", "2A. Roth Conversion"]
            if sheet in wb.sheetnames
            for row in wb[sheet].iter_rows(values_only=True)
            for cell in row
            if cell is not None
        )
        for forbidden in snap["forbidden_roth_phrases"]:
            self.assertNotIn(forbidden, combined)


if __name__ == "__main__":
    unittest.main()
