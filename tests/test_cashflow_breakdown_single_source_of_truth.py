"""The cash-flow breakdown is a single source of truth.

`deterministic_engine` attaches a read-only `row['cashflow_breakdown']` to every
projection row, and all four cash-flow consumers read it verbatim instead of
each re-deriving their own itemization:

  * src/reporting/sheets_projection_cashflow.py::build_sheet6  (Excel "1C. Cash Flow")
  * src/reporting/sheets_projection_charts.py::build_sheet8    (Excel "Charts Dashboard")
  * src/results_model.py::_cashflow_page                       (UI "1C. Cash Flow")
  * src/results_model.py::_chart_page                          (UI cash-flow charts)

Before this, the two charts (#3/#4) itemized a "Surplus (Reinvested)" plug from
their OWN locally re-derived totals, which disagreed with the worksheet's
row['surplus'] by tens of thousands of dollars in many years (the residual the
user reported). These tests pin the guarantee that every consumer now shows the
same income / spend / tax / surplus figures for a given year.

Empirically (live household, frozen pricing) the breakdown's four sub-totals
reconcile against surplus/unfunded_gap with a *zero* residual -- the itemization
categories were designed to sum exactly to the engine's own total_spend /
total_tax, and surplus is read authoritatively rather than re-derived -- so the
tolerances below are tiny (rounding only), not the multi-thousand-dollar gaps
the plug produced.
"""
from pathlib import Path

from openpyxl import Workbook

from src.data_io import load_csv, parse_client
from src.plan_config import ensure_engine_config
from src.planning_engines import project
from src.reporting.sheets_projection_cashflow import build_sheet6
from src.reporting.sheets_projection_charts import build_sheet8
from src.results_model import _cashflow_page, _chart_page, _row_cashflow_breakdown
from tests.golden_pricing import FROZEN_GOLDEN_MASTER_PRICES, frozen_holdings_prices

ROOT = Path(__file__).resolve().parents[1]


def _real_rows():
    c = ensure_engine_config(parse_client(load_csv(ROOT / 'input' / 'client_data.csv'), ''), source='test')
    with frozen_holdings_prices(FROZEN_GOLDEN_MASTER_PRICES):
        rows = project(c)
    return c, rows


def _col_by_header(ws, header_row, text):
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=header_row, column=col).value == text:
            return col
    raise AssertionError(f"header {text!r} not found in row {header_row}")


# ── 1. Breakdown exists and reconciles exactly ──────────────────────────────

def test_breakdown_present_and_reconciles_to_surplus():
    _c, rows = _real_rows()
    assert rows
    for r in rows:
        assert 'cashflow_breakdown' in r, f"year {r['year']} missing cashflow_breakdown"
        b = r['cashflow_breakdown']
        inc = sum(b['income'].values())
        draws = sum(b['draws'].values())
        exp = sum(b['expense'].values())     # includes other_cash_need
        tax = sum(b['tax'].values())
        # The itemized tax dict sums to the engine's authoritative total_tax
        # (exactly, by construction -- the 'other' key is the reconciling
        # remainder for equity-comp AMT/LTCG etc.).
        assert abs(tax - r['total_tax']) < 1e-6, f"year {r['year']}: tax {tax} != total_tax {r['total_tax']}"
        # Spend components (all expense keys except other_cash_need) sum to the
        # engine's total_spend need.
        exp_spend = exp - b['expense']['other_cash_need']
        assert abs(exp_spend - r['total_spend']) < 1e-6, f"year {r['year']}: spend mismatch"
        # Master identity: income + draws - (spend + tax) == surplus - unfunded.
        residual = (inc + draws) - (exp + tax) - (b['surplus'] - b['unfunded_gap'])
        assert abs(residual) < 1e-6, f"year {r['year']}: reconciliation residual {residual}"
        # surplus is the authoritative engine value, not re-derived.
        assert b['surplus'] == r['surplus']
        assert b['unfunded_gap'] == r['unfunded_gap']


# ── 2. UI worksheet and UI charts agree on every overlapping figure ─────────

def test_ui_worksheet_surplus_and_income_match_engine():
    """The UI worksheet's Σ Income and Surplus columns are now sourced from the
    engine breakdown, so they equal the engine's own figures. The Surplus
    column is exactly what disagreed with the chart before -- the chart used a
    locally re-derived plug; both now read row['surplus'] verbatim."""
    c, rows = _real_rows()
    cf_page = _cashflow_page(c, rows)
    section = cf_page['sections'][0]
    data_rows = section['rows'][2:]              # rows[0]=groups, rows[1]=headers
    hdr = [hc['value'] for hc in section['rows'][1]['cells']]
    inc_idx = hdr.index('Σ Income')
    surplus_idx = hdr.index('Surplus')
    for i, r in enumerate(rows):
        bd = r['cashflow_breakdown']
        ws_inc = data_rows[i]['cells'][inc_idx]['value']
        ws_surplus = data_rows[i]['cells'][surplus_idx]['value']
        assert abs(ws_inc - sum(bd['income'].values())) <= 1, f"year {r['year']}: Σ Income"
        assert abs(ws_surplus - r['surplus']) <= 1, f"year {r['year']}: surplus"


def test_ui_cashflow_bars_reconcile_via_authoritative_surplus():
    """Income & Portfolio Draws bar total == Spending & Taxes bar total for
    every year, using the engine's authoritative surplus/unfunded_gap (not a
    chart-local plug). This holds ONLY because the expense chart now itemizes
    the full tax -- including the "LTCG Tax" series carrying portfolio LTCG
    beyond the home-sale portion, and an "Other Cash Need" series. Without
    those (the pre-fix state), a year with taxable trust draws would leave the
    two bars differing by the un-itemized LTCG (up to ~$349 for this
    household), far exceeding this per-series rounding tolerance -- that is
    precisely the drift the single-source fix removes. Compaction preserves
    per-year totals, so this holds even though both bars fold small series into
    'Other'. The <=10 tolerance is only per-series $-rounding across ~30
    independently-rounded series (matches the sibling reconciliation test in
    tests/test_cash_flow_chart_series_compaction.py)."""
    c, rows = _real_rows()
    page, _ = _chart_page(c, rows, None)
    inc_chart = next(ch for ch in page['charts'] if 'Portfolio Draws' in ch['title'])
    exp_chart = next(ch for ch in page['charts'] if 'Spending & Taxes' in ch['title'])
    n = min(len(inc_chart['x']), len(exp_chart['x']))
    for i in range(n):
        inc_t = sum(s['values'][i] for s in inc_chart['series'])
        exp_t = sum(s['values'][i] for s in exp_chart['series'])
        assert abs(inc_t - exp_t) <= 10, f"year index {i}: income {inc_t} vs spend+surplus {exp_t}"


# ── 3. Excel sheets agree with each other and with the UI ───────────────────

def test_excel_sheets_agree_with_ui_on_income_and_surplus():
    c, rows = _real_rows()

    wb6 = Workbook()
    build_sheet6(wb6.active, c, rows)
    ws6 = wb6.active
    c6_inc = _col_by_header(ws6, 2, 'Σ Income')
    c6_gap = _col_by_header(ws6, 2, 'Reinvested Surplus (forced income > need)')
    c6_tcn = _col_by_header(ws6, 2, 'Total Cash Need')

    wb8 = Workbook()
    ws8_visible = wb8.active
    ws8_visible.title = '1E. Charts'
    build_sheet8(ws8_visible, c, rows)
    data8 = wb8['_Chart Dashboard Data']
    c8_inc = _col_by_header(data8, 4, 'Σ Income')
    c8_surplus = _col_by_header(data8, 4, 'Surplus (Reinvested)')

    for i, r in enumerate(rows):
        bd = r['cashflow_breakdown']
        engine_streams = sum(bd['income'].values())                      # income only
        engine_income_and_draws = engine_streams + sum(bd['draws'].values())
        engine_surplus = round(bd['surplus'])
        engine_tcn = round(sum(bd['expense'].values()) + sum(bd['tax'].values()))

        # build_sheet6 "Σ Income" is income streams only (draws are a separate
        # section); build_sheet8 "Σ Income" is the Income & Portfolio Draws bar
        # total (streams + positive draws). Each matches the engine's own
        # corresponding total.
        s6_inc = ws6.cell(row=3 + i, column=c6_inc).value
        s6_gap = ws6.cell(row=3 + i, column=c6_gap).value          # = -surplus (+unfunded)
        s6_tcn = ws6.cell(row=3 + i, column=c6_tcn).value
        s8_inc = data8.cell(row=5 + i, column=c8_inc).value
        s8_surplus = data8.cell(row=5 + i, column=c8_surplus).value

        assert abs(s6_inc - engine_streams) <= 1, f"year {r['year']}: sheet6 Σ Income"
        # build_sheet8's bar total sums 15 independently-rounded display
        # components (10 income streams + 5 draw types), unlike sheet6's single
        # raw-sum-then-format cell -- worst-case cumulative rounding drift is
        # 15 * 0.5 = 7.5, not the ~$1-2 seen on other years' data.
        assert abs(s8_inc - engine_income_and_draws) <= 8, f"year {r['year']}: sheet8 Σ Income"
        # Surplus agrees: build_sheet8's series and build_sheet6's (negated)
        # cash-bridge gap both equal the engine's authoritative value.
        assert abs(s8_surplus - engine_surplus) <= 1, f"year {r['year']}: sheet8 surplus"
        assert abs((-s6_gap) - engine_surplus) <= 1, f"year {r['year']}: sheet6 cash-bridge gap vs surplus"
        # Total cash need agrees between build_sheet6 and the engine breakdown.
        assert abs(s6_tcn - engine_tcn) <= 1, f"year {r['year']}: sheet6 total cash need"


# ── 4. Fallback + breakdown-wins (would have caught the prior drift) ────────

def test_consumers_prefer_breakdown_over_stale_legacy_fields():
    """Guard: a consumer must read row['cashflow_breakdown'], not the legacy
    per-field row values. We poison a row's legacy 'surplus'/'fed_tax' fields
    while leaving an intact breakdown, and confirm the UI still reports the
    breakdown's numbers. Under the pre-fix code (which read row.get('surplus')
    and re-derived a plug from raw fields) this would surface the poisoned
    value; now the breakdown wins."""
    c, rows = _real_rows()
    # Deep-ish copy of a single mid-horizon row, then corrupt legacy scalars.
    victim = dict(rows[5])
    good_surplus = victim['cashflow_breakdown']['surplus']
    victim['surplus'] = good_surplus + 999_999.0     # legacy field corrupted
    victim['fed_tax'] = victim.get('fed_tax', 0.0) + 999_999.0
    one = [victim]

    page = _cashflow_page(c, one)
    section = page['sections'][0]
    hdr = [hc['value'] for hc in section['rows'][1]['cells']]
    surplus_idx = hdr.index('Surplus')
    shown = section['rows'][2]['cells'][surplus_idx]['value']
    assert abs(shown - good_surplus) <= 1, f"cashflow page showed {shown}, expected breakdown surplus {good_surplus}"


def test_fallback_reconstructs_breakdown_for_engineless_row():
    """A hand-built row lacking 'cashflow_breakdown' must not crash the UI:
    _row_cashflow_breakdown reconstructs an equivalent dict from legacy fields.
    Guards the defensive fallback that build_sheet6 / build_sheet8 / the UI all
    rely on for rows not produced by the full deterministic engine."""
    synthetic = {
        'year': 2030, 'earned': 100_000.0, 'h_ss': 0.0, 'w_ss': 0.0, 'pension': 0.0,
        'spend_base_yr': 40_000.0, 'mortgage': 20_000.0, 'fed_tax': 15_000.0,
        'state_tax': 5_000.0, 'ltcg_tax': 1_000.0, 'home_sale_tax': 0.0,
        'surplus': 20_000.0, 'unfunded_gap': 0.0,
    }
    bd = _row_cashflow_breakdown(synthetic)
    assert bd['income']['earned'] == 100_000.0
    assert bd['tax']['federal'] == 15_000.0
    assert bd['tax']['ltcg'] == 1_000.0            # ltcg - home_sale
    assert bd['surplus'] == 20_000.0
    # The UI cash-flow page (which reads every field via .get) renders the
    # engineless row without crashing, using the reconstructed breakdown.
    page = _cashflow_page({'h_name': 'A', 'w_name': 'B', 'state': 'IL'}, [synthetic])
    hdr = [hc['value'] for hc in page['sections'][0]['rows'][1]['cells']]
    surplus_idx = hdr.index('Surplus')
    assert page['sections'][0]['rows'][2]['cells'][surplus_idx]['value'] == 20_000.0
