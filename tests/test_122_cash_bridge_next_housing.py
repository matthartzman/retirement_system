from openpyxl import Workbook

from src.reporting.sheets_projection import build_sheet5, build_sheet6


def _config():
    return {
        'plan_start': 2026,
        'plan_end': 2026,
        'post_sale_rent_mo': 0,
        'ltc_enabled': False,
        'ltc_annual_prem': 0,
        'state': 'IL',
        'inf': 0.025,
        'h_name': 'Matthew',
        'w_name': 'Patricia',
        'account_registry': [],
        'pre_tax_ids': [],
        'roth_ids': [],
        'taxable_ids': [],
        'hsa_ids': [],
        'cash_ids': [],
    }


def _row():
    return {
        'year': 2026,
        'h_age': 60,
        'w_age': 60,
        'earned': 1000,
        'h_ss': 0,
        'w_ss': 0,
        'pension': 0,
        'wife_single_ann': 0,
        'wife_joint_ann': 0,
        'h_single_ann': 0,
        'h_joint_ann': 0,
        'note_princ': 0,
        'note_int': 0,
        'rmd_total': 0,
        'portfolio_income_total': 50,
        'roth_conv': 0,
        'agi': 1000,
        'taxable_inc': 0,
        'fed_tax': 100,
        'state_tax': 20,
        'niit': 0,
        'irmaa': 0,
        'total_tax': 120,
        'spend_base_yr': 1000,
        'mortgage': 0,
        'mortgage_payment_yr': 0,
        'real_estate_tax_yr': 0,
        'home_improvement_yr': 0,
        'rent_yr': 0,
        'housing_utilities_yr': 0,
        'housing_maintenance_yr': 0,
        'housing_other_yr': 0,
        'housing_total_yr': 0,
        'wellness_base_yr': 0,
        'wellness_premiums_yr': 0,
        'wellness_medical_yr': 0,
        'wellness_dental_yr': 0,
        'wellness_vision_yr': 0,
        'wellness_rx_otc_yr': 0,
        'wellness_other_yr': 0,
        'ltc_prem_yr': 0,
        'rec_extra': 0,
        'lump': 0,
        'heloc_interest': 0,
        'heloc_repayment_principal': 0,
        'other_cash_need_yr': 200,
        'total_cash_need': 1320,
        'income_funding': 1050,
        'h_trust_wd': 270,
        'w_trust_wd': 0,
        'hsa_wd': 0,
        'h_roth_wd': 0,
        'w_roth_wd': 0,
        'rmd_h': 0,
        'h_ira_elective': 0,
        'rmd_w': 0,
        'w_ira_elective': 0,
        'heloc_draw': 0,
        'heloc_balance': 0,
        'surplus': 0,
        'unfunded_gap': 0,
        'total_nw': 10000,
        'other_nw': 8000,
        'pension_pv': 0,
        'w_single_pv': 0,
        'w_joint_pv': 0,
        'h_single_pv': 0,
        'h_joint_pv': 0,
        'ann_nw': 0,
        'pretax_nw': 0,
        'roth_nw': 0,
        'trust_nw': 0,
        'hsa_nw': 0,
        'cash_nw': 0,
        'home_equity': 5000,
        'next_housing_equity': 3000,
        'startup_val': 0,
        'autos_val': 0,
        'note_bal': 0,
    }


def test_cash_flow_has_explicit_cash_bridge_columns():
    wb = Workbook()
    ws = wb.active
    build_sheet6(ws, _config(), [_row()])
    headers = [ws.cell(2, col).value for col in range(1, ws.max_column + 1)]
    for header in [
        'Total Taxes', 'Total Cash Need', 'Income Funding',
        'Other Funding', 'Elective Portfolio Withdrawals (excl. RMDs)', 'Reinvested Surplus (forced income > need)',
        'Other Cash Need'
    ]:
        assert header in headers


def test_net_worth_exposes_next_housing_equity():
    wb = Workbook()
    ws = wb.active
    build_sheet5(ws, _config(), [_row()])
    headers = [ws.cell(2, col).value for col in range(1, ws.max_column + 1)]
    assert 'Next Housing Equity' in headers
