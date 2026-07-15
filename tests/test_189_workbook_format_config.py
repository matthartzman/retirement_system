"""Item 189 — Settings → Workbook Formatting column-width editor.

Covers the shared config module (structure tree, override round-trip, and
generation-time application) and the presence of the UI + route wiring.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from src.reporting import workbook_format_config as wf

ROOT = Path(__file__).resolve().parents[1]
BUILT_WORKBOOK = ROOT / "output" / "retirement_plan.xlsx"


def _make_workbook(tmp_path: Path) -> Path:
    """A tiny workbook with one multi-table sheet and one flat sheet."""
    wb = openpyxl.Workbook()
    multi = wb.active
    multi.title = "Multi"
    # Two side-by-side merged banners in row 1 => two tables.
    multi["A1"] = "GROUP ONE"
    multi.merge_cells("A1:B1")
    multi["C1"] = "GROUP TWO"
    multi.merge_cells("C1:D1")
    for col, title in zip("ABCD", ["Year", "Value", "Rate", "Total"]):
        multi[f"{col}2"] = title
    multi.column_dimensions["A"].width = 10
    multi.column_dimensions["B"].width = 12
    multi.column_dimensions["C"].width = 8
    multi.column_dimensions["D"].width = 14

    flat = wb.create_sheet("Flat")
    flat["A1"] = "FLAT REPORT"
    flat.merge_cells("A1:C1")
    for col, title in zip("ABC", ["Label", "Amount", "Note"]):
        flat[f"{col}2"] = title
    for col in "ABC":
        flat.column_dimensions[col].width = 15

    path = tmp_path / "wb.xlsx"
    wb.save(path)
    return path


def test_tree_detects_tables_and_single_table(tmp_path):
    path = _make_workbook(tmp_path)
    tree = wf.build_format_tree(path)
    assert tree["available"] is True
    by_name = {s["sheet"]: s for s in tree["sheets"]}

    multi = by_name["Multi"]
    assert multi["single_table"] is False
    names = [t["name"] for t in multi["tables"]]
    assert "GROUP ONE" in names and "GROUP TWO" in names
    g1 = next(t for t in multi["tables"] if t["name"] == "GROUP ONE")
    titles = {c["col"]: c["title"] for c in g1["columns"]}
    assert titles["A"] == "Year" and titles["B"] == "Value"

    flat = by_name["Flat"]
    assert flat["single_table"] is True
    assert len(flat["tables"]) == 1
    assert flat["tables"][0]["name"] is None
    flat_titles = {c["col"]: c["title"] for c in flat["tables"][0]["columns"]}
    assert flat_titles["A"] == "Label" and flat_titles["B"] == "Amount"


def test_missing_workbook_reports_unavailable(tmp_path):
    tree = wf.build_format_tree(tmp_path / "does_not_exist.xlsx")
    assert tree["available"] is False
    assert tree["sheets"] == []


def test_overrides_round_trip_and_sanitize(tmp_path):
    saved = wf.save_overrides(
        {
            "Multi": {"a": 20, "B": 15.5, "zz": -3, "bad!": 9, "C": "not a number"},
            "": {"A": 5},  # empty sheet name dropped
        },
        input_dir=tmp_path,
    )
    assert saved == {"Multi": {"A": 20.0, "B": 15.5}}
    assert wf.load_overrides(input_dir=tmp_path) == {"Multi": {"A": 20.0, "B": 15.5}}


def test_overrides_clamped(tmp_path):
    saved = wf.save_overrides({"S": {"A": 9999, "B": 0.01}}, input_dir=tmp_path)
    assert saved["S"]["A"] == wf.MAX_WIDTH
    assert saved["S"]["B"] == wf.MIN_WIDTH


def test_apply_overrides_sets_widths(tmp_path):
    path = _make_workbook(tmp_path)
    wf.save_overrides({"Multi": {"A": 42.0}}, input_dir=tmp_path)
    wb = openpyxl.load_workbook(path)
    assert wb["Multi"].column_dimensions["A"].width == 10
    wf.apply_overrides(wb, input_dir=tmp_path)
    assert wb["Multi"].column_dimensions["A"].width == 42.0


def test_overridden_flag_reflects_saved(tmp_path):
    path = _make_workbook(tmp_path)
    overrides = {"Multi": {"C": 30.0}}
    tree = wf.build_format_tree(path, overrides)
    multi = next(s for s in tree["sheets"] if s["sheet"] == "Multi")
    flags = {
        c["col"]: c["overridden"] for t in multi["tables"] for c in t["columns"]
    }
    assert flags["C"] is True
    assert flags["A"] is False


@pytest.mark.skipif(not BUILT_WORKBOOK.exists(), reason="no built workbook present")
def test_real_workbook_multi_table_sheets():
    tree = wf.build_format_tree(BUILT_WORKBOOK)
    assert tree["available"] is True
    by_name = {s["sheet"]: s for s in tree["sheets"]}
    # Net Worth and Cash Flow are the wide matrix sheets with grouped columns.
    if "1B. Net Worth" in by_name:
        assert by_name["1B. Net Worth"]["single_table"] is False
    # Executive Summary is a single-table narrative sheet.
    if "1A. Executive Summary" in by_name:
        assert by_name["1A. Executive Summary"]["single_table"] is True


def test_generation_applies_overrides_hook_present():
    src = (ROOT / "src" / "reporting" / "workbook_builder.py").read_text(encoding="utf-8")
    assert "apply_overrides as _apply_format_overrides" in src
    assert "_apply_format_overrides(wb)" in src


def test_routes_and_ui_wired():
    routes = (ROOT / "src" / "server" / "workbook_routes.py").read_text(encoding="utf-8")
    assert '"/api/workbook-format", methods=["GET"]' in routes
    assert '"/api/workbook-format", methods=["POST"]' in routes
    js = (ROOT / "frontend" / "js" / "dashboard.js").read_text(encoding="utf-8")
    assert "function renderWorkbookFormatting" in js
    assert "data-step-id=\"workbook_formatting\"" in js
    assert "/api/workbook-format" in js
    nav = (ROOT / "frontend" / "js" / "navigation.js").read_text(encoding="utf-8")
    assert "workbook_formatting" in nav
