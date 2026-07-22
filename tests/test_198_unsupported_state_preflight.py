"""Item 1.11: unrecognized residence_state must fail loudly, not silently
borrow Illinois' 4.95% flat rate + retirement-income exemption, and the
State Residency sheet must caveat any comparison built against an unmodeled
client state instead of quietly presenting Illinois numbers as the client's
own.

See PROJECT_MANIFEST.md item 1.11 background: src/core.py used to do
``STATE_TAX_RULES.get(state, STATE_TAX_RULES.get('Illinois', ...))`` with no
warning anywhere in the 25-sheet workbook.
"""
import pytest
from openpyxl import Workbook

from src.core import state_income_tax, supported_states, STATE_TAX_RULES
from src.reporting.sheets_strategy import build_sheet13

SUPPORTED = (
    'Arizona', 'California', 'Colorado', 'Florida', 'Illinois', 'Indiana',
    'Nevada', 'New York', 'North Carolina', 'South Dakota', 'Tennessee',
    'Texas', 'Wyoming',
)


def test_supported_states_is_derived_from_state_tax_rules():
    # Single source of truth: no second hardcoded list to drift out of sync.
    assert supported_states() == tuple(sorted(STATE_TAX_RULES.keys()))
    assert set(supported_states()) == set(SUPPORTED)
    assert len(supported_states()) == 13


def test_unsupported_state_raises_readable_preflight_error():
    with pytest.raises(ValueError) as exc_info:
        state_income_tax('Minnesota', 50_000, 40_000, 20_000, 5_000, 0, 0, 2026)
    message = str(exc_info.value)
    # Names the offending state...
    assert "Minnesota" in message
    # ...and lists the supported set so the user knows what to change to.
    for state in SUPPORTED:
        assert state in message
    # ...and points at the extension path instead of leaving the user stuck.
    assert 'reference_data/state_tax.csv' in message


def test_unsupported_state_error_is_a_plain_valueerror():
    # ValueError is the established config-error convention in this codebase
    # (see plan_config.normalize_engine_config) and is what the build-job
    # failure-message extractor (build_job_service.extract_build_failure_message)
    # scans stderr for via the "SomeError: message" pattern — this is how the
    # error reaches the user as a clean message instead of a raw traceback.
    with pytest.raises(ValueError):
        state_income_tax('New Jersey', 0, 0, 0, 0, 0, 0, 2026)


def test_illinois_happy_path_unaffected():
    # The live plan uses Illinois; behavior must be byte-for-byte the same
    # as before this item (Illinois exempts retirement income and SS, flat
    # 4.95% on everything else).
    tax = state_income_tax(
        state='Illinois', earned=100_000, retirement_dist=40_000,
        ss_taxable=20_000, investment_inc=10_000, nonqual_annuity=5_000,
        roth_conv=0, year=2026, age_over_65=True, filing='MFJ',
    )
    # Illinois exempts retirement distributions and SS: only earned +
    # investment + nonqual annuity (115,000) is taxed at the flat 4.95% rate.
    expected = (100_000 + 10_000 + 5_000) * 0.0495
    assert tax == pytest.approx(expected)


def test_blank_state_still_falls_back_silently_not_bricked():
    # residence_state is a required Plan Data field (reference_data/schema.csv)
    # already gated by the separate "missing required field" preflight check.
    # A blank string reaching state_income_tax (e.g. an in-progress autosave
    # backup) must not newly hard-fail here — that would brick loading of
    # existing incomplete snapshots for a problem this item isn't meant to
    # police. It keeps borrowing Illinois' rate exactly as before.
    tax_blank = state_income_tax('', 100_000, 0, 0, 0, 0, 0, 2026)
    tax_illinois = state_income_tax('Illinois', 100_000, 0, 0, 0, 0, 0, 2026)
    assert tax_blank == pytest.approx(tax_illinois)


def _sheet13_config(state):
    return {
        'plan_start': 2026, 'plan_end': 2027,
        'home_val': 500_000, 'home_appr': 0.03,
        'il_exempt': 4_000_000,
        'state': state, 'residency_target_state': '',
        'h_name': 'Matthew', 'w_name': 'Patricia',
    }


def _sheet13_rows():
    return [
        {'year': 2026, 'state_earned_net': 0, 'state_retirement': 50_000,
         'state_ss_taxable': 20_000, 'state_investment': 5_000,
         'state_nonqual_ann': 0, 'state_roth_conv': 0, 'agi': 75_000,
         'total_nw': 1_000_000, 'spend_base_yr': 60_000},
        {'year': 2027, 'state_earned_net': 0, 'state_retirement': 52_000,
         'state_ss_taxable': 21_000, 'state_investment': 5_200,
         'state_nonqual_ann': 0, 'state_roth_conv': 0, 'agi': 78_000,
         'total_nw': 1_050_000, 'spend_base_yr': 61_000},
    ]


def test_state_residency_sheet_caveats_unmodeled_client_state():
    ws = Workbook().active
    build_sheet13(ws, _sheet13_config('Minnesota'), _sheet13_rows())
    caveat = ws.cell(row=3, column=1).value
    assert 'Minnesota' in caveat
    assert 'no modeled state-tax rules' in caveat
    assert 'reference_data/state_tax.csv' in caveat
    # Naming every supported state in the caveat lets the reader see the
    # comparison table below is real, just not for their own state.
    for state in SUPPORTED:
        assert state in caveat
    # The sheet still renders (refuse-vs-caveat: this item chose caveat, not
    # a hard stop) — the normal Section A header follows two rows down.
    assert ws.cell(row=5, column=1).value == 'Lifetime Income Components (State Tax Basis)'


def test_state_residency_sheet_no_caveat_for_supported_state():
    ws = Workbook().active
    build_sheet13(ws, _sheet13_config('Illinois'), _sheet13_rows())
    # No caveat inserted: Section A header is the very first content row.
    assert ws.cell(row=3, column=1).value == 'Lifetime Income Components (State Tax Basis)'


def test_state_residency_sheet_does_not_crash_building_full_comparison_table():
    # Even with an unmodeled client state, Section B only ever iterates the
    # eleven STATE_TAX_RULES keys (never the client's own unmodeled state
    # string), so it must not raise.
    ws = Workbook().active
    build_sheet13(ws, _sheet13_config('Minnesota'), _sheet13_rows())
    # 'Lifetime Tax Burden by State' header should appear somewhere below.
    values = [ws.cell(row=r, column=1).value for r in range(1, 20)]
    assert any(v and 'Lifetime Tax Burden by State' in str(v) for v in values)
