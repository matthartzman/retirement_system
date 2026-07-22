"""T1b (system review 2026-07-21, P5): the State Residency 'Delta vs IL'
baseline must anchor to the client's actual current state (c['state']), not
to whichever state happens to iterate first in STATE_TAX_RULES."""
import unittest

from openpyxl import Workbook

from src.reporting.sheets_strategy import build_sheet13

CONFIG = {
    'plan_start': 2026, 'plan_end': 2027,
    'home_val': 500_000, 'home_appr': 0.03,
    'il_exempt': 4_000_000,
    'residency_target_state': '',
    'h_name': 'Matthew', 'w_name': 'Patricia',
}

ROWS = [
    {'year': 2026, 'state_earned_net': 0, 'state_retirement': 50_000,
     'state_ss_taxable': 20_000, 'state_investment': 5_000,
     'state_nonqual_ann': 0, 'state_roth_conv': 0, 'agi': 75_000,
     'total_nw': 1_000_000, 'spend_base_yr': 60_000},
    {'year': 2027, 'state_earned_net': 0, 'state_retirement': 52_000,
     'state_ss_taxable': 21_000, 'state_investment': 5_200,
     'state_nonqual_ann': 0, 'state_roth_conv': 0, 'agi': 78_000,
     'total_nw': 1_050_000, 'spend_base_yr': 61_000},
]

# Columns in Section B: State, Income Rate, Income Tax, Property Tax,
# Sales Tax, Estate Tax, Total Tax, Delta vs IL, Retirement Income Taxed.
COL_STATE, COL_TOTAL, COL_DELTA = 1, 7, 8


def _section_b_rows(ws):
    header_row = next(
        r for r in range(1, ws.max_row + 1)
        if ws.cell(row=r, column=COL_STATE).value == 'State'
    )
    out = []
    r = header_row + 1
    while ws.cell(row=r, column=COL_STATE).value not in (None, ''):
        out.append({
            'state': ws.cell(row=r, column=COL_STATE).value,
            'total': ws.cell(row=r, column=COL_TOTAL).value,
            'delta': ws.cell(row=r, column=COL_DELTA).value,
        })
        r += 1
    return out


class StateResidencyBaselineAnchorTests(unittest.TestCase):
    def test_baseline_anchors_to_clients_actual_state_not_dict_order(self):
        # Texas is not first in STATE_TAX_RULES insertion order (Illinois is),
        # so this reproduces the old bug where the baseline silently used
        # whichever state iterated first instead of the client's own state.
        c = dict(CONFIG, state='Texas')
        ws = Workbook().active
        build_sheet13(ws, c, ROWS)
        section_b = _section_b_rows(ws)

        current = next(row for row in section_b if 'Texas' in row['state'])
        self.assertEqual(current['delta'], 'Baseline')

        for row in section_b:
            if 'Texas' in row['state']:
                continue
            self.assertAlmostEqual(row['delta'], row['total'] - current['total'], places=2)

    def test_illinois_client_baseline_unaffected(self):
        # Illinois is already first in dict order, so this locks in the
        # unchanged happy path.
        c = dict(CONFIG, state='Illinois')
        ws = Workbook().active
        build_sheet13(ws, c, ROWS)
        section_b = _section_b_rows(ws)

        current = next(row for row in section_b if 'Illinois' in row['state'])
        self.assertEqual(current['delta'], 'Baseline')
        for row in section_b:
            if 'Illinois' in row['state']:
                continue
            self.assertAlmostEqual(row['delta'], row['total'] - current['total'], places=2)


if __name__ == '__main__':
    unittest.main()
