from openpyxl import Workbook

from src.reporting.sheets_projection import build_sheet5


def _row():
    return {
        "year": 2026,
        "h_age": 60,
        "w_age": 59,
        "pension_pv": 0,
        "w_single_pv": 0,
        "w_joint_pv": 0,
        "h_single_pv": 0,
        "h_joint_pv": 0,
        "Husband_401k": 20,
        "Husband_IRA": 100,
        "Wife_IRA": 300,
        "Husband_Roth": 40,
        "Wife_Roth": 50,
        "Husband_Trust": 60,
        "Wife_Trust": 70,
        "Husband_HSA": 80,
        "home_equity": 90,
        "startup_val": 0,
        "autos_val": 0,
        "note_bal": 0,
        "cash_other": 10,
        "ann_nw": 0,
        "pretax_nw": 420,
        "roth_nw": 90,
        "trust_nw": 130,
        "hsa_nw": 80,
        "other_nw": 100,
        "total_nw": 820,
    }


def test_net_worth_account_labels_follow_account_ids_not_sorted_bucket_order():
    # This reproduces the packaged registry order: Husband_401k sorts before
    # Husband_IRA inside pre_tax_ids. The workbook must still bind the Husband
    # IRA label to the Husband_IRA value, not to the first pre_tax_ids item.
    cfg = {
        "plan_start": 2026,
        "plan_end": 2026,
        "pre_tax_ids": ["Husband_401k", "Husband_IRA", "Wife_IRA"],
        "roth_ids": ["Husband_Roth", "Wife_Roth"],
        "taxable_ids": ["Husband_Trust", "Wife_Trust"],
        "hsa_ids": ["Husband_HSA"],
        "account_registry": [
            {"id": "Husband_401k", "owner_idx": 0, "owner_name": "Matthew", "acct_type": "401k", "tax": "pre_tax"},
            {"id": "Husband_IRA", "owner_idx": 0, "owner_name": "Matthew", "acct_type": "traditional_ira", "tax": "pre_tax"},
            {"id": "Wife_IRA", "owner_idx": 1, "owner_name": "Patricia", "acct_type": "traditional_ira", "tax": "pre_tax"},
            {"id": "Husband_Roth", "owner_idx": 0, "owner_name": "Matthew", "acct_type": "roth_ira", "tax": "roth"},
            {"id": "Wife_Roth", "owner_idx": 1, "owner_name": "Patricia", "acct_type": "roth_ira", "tax": "roth"},
            {"id": "Husband_Trust", "owner_idx": 0, "owner_name": "Matthew", "acct_type": "trust", "tax": "taxable"},
            {"id": "Wife_Trust", "owner_idx": 1, "owner_name": "Patricia", "acct_type": "trust", "tax": "taxable"},
            {"id": "Husband_HSA", "owner_idx": 0, "owner_name": "Matthew", "acct_type": "hsa", "tax": "hsa"},
        ],
    }
    wb = Workbook()
    ws = wb.active

    build_sheet5(ws, cfg, [_row()])

    assert ws.cell(2, 10).value == "Husband IRA"
    assert ws.cell(3, 10).value == 100
    assert ws.cell(2, 11).value == "Husband 401k"
    assert ws.cell(3, 11).value == 20
    assert ws.cell(2, 12).value == "Wife IRA"
    assert ws.cell(3, 12).value == 300
    assert ws.cell(2, 14).value == "Husband Roth"
    assert ws.cell(3, 14).value == 40
    assert ws.cell(2, 17).value == "Husband Trust"
    assert ws.cell(3, 17).value == 60
