"""T1a (system review 2026-07-21, P4): the Federal Estate Tax sheet must not
print "no federal tax likely" when a positive federal estate tax is computed
directly above it."""
import unittest
from pathlib import Path

from openpyxl import Workbook

from src.data_io import load_csv, parse_client
from src.plan_config import ensure_engine_config
from src.planning_engines import project
from src.reporting.sheets_strategy import build_sheet14

ROOT = Path(__file__).resolve().parents[1]


def sample_config():
    data = load_csv(ROOT / 'input' / 'client_data.csv')
    c = parse_client(data, '')
    c['roth_policy'] = 'none'
    c['mc_paths'] = 5
    c['mc_sensitivity_sims'] = 1
    return ensure_engine_config(c, source='test')


def federal_estate_section_texts(ws):
    """Cell text between the 'Federal Estate Tax' header and the next
    section header (e.g. Illinois Estate Tax), so state-estate-tax
    ACTION REQUIRED text doesn't leak into the federal-section assertions."""
    values = [
        (r, str(ws.cell(row=r, column=1).value))
        for r in range(1, ws.max_row + 1)
        if ws.cell(row=r, column=1).value is not None
    ]
    start = next(r for r, v in values if v == 'Federal Estate Tax')
    end = next((r for r, v in values if r > start and 'Illinois Estate Tax' in v), ws.max_row + 1)
    return [
        str(ws.cell(row=r, column=col).value)
        for r in range(start, end)
        for col in range(1, 5)
        if ws.cell(row=r, column=col).value is not None
    ]


class EstateNoteConditionalTests(unittest.TestCase):
    def test_below_exemption_shows_no_tax_note(self):
        c = sample_config()
        rows = project(c)
        c['fed_exempt'] = 1_000_000_000.0  # dwarfs any projected estate
        wb = Workbook()
        ws = wb.active
        build_sheet14(ws, c, rows)
        texts = federal_estate_section_texts(ws)
        self.assertTrue(any('below' in t and 'exemption' in t and 'no federal tax likely' in t for t in texts))
        self.assertFalse(any('ACTION REQUIRED' in t for t in texts))

    def test_above_exemption_shows_warning_not_no_tax_note(self):
        c = sample_config()
        rows = project(c)
        c['fed_exempt'] = 1.0  # forces a positive computed federal estate tax
        wb = Workbook()
        ws = wb.active
        build_sheet14(ws, c, rows)
        texts = federal_estate_section_texts(ws)
        self.assertTrue(any('ACTION REQUIRED' in t for t in texts))
        self.assertFalse(any('no federal tax likely' in t for t in texts))


if __name__ == '__main__':
    unittest.main()
