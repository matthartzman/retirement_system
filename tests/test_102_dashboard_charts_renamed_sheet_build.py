from pathlib import Path

from openpyxl import Workbook, load_workbook

from src.detailed_results import workbook_detailed_index, workbook_detailed_sheet
from src.reporting.dashboard import _find_workbook_charts_sheet, build_html_dashboard


def _make_refactored_charts_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = '1E. Charts'
    ws['A1'] = 'Charts'
    helper = wb.create_sheet('_Chart Dashboard Data')
    helper.sheet_state = 'hidden'
    # Net worth block: A:I, headers on row 4, data from row 5.
    helper.cell(4, 1).value = 'Year'
    helper.cell(4, 2).value = 'Pre-Tax IRA/401k'
    helper.cell(4, 3).value = 'Roth'
    helper.cell(4, 9).value = 'TOTAL NW'
    helper.cell(5, 1).value = 2026
    helper.cell(5, 2).value = 100000
    helper.cell(5, 3).value = 50000
    helper.cell(5, 9).value = 150000
    # Income block: K:AA.
    helper.cell(4, 11).value = 'Year'
    helper.cell(4, 12).value = 'Earned Income'
    helper.cell(4, 27).value = 'Total Income'
    helper.cell(5, 11).value = 2026
    helper.cell(5, 12).value = 10000
    helper.cell(5, 27).value = 10000
    # Expense block: AC:AM.
    helper.cell(4, 29).value = 'Year'
    helper.cell(4, 30).value = 'Base Spending'
    helper.cell(4, 39).value = 'Total Expense'
    helper.cell(5, 29).value = 2026
    helper.cell(5, 30).value = 8000
    helper.cell(5, 39).value = 8000
    wb.save(path)


def test_html_dashboard_accepts_refactored_1e_charts_sheet(tmp_path):
    xlsx = tmp_path / 'retirement_plan.xlsx'
    html = tmp_path / 'retirement_dashboard.html'
    _make_refactored_charts_workbook(xlsx)

    wb = load_workbook(xlsx, data_only=True)
    assert _find_workbook_charts_sheet(wb).title == '1E. Charts'
    wb.close()

    build_html_dashboard(
        xlsx,
        html,
        rows=[{'year': 2026, 'total_nw': 150000, 'spend_base_yr': 8000}],
        c={'plan_start': 2026, 'plan_end': 2026, 'state': 'PA', 'positions': {}},
    )
    text = html.read_text(encoding='utf-8')
    assert 'Retirement Plan Dashboard' in text
    assert 'Pre-Tax IRA/401k' in text


def test_results_explorer_treats_1e_charts_as_chart_dashboard(tmp_path):
    xlsx = tmp_path / 'retirement_plan.xlsx'
    _make_refactored_charts_workbook(xlsx)
    index = workbook_detailed_index(xlsx)
    sheet = next(s for s in index['sheets'] if s['name'] == '1E. Charts')
    assert sheet['kind'] == 'chart_dashboard'
    detail = workbook_detailed_sheet(xlsx, '1E. Charts')
    assert detail['sheet']['kind'] == 'chart_dashboard'
