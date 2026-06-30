from openpyxl import Workbook

from src.reporting.workbook_builder import build_workbook_section_divider


def test_section_divider_builds_without_unsupported_write_cell_kwargs():
    wb = Workbook()
    ws = wb.active
    ws.title = "Reports"

    build_workbook_section_divider(ws, {
        "section": "Reports",
        "code": "reports",
        "description": "Generated reporting outputs.",
        "sheets": ["1. Executive Summary"],
    })

    note_values = [cell.value for row in ws.iter_rows() for cell in row if cell.value]
    assert any("output/reporting surface" in str(v) for v in note_values)
    note_cell = next(cell for row in ws.iter_rows() for cell in row if cell.value and "output/reporting surface" in str(cell.value))
    assert note_cell.font.italic is True
