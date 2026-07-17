"""Advanced planning modules (Phase 1, report-only): 529 education, existing
life insurance, disability income, P&C/umbrella, equity comp, special-needs.

Covers the three layers each module touches without a full workbook build:
parsing (data_io.parse_advanced_modules), the sheet builders, and the
optional-module registration that gates/renames them.
"""
import os

import pytest
from openpyxl import Workbook

from src.data_io import load_csv, parse_advanced_modules
from src.report_compute import prepare_config_from_sectioned_data
from src.planning_engines import project
from src.reporting.workbook_common import OPTIONAL_MODULE_SHEETS
from src.reporting.workbook_builder import FINAL_SHEET_RENAMES
from src.reporting.sheets_protection import build_existing_life, build_disability, build_pc_umbrella
from src.reporting.sheets_wealth import (build_education_funding, build_equity_comp,
                                         build_special_needs, build_business_succession)

os.environ.setdefault("RETIREMENT_SYSTEM_DISABLE_LIVE_PRICE_PROVIDERS", "1")

# (toggle key, legacy build-time sheet, final presentation tab, builder)
MODULES = [
    ("education_funding_529",       "30. Education Funding",       "2J. Education Funding",       build_education_funding),
    ("equity_compensation",         "35. Equity Compensation",     "2K. Equity Compensation",     build_equity_comp),
    ("special_needs_planning",      "36. Special-Needs Planning",  "2L. Special-Needs Planning",  build_special_needs),
    ("existing_life_insurance",     "31. Existing Life Insurance", "3D. Existing Life Insurance", build_existing_life),
    ("disability_income_insurance", "32. Disability Income",       "3E. Disability Income",       build_disability),
    ("property_casualty_umbrella",  "33. P&C Umbrella",            "3F. P&C Umbrella",            build_pc_umbrella),
    ("business_succession",         "34. Business Succession",     "2M. Business Succession",     build_business_succession),
]


@pytest.fixture(scope="module")
def cfg_rows():
    data = load_csv("input/client_data.csv")
    c = prepare_config_from_sectioned_data(data)
    return c, project(c)


def test_parse_returns_all_module_keys():
    data = load_csv("input/client_data.csv")
    m = parse_advanced_modules(data)
    for key in ["edu_funding", "life_policies", "disability", "pc_umbrella",
                "equity_comp", "special_needs", "business_succession"]:
        assert key in m, f"{key} missing from parsed advanced modules"
    # Containers are always present even when a section is absent.
    assert isinstance(m["edu_funding"]["accounts"], list)
    assert isinstance(m["life_policies"], list)


def test_parse_classifies_insurance_by_kind():
    data = load_csv("input/client_data.csv")
    m = parse_advanced_modules(data)
    # A life policy carries a face amount; a DI policy a monthly benefit; a P&C
    # policy a coverage limit. Classification must not cross-contaminate.
    for p in m["life_policies"]:
        assert "face_amount" in p
    for p in m["disability"]["policies"]:
        assert p["monthly_benefit"] >= 0 and "coverage_limit" not in p
    for p in m["pc_umbrella"]["policies"]:
        assert "coverage_limit" in p


def test_parse_config_exposes_keys_on_c(cfg_rows):
    c, _ = cfg_rows
    for key in ["edu_funding", "life_policies", "disability", "pc_umbrella",
                "equity_comp", "special_needs"]:
        assert key in c


@pytest.mark.parametrize("toggle,legacy,final,builder", MODULES,
                         ids=[m[0] for m in MODULES])
def test_builder_runs_without_error(cfg_rows, toggle, legacy, final, builder):
    c, rows = cfg_rows
    ws = Workbook().create_sheet("t")
    builder(ws, c, rows)
    # A real report was written (title + at least one data row).
    assert ws.max_row >= 3
    assert ws.max_column >= 1


@pytest.mark.parametrize("toggle,legacy,final,builder", MODULES,
                         ids=[m[0] for m in MODULES])
def test_module_is_registered_for_gating_and_rename(toggle, legacy, final, builder):
    assert OPTIONAL_MODULE_SHEETS.get(toggle) == [legacy], \
        f"{toggle} not registered to {legacy} in OPTIONAL_MODULE_SHEETS"
    assert FINAL_SHEET_RENAMES.get(legacy) == final, \
        f"{legacy} not renamed to {final} in FINAL_SHEET_RENAMES"


def test_empty_sections_render_placeholder():
    # A builder given an empty module must still produce a graceful placeholder
    # sheet rather than raising.
    c = {"plan_start": 2026, "edu_funding": {}, "equity_comp": [], "special_needs": {},
         "life_policies": [], "disability": {}, "pc_umbrella": {}, "business_succession": [],
         "earned": 0, "mort_bal": 0}
    for _, _, _, builder in MODULES:
        ws = Workbook().create_sheet("e")
        builder(ws, c, [])
        assert ws.max_row >= 1
