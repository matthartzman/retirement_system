"""Wave 3.4 (system review 2026-08-04, finding C5 / reporting-nominal-dollars-only):
dual-column nominal + today's-dollars reporting on Sheets 1/5/6/7/15 and the
forecast API, so a 2056 dollar is never presented as directly comparable to a
2026 one without a purchasing-power-adjusted companion nearby.
"""
from __future__ import annotations

from openpyxl import Workbook

from src.core import deflate_to_present
from src.data_io import load_csv, parse_client
from src.plan_config import ensure_engine_config
from src.planning_engines import project
from src.reporting.sheets_summary_builder import build_sheet1
from src.reporting.sheets_projection_net_worth import build_sheet5
from src.reporting.sheets_projection_cashflow import build_sheet6
from src.reporting.sheets_projection_tax import build_sheet7
from tests.golden_pricing import FROZEN_GOLDEN_MASTER_PRICES, frozen_holdings_prices

from conftest import TEST_INPUT_DIR


def sample_config_and_rows():
    c = parse_client(load_csv(TEST_INPUT_DIR / "client_data.csv"), "")
    c["roth_policy"] = "none"
    c["mc_paths"] = 5
    c["mc_sensitivity_sims"] = 1
    c = ensure_engine_config(c, source="test")
    with frozen_holdings_prices(FROZEN_GOLDEN_MASTER_PRICES):
        rows = project(c)
    return c, rows


def test_deflate_to_present_matches_zero_at_plan_start():
    c, rows = sample_config_and_rows()
    assert deflate_to_present(1000.0, c["plan_start"], c) == 1000.0
    later = deflate_to_present(1000.0, c["plan_start"] + 10, c)
    assert later < 1000.0


def _sheet_text(ws):
    cells = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cells.append(str(cell.value))
    return cells


def test_sheet1_shows_a_today_referenced_companion_section():
    # #293 (independent of this item, landed on main): Sheet 1's headline
    # rows were replaced (Terminal Net Worth "(Yn, Today's $)" -> Expected
    # After-Tax LCV / NPV of Future Taxes / Worst-Case Ending Wealth), but
    # the dual-column intent this item exists for is preserved by the
    # "Forward-Looking Metrics (From Today)" section (FCV/EFTR), which is
    # explicitly present-valued from today rather than plan_start.
    c, rows = sample_config_and_rows()
    wb = Workbook()
    ws = wb.active
    with frozen_holdings_prices(FROZEN_GOLDEN_MASTER_PRICES):
        build_sheet1(ws, c, rows, {}, ss_sweep=None)
    text = _sheet_text(ws)
    assert any("From Today" in t for t in text)
    assert any("from today" in t for t in text)


def test_sheet5_summary_has_todays_dollars_column_header():
    c, rows = sample_config_and_rows()
    wb = Workbook()
    ws = wb.active
    build_sheet5(ws, c, rows)
    assert any("Today's $" in t for t in _sheet_text(ws))


def test_sheet6_has_lifetime_totals_todays_dollars_block():
    c, rows = sample_config_and_rows()
    wb = Workbook()
    ws = wb.active
    build_sheet6(ws, c, rows)
    assert any("Today's $" in t for t in _sheet_text(ws))


def test_sheet7_has_todays_dollars_lifetime_total_rows():
    c, rows = sample_config_and_rows()
    wb = Workbook()
    ws = wb.active
    build_sheet7(ws, c, rows)
    assert any("Today's $" in t for t in _sheet_text(ws))
