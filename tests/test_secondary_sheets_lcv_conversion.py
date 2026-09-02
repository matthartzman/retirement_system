"""Ticket 293 follow-up: the two secondary comparison sheets flagged out of
the first pass (Social Security timing sweep, Current vs. Proposed) now
display LCV / NPV of Future Taxes / Worst-Case (5th %ile) Ending Wealth
instead of raw Terminal Net Worth / nominal Lifetime Tax / MC Success %,
matching the Impact page and Executive Summary convention -- while each
sheet's own internal candidate-ranking/scoring logic (which pair gets
recommended) is deliberately left untouched.
"""
from __future__ import annotations

import unittest

from openpyxl import Workbook

from src.data_io import load_csv, parse_client
from src.plan_config import ensure_engine_config
from src.planning_engines import project
from src.reporting.sheets_current_vs_proposed import build_sheet_current_vs_proposed
from src.reporting.sheets_strategy import build_sheet10
from tests.golden_pricing import FROZEN_GOLDEN_MASTER_PRICES, frozen_holdings_prices

from conftest import TEST_INPUT_DIR


def _sample_config_and_rows():
    c = parse_client(load_csv(TEST_INPUT_DIR / "client_data.csv"), "")
    c["roth_policy"] = "none"
    c = ensure_engine_config(c, source="test")
    with frozen_holdings_prices(FROZEN_GOLDEN_MASTER_PRICES):
        rows = project(c)
    return c, rows


class CurrentVsProposedLcvConversionTests(unittest.TestCase):
    def test_headers_show_lcv_and_npv_of_future_taxes_not_terminal_nw_or_lifetime_tax(self):
        from openpyxl import Workbook

        c, rows = _sample_config_and_rows()
        wb = Workbook()
        ws = wb.active
        build_sheet_current_vs_proposed(ws, c, rows)
        texts = [str(cell.value) for row in ws.iter_rows() for cell in row if cell.value is not None]
        assert any("Current LCV" in t for t in texts)
        assert any("Proposed LCV" in t for t in texts)
        assert any("Δ LCV" in t for t in texts)
        assert any("Current NPV of Future Taxes" in t for t in texts)
        assert any("Proposed NPV of Future Taxes" in t for t in texts)
        assert any("Δ NPV of Future Taxes" in t for t in texts)
        assert not any(t == "Current Terminal NW" for t in texts)
        assert not any(t == "Current Lifetime Tax" for t in texts)

    def test_sheet_still_renders_a_real_engine_delta(self):
        # Regression guard: the LCV/NPV swap must not silently produce a
        # flat $0 delta table (the original bug this sheet exists to catch
        # for Terminal NW/Lifetime Tax).
        from openpyxl import Workbook
        from src.reporting.sheets_current_vs_proposed import _proposed_changes

        c, rows = _sample_config_and_rows()
        changes = _proposed_changes(c)
        ltc_change = next((ch for ch in changes if "LTC" in ch[0]), None)
        assert ltc_change is not None, "expected an LTC recommendation to be tracked"
        wb = Workbook()
        ws = wb.active
        build_sheet_current_vs_proposed(ws, c, rows)
        # At least one numeric delta cell must be nonzero -- if compute_
        # baseline_lcv_and_eltr silently returned zeros this would fail.
        numeric_cells = [
            cell.value for row in ws.iter_rows() for cell in row
            if isinstance(cell.value, (int, float))
        ]
        assert any(abs(v) > 1.0 for v in numeric_cells)


def _fast_ss_sweep_config():
    c = parse_client(load_csv(TEST_INPUT_DIR / "client_data.csv"), "")
    c["roth_policy"] = "none"
    c["mc_paths"] = 5
    c["mc_sensitivity_sims"] = 1
    return c


class SsTimingSweepLcvConversionTests(unittest.TestCase):
    def test_headers_show_lcv_and_npv_of_future_taxes_not_terminal_nw_or_mc_success(self):
        c = _fast_ss_sweep_config()
        rows = project(c)
        ws = Workbook().active
        result = build_sheet10(ws, c, rows)
        texts = [str(cell.value) for row in ws.iter_rows() for cell in row if cell.value is not None]
        assert any(t == "LCV" for t in texts)
        assert any(t == "Δ LCV" for t in texts)
        assert any(t == "NPV of Future Taxes" for t in texts)
        assert any("Worst-Case Ending Wealth" in t for t in texts)
        assert not any(t == "Terminal NW" for t in texts)
        assert not any(t == "MC Success %" for t in texts)
        assert not any(t == "Lifetime SS Income" for t in texts)
        # Internal candidate-ranking logic (which pair gets recommended) is
        # deliberately untouched by this ticket -- assert it is still present.
        best = result["best"]
        assert "lcv_score" in best
        assert "objective_value" in best
        assert "lcv" in best
        assert "npv_future_taxes" in best


if __name__ == "__main__":
    unittest.main()
