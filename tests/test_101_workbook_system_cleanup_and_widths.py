import pytest
from openpyxl import load_workbook

from src.reporting.workbook_common import TEMPLATE_LAYOUT, _needed_number_width

pytestmark = pytest.mark.slow


def _visible_wb(workbook_path):
    assert workbook_path.exists(), 'Generated workbook is missing'
    return load_workbook(workbook_path, read_only=False, data_only=False)


def test_system_section_uses_clean_sheet_sequence_without_feature_toggle(built_workbook_path):
    wb = _visible_wb(built_workbook_path)
    visible = [ws.title for ws in wb.worksheets if ws.sheet_state == 'visible']
    expected = [
        '1. Reports','1A. Executive Summary','1B. Net Worth','1C. Cash Flow','1D. Balance Sheet','1E. Charts','1F. Lifetime Taxes',
        # #221: Core Spending merged into Spending Summary -- densely 1G now.
        '1G. Spending Summary',
        # #209/#210/#212/#228: 2J (not the old static 2N) -- this fixture's
        # plan has the advanced modules off, so Gain Harvesting fills the gap
        # densely instead of leaving one.
        '2. Optimizers','2A. Roth Conversion','2B. Asset Allocation','2C. State Residency','2D. Social Security','2E. S-Corp vs LLC','2F. Charitable Giving','2G. Estate & Legacy Planning','2I. Tax-Loss Harvesting','2J. Gain Harvesting',
        '3. Risk & Stress Tests','3A. Monte Carlo','3B. Survivor','3C. LTC + Life Insurance',
        '4. System','4A. Plan Data','4B. Assumptions','2H. Planning Levers','4C. Account Reconciliation','4D. Quality Control','4E. RMD Audit','4F. Methodology','4G. Glossary',
    ]
    assert visible[:len(expected)] == expected
    assert '4D. Feature Toggle' not in visible
    assert '4A. Plan Scope' not in visible


def test_visible_workbook_has_no_stale_feature_or_plan_scope_labels(built_workbook_path):
    wb = _visible_wb(built_workbook_path)
    banned = ['Feature Toggle', 'Feature Toggles', 'FEATURE TOGGLES', 'Feature / Toggle', 'Plan Scope', 'Charts Dashboard', 'System Configuration']
    hits = []
    for ws in wb.worksheets:
        if ws.sheet_state != 'visible':
            continue
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    for term in banned:
                        if term in cell.value:
                            hits.append((ws.title, cell.coordinate, term, cell.value))
    assert not hits


def _expected_width(stable_sheet, col, heuristic_cap):
    """A column's expected width: the reference formatting workbook's exact
    value if it pins this column, else the heuristic cap.

    #209/#210/#212/#228: TEMPLATE_LAYOUT is keyed by each sheet's STABLE
    (build-time) name, not its final letter -- pass the stable name.
    """
    pinned = TEMPLATE_LAYOUT.get(stable_sheet, {}).get('cols', {}).get(col)
    return pinned if pinned is not None else heuristic_cap


def test_column_width_caps_are_applied_without_header_driven_expansion(built_workbook_path):
    wb = _visible_wb(built_workbook_path)
    # The generated layout pass uses Excel character widths approximating the requested pixel caps,
    # except where the reference formatting workbook (template for column widths and height.xlsx)
    # pins an exact width for that column — those exact values win at generation time.
    max_text_width = round((200 - 5) / 7, 1) + 0.1
    max_dollar_width = round((71 - 5) / 7, 1) + 0.1
    max_int_width = round((40 - 5) / 7, 1) + 0.1
    assert wb['4F. Methodology'].column_dimensions['A'].width <= _expected_width('23. Methodology', 'A', max_text_width)

    # RMD Audit column G holds account balances, which can genuinely need
    # more than the hand-tuned template's pinned width (e.g. a 7-figure IRA
    # balance) -- widen_overflowing_number_columns() then grows it past the
    # cap so Excel shows the real value instead of "#####". Allow the width
    # up to what the sheet's actual largest value needs; anything beyond that
    # would signal header-driven (not data-driven) expansion, which this test
    # still guards against.
    rmd_ws = wb['4E. RMD Audit']
    rmd_g_cap = _expected_width('20. RMD Audit', 'G', max_dollar_width)
    g_cells = [
        cell for row in rmd_ws.iter_rows(min_col=7, max_col=7)
        for cell in row if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool)
    ]
    needed_for_data = max((_needed_number_width(c.value, c.number_format) or 0 for c in g_cells), default=0)
    assert rmd_ws.column_dimensions['G'].width <= max(rmd_g_cap, needed_for_data) + 1.0

    assert wb['4E. RMD Audit'].column_dimensions['C'].width <= _expected_width('20. RMD Audit', 'C', max_int_width)
