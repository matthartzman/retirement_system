from __future__ import annotations

"""A home-purchase down payment is real cash the plan must fund, exposed by
the engine as row['cashflow_breakdown']['expense']['other_cash_need'] (see
deterministic_engine._next_housing_for_year). The engine itself and the UI's
_chart_page (results_model.py) already itemized this correctly -- the bug
was in the two workbook-side consumers:

  * sheets_projection_charts.py::build_sheet8 (Excel "Charts Dashboard")
    never added other_cash_need into its expense total (exp_raw) or its
    EXP_SER series list at all, so a home-purchase year's "Cash Flow --
    Spending & Taxes" bar (and its Sigma Spend+Tax total) fell short of the
    "Income & Portfolio Draws" bar by the full down payment.

  * dashboard.py's HTML dashboard read that same helper sheet through
    _extract_chart_block(..., total_col=39), a stale constant left over from
    before EXP_SER grew past 9 series -- silently dropping IRMAA, Home Sale
    Tax, Surplus, HELOC P&I, and (once added) Other Cash Need from the HTML
    chart even after the Excel chart itself was fixed. exp_colors was
    similarly stale at 9 entries.

Originally exercised via the live input/client_spending.csv Housing
next_step_1 row, which at the time configured a real purchase scenario
(Florida, $1,000,000 @ 80% down = an $800,000 down payment in 2038). That
made this suite depend on whatever the advisor's real household happened to
have configured -- the frozen fixture (tests/fixtures/sample_plan_frozen/)
never had an equivalent scenario, so these tests failed outright once the
suite was migrated onto it. Added a fictional purchase scenario directly to
the frozen fixture's Housing next_step_1 row (Texas, $400,000 @ 27% down =
$108,000 down payment in 2036) rather than constructing an ad hoc synthetic
config, preserving the original intent of exercising a real-shaped
household config rather than a hand-built stub.
"""

import re
from pathlib import Path

from openpyxl import Workbook

from src.data_io import load_csv, parse_client
from src.plan_config import ensure_engine_config
from src.planning_engines import project
from src.reporting.dashboard import build_html_dashboard
from src.reporting.sheets_projection_charts import build_sheet8
from tests.golden_pricing import FROZEN_GOLDEN_MASTER_PRICES, frozen_holdings_prices

ROOT = Path(__file__).resolve().parents[1]

from conftest import TEST_INPUT_DIR


def _real_config_and_rows():
    c = ensure_engine_config(parse_client(load_csv(TEST_INPUT_DIR / "client_data.csv"), ""), source="test")
    with frozen_holdings_prices(FROZEN_GOLDEN_MASTER_PRICES):
        rows = project(c)
    return c, rows


def _purchase_year(c):
    steps = c.get("next_housing_steps") or []
    step = next(s for s in steps if s.get("type") == "purchase" and float(s.get("purchase_price", 0) or 0) > 0)
    return int(step["start_year"])


def _built_chart_dashboard_data(c, rows):
    wb = Workbook()
    ws8 = wb.active
    ws8.title = "8. Charts Dashboard"
    with frozen_holdings_prices(FROZEN_GOLDEN_MASTER_PRICES):
        build_sheet8(ws8, c, rows)
    return wb, wb["_Chart Dashboard Data"]


def test_fixture_actually_has_a_home_purchase_scenario():
    # Sanity check that this test isn't accidentally vacuous.
    c, rows = _real_config_and_rows()
    year = _purchase_year(c)
    row = next(r for r in rows if r["year"] == year)
    assert row["cashflow_breakdown"]["expense"]["other_cash_need"] > 100_000


def test_excel_expense_chart_itemizes_the_down_payment():
    c, rows = _real_config_and_rows()
    year = _purchase_year(c)
    row = next(r for r in rows if r["year"] == year)
    engine_down_payment = row["cashflow_breakdown"]["expense"]["other_cash_need"]

    _wb, data = _built_chart_dashboard_data(c, rows)
    header_row = 4
    year_col = 29
    cols = {data.cell(row=header_row, column=col).value: col for col in range(30, 46)}
    assert "Other Cash Need" in cols, "Other Cash Need series missing from the Excel expense chart table"

    year_row = next(r for r in range(5, data.max_row + 1) if data.cell(row=r, column=year_col).value == year)
    shown = data.cell(row=year_row, column=cols["Other Cash Need"]).value
    assert abs(shown - engine_down_payment) <= 1


def test_excel_expense_and_income_bars_reconcile_in_the_purchase_year():
    """Before the fix, omitting other_cash_need from exp_raw left the expense
    bar short by the full down payment (~$800K on the original real
    household this was written against). A small residual is still expected
    -- build_sheet8's fixed column layout doesn't itemize portfolio LTCG
    beyond the home-sale portion (documented, pre-existing, unrelated to
    this bug). On the frozen fixture's $108,000 down payment, funding it
    draws ~$180K from the taxable Trust, realizing enough LTCG (~$1,666 tax
    at this household's income level) to need a wider tolerance than the
    original $1,000 -- still nearly two orders of magnitude tighter than the
    ~$800K bug this guards against, just not three.
    """
    c, rows = _real_config_and_rows()
    year = _purchase_year(c)
    _wb, data = _built_chart_dashboard_data(c, rows)
    year_row = next(r for r in range(5, data.max_row + 1) if data.cell(row=r, column=29).value == year)
    inc_total = data.cell(row=year_row, column=27).value
    exp_total = data.cell(row=year_row, column=44).value
    assert abs(inc_total - exp_total) <= 2000, (
        f"year {year}: Income & Portfolio Draws ({inc_total}) vs Spending & Taxes "
        f"({exp_total}) differ by {inc_total - exp_total}, far more than the small "
        "residual from unitemized LTCG-beyond-home-sale tax"
    )


def test_dashboard_expense_extraction_total_col_matches_actual_sheet_layout():
    """Guards the exact bug class dashboard.py's total_col=39 was: a magic
    number that must track build_sheet8's real Sigma Spend+Tax column
    (EXP_TOTAL_COL) but can silently drift out of sync as EXP_SER grows.
    Computed dynamically from the built sheet, not hardcoded, so this fails
    again the next time a series is added to EXP_SER without updating
    dashboard.py.
    """
    c, rows = _real_config_and_rows()
    _wb, data = _built_chart_dashboard_data(c, rows)
    header_row = 4
    total_col = next(
        col for col in range(30, 80)
        if data.cell(row=header_row, column=col).value == "Σ Spend+Tax"
    )
    src = (ROOT / "src/reporting/dashboard.py").read_text(encoding="utf-8")
    m = re.search(r"exp_years, exp_labels, exp_ser = _extract_chart_block\([^)]*total_col=(\d+)", src)
    assert m, "could not find dashboard.py's expense _extract_chart_block call"
    assert int(m.group(1)) == total_col, (
        f"dashboard.py's expense total_col={m.group(1)} does not match the actual "
        f"Sigma Spend+Tax column ({total_col}) -- series between them are silently "
        "dropped from the HTML dashboard chart"
    )


def test_html_dashboard_expense_chart_includes_all_series_with_matching_colors(tmp_path):
    c, rows = _real_config_and_rows()
    year = _purchase_year(c)
    wb, _data = _built_chart_dashboard_data(c, rows)
    xlsx_path = tmp_path / "wb.xlsx"
    wb.save(xlsx_path)
    html_path = tmp_path / "dashboard.html"
    with frozen_holdings_prices(FROZEN_GOLDEN_MASTER_PRICES):
        build_html_dashboard(str(xlsx_path), str(html_path), rows, c)
    html = html_path.read_text(encoding="utf-8")

    m_labels = re.search(r"const EXP_L=(\[[^\]]*\]);", html)
    m_colors = re.search(r"const EXP_C=(\[[^\]]*\]);", html)
    m_data = re.search(r"const YEARS=(\[[^\]]*\]);", html)
    assert m_labels and m_colors and m_data
    import json
    labels = json.loads(m_labels.group(1))
    colors = json.loads(m_colors.group(1))
    years = json.loads(m_data.group(1))

    assert "Other Cash Need" in labels
    # Every label must have its own real color -- not fewer colors than
    # labels, which would silently fall back every extra series to the same
    # generic gray in makeStackedBarSvg/makeLegend.
    assert len(colors) >= len(labels)
    assert len(set(colors[: len(labels)])) == len(labels), "expense series colors are not all distinct"

    m_exp_d = re.search(r"const EXP_D=(\[.*?\]\]);", html)
    assert m_exp_d
    exp_d = json.loads(m_exp_d.group(1))
    ocn_idx = labels.index("Other Cash Need")
    year_idx = years.index(year)
    assert exp_d[ocn_idx][year_idx] > 100_000
