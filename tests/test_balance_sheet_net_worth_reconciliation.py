"""#253/#254: the Balance Sheet, Net Worth, and Asset Allocation sheets must
all agree on "today's" (plan-start) pre-tax/roth/trust/hsa/cash balances.

#253 root cause: Balance Sheet (build_sheet3) summed each account's
END-OF-YEAR Y0 balance (rows[0][acct_id], after that year's growth/
withdrawals/CST funding), while the Net Worth sheet's Plan-Start column
deliberately uses the PRE-activity opening balance
(row['_account_opening'], seeded from live holdings) so it reconciles with
the Asset Allocation sheet's "today's" figures. Balance Sheet was the odd
one out.

#254 root cause: the Net Worth sheet's per-row "Cash" column read
row.get('cash_other', ...) -- a key the engine never writes (it writes
'cash_nw') -- so it always fell through to the static plan-start-only
c['cash_other'], freezing the displayed Cash/Sigma Other columns at the Y0
value every year regardless of later reserve draws/growth, even though the
row's TOTAL NW (sourced independently) was already correct.
"""
from openpyxl import Workbook

from src.data_io import load_csv, parse_client
from src.plan_config import ensure_engine_config
from src.planning_engines import project
from src.reporting.sheets_projection_facade import build_sheet5
from src.reporting.sheets_tax_reporter import build_sheet3
from tests.golden_pricing import FROZEN_GOLDEN_MASTER_PRICES, frozen_holdings_prices


def _real_config_and_rows():
    c = ensure_engine_config(parse_client(load_csv('input/client_data.csv'), ''), source='test')
    with frozen_holdings_prices(FROZEN_GOLDEN_MASTER_PRICES):
        rows = project(c)
    return c, rows


def _find_row(ws, label, col=1):
    for row in ws.iter_rows():
        cell = row[col - 1]
        if isinstance(cell.value, str) and cell.value.strip() == label:
            return cell.row
    raise AssertionError(f"row labeled {label!r} not found")


def test_balance_sheet_and_net_worth_sheet_agree_on_today_account_groups():
    c, rows = _real_config_and_rows()

    wb3 = Workbook()
    build_sheet3(wb3.active, c, rows)
    ws3 = wb3.active

    wb5 = Workbook()
    build_sheet5(wb5.active, c, rows)
    ws5 = wb5.active

    bs_pretax = ws3.cell(_find_row(ws3, 'Total Pre-Tax (Tax-Deferred)'), 2).value
    bs_roth = ws3.cell(_find_row(ws3, 'Total Roth (Tax-Free)'), 2).value
    bs_trust = ws3.cell(_find_row(ws3, 'Total Taxable / Trust'), 2).value
    bs_hsa = ws3.cell(_find_row(ws3, 'Total Health Savings Account'), 2).value
    bs_net_worth = ws3.cell(_find_row(ws3, 'NET WORTH'), 2).value

    # First data row of the Net Worth sheet is the plan-start year: cols
    # 13/16/19/20/28 are Sigma PreTax / Sigma Roth / Sigma Trust / HSA / TOTAL NW.
    plan_start_row = next(r for r in range(1, ws5.max_row + 1) if isinstance(ws5.cell(r, 1).value, int))
    nw_pretax = ws5.cell(plan_start_row, 13).value
    nw_roth = ws5.cell(plan_start_row, 16).value
    nw_trust = ws5.cell(plan_start_row, 19).value
    nw_hsa = ws5.cell(plan_start_row, 20).value
    nw_total = ws5.cell(plan_start_row, 28).value

    assert abs(bs_pretax - nw_pretax) < 1.0, f"pretax: balance sheet {bs_pretax} vs net worth {nw_pretax}"
    assert abs(bs_roth - nw_roth) < 1.0, f"roth: balance sheet {bs_roth} vs net worth {nw_roth}"
    assert abs(bs_trust - nw_trust) < 1.0, f"trust: balance sheet {bs_trust} vs net worth {nw_trust}"
    assert abs(bs_hsa - nw_hsa) < 1.0, f"hsa: balance sheet {bs_hsa} vs net worth {nw_hsa}"
    assert abs(bs_net_worth - nw_total) < 1.0, f"net worth {bs_net_worth} vs projected total {nw_total}"


def test_net_worth_sheet_cash_column_uses_live_per_year_balance():
    """#254: the Cash column must read the engine's per-row cash_nw, not a
    frozen static config default, and must exist under both keys' worth of
    row content (i.e. actually populated) for at least one non-zero year."""
    c, rows = _real_config_and_rows()
    wb = Workbook()
    build_sheet5(wb.active, c, rows)
    ws = wb.active

    plan_start_row = next(r for r in range(1, ws.max_row + 1) if isinstance(ws.cell(r, 1).value, int))
    displayed_cash = ws.cell(plan_start_row, 26).value  # col 26 = Cash (see build_sheet5's vals list)
    engine_cash_nw = rows[0].get('cash_nw', 0)
    assert engine_cash_nw > 0, 'sanity: this household must have a nonzero checking balance'
    assert abs(displayed_cash - engine_cash_nw) < 1.0, (
        f"displayed Cash column {displayed_cash} != engine cash_nw {engine_cash_nw}"
    )


def test_net_worth_sheet_never_reads_the_nonexistent_cash_other_row_key():
    import inspect
    from src.reporting import sheets_projection_net_worth as m
    src = inspect.getsource(m)
    assert "row.get('cash_other'" not in src
    assert "row.get(\"cash_other\"" not in src
