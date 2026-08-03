"""#255: expand the State Residency sheet's tax-comparison table to also
carry the geographic cost-of-living expense deltas (auto/homeowners
insurance, utilities, home maintenance), merging what used to be two
separate tables (Section B "Lifetime Tax Burden by State" and Section C
"Geographic Cost-of-Living Delta", which only ever compared ONE target
state) into a single "Lifetime Tax and Expenses by State" table that shows
every state's tax burden AND its expense delta side by side. The expense
factors (STATE_COL_FACTORS) are indexed with Illinois = 1.00 for every
category, so Illinois is the fixed basis those deltas are computed from.

Also #255: the "State residency" quick-nav button is removed from the
Planning Levers page's "Strategy - decide" card -- the merged reference
table above replaces its role there. The State Residency page itself (and
its own nav step) is unchanged and still reachable normally.
"""
from pathlib import Path

from openpyxl import Workbook

from src.reporting.sheets_strategy import build_sheet13
from src.reporting.workbook_common import STATE_TAX_RULES
from src.taxes import STATE_COL_FACTORS

ROOT = Path(__file__).resolve().parents[1]

CONFIG = {
    'plan_start': 2026, 'plan_end': 2027,
    'home_val': 500_000, 'home_appr': 0.03,
    'il_exempt': 4_000_000,
    'residency_target_state': '',
    'h_name': 'Matthew', 'w_name': 'Patricia',
    'state': 'Illinois',
    'current_auto_insurance_annual': 2_000,
    'current_homeowners_insurance_annual': 1_800,
    'current_home_utilities_annual': 3_600,
    'current_home_maintenance_annual': 4_000,
}

ROWS = [
    {'year': 2026, 'state_earned_net': 0, 'state_retirement': 50_000,
     'state_ss_taxable': 20_000, 'state_investment': 5_000,
     'state_nonqual_ann': 0, 'state_roth_conv': 0, 'agi': 75_000,
     'total_nw': 1_000_000, 'spend_base_yr': 60_000},
    {'year': 2027, 'state_earned_net': 0, 'state_retirement': 52_000,
     'state_ss_taxable': 21_000, 'state_investment': 5_200,
     'state_nonqual_ann': 0, 'state_roth_conv': 0, 'agi': 78_000,
     'total_nw': 1_050_000, 'spend_base_yr': 61_000},
]

HEADER_ROW = ['State', 'Income Rate', 'Income Tax', 'Property Tax', 'Sales Tax',
              'Estate Tax', 'Total Tax', 'Tax Delta vs Current', 'Auto Ins. Delta',
              'Home Ins. Delta', 'Utilities Delta', 'Maintenance Delta',
              'Total Delta (Tax + Expenses)', 'Retirement Income Taxed']


def _table_rows(ws):
    header_row = next(
        r for r in range(1, ws.max_row + 1)
        if ws.cell(row=r, column=1).value == 'State'
    )
    assert [ws.cell(row=header_row, column=i + 1).value for i in range(len(HEADER_ROW))] == HEADER_ROW
    out = []
    r = header_row + 1
    while ws.cell(row=r, column=1).value not in (None, ''):
        out.append({
            'state': ws.cell(row=r, column=1).value,
            'total_tax': ws.cell(row=r, column=7).value,
            'tax_delta': ws.cell(row=r, column=8).value,
            'auto_delta': ws.cell(row=r, column=9).value,
            'home_ins_delta': ws.cell(row=r, column=10).value,
            'utilities_delta': ws.cell(row=r, column=11).value,
            'maintenance_delta': ws.cell(row=r, column=12).value,
            'combined_delta': ws.cell(row=r, column=13).value,
        })
        r += 1
    return out


def test_illinois_is_the_fixed_basis_for_every_states_expense_factors():
    assert STATE_COL_FACTORS['Illinois'] == {
        'auto': 1.0, 'home_ins': 1.0, 'utilities': 1.0, 'maintenance': 1.0,
    }
    for state, factors in STATE_COL_FACTORS.items():
        if state == 'Illinois':
            continue
        assert set(factors) == {'auto', 'home_ins', 'utilities', 'maintenance'}


def test_tax_and_expense_tables_are_merged_into_one():
    ws = Workbook().active
    build_sheet13(ws, CONFIG, ROWS)
    rows = _table_rows(ws)
    assert len(rows) == len(STATE_TAX_RULES)  # every STATE_TAX_RULES entry, one merged row each

    illinois = next(r for r in rows if 'Illinois' in r['state'])
    assert illinois['tax_delta'] == 'Baseline'
    assert illinois['combined_delta'] == 'Baseline'
    # Illinois is both the household's current state AND the expense-factor
    # basis, so its own expense deltas must be exactly zero.
    assert illinois['auto_delta'] == 0.0
    assert illinois['home_ins_delta'] == 0.0
    assert illinois['utilities_delta'] == 0.0
    assert illinois['maintenance_delta'] == 0.0

    florida = next(r for r in rows if 'Florida' in r['state'])
    # Florida's homeowners-insurance factor (2.10) is much higher than
    # Illinois's (1.00), so relocating there must show a positive delta.
    assert florida['home_ins_delta'] > 0
    # combined_delta must equal tax_delta + expense deltas summed (lifetime).
    yrs = max(1, CONFIG['plan_end'] - CONFIG['plan_start'])
    expected_combined = florida['tax_delta'] + florida['auto_delta'] + florida['home_ins_delta'] \
        + florida['utilities_delta'] + florida['maintenance_delta']
    assert abs(florida['combined_delta'] - expected_combined) < 0.01
    for r in rows:
        if r['tax_delta'] == 'Baseline':
            continue
        assert isinstance(r['auto_delta'], float)


def test_no_separate_geographic_cost_of_living_table_remains():
    ws = Workbook().active
    build_sheet13(ws, CONFIG, ROWS)
    values = [ws.cell(row=r, column=c).value for r in range(1, ws.max_row + 1) for c in range(1, 6)]
    assert not any(v and 'Geographic Cost-of-Living Delta' in str(v) for v in values)


def test_state_residency_button_removed_from_planning_levers_decide_card():
    js = (ROOT / 'frontend/js/dashboard.js').read_text(encoding='utf-8')
    start = js.index('function renderPlanningLevers(')
    end = js.index('\nfunction chatMessageHtml', start)
    fn = js[start:end]
    assert 'data-step-id="state_residency">State residency</button>' not in fn
    # The State Residency page/nav step itself must still exist elsewhere --
    # this only removes the lever quick-nav shortcut, not the whole feature.
    assert 'id: "state_residency"' in js
    assert 'function renderStateResidency' in js
