from __future__ import annotations

"""minimize_row_heights() previously sized a wrapped cell's row to exactly
`lines * (font_size + LINE_PAD)`, with no separate top/bottom margin term --
LINE_PAD only adds space between/around individual lines, so a cell could
compute to precisely the height its text occupies with zero cushion above
the first line or below the last. CELL_VPAD adds a small fixed margin once
per cell (not per line, so it doesn't compound with line count), including
for text merged across multiple rows, where the combined block's total
needed height (lines*line_pt + CELL_VPAD) is what gets spread evenly across
the spanned rows.

Separately, the line-wrap estimate itself used a single flat
CHARS_PER_WIDTH_UNIT ratio (1.3, calibrated against generic lowercase
prose) for every character. This workbook's cells are packed with dollar
figures and short ALL-CAPS/Title-Case labels -- digits, uppercase letters,
and financial symbols ($%#&) render measurably wider than the average
lowercase character that ratio assumes, so a flat ratio under-counts lines
for number/caps-heavy text. Real example that motivated the fix: Executive
Summary cell C26 ("Face value $250K-$500K covers facility care risk;
avoids $113K-$213K annual deficit in worst case", 97 characters at a
~81-width-unit column) was predicted to fit on one line under the flat
ratio and got clipped to a single-line row, while neighboring rows with
similar or even shorter text correctly wrapped to two lines. Digit/
uppercase/financial-symbol characters are now weighted at 1.15
width-units each instead of folding into the narrow-prose ratio.
"""

import math
import string
from pathlib import Path

import openpyxl
import pytest
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from src.reporting.workbook_common import minimize_row_heights

LINE_PAD = 4.0
CELL_VPAD = 3.0
CHARS_PER_WIDTH_UNIT = 1.3
WIDE_CHAR_WIDTH_UNITS = 1.15
WIDE_CHARS = set(string.digits) | set(string.ascii_uppercase) | set('$%#&')


def _weighted_width(text):
    narrow_width = 1.0 / CHARS_PER_WIDTH_UNIT
    return sum(WIDE_CHAR_WIDTH_UNITS if ch in WIDE_CHARS else narrow_width for ch in text)


def _expected_lines(text, width, font_size=10.0):
    return sum(max(1, math.ceil(_weighted_width(line) / max(width, 1.0))) for line in (text.splitlines() or ['']))


def test_single_row_wrapped_cell_includes_fixed_vpad_beyond_line_leading():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.column_dimensions['A'].width = 20.0
    text = 'a ' * 120  # long enough to wrap across multiple lines at width 20
    cell = ws.cell(1, 1, value=text)
    cell.font = Font(size=10)
    cell.alignment = Alignment(wrap_text=True)

    minimize_row_heights(wb)

    lines = _expected_lines(text, 20.0)
    assert lines > 1, 'test text should wrap to more than one line'
    expected = lines * (10.0 + LINE_PAD) + CELL_VPAD
    assert ws.row_dimensions[1].height == round(expected, 1)


def test_row_spanning_merged_wrapped_cell_distributes_vpad_across_the_span():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.column_dimensions['A'].width = 20.0
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    text = 'a ' * 120
    cell = ws.cell(1, 1, value=text)
    cell.font = Font(size=10)
    cell.alignment = Alignment(wrap_text=True)

    minimize_row_heights(wb)

    lines = _expected_lines(text, 20.0)
    total_expected = lines * (10.0 + LINE_PAD) + CELL_VPAD
    per_row_expected = round(total_expected / 2, 1)
    assert ws.row_dimensions[1].height == per_row_expected
    assert ws.row_dimensions[2].height == per_row_expected
    # The two rows together give the wrapped text at least as much room as a
    # single unmerged row would have needed -- CELL_VPAD isn't lost by
    # splitting across the span.
    combined = (ws.row_dimensions[1].height or 0) + (ws.row_dimensions[2].height or 0)
    assert combined >= total_expected - 0.5


def test_single_line_cell_still_gets_the_fixed_vpad_margin():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.column_dimensions['A'].width = 30.0
    cell = ws.cell(1, 1, value='short label')
    cell.font = Font(size=10)
    cell.alignment = Alignment(wrap_text=True)

    minimize_row_heights(wb)

    expected = 1 * (10.0 + LINE_PAD) + CELL_VPAD
    assert ws.row_dimensions[1].height == round(expected, 1)


def test_dollar_and_digit_heavy_text_wraps_to_more_lines_than_a_flat_ratio_predicts():
    """Reproduces the Executive Summary C26 bug directly: this exact text at
    this exact column width was clipped to a single-line row because the old
    flat-ratio estimate (97 raw chars vs. ~105-char flat-ratio capacity)
    predicted it would fit on one line."""
    wb = openpyxl.Workbook()
    ws = wb.active
    width = 81.14
    ws.column_dimensions['A'].width = width
    text = ('Face value $250K–$500K covers facility care risk; avoids '
             '$113K–$213K annual deficit in worst case')
    cell = ws.cell(1, 1, value=text)
    cell.font = Font(name='Arial', size=10)
    cell.alignment = Alignment(wrap_text=True)

    old_flat_lines = max(1, math.ceil(len(text) / max(width * CHARS_PER_WIDTH_UNIT, 1.0)))
    assert old_flat_lines == 1, 'this text should be the exact case the old flat ratio got wrong'

    minimize_row_heights(wb)

    lines = _expected_lines(text, width)
    assert lines == 2, 'digit/uppercase/$-heavy text should now be predicted to wrap to 2 lines'
    expected = lines * (10.0 + LINE_PAD) + CELL_VPAD
    actual = ws.row_dimensions[1].height
    assert actual == round(expected, 1)
    assert actual > 17.0, 'row must be taller than the old (clipped) single-line height'


def test_all_caps_label_needs_more_room_than_equal_length_lowercase_text():
    wb = openpyxl.Workbook()
    ws = wb.active
    width = 6.0  # narrow enough that the extra uppercase width tips it to 2 lines
    ws.column_dimensions['A'].width = width
    caps_cell = ws.cell(1, 1, value='OTHER ASSETS')
    caps_cell.font = Font(size=10)
    caps_cell.alignment = Alignment(wrap_text=True)
    lower_cell = ws.cell(2, 1, value='other assets')
    lower_cell.font = Font(size=10)
    lower_cell.alignment = Alignment(wrap_text=True)

    minimize_row_heights(wb)

    caps_lines = _expected_lines('OTHER ASSETS', width)
    lower_lines = _expected_lines('other assets', width)
    assert caps_lines > lower_lines, 'same text length, but ALL-CAPS should need more room'
    assert ws.row_dimensions[1].height > ws.row_dimensions[2].height, (
        'production row heights must reflect that ALL-CAPS needs more room than lowercase, '
        'not just the test helper'
    )


@pytest.mark.slow
def test_real_executive_summary_c26_is_no_longer_clipped_to_one_line(built_workbook_path):
    """The actual reported cell: Executive Summary C26 must render at a
    height tall enough for its real 2-line wrap, not the clipped 1-line
    height the flat-ratio bug produced."""
    wb = load_workbook(built_workbook_path, data_only=False)
    ws = wb['1A. Executive Summary']
    row = next(
        r for r in range(1, ws.max_row + 1)
        if isinstance(ws.cell(r, 3).value, str) and 'terminal net worth alone' in ws.cell(r, 3).value
    )
    height = ws.row_dimensions[row].height
    assert height is not None and height > 17.0, (
        f'row {row} (Executive Summary C{row}) height {height} looks clipped to a single line'
    )
