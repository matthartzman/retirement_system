"""Unit tests for item 2.11 (finding F2 generalization): the Executive
Summary's QTIP recommendation row is now gated on federal estate
materiality, not fired purely on ``not c.get('qtip_enabled')``.
"""
from __future__ import annotations

from src.reporting.summary_figures import federal_estate_materiality


def _base_config(**overrides):
    c = {
        "plan_start": 2026,
        "plan_end": 2026,
        "fed_exempt": 13_990_000.0,
        "brk_inf": 0.02,
        "state": "Texas",
    }
    c.update(overrides)
    return c


def test_federal_estate_materiality_true_for_estate_at_or_above_exemption():
    c = _base_config()
    rows = [{"year": 2026, "total_nw": 20_000_000.0}]
    estate, exempt, exposed = federal_estate_materiality(c, rows)
    assert estate == 20_000_000.0
    assert exposed is True


def test_federal_estate_materiality_false_for_estate_well_below_exemption():
    c = _base_config()
    rows = [{"year": 2026, "total_nw": 500_000.0}]
    estate, exempt, exposed = federal_estate_materiality(c, rows)
    assert exposed is False


def test_federal_estate_materiality_nets_out_lifetime_exemption_used():
    # Item 2.6: a household that has already consumed most of its lifetime
    # exemption on gifts has a much smaller REMAINING exemption, so an
    # estate that looks comfortably under the nominal exemption can still
    # be exposed once lifetime gift usage is netted out.
    c = _base_config(fed_exempt=13_990_000.0)
    rows = [{
        "year": 2026,
        "total_nw": 13_000_000.0,
        "lifetime_exemption_used_cumulative": 12_500_000.0,
    }]
    estate, exempt, exposed = federal_estate_materiality(c, rows)
    assert exempt == 13_990_000.0 - 12_500_000.0
    assert exposed is True


def test_federal_estate_materiality_none_with_no_projection():
    c = _base_config()
    estate, exempt, exposed = federal_estate_materiality(c, rows=None)
    assert estate is None
    assert exposed is False


def _frozen_config_and_rows():
    from src.data_io import load_csv, parse_client
    from src.plan_config import ensure_engine_config
    from src.planning_engines import project
    from tests.golden_pricing import FROZEN_GOLDEN_MASTER_PRICES, frozen_holdings_prices
    from conftest import TEST_INPUT_DIR

    c = parse_client(load_csv(TEST_INPUT_DIR / "client_data.csv"), "")
    c["roth_policy"] = "none"
    c = ensure_engine_config(c, source="test")
    with frozen_holdings_prices(FROZEN_GOLDEN_MASTER_PRICES):
        rows = project(c)
    return c, rows


def test_qtip_row_suppressed_when_estate_not_material():
    from openpyxl import Workbook
    from src.reporting.sheets_summary_builder import build_sheet1

    c, rows = _frozen_config_and_rows()
    c = dict(c)
    c["qtip_enabled"] = False
    rows = [dict(rows[-1])]
    rows[-1]["total_nw"] = 200_000.0  # well under any federal exemption
    wb = Workbook()
    ws = wb.active
    build_sheet1(ws, c, rows, {})
    texts = [str(cell.value) for row in ws.iter_rows() for cell in row if cell.value is not None]
    assert not any("QTIP Trust to Manage Annuity Post-First-Death" in t for t in texts)


def test_qtip_row_shown_when_estate_is_material():
    from openpyxl import Workbook
    from src.reporting.sheets_summary_builder import build_sheet1

    c, rows = _frozen_config_and_rows()
    c = dict(c)
    c["qtip_enabled"] = False
    c["fed_exempt"] = 13_990_000.0  # override the fixture's own high configured exemption
    rows = [dict(rows[-1])]
    rows[-1]["lifetime_exemption_used_cumulative"] = 0.0
    rows[-1]["total_nw"] = 30_000_000.0  # comfortably above the federal exemption
    wb = Workbook()
    ws = wb.active
    build_sheet1(ws, c, rows, {})
    texts = [str(cell.value) for row in ws.iter_rows() for cell in row if cell.value is not None]
    assert any("QTIP Trust to Manage Annuity Post-First-Death" in t for t in texts)
