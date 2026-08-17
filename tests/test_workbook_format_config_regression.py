"""Item 189 — Settings → Workbook Formatting column-width editor.

Covers the shared config module (structure tree, override round-trip, and
generation-time application) and the presence of the UI + route wiring.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from src.reporting import workbook_format_config as wf
from tests._decomp_dashboard import dashboard_js_text

ROOT = Path(__file__).resolve().parents[1]
BUILT_WORKBOOK = ROOT / "output" / "retirement_plan.xlsx"


def _make_workbook(tmp_path: Path) -> Path:
    """A tiny workbook with one multi-table sheet and one flat sheet."""
    wb = openpyxl.Workbook()
    multi = wb.active
    multi.title = "Multi"
    # Two side-by-side merged banners in row 1 => two tables.
    multi["A1"] = "GROUP ONE"
    multi.merge_cells("A1:B1")
    multi["C1"] = "GROUP TWO"
    multi.merge_cells("C1:D1")
    for col, title in zip("ABCD", ["Year", "Value", "Rate", "Total"]):
        multi[f"{col}2"] = title
    multi.column_dimensions["A"].width = 10
    multi.column_dimensions["B"].width = 12
    multi.column_dimensions["C"].width = 8
    multi.column_dimensions["D"].width = 14

    flat = wb.create_sheet("Flat")
    flat["A1"] = "FLAT REPORT"
    flat.merge_cells("A1:C1")
    for col, title in zip("ABC", ["Label", "Amount", "Note"]):
        flat[f"{col}2"] = title
    for col in "ABC":
        flat.column_dimensions[col].width = 15

    path = tmp_path / "wb.xlsx"
    wb.save(path)
    return path


def test_tree_detects_tables_and_single_table(tmp_path):
    path = _make_workbook(tmp_path)
    tree = wf.build_format_tree(path)
    assert tree["available"] is True
    by_name = {s["sheet"]: s for s in tree["sheets"]}

    multi = by_name["Multi"]
    assert multi["single_table"] is False
    names = [t["name"] for t in multi["tables"]]
    assert "GROUP ONE" in names and "GROUP TWO" in names
    g1 = next(t for t in multi["tables"] if t["name"] == "GROUP ONE")
    titles = {c["col"]: c["title"] for c in g1["columns"]}
    assert titles["A"] == "Year" and titles["B"] == "Value"

    flat = by_name["Flat"]
    assert flat["single_table"] is True
    assert len(flat["tables"]) == 1
    assert flat["tables"][0]["name"] is None
    flat_titles = {c["col"]: c["title"] for c in flat["tables"][0]["columns"]}
    assert flat_titles["A"] == "Label" and flat_titles["B"] == "Amount"


def test_missing_workbook_reports_unavailable(tmp_path):
    tree = wf.build_format_tree(tmp_path / "does_not_exist.xlsx")
    assert tree["available"] is False
    assert tree["sheets"] == []


def test_overrides_round_trip_and_sanitize(tmp_path):
    saved = wf.save_overrides(
        {
            "Multi": {"a": 20, "B": 15.5, "zz": -3, "bad!": 9, "C": "not a number"},
            "": {"A": 5},  # empty sheet name dropped
        },
        input_dir=tmp_path,
    )
    assert saved == {"Multi": {"A": 20.0, "B": 15.5}}
    assert wf.load_overrides(input_dir=tmp_path) == {"Multi": {"A": 20.0, "B": 15.5}}


def test_overrides_clamped(tmp_path):
    saved = wf.save_overrides({"S": {"A": 9999, "B": 0.01}}, input_dir=tmp_path)
    assert saved["S"]["A"] == wf.MAX_WIDTH
    assert saved["S"]["B"] == wf.MIN_WIDTH


def test_merge_overrides_does_not_wipe_other_sheets(tmp_path):
    """The live Settings -> Workbook Formatting UI autosaves one column patch
    at a time. save_overrides() replaces the whole file with exactly what it's
    given, so patching one sheet through it would silently delete every other
    sheet's saved widths -- the "changes keep reverting" bug. merge_overrides()
    must instead layer the patch on top of whatever was already persisted."""
    wf.save_overrides({"Sheet A": {"A": 20.0}, "Sheet B": {"C": 30.0}}, input_dir=tmp_path)
    merged = wf.merge_overrides({"Sheet A": {"B": 15.0}}, input_dir=tmp_path)
    assert merged == {"Sheet A": {"A": 20.0, "B": 15.0}, "Sheet B": {"C": 30.0}}
    assert wf.load_overrides(input_dir=tmp_path) == merged


def test_merge_overrides_deletes_only_the_named_column(tmp_path):
    wf.save_overrides({"Sheet A": {"A": 20.0, "B": 15.0}}, input_dir=tmp_path)
    merged = wf.merge_overrides({"Sheet A": {"B": 0}}, input_dir=tmp_path)
    assert merged == {"Sheet A": {"A": 20.0}}
    # Deleting the sheet's last remaining column drops the empty sheet entry.
    merged = wf.merge_overrides({"Sheet A": {"A": -1}}, input_dir=tmp_path)
    assert merged == {}


def test_merge_alignments_does_not_wipe_other_sheets(tmp_path):
    wf.save_alignments({"Sheet A": {"A": "left"}, "Sheet B": {"C": "right"}}, input_dir=tmp_path)
    merged = wf.merge_alignments({"Sheet A": {"B": "center"}}, input_dir=tmp_path)
    assert merged == {"Sheet A": {"A": "left", "B": "center"}, "Sheet B": {"C": "right"}}
    assert wf.load_alignments(input_dir=tmp_path) == merged


def test_merge_alignments_deletes_only_the_named_column(tmp_path):
    wf.save_alignments({"Sheet A": {"A": "left", "B": "center"}}, input_dir=tmp_path)
    merged = wf.merge_alignments({"Sheet A": {"B": ""}}, input_dir=tmp_path)
    assert merged == {"Sheet A": {"A": "left"}}


def test_route_uses_merge_not_replace():
    """Regression guard: the POST route must call merge_overrides/
    merge_alignments (non-destructive), never save_overrides/save_alignments
    (whole-file replace) -- reverting to the replace-based calls silently
    reintroduces the "editing one sheet deletes every other sheet" bug."""
    src = (ROOT / "src" / "server" / "workbook_routes.py").read_text(encoding="utf-8")
    handler = src.split("def save_workbook_format", 1)[1]
    assert "_wf.merge_overrides(" in handler
    assert "_wf.merge_alignments(" in handler
    assert "_wf.save_overrides(" not in handler
    assert "_wf.save_alignments(" not in handler


def test_apply_overrides_sets_widths(tmp_path):
    path = _make_workbook(tmp_path)
    wf.save_overrides({"Multi": {"A": 42.0}}, input_dir=tmp_path)
    wb = openpyxl.load_workbook(path)
    assert wb["Multi"].column_dimensions["A"].width == 10
    wf.apply_overrides(wb, input_dir=tmp_path)
    assert wb["Multi"].column_dimensions["A"].width == 42.0


def test_overridden_flag_reflects_saved(tmp_path):
    path = _make_workbook(tmp_path)
    overrides = {"Multi": {"C": 30.0}}
    tree = wf.build_format_tree(path, overrides)
    multi = next(s for s in tree["sheets"] if s["sheet"] == "Multi")
    flags = {
        c["col"]: c["overridden"] for t in multi["tables"] for c in t["columns"]
    }
    assert flags["C"] is True
    assert flags["A"] is False


@pytest.mark.skipif(not BUILT_WORKBOOK.exists(), reason="no built workbook present")
def test_real_workbook_multi_table_sheets():
    tree = wf.build_format_tree(BUILT_WORKBOOK)
    assert tree["available"] is True
    # #209/#210/#212/#228: "sheet" is now the stable (build-time) key so it
    # survives letters shifting; "display" is this build's actual final title.
    by_name = {s["sheet"]: s for s in tree["sheets"]}
    # Net Worth and Cash Flow are the wide matrix sheets with grouped columns.
    if "5. Net Worth Projection" in by_name:
        node = by_name["5. Net Worth Projection"]
        assert node["single_table"] is False
        assert node["display"].endswith("Net Worth")
    # Executive Summary is a single-table narrative sheet.
    if "1. Executive Summary" in by_name:
        node = by_name["1. Executive Summary"]
        assert node["single_table"] is True
        assert node["display"].endswith("Executive Summary")


def test_generation_applies_overrides_hook_present():
    src = (ROOT / "src" / "reporting" / "workbook_builder.py").read_text(encoding="utf-8")
    assert "apply_overrides as _apply_format_overrides" in src
    # #209/#210/#212/#228: overrides are keyed by stable sheet name, so the
    # call passes this build's live {stable: final_title} map.
    assert "_apply_format_overrides(wb, sheet_renames=FINAL_SHEET_RENAMES)" in src


def test_alignments_round_trip_and_sanitize(tmp_path):
    saved = wf.save_alignments(
        {
            "Multi": {"a": "L", "B": "Center", "C": "right", "bad!": "left", "D": "diagonal"},
            "": {"A": "left"},  # empty sheet name dropped
        },
        input_dir=tmp_path,
    )
    assert saved == {"Multi": {"A": "left", "B": "center", "C": "right"}}
    assert wf.load_alignments(input_dir=tmp_path) == {"Multi": {"A": "left", "B": "center", "C": "right"}}


def test_apply_alignments_sets_horizontal_on_data_rows_only(tmp_path):
    path = _make_workbook(tmp_path)
    wf.save_alignments({"Multi": {"A": "right"}}, input_dir=tmp_path)
    wb = openpyxl.load_workbook(path)
    ws = wb["Multi"]
    ws["A3"] = 5  # a data row below the header band
    header_align_before = ws["A2"].alignment.horizontal
    wf.apply_alignments(wb, input_dir=tmp_path)
    assert ws["A3"].alignment.horizontal == "right"
    # Header band is left untouched.
    assert ws["A2"].alignment.horizontal == header_align_before


def test_align_node_reflects_saved_override(tmp_path):
    path = _make_workbook(tmp_path)
    tree = wf.build_format_tree(path, alignments={"Multi": {"B": "center"}})
    multi = next(s for s in tree["sheets"] if s["sheet"] == "Multi")
    flags = {c["col"]: (c["align"], c["align_overridden"]) for t in multi["tables"] for c in t["columns"]}
    assert flags["B"] == ("center", True)
    assert flags["A"][1] is False


def test_routes_and_ui_wired():
    routes = (ROOT / "src" / "server" / "workbook_routes.py").read_text(encoding="utf-8")
    assert '"/api/workbook-format", methods=["GET"]' in routes
    assert '"/api/workbook-format", methods=["POST"]' in routes
    # renderWorkbookFormatting/setWorkbookColAlign live in
    # dashboard_decomp_workbook_formatting.js, a sibling module loaded alongside
    # dashboard.js (see frontend/index.html); the STEPS nav entry and dispatch
    # stayed in dashboard.js itself.
    js = dashboard_js_text()
    js += (ROOT / "frontend" / "js" / "dashboard_decomp_workbook_formatting.js").read_text(encoding="utf-8")
    assert "function renderWorkbookFormatting" in js
    # Item 192 (Option 4 Phase 2): Workbook Formatting is a first-class Settings
    # nav page rather than a card button in the Settings hub.
    assert 'id: "workbook_formatting"' in js
    assert 'activeStep === "workbook_formatting"' in js
    assert "/api/workbook-format" in js
    assert "alignments" in routes
    assert "function setWorkbookColAlign" in js
    nav = (ROOT / "frontend" / "js" / "navigation.js").read_text(encoding="utf-8")
    assert "workbook_formatting" in nav


def _seed_built_workbook(output: Path) -> Path:
    output.mkdir(parents=True, exist_ok=True)
    wb_path = output / "retirement_plan.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Sheet"
    ws["A1"] = "TITLE"
    ws["A2"] = "Header"
    ws.column_dimensions["A"].width = 8.71
    wb.save(wb_path)
    return wb_path


def test_get_workbook_format_route_flags_overrides_saved_after_last_build(monkeypatch, tmp_path):
    """Reported live: a saved column-width override can be edited without an
    intervening rebuild, so the "Automatic"/"Last built" width shown on the
    page is really just whatever happened to be baked into the last .xlsx --
    stale relative to the override the user just saved, with nothing on the
    page saying so. overrides_stale compares the overrides file's mtime
    against the built workbook's mtime and must flip true the moment an
    override is saved after that build, and back to false once a fresh build
    (a newer workbook file) supersedes it."""
    import time

    from src.server import app
    import src.server.workbook_routes as workbook_routes

    output = tmp_path / "output"
    wb_path = _seed_built_workbook(output)
    monkeypatch.setattr(workbook_routes, "_workspace_output", lambda: output)
    monkeypatch.setattr(wf, "overrides_path", lambda input_dir=None: tmp_path / "workbook_format_overrides.json")
    monkeypatch.setattr(wf, "alignments_path", lambda input_dir=None: tmp_path / "workbook_format_alignments.json")

    client = app.test_client()
    headers = {"X-User-Role": "admin"}

    # No overrides saved yet -- not stale.
    resp = client.get("/api/workbook-format", headers=headers)
    assert resp.status_code == 200
    assert resp.get_json()["overrides_stale"] is False

    # Save an override -- its file's mtime is now newer than the workbook's.
    # The sleep guards against filesystem mtime resolution ties: back-to-back
    # writes within the same tick would otherwise compare equal, not >.
    time.sleep(0.05)
    resp = client.post(
        "/api/workbook-format",
        headers=headers,
        json={"overrides": {"Test Sheet": {"A": 9.43}}},
    )
    assert resp.status_code == 200

    resp = client.get("/api/workbook-format", headers=headers)
    payload = resp.get_json()
    assert payload["overrides_stale"] is True
    col_a = payload["sheets"][0]["tables"][0]["columns"][0]
    assert col_a["col"] == "A"
    assert col_a["width"] == 8.71  # still the stale, pre-rebuild width
    assert col_a["overridden"] is True  # the save itself did land

    # A rebuild (a newer workbook file) supersedes the pending override.
    time.sleep(0.05)
    wb2 = openpyxl.load_workbook(wb_path)
    wb2["Test Sheet"].column_dimensions["A"].width = 9.43
    wb2.save(wb_path)

    resp = client.get("/api/workbook-format", headers=headers)
    assert resp.get_json()["overrides_stale"] is False
