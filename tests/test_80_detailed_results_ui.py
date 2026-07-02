from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from src.version import VERSION
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import SeriesLabel

from src.detailed_results import workbook_detailed_results, workbook_detailed_index, workbook_detailed_sheet, _clean_sheet_title

ROOT = Path(__file__).resolve().parents[1]


def test_detailed_results_parser_groups_workbook_sections(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = '6. Cash Flow Projection'
    ws['A1'] = 'CASH FLOW PROJECTION'
    ws['A3'] = 'Year'
    ws['B3'] = 'Income'
    ws['A4'] = 2026
    ws['B4'] = 100000
    ws['A6'] = 'Portfolio Withdrawals'
    ws['A7'] = 'Year'
    ws['B7'] = 'Withdrawal'
    ws['A8'] = 2026
    ws['B8'] = 20000
    ws2 = wb.create_sheet('22. Glossary')
    ws2['A1'] = 'Term'
    ws2['B1'] = 'Definition'
    ws2['A2'] = 'RMD'
    ws2['B2'] = 'Required minimum distribution'
    path = tmp_path / 'retirement_plan.xlsx'
    wb.save(path)

    out = workbook_detailed_results(path)

    assert out['success'] is True
    assert out['sheet_count'] == 2
    cash = next(s for s in out['sheets'] if s['name'] == '6. Cash Flow Projection')
    assert cash['category'] == 'Reports'
    assert len(cash['sections']) == 2
    assert cash['sections'][0]['title'] == 'CASH FLOW PROJECTION'
    assert cash['sections'][0]['rows'][1]['cells'][0]['display'] == 'Year'
    assert any(c['name'] == 'Reports' for c in out['categories'])


def test_clean_sheet_title_strips_hierarchical_ordinal_prefixes():
    """Sheet tabs use hierarchical prefixes like '1A.'/'1H.'/'2B.', not just
    plain numbers like '22.' - the ordinal can include trailing letters."""
    assert _clean_sheet_title('1H. Spending Summary') == 'Spending Summary'
    assert _clean_sheet_title('1A. Executive Summary') == 'Executive Summary'
    assert _clean_sheet_title('2B. Asset Allocation') == 'Asset Allocation'
    assert _clean_sheet_title('22. Glossary') == 'Glossary'
    assert _clean_sheet_title('No Prefix Sheet') == 'No Prefix Sheet'


def test_detailed_results_merges_untitled_trailing_blocks_into_prior_named_section(tmp_path):
    """Reported UI bug: a sheet whose title row is followed by blank-row-
    separated blocks with no title of their own (a metrics block, a data
    table, a total row) previously showed one real heading plus two bare
    "Spending Summary"-style duplicate headings - matching build_sheet_
    spending_summary's structure (title row, blank, headers+data, blank,
    total row). Untitled trailing blocks should merge into the preceding
    named section instead of showing a redundant generic heading."""
    wb = Workbook()
    ws = wb.active
    ws.title = '1H. Spending Summary'
    ws['A1'] = 'SPENDING SUMMARY — 2026 YTD  (182 days elapsed)'
    ws['A2'] = 'YTD total spending'
    ws['B2'] = 178000
    # blank row 3
    ws['A5'] = 'Tracking Type / Group / Category'
    ws['B5'] = 'YTD Actual'
    ws['C5'] = 'Annualized'
    ws['D5'] = 'Budget'
    ws['E5'] = 'vs Budget'
    ws['F5'] = '% of Total'
    ws['A6'] = 'Core Expenses'
    ws['B6'] = 178000
    ws['C6'] = 357000
    ws['D6'] = 336000
    ws['E6'] = 21000
    ws['F6'] = 1.0
    # blank row 7
    ws['A8'] = 'TOTAL SPENDING'
    ws['B8'] = 178000
    ws['C8'] = 357000
    ws['D8'] = 336000
    ws['E8'] = 21000
    ws['F8'] = 1.0
    path = tmp_path / 'retirement_plan.xlsx'
    wb.save(path)

    sheet = workbook_detailed_sheet(path, '1H. Spending Summary')
    assert sheet['success'] is True
    sections = sheet['sheet']['sections']
    assert len(sections) == 1
    assert sections[0]['title'] == 'SPENDING SUMMARY — 2026 YTD  (182 days elapsed)'
    assert sections[0]['row_count'] == 5


def test_detailed_results_index_and_sheet_load_on_demand(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = '6. Cash Flow Projection'
    ws['A1'] = 'CASH FLOW PROJECTION'
    ws['A3'] = 'Year'
    ws['B3'] = 'Income'
    ws['A4'] = 2026
    ws['B4'] = 100000
    path = tmp_path / 'retirement_plan.xlsx'
    wb.save(path)

    index = workbook_detailed_index(path)
    assert index['success'] is True
    assert index['mode'] == 'index'
    assert index['sheets'][0]['name'] == '6. Cash Flow Projection'
    assert index['sheets'][0]['loaded'] is False
    assert 'sections' not in index['sheets'][0]

    sheet = workbook_detailed_sheet(path, '6. Cash Flow Projection')
    assert sheet['success'] is True
    assert sheet['mode'] == 'sheet'
    assert sheet['sheet']['name'] == '6. Cash Flow Projection'
    assert sheet['sheet']['sections'][0]['title'] == 'CASH FLOW PROJECTION'



def test_detailed_results_formats_currency_cells_as_rounded_k(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = '6. Cash Flow Projection'
    ws['A1'] = 'CASH FLOW PROJECTION'
    ws['A3'] = 'Year'
    ws['B3'] = 'Income'
    ws['A4'] = 2026
    ws['B4'] = 1234567
    ws['B4'].number_format = '$#,##0'
    path = tmp_path / 'retirement_plan.xlsx'
    wb.save(path)

    sheet = workbook_detailed_sheet(path, '6. Cash Flow Projection')
    cells = sheet['sheet']['sections'][0]['rows'][2]['cells']

    assert cells[1]['kind'] == 'currency'
    assert cells[1]['display'] == '$1,235K'


def test_chart_dashboard_sheet_is_chart_only_and_uses_hidden_source_data(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = '8. Charts Dashboard'
    ws['A1'] = 'CHART DASHBOARD'
    hidden = wb.create_sheet('_Chart Dashboard Data')
    hidden.sheet_state = 'hidden'
    hidden['A4'] = 'Year'
    hidden['B4'] = 'Trust'
    hidden['C4'] = 'Roth'
    hidden['A5'] = 2026
    hidden['B5'] = 100000
    hidden['C5'] = 50000
    hidden['A6'] = 2027
    hidden['B6'] = 120000
    hidden['C6'] = 60000
    hidden['K4'] = 'Year'
    hidden['L4'] = 'Earned Income'
    hidden['K5'] = 2026
    hidden['L5'] = 10000
    hidden['AC4'] = 'Year'
    hidden['AD4'] = 'Base Spending'
    hidden['AC5'] = 2026
    hidden['AD5'] = 9000
    hidden['AQ4'] = 'Year'
    hidden['AR4'] = 'P10'
    hidden['AQ5'] = 2026
    hidden['AR5'] = 80000
    hidden['A20'] = 'Asset Class'
    hidden['B20'] = 'Before'
    hidden['C20'] = 'After'
    hidden['A21'] = 'Stocks'
    hidden['B21'] = 70000
    hidden['C21'] = 80000
    path = tmp_path / 'retirement_plan.xlsx'
    wb.save(path)

    index = workbook_detailed_index(path)
    names = [s['name'] for s in index['sheets']]
    assert '8. Charts Dashboard' in names
    assert '_Chart Dashboard Data' not in names
    indexed = next(s for s in index['sheets'] if s['name'] == '8. Charts Dashboard')
    assert indexed['kind'] == 'chart_dashboard'
    assert indexed['preview'] is False

    sheet = workbook_detailed_sheet(path, '8. Charts Dashboard')['sheet']
    assert sheet['kind'] == 'chart_dashboard'
    assert 'sections' not in sheet
    assert sheet['charts'][0]['title'] == 'Net Worth by Component'
    assert sheet['charts'][0]['series'][0]['label'] == 'Trust'
    assert 'hidden' in sheet['chart_note'].lower()



def test_chart_dashboard_falls_back_to_embedded_excel_chart_ranges(tmp_path):
    wb = Workbook()
    chart_ws = wb.active
    chart_ws.title = '8. Charts Dashboard'
    data = wb.create_sheet('Legacy Chart Data')
    data['A1'] = 'Year'
    data['B1'] = 'Trust'
    data['C1'] = 'Roth'
    data['A2'] = 2026
    data['B2'] = 100000
    data['C2'] = 50000
    data['A3'] = 2027
    data['B3'] = 120000
    data['C3'] = 60000
    data['E1'] = 'Asset Class'
    data['F1'] = 'Value'
    data['E2'] = 'Stocks'
    data['F2'] = 70000
    data['E3'] = 'Bonds'
    data['F3'] = 30000

    bar = BarChart()
    bar.type = 'col'
    bar.grouping = 'stacked'
    bar.title = 'Legacy Net Worth Chart'
    for idx, col in enumerate([2, 3]):
        bar.add_data(Reference(data, min_col=col, min_row=2, max_row=3))
        bar.series[idx].title = SeriesLabel(v=data.cell(1, col).value)
    bar.set_categories(Reference(data, min_col=1, min_row=2, max_row=3))
    chart_ws.add_chart(bar, 'A4')

    pie = PieChart()
    pie.title = 'Legacy Allocation Chart'
    pie.add_data(Reference(data, min_col=6, min_row=2, max_row=3), titles_from_data=False)
    pie.set_categories(Reference(data, min_col=5, min_row=2, max_row=3))
    chart_ws.add_chart(pie, 'A20')

    path = tmp_path / 'retirement_plan.xlsx'
    wb.save(path)

    sheet = workbook_detailed_sheet(path, '8. Charts Dashboard')['sheet']

    assert sheet['kind'] == 'chart_dashboard'
    assert len(sheet['charts']) == 2
    assert sheet['charts'][0]['title'] == 'Legacy Net Worth Chart'
    assert sheet['charts'][0]['series'][0]['label'] == 'Trust'
    assert sheet['charts'][0]['x'] == [2026, 2027]
    assert sheet['charts'][1]['type'] == 'pie'
    assert sheet['charts'][1]['slices'][0]['label'] == 'Stocks'
    assert 'embedded Excel chart source ranges' in sheet['chart_note']
    assert 'could not be reconstructed' not in sheet['chart_note']

def test_detailed_results_ui_nav_and_endpoint_are_present():
    js = (ROOT / 'frontend/js/dashboard.js').read_text(encoding='utf-8')
    rui = (ROOT / 'frontend/js/reports_ui.js').read_text(encoding='utf-8')
    css = (ROOT / 'frontend/css/dashboard.css').read_text(encoding='utf-8')
    routes = (ROOT / 'src/server/workbook_routes.py').read_text(encoding='utf-8')
    assert "id:'detailed_results'" in js
    assert 'Results Explorer' in js
    assert 'Workbook Results Explorer' not in js
    assert 'Workbook result explorer' not in js
    assert 'View Detailed Results' not in js
    assert 'renderDetailedResultsNav' in js
    assert 'renderDetailedResults' in js
    assert 'startDetailedResultsProgress' in js
    assert 'detailedResultsNavOpen' in js
    assert 'retirementDetailedResultsNavOpen' in js
    assert 'setDetailedResultsNavOpen(this.open)' in js
    assert 'window.setDetailedResultsNavOpen=setDetailedResultsNavOpen' in js
    assert 'timeoutMs:30000' in js
    assert 'isAssetAllocationSheet' in js
    assert 'timeoutMs:isChartDashboardSheet?20000:(isAssetAllocationSheet?30000:60000)' in js
    assert 'Chart Dashboard loading timed out' in js
    assert 'Asset Allocation loading timed out' in js
    assert 'detailedResultSheetSeq' in js
    assert 'detailedResultSheetInFlight' in js
    assert 'detailedResultsIndexInFlight' in js
    assert 'async function api' in js
    assert 'async function loadAll' in js
    assert 'async function saveAll' in js
    assert 'async function runBuild' in js
    assert 'detail-progress-bar' in css
    assert 'DETAIL_MONEY_TERMS' in rui
    assert 'detailCurrencyK' in rui
    assert 'detailHeaderRowIndex' in rui
    assert 'detailedProgressHtml(true)' not in js
    assert 'detail-cell-currency' in css
    assert "if(sheet.kind==='chart_dashboard'&&Array.isArray(sheet.charts))" in rui
    assert 'renderChartDashboardSheet' in rui
    assert 'detail-chart-grid' in css
    assert 'Show / hide columns' in rui
    assert 'toggleDetailColumnGroup' in js
    assert 'Workbook row' not in js
    assert 'detailColumnLabel' not in js
    assert 'data-detail-sheet' in js
    assert "detailed_results" in js
    assert "planLoaded" in js
    assert '/api/detailed-results?index=1' in js
    assert '/api/detailed-results?sheet=' in js
    assert 'detailed-results-nav' in css
    assert 'detailed-result-section' in css
    service = (ROOT / 'src/server_services/report_service.py').read_text(encoding='utf-8')
    assert '@app.route("/api/detailed-results"' in routes
    assert 'report_service.detailed_results_payload(' in routes
    assert 'workbook_detailed_results' in service
    assert 'workbook_detailed_index' in service
    assert 'workbook_detailed_sheet' in service


def test_results_explorer_uses_human_headings_not_measure_fallbacks():
    js = (ROOT / 'frontend/js/dashboard.js').read_text(encoding='utf-8')
    rui = (ROOT / 'frontend/js/reports_ui.js').read_text(encoding='utf-8')
    css = (ROOT / 'frontend/css/dashboard.css').read_text(encoding='utf-8')

    assert "function detailProgressState" in js
    assert "setInterval(()=>" in js and ",250)" in js
    assert "Detecting sticky heading rows and human-readable column groups" in js
    assert "function detailLabelForColumn" in rui
    assert "function detailGroupLabel" in rui
    assert "Years ${first}–${last}" in rui
    assert "Measure ${i+1}" not in js
    assert "detailIdentifierRowIndex" in rui
    assert "detail-super-head" in rui
    assert "has-super-head" in rui
    assert "detailCleanSectionTitle" in rui
    assert ".detail-result-table.has-super-head" in css


def test_chart_dashboard_derives_browser_charts_from_projection_sheets_when_sources_missing(tmp_path):
    wb = Workbook()
    chart_ws = wb.active
    chart_ws.title = '8. Charts Dashboard'
    chart_ws['A1'] = 'CHART DASHBOARD'

    nw = wb.create_sheet('5. Net Worth Projection')
    nw.append(['Identifiers', None, None, 'ANNUITIES & PENSION', None, None, None, None, 'PRE-TAX'])
    nw.append(['Year', 'H Age', 'W Age', 'Σ Ann', 'Σ PreTax', 'Σ Roth', 'Σ Trust', 'HSA', 'Home Equity', 'Σ Other', 'TOTAL NW'])
    nw.append([2026, 50, 50, 100000, 200000, 50000, 25000, 10000, 500000, 40000, 925000])
    nw.append([2027, 51, 51, 105000, 220000, 60000, 30000, 11000, 510000, 42000, 978000])

    cash = wb.create_sheet('6. Cash Flow Projection')
    cash.append(['Identifiers', None, None, 'INCOME'])
    cash.append(['Year', 'H Age', 'W Age', 'Earned', 'Matthew SS', 'Patricia SS', 'Pension', 'W Single Ann', 'W Joint Ann', 'H Single Ann', 'H Joint Ann', 'Note P+I', 'RMD Dist', 'Σ Income', 'Roth Conv', 'AGI', 'Taxable Inc', 'Fed Tax', 'State Tax', 'NIIT', 'Spend Base', 'Rec Extra', 'Lump', 'Mortgage + RE Tax', 'Rent', 'Σ Spend', 'H Trust WD', 'W Trust WD', 'Σ Trust', 'HSA WD', 'H Roth WD', 'W Roth WD', 'Σ Roth', 'H IRA RMD', 'H IRA Elec', 'H IRA Total', 'W IRA RMD', 'W IRA Elec', 'W IRA Total', 'Home Eq Tap', 'Σ Draws', 'Surplus', 'NW Check'])
    cash.append([2026, 50, 50, 100000, 0, 0, 12000, 0, 0, 0, 0, 5000, 0, 117000, 0, 117000, 100000, 10000, 3000, 0, 80000, 5000, 0, 36000, 0, 121000, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, -4000, 925000])
    cash.append([2027, 51, 51, 90000, 0, 0, 12000, 0, 0, 0, 0, 5000, 0, 107000, 0, 107000, 95000, 9500, 2800, 0, 83000, 5000, 0, 36000, 0, 124000, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, -17000, 978000])

    path = tmp_path / 'retirement_plan.xlsx'
    wb.save(path)

    sheet = workbook_detailed_sheet(path, '8. Charts Dashboard')['sheet']

    assert sheet['kind'] == 'chart_dashboard'
    assert len(sheet['charts']) >= 3
    assert 'projection result sheets' in sheet['chart_note']
    assert 'does not expose readable chart source data' not in sheet['chart_note']
    assert 'Rebuild the workbook' not in sheet['chart_note']
    assert sheet['charts'][0]['title'] == 'Net Worth by Component'


def test_results_model_sidecar_populates_index_sheets_with_source(tmp_path):
    from src.results_model import write_result_explorer_model, RESULTS_MODEL_FILENAME

    workbook = tmp_path / 'retirement_plan.xlsx'
    wb = Workbook()
    ws = wb.active
    ws.title = '6. Cash Flow Projection'
    ws['A1'] = 'OLD EXCEL CASHFLOW'
    wb.save(workbook)

    c = {'h_name': 'Matt', 'w_name': 'Pat', 'plan_start': 2026, 'plan_end': 2027, 'state': 'IL'}
    rows = [
        {'year': 2026, 'h_age': 50, 'w_age': 49, 'earned': 100000, 'h_ss': 0, 'w_ss': 0, 'pension': 0, 'wife_single_ann': 0, 'wife_joint_ann': 0, 'h_single_ann': 0, 'h_joint_ann': 0, 'note_princ': 0, 'note_int': 0, 'rmd_total': 0, 'spend_base_yr': 80000, 'rec_extra': 0, 'lump': 0, 'mortgage': 36000, 'rent_yr': 0, 'fed_tax': 10000, 'state_tax': 3000, 'niit': 0, 'total_nw': 1000000, 'ann_nw': 0, 'pretax_nw': 100000, 'roth_nw': 50000, 'trust_nw': 25000, 'hsa_nw': 10000, 'home_equity': 500000, 'other_nw': 0, 'total_tax': 13000, 'agi': 100000, 'taxable_inc': 90000, 'roth_conv': 0, 'surplus': -29000},
        {'year': 2027, 'h_age': 51, 'w_age': 50, 'earned': 100000, 'h_ss': 0, 'w_ss': 0, 'pension': 0, 'wife_single_ann': 0, 'wife_joint_ann': 0, 'h_single_ann': 0, 'h_joint_ann': 0, 'note_princ': 0, 'note_int': 0, 'rmd_total': 0, 'spend_base_yr': 83000, 'rec_extra': 0, 'lump': 0, 'mortgage': 36000, 'rent_yr': 0, 'fed_tax': 9500, 'state_tax': 3000, 'niit': 0, 'total_nw': 1050000, 'ann_nw': 0, 'pretax_nw': 110000, 'roth_nw': 60000, 'trust_nw': 30000, 'hsa_nw': 11000, 'home_equity': 510000, 'other_nw': 0, 'total_tax': 12500, 'agi': 100000, 'taxable_inc': 90000, 'roth_conv': 0, 'surplus': -31500},
    ]
    write_result_explorer_model(tmp_path / RESULTS_MODEL_FILENAME, c, rows, {'success_rate': 0.98})

    index = workbook_detailed_index(workbook)
    assert index['version'] == VERSION
    assert any(s['source'] == 'semantic_results_model' for s in index['sheets'])


def test_results_model_chart_dashboard_renders_from_model(tmp_path):
    from src.results_model import write_result_explorer_model, RESULTS_MODEL_FILENAME, model_sheet, read_result_explorer_model

    workbook = tmp_path / 'retirement_plan.xlsx'
    wb = Workbook()
    ws = wb.active
    ws.title = '8. Charts Dashboard'
    ws['A1'] = 'No chart sources here'
    wb.save(workbook)

    c = {'h_name': 'Matt', 'w_name': 'Pat', 'plan_start': 2026, 'plan_end': 2027, 'state': 'IL'}
    rows = [
        {'year': 2026, 'h_age': 50, 'w_age': 49, 'earned': 100000, 'spend_base_yr': 80000, 'mortgage': 36000, 'fed_tax': 10000, 'state_tax': 3000, 'total_nw': 1000000, 'pretax_nw': 100000, 'roth_nw': 50000, 'trust_nw': 25000, 'hsa_nw': 10000, 'home_equity': 500000, 'other_nw': 50000},
        {'year': 2027, 'h_age': 51, 'w_age': 50, 'earned': 100000, 'spend_base_yr': 83000, 'mortgage': 36000, 'fed_tax': 9500, 'state_tax': 3000, 'total_nw': 1050000, 'pretax_nw': 110000, 'roth_nw': 60000, 'trust_nw': 30000, 'hsa_nw': 11000, 'home_equity': 510000, 'other_nw': 52000},
    ]
    write_result_explorer_model(tmp_path / RESULTS_MODEL_FILENAME, c, rows, {'success_rate': 0.98})

    model_payload = read_result_explorer_model(tmp_path / RESULTS_MODEL_FILENAME)
    sheet = model_sheet(model_payload, '1E. Charts')
    assert sheet is not None
    assert sheet['source'] == 'semantic_results_model'
    assert sheet['kind'] == 'chart_dashboard'
    assert len(sheet['charts']) >= 3
    assert 'semantic results model' in sheet['chart_note'].lower()
