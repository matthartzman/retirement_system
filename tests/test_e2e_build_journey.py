"""Real end-to-end build journey (system review Q2, Wave 3 item 3.14).

tests/test_161_phase2_workflow_route_plumbing.py (previously named
..._live_workflow_journeys, which claimed coverage it didn't provide) proves
the routes wire together, but replaces `_run_build_progress_job` with a fake
that never runs the workbook builder, and `report_service.detailed_results_payload`
with a hand-written dict. No test before this one posted /api/build/start
and polled it to real completion against real sheet content.

This is that test: a genuine subprocess build via the actual HTTP routes
(POST /api/build/start -> poll GET /api/build/progress/<job_id> -> GET
/api/detailed-results), asserting the resulting payload is real workbook
content, not a fixture. Marked slow (registered in pyproject.toml) since it
runs a real tools/build_workbook.py subprocess end to end - skip locally
with `-m "not slow"`.
"""
from __future__ import annotations

import io
import time
from urllib.parse import urlencode

import pytest
from openpyxl import load_workbook

from src.server import app

HEADERS = {"X-User-Role": "admin"}
BUILD_TIMEOUT_SECONDS = 180


def _poll_until_done(client, job_id: str) -> dict:
    job: dict = {}
    deadline = time.monotonic() + BUILD_TIMEOUT_SECONDS
    while time.monotonic() < deadline:
        progress = client.get(f"/api/build/progress/{job_id}", headers=HEADERS)
        assert progress.status_code == 200
        job = progress.get_json()["job"]
        if job.get("status") in {"done", "failed"}:
            return job
        time.sleep(0.5)
    raise AssertionError(f"build job {job_id} did not finish within {BUILD_TIMEOUT_SECONDS}s: {job}")


@pytest.mark.slow
def test_real_build_journey_start_to_real_detailed_results_and_download(monkeypatch, tmp_path):
    # /api/build/start resolves its own output_dir as workspace_output_dir
    # (workspace_id, BASE_DIR) - literally the package root, not
    # tests/conftest.py's WORKSPACE_ROOT redirect - so the route (and the
    # later /api/detailed-results and /api/xlsx reads, which use the same
    # resolution) would otherwise look for this build's output in the real,
    # live output/ directory even though the spawned subprocess itself (via
    # platform_runtime.workspace_root(), which DOES honor the redirect)
    # writes into the throwaway workspace. RETIREMENT_SYSTEM_OUTPUT_DIR is
    # the override both sides already understand - same technique
    # conftest.py's built_workbook_dir fixture and test_192 use for their own
    # direct subprocess calls, applied here to the real HTTP-driven job.
    monkeypatch.setenv("RETIREMENT_SYSTEM_OUTPUT_DIR", str(tmp_path))

    client = app.test_client()

    started = client.post("/api/build/start", headers=HEADERS)
    assert started.status_code == 200, started.get_data(as_text=True)
    start_payload = started.get_json()
    assert start_payload["success"] is True
    job_id = start_payload["job_id"]
    assert job_id

    job = _poll_until_done(client, job_id)
    assert job.get("status") == "done", f"real build failed: {job}"
    assert job.get("progress") == 100

    # Real sheet index - not test_161's {"categories": [...], "sheets": []} fake.
    index = client.get("/api/detailed-results?index=1", headers=HEADERS)
    assert index.status_code == 200
    index_payload = index.get_json()
    assert index_payload["success"] is True
    sheets = index_payload.get("sheets") or []
    assert len(sheets) > 10, f"expected a real multi-sheet workbook index, got {len(sheets)} sheets"
    sheet_names = [s.get("name") or "" for s in sheets]
    exec_summary = next((n for n in sheet_names if "Executive Summary" in n), None)
    assert exec_summary, f"no Executive Summary sheet in real index: {sheet_names}"
    assert all(int(s.get("row_count") or 0) >= 0 for s in sheets)

    # Real per-sheet content: the Executive Summary should have actual
    # sections with actual rows, sourced from the just-built xlsx file. A
    # modern build writes a results_explorer_model.json sidecar, so this
    # goes through workbook_detailed_sheet() -> model_sheet(), whose success
    # shape is the page dict itself (name/sections/... at the top level, via
    # `dict(page, success=True, ...)`) - NOT wrapped in a "sheet" key like the
    # older Excel-parser fallback shape. Handle both so this doesn't depend
    # on which path today's build happens to take.
    detail = client.get(
        "/api/detailed-results?" + urlencode({"sheet": exec_summary}), headers=HEADERS
    )
    assert detail.status_code == 200
    detail_payload = detail.get_json()
    assert detail_payload.get("success") is True, detail_payload
    sheet = detail_payload.get("sheet") or detail_payload
    assert sheet.get("name") == exec_summary, detail_payload
    sections = sheet.get("sections") or []
    assert sections, "Executive Summary should have real section content, not an empty fixture"
    total_rows = sum(len(sec.get("rows") or []) for sec in sections)
    assert total_rows > 5, f"expected real row content from the built workbook, got {total_rows} rows"
    # At least one real cell value somewhere, so this isn't just blank rows.
    all_cells = [c for sec in sections for row in (sec.get("rows") or []) for c in (row.get("cells") or [])]
    assert any((c.get("value") not in (None, "")) for c in all_cells), "expected at least one non-empty real cell value"

    # The real xlsx this all came from should be downloadable and non-trivial.
    xlsx = client.get("/api/xlsx", headers=HEADERS)
    assert xlsx.status_code == 200
    xlsx_bytes = xlsx.get_data()
    assert len(xlsx_bytes) > 10_000, f"downloaded workbook suspiciously small ({len(xlsx_bytes)} bytes)"


@pytest.mark.slow
def test_real_build_journey_reflects_a_user_edited_input(monkeypatch, tmp_path):
    """Q4 (system review 2026-07-21): closes the gap this file's own docstring
    calls out -- the build-journey test above proves the HTTP build/results
    path works, but only against whatever input already contains; it never
    proves a user-edited input flows through the real plan-data save route
    into the built workbook. This posts a changed value to POST
    /api/config/rows (the same route dashboard.js's Save calls), builds a
    real workbook, then scans every sheet of the downloaded xlsx for the new
    value -- not one hardcoded sheet/cell, so this doesn't become brittle
    against sheet-layout changes unrelated to the input-edit path itself.
    """
    monkeypatch.setenv("RETIREMENT_SYSTEM_OUTPUT_DIR", str(tmp_path))

    client = app.test_client()

    rows_resp = client.get("/api/config/rows", headers=HEADERS)
    assert rows_resp.status_code == 200
    rows = rows_resp.get_json()["rows"]

    def _row_index(section, subsection, label):
        return next(
            r["row_index"] for r in rows
            if r["section"] == section and r["subsection"] == subsection and r["label"] == label
        )

    home_value_row = _row_index("Other Assets", "Home", "value_as_of_plan_start")
    appreciation_row = _row_index("Other Assets", "Home", "appreciation_rate")

    # A distinctive figure vanishingly unlikely to arise from any unrelated
    # computation. Appreciation is pinned to 0% in the same save so the
    # value can't drift across projection years regardless of which year a
    # sheet happens to display home value from.
    NEW_HOME_VALUE = 1_847_213
    saved = client.post(
        "/api/config/rows",
        json={
            "updates": [
                {"row_index": home_value_row, "value": f"${NEW_HOME_VALUE:,}"},
                {"row_index": appreciation_row, "value": "0.00%"},
            ],
            "sync": True,
        },
        headers=HEADERS,
    )
    assert saved.status_code == 200, saved.get_data(as_text=True)
    saved_payload = saved.get_json()
    assert saved_payload["success"] is True, saved_payload
    assert saved_payload["updated"] == 2, saved_payload

    started = client.post("/api/build/start", headers=HEADERS)
    assert started.status_code == 200, started.get_data(as_text=True)
    job_id = started.get_json()["job_id"]

    job = _poll_until_done(client, job_id)
    assert job.get("status") == "done", f"real build failed: {job}"

    xlsx = client.get("/api/xlsx", headers=HEADERS)
    assert xlsx.status_code == 200
    wb = load_workbook(io.BytesIO(xlsx.get_data()), data_only=True)

    found_at = None
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, (int, float)) and not isinstance(v, bool) and abs(v - NEW_HOME_VALUE) < 1.0:
                    found_at = (ws.title, cell.coordinate, v)
                elif isinstance(v, str) and str(NEW_HOME_VALUE) in v.replace(",", ""):
                    found_at = (ws.title, cell.coordinate, v)
                if found_at:
                    break
            if found_at:
                break
        if found_at:
            break

    if not found_at:
        for ws in wb.worksheets:
            if "BALANCE" in ws.title.upper():
                print(f"--- {ws.title} ---")
                for row in ws.iter_rows(max_row=30):
                    for cell in row:
                        if cell.value not in (None, ""):
                            print(cell.coordinate, repr(cell.value))
        print("job:", job)

    assert found_at, (
        f"edited home value {NEW_HOME_VALUE} never appeared anywhere in the "
        f"built workbook ({len(wb.worksheets)} sheets) -- the user-input-edit "
        f"path from /api/config/rows through the real build appears broken"
    )


@pytest.mark.slow
def test_detailed_results_read_routes_against_the_canonical_built_workbook(monkeypatch, built_workbook_dir):
    """Read-side coverage against `built_workbook_dir` (root conftest.py) - a
    real workbook other tests in this session already pay to build once, not
    a mocked payload. Marked slow (like every other `built_workbook_dir`/
    `built_workbook_path` consumer) so `-m "not slow"` never triggers a build
    by being the first test in a run to request the fixture; when the full
    suite runs, this still just reuses whatever build happened first.
    """
    monkeypatch.setenv("RETIREMENT_SYSTEM_OUTPUT_DIR", str(built_workbook_dir))

    client = app.test_client()

    index = client.get("/api/detailed-results?index=1", headers=HEADERS)
    assert index.status_code == 200
    index_payload = index.get_json()
    assert index_payload["success"] is True
    sheets = index_payload.get("sheets") or []
    assert len(sheets) > 10
    sheet_names = [s.get("name") or "" for s in sheets]

    # Spot-check several distinct real sheets, not just one, across a mix of
    # table and chart-dashboard kinds.
    checked = 0
    for name in sheet_names:
        if checked >= 4:
            break
        detail = client.get("/api/detailed-results?" + urlencode({"sheet": name}), headers=HEADERS)
        assert detail.status_code == 200, name
        payload = detail.get_json()
        assert payload.get("success") is True, (name, payload)
        page = payload.get("sheet") or payload
        assert page.get("name") == name, (name, payload)
        checked += 1
    assert checked >= 4, f"expected to check several real sheets, only found {checked} in {sheet_names}"

    pdf = client.get("/api/pdf", headers=HEADERS)
    assert pdf.status_code == 200
    assert len(pdf.get_data()) > 1000, "downloaded PDF suspiciously small"
