"""Wave 4 item 4.5 (system review P12 second half): a per-member Form SSA-44
life-changing-event appeal suppresses the IRMAA *surcharge* (never the base
Part B/D/G premium) from a stated year onward. The historical-MAGI seed half
of P12 already landed as item 2.6 (`irmaa_actual_magi_2yr_prior`/`1yr_prior`).
"""
from __future__ import annotations

import copy
import unittest
from pathlib import Path

from src.data_io import load_csv, parse_client
from src.planning_engines import project

ROOT = Path(__file__).resolve().parents[1]


def _config():
    c = parse_client(load_csv(ROOT / "input" / "client_data.csv"), "")
    c["roth_policy"] = "none"
    return c


def _irmaa_by_year(rows):
    return {int(r["year"]): float(r.get("irmaa") or 0.0) for r in rows}


class Ssa44IrmaaReliefTests(unittest.TestCase):
    def test_schema_declares_both_optional_relief_year_inputs(self):
        schema = (ROOT / "reference_data" / "schema.csv").read_text(encoding="utf-8")
        self.assertIn("h_ssa44_relief_year", schema)
        self.assertIn("w_ssa44_relief_year", schema)

    def test_blank_relief_year_parses_to_none_and_changes_nothing(self):
        c = _config()
        self.assertIsNone(c.get("h_ssa44_relief_year"))
        self.assertIsNone(c.get("w_ssa44_relief_year"))
        rows = project(c)
        self.assertFalse(any(r.get("irmaa_ssa44_relief_active") for r in rows))

    def test_relief_year_suppresses_surcharge_from_that_year_forward_only(self):
        baseline = _irmaa_by_year(project(_config()))
        irmaa_years = sorted(y for y, v in baseline.items() if v > 0)
        self.assertGreaterEqual(len(irmaa_years), 2, "fixture should already trigger IRMAA in 2+ years")
        relief_year = irmaa_years[len(irmaa_years) // 2]

        c = _config()
        c["h_ssa44_relief_year"] = relief_year
        relieved = _irmaa_by_year(project(c))

        for y in irmaa_years:
            if y < relief_year:
                self.assertAlmostEqual(relieved[y], baseline[y], places=2, msg=f"year {y} predates relief and should be unchanged")
            else:
                self.assertLessEqual(relieved[y], baseline[y] + 1e-6, msg=f"year {y} is post-relief and should not exceed baseline")

    def test_relief_never_touches_base_medicare_premium(self):
        c_base = _config()
        c_relief = _config()
        irmaa_years = sorted(y for y, v in _irmaa_by_year(project(c_base)).items() if v > 0)
        c_relief["h_ssa44_relief_year"] = irmaa_years[0]
        rows_base = project(c_base)
        rows_relief = project(c_relief)
        premium_base = {int(r["year"]): float(r.get("medicare_base_premium") or 0.0) for r in rows_base}
        premium_relief = {int(r["year"]): float(r.get("medicare_base_premium") or 0.0) for r in rows_relief}
        for y in premium_base:
            self.assertAlmostEqual(premium_base[y], premium_relief[y], places=2, msg=f"base Medicare premium must be unaffected by SSA-44 relief in {y}")

    def test_qc_sheet_caveats_an_active_relief_flag_as_granted_not_guaranteed(self):
        import openpyxl
        from src.reporting.sheets_qc_reference import build_sheet21

        c = _config()
        irmaa_years = sorted(y for y, v in _irmaa_by_year(project(c)).items() if v > 0)
        c["h_ssa44_relief_year"] = irmaa_years[0]
        rows = project(c)

        wb = openpyxl.Workbook()
        ws = wb.active
        build_sheet21(ws, [], rows, c)
        text = "\n".join(str(cell.value) for row in ws.iter_rows() for cell in row if cell.value is not None)
        self.assertIn("SSA-44", text)
        self.assertIn("never guaranteed", text)

        wb2 = openpyxl.Workbook()
        ws2 = wb2.active
        build_sheet21(ws2, [], project(_config()), _config())
        text2 = "\n".join(str(cell.value) for row in ws2.iter_rows() for cell in row if cell.value is not None)
        self.assertNotIn("SSA-44", text2)

    def test_backfill_entry_wires_ssa44_rows_into_client_policy_csv(self):
        import src.server.app_core as app_core
        entry = next(e for e in app_core.PLAN_DATA_BACKFILL_ENTRIES if e.rows is app_core.SSA44_UI_PLAN_DATA_ROWS)
        self.assertEqual(entry.file_name, "client_policy.csv")


if __name__ == "__main__":
    unittest.main()
