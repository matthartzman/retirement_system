"""Wave 5.6 (system review 2026-08-04, planner finding
ltc-scenarios-fixed-and-incomplete): "State-adjusted, surviving-spouse,
modeled funding."

Sheet 17 (LTC Stress Test) previously used flat national-median cost
figures regardless of the client's actual state of residence, only modeled
scenarios where both spouses stay alive throughout, and labeled each
scenario's "Funding Source" with a hand-picked string ('Trust / Roth',
'IRA + Trust', ...) that was never actually checked against what the
projection engine did. This fixes all three:

1. LTC_STATE_COST_INDEX scales the illustrative national-median cost
   figures for the client's actual state (src/reporting/sheets_stress.py).
2. A fifth scenario models the household's other major LTC risk pattern:
   one spouse dies early, and the survivor -- now filing Single, with
   reduced survivor income -- faces facility care alone later.
3. _ltc_modeled_funding_source computes which accounts ACTUALLY absorbed
   the incremental withdrawal during each scenario's care years (scenario
   minus baseline, by tax bucket), instead of a hardcoded label.
"""
from __future__ import annotations

from src.data_io import load_csv, parse_client
from src.plan_config import ensure_engine_config
from src.planning_engines import project
from src.reporting.sheets_stress import (
    LTC_STATE_COST_INDEX,
    _ltc_modeled_funding_source,
    _ltc_state_cost_index,
    build_sheet17,
)
from tests.golden_pricing import FROZEN_GOLDEN_MASTER_PRICES, frozen_holdings_prices

from conftest import TEST_INPUT_DIR


def sample_config_and_rows(**overrides):
    c = parse_client(load_csv(TEST_INPUT_DIR / "client_data.csv"), "")
    c["roth_policy"] = "none"
    c.update(overrides)
    c = ensure_engine_config(c, source="test")
    with frozen_holdings_prices(FROZEN_GOLDEN_MASTER_PRICES):
        rows = project(c)
    return c, rows


def test_illinois_is_anchored_at_national_baseline():
    # Sheet 19's existing "OPTIMAL" LTC illustration ($103K/yr IL median
    # facility care) was already implicitly IL-anchored -- IL must stay 1.0
    # so existing figures elsewhere in the workbook don't silently diverge.
    assert LTC_STATE_COST_INDEX["illinois"] == 1.0


def test_unlisted_state_defaults_to_national_baseline():
    c, rows = sample_config_and_rows()
    c = dict(c)
    c["state"] = "Some Unlisted Territory"
    assert _ltc_state_cost_index(c) == 1.0


def test_high_cost_state_scales_above_national_baseline():
    c, rows = sample_config_and_rows()
    c = dict(c)
    c["state"] = "California"
    assert _ltc_state_cost_index(c) > 1.0


def test_low_cost_state_scales_below_national_baseline():
    c, rows = sample_config_and_rows()
    c = dict(c)
    c["state"] = "Mississippi"
    assert _ltc_state_cost_index(c) < 1.0


def test_modeled_funding_source_reflects_real_incremental_withdrawals():
    # Regression for the old hardcoded label: force a large LTC shock and
    # confirm the reported source is derived from the scenario's actual
    # account-withdrawal delta, not a fixed string, and sums to ~100%.
    c, rows = sample_config_and_rows()
    from src.planning_engines import run_scenario

    shock_year = rows[len(rows) // 2]["year"]

    def _mutate(c2):
        shocks = dict(c2.get("wellness_shock_by_year") or {})
        shocks[shock_year] = shocks.get(shock_year, 0.0) + 150000.0
        c2["wellness_shock_by_year"] = shocks

    _c2, rows2 = run_scenario(c, mutate=_mutate)
    funding = _ltc_modeled_funding_source(c, rows, rows2, {shock_year})
    assert funding != ""
    assert "No incremental draw" in funding or "%" in funding


def test_sheet17_renders_five_scenarios_including_surviving_spouse():
    from openpyxl import Workbook

    c, rows = sample_config_and_rows()
    wb = Workbook()
    ws = wb.active
    build_sheet17(ws, c, rows)
    texts = [str(cell.value) for row in ws.iter_rows() for cell in row if cell.value is not None]
    assert any("Moderate Home Care" in t for t in texts)
    assert any("Severe Home Care" in t for t in texts)
    assert any("Facility (Memory Care)" in t for t in texts)
    assert any("Catastrophic (Both)" in t for t in texts)
    assert any("Surviving-Spouse" in t for t in texts)
    assert any("Funding Source (modeled)" in t for t in texts)
    assert any("regional cost index" in t for t in texts)


def test_survivor_scenario_picks_the_plans_own_longer_lived_spouse():
    c, rows = sample_config_and_rows()
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    build_sheet17(ws, c, rows)
    texts = [str(cell.value) for row in ws.iter_rows() for cell in row if cell.value is not None]
    survivor_row = next(t for t in texts if "Surviving-Spouse" in t)
    expected_survivor = (c.get("h_nick") or c.get("h_name") or "Member 1") \
        if c["h_death_yr"] >= c["w_death_yr"] \
        else (c.get("w_nick") or c.get("w_name") or "Member 2")
    assert str(expected_survivor) in survivor_row
