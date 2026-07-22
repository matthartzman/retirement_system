"""T3e (system review 2026-07-21, D3 follow-up): src/glossary.py is the one
canonical source both the workbook's Glossary sheet and the front end (via
GET /api/glossary) render from. Before this, dashboard.js's
ACRONYM_DEFINITIONS and sheets_qc_reference.py's hardcoded 22-term list were
independently maintained and had already drifted (divergent IRMAA wording).
"""
from openpyxl import Workbook

from src.glossary import build_glossary, GLOSSARY
from src.reporting.sheets_qc_reference import build_sheet22
from src.server import app

HEADERS = {"X-User-Role": "admin"}


def test_build_glossary_computes_a_fresh_salt_cap_definition():
    terms = build_glossary(2026)
    assert "SALT Cap" in terms
    assert "2026" in terms["SALT Cap"]
    assert "$" in terms["SALT Cap"]
    # Every other term is a static passthrough of GLOSSARY.
    assert terms["IRMAA"] == GLOSSARY["IRMAA"]


def test_workbook_glossary_sheet_renders_from_the_canonical_source():
    ws = Workbook().active
    build_sheet22(ws)
    rows = {
        ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value
        for r in range(3, ws.max_row + 1)
        if ws.cell(row=r, column=1).value
    }
    canonical = build_glossary()
    assert len(rows) == len(canonical)
    assert rows["IRMAA"] == canonical["IRMAA"]
    assert rows["SALT Cap"] == canonical["SALT Cap"]


def test_glossary_api_route_serves_the_same_canonical_terms():
    client = app.test_client()
    resp = client.get("/api/glossary", headers=HEADERS)
    assert resp.status_code == 200
    payload = resp.get_json()
    assert payload["success"] is True
    assert payload["schema"] == "glossary_v1"
    assert payload["terms"]["IRMAA"] == GLOSSARY["IRMAA"]
    assert payload["terms"] == build_glossary()


def test_irmaa_definition_is_reconciled_not_divergent():
    # The two independently-maintained wordings this item reconciled:
    # dashboard.js's old short form ("Income-related monthly adjustment
    # amount") vs the workbook's fuller form. The canonical source uses the
    # fuller, more informative wording.
    assert "Medicare" in GLOSSARY["IRMAA"]
    assert "MAGI" in GLOSSARY["IRMAA"]
