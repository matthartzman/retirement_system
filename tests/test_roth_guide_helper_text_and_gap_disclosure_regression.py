"""Ticket 289, Step 8.2: enhanced Roth field helper text, and disclosure of
the two gaps the Roth Conversion Modeling Guide names that this engine does
not implement (conversion tax payment source; asset-location-aware
conversion) -- in both the UI and the workbook, matching the P4/P7 house
style ("state what the model does do, then name the lever it does not
have").

See docs/superpowers/plans/2026-08-17-roth-guide-audit.md for the audit
these enhancements are drawn from.
"""
from __future__ import annotations

import csv
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
SCHEMA = ROOT / "reference_data" / "schema.csv"


def _schema_row(label):
    with SCHEMA.open(newline="", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            if row.get("label") == label:
                return row
    return None


# --------------------------------------------------------------------------
# Step 8.2: helper text for the four priority fields
# --------------------------------------------------------------------------


def test_tax_discount_rate_help_states_the_bracket_rule_and_the_why():
    row = _schema_row("roth_tax_discount_rate")
    assert row is not None
    desc = row["description"]
    assert "10-12%" in desc and "22-24%" in desc and "32%+" in desc, (
        "must state the guide's §1C bracket-conditioned rule of thumb"
    )
    assert "understates Roth value" in desc, (
        "must explain WHY a too-low rate is wrong, not just state the recommended value"
    )


def test_target_bracket_rate_help_states_the_decision_rule():
    row = _schema_row("roth_target_bracket_rate")
    assert row is not None
    desc = row["description"]
    assert "10-12%" in desc and "22-24%" in desc and "32%" in desc


def test_irmaa_target_tier_help_states_two_year_lookback_and_both_spouses():
    row = _schema_row("roth_irmaa_target_tier")
    assert row is not None
    desc = row["description"]
    assert "2-year lookback" in desc or "2 year lookback" in desc.replace("-", " ")
    assert "BOTH spouses" in desc or "both spouses" in desc.lower()


def test_terminal_and_lifetime_weight_help_explain_the_tradeoff():
    terminal = _schema_row("roth_optimize_terminal_weight")
    lifetime = _schema_row("roth_optimize_lifetime_tax_weight")
    assert terminal is not None and lifetime is not None
    assert "terminal net worth" in terminal["description"]
    assert "lifetime tax" in lifetime["description"].lower()
    # Each must explain what trading one for the other actually does, not just
    # restate its own name.
    assert "biases the optimizer" in terminal["description"]
    assert "biases the optimizer" in lifetime["description"]


def test_schema_csv_is_still_well_formed():
    """The four edits above are hand-written CSV; confirm no row lost a
    column or gained a stray one."""
    with SCHEMA.open(newline="", encoding="utf-8-sig") as f:
        rows = list(csv.DictReader(f))
    for label in (
        "roth_tax_discount_rate", "roth_target_bracket_rate",
        "roth_irmaa_target_tier", "roth_optimize_terminal_weight",
        "roth_optimize_lifetime_tax_weight",
    ):
        matches = [r for r in rows if r.get("label") == label]
        assert len(matches) == 1, f"{label} must appear exactly once"
        assert matches[0].get("section") == "Withdrawal Policy"
        assert matches[0].get("subsection") == "Roth Conversion"


# --------------------------------------------------------------------------
# UI disclosure
# --------------------------------------------------------------------------


def test_ui_discloses_both_gaps_near_the_conversion_controls():
    js = (ROOT / "frontend" / "js" / "dashboard_decomp_allocation_optimizer.js").read_text(encoding="utf-8")
    assert "how</b> conversion taxes are paid" in js
    assert "asset-location-aware conversion" in js
    # Must appear before the policy control fields render, i.e. near the top
    # of renderRothConversion, not buried after every field -- "near the
    # conversion policy controls" per the brief.
    fn_start = js.index("export function renderRothConversion()")
    disclosure_pos = js.index("how</b> conversion taxes are paid", fn_start)
    control_render_pos = js.index('${control.map(fieldHtml).join("")}', fn_start)
    assert disclosure_pos < control_render_pos


def test_ui_disclosure_is_gated_on_the_features_absence():
    """Ticket 289 requires the disclosure to remove itself automatically if
    either feature is later built, rather than becoming a stale claim."""
    js = (ROOT / "frontend" / "js" / "dashboard_decomp_allocation_optimizer.js").read_text(encoding="utf-8")
    assert 'rowByNormLabel("roth_conversion_tax_source")' in js
    assert 'rowByNormLabel("roth_conversion_asset_location_aware")' in js


# --------------------------------------------------------------------------
# Workbook disclosure
# --------------------------------------------------------------------------


def _build_minimal_sheet11():
    from openpyxl import Workbook
    from src.reporting.sheets_strategy import build_sheet11

    wb = Workbook()
    ws = wb.active
    c = {
        "roth_target_rate": 0.24, "roth_max_conversion_years": 10,
        "roth_headroom_usage_pct": 0.95, "roth_irmaa_headroom_usage_pct": 0.95,
        "roth_max_annual_conversion_pct_of_traditional_ira": 0.20,
        "roth_heir_filing_status": "Single",
    }
    build_sheet11(ws, c, [])
    return ws


def _sheet_text(ws):
    return "\n".join(
        str(cell) for row in ws.iter_rows(values_only=True) for cell in row if cell is not None
    )


def test_workbook_discloses_conversion_tax_source_gap():
    text = _sheet_text(_build_minimal_sheet11())
    assert "does not choose HOW" in text
    assert "the tax is paid" in text


def test_workbook_discloses_asset_location_gap():
    text = _sheet_text(_build_minimal_sheet11())
    assert "asset-location-aware selection" in text
    assert "LOCATION" in text and "not in-account sleeve variance" in text


def test_workbook_disclosure_is_gated_on_the_features_absence():
    """Setting either future plan-data key must remove its disclosure --
    proves the gate is live, not decorative."""
    from openpyxl import Workbook
    from src.reporting.sheets_strategy import build_sheet11

    wb = Workbook()
    ws = wb.active
    c = {
        "roth_target_rate": 0.24, "roth_max_conversion_years": 10,
        "roth_headroom_usage_pct": 0.95, "roth_irmaa_headroom_usage_pct": 0.95,
        "roth_max_annual_conversion_pct_of_traditional_ira": 0.20,
        "roth_heir_filing_status": "Single",
        "roth_conversion_tax_source": "taxable_cash",
        "roth_conversion_asset_location_aware": True,
    }
    build_sheet11(ws, c, [])
    text = _sheet_text(ws)
    assert "does not choose HOW" not in text
    assert "asset-location-aware selection" not in text
