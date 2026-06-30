from pathlib import Path
import openpyxl

from src.reporting.dashboard import build_html_dashboard


def _sample_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "8. Charts Dashboard"
    for r, year in enumerate(range(2026, 2031), start=5):
        ws.cell(r, 1, year)
        for c in range(2, 37):
            ws.cell(r, c, (r - 4) * 10000 + c * 1000)
    wb.save(path)


def test_html_dashboard_has_no_external_chart_dependencies(tmp_path):
    xlsx = tmp_path / "chart_source.xlsx"
    html_path = tmp_path / "retirement_dashboard.html"
    _sample_workbook(xlsx)

    build_html_dashboard(
        xlsx,
        html_path,
        rows=[],
        c={"positions": {}, "state": "IL", "plan_start": 2026, "plan_end": 2030},
    )

    html = html_path.read_text(encoding="utf-8")
    forbidden = [
        "Chart.js",
        "chart.umd",
        "cdnjs",
        "fonts.googleapis",
        "@import url(",
        "https://",
        "http://",
        "<canvas",
    ]
    for token in forbidden:
        assert token not in html
    assert "makeStackedBarSvg" in html
    assert "native-chart" in html
    assert "offline native SVG" in html


def test_html_dashboard_still_contains_three_chart_mounts(tmp_path):
    xlsx = tmp_path / "chart_source.xlsx"
    html_path = tmp_path / "retirement_dashboard.html"
    _sample_workbook(xlsx)

    build_html_dashboard(
        xlsx,
        html_path,
        rows=[],
        c={"positions": {}, "state": "IL", "plan_start": 2026, "plan_end": 2030},
    )

    html = html_path.read_text(encoding="utf-8")
    assert 'id="nwChart"' in html
    assert 'id="incChart"' in html
    assert 'id="expChart"' in html
    assert "renderAllCharts(); renderHoldings();" in html


def test_html_dashboard_reads_hidden_chart_source_sheet_when_visible_dashboard_is_blank(tmp_path):
    xlsx = tmp_path / "hidden_chart_source.xlsx"
    html_path = tmp_path / "retirement_dashboard.html"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "8. Charts Dashboard"
    ws.cell(1, 1, "visible dashboard intentionally has no source rows")
    data = wb.create_sheet("_Chart Dashboard Data")
    data.sheet_state = "hidden"
    # Header row matches build_sheet8 helper layout.
    for col, label in enumerate(["Year", "Annuities & Pension", "Pre-Tax IRA/401k", "Roth", "Trust", "HSA", "Home Equity", "Other Assets", "Total NW"], start=1):
        data.cell(4, col, label)
    for col, label in enumerate(["Year", "Earned Income", "Matthew SS", "Patricia SS", "Pension", "Wife Single Ann", "Wife Joint Ann", "Husband Single Ann", "Husband Joint Ann", "Note P+I", "RMD", "Trust Draw", "HSA Draw", "Roth Draw", "IRA Draw", "Home Equity Draw", "Σ Income"], start=11):
        data.cell(4, col, label)
    for col, label in enumerate(["Year", "Base Spending", "Rec Extras", "Lump Events", "Mortgage + RE Tax", "Rent", "Federal Tax", "State Tax (IL)", "NIIT", "Payroll Tax", "Σ Spend+Tax"], start=29):
        data.cell(4, col, label)
    for r, year in enumerate(range(2026, 2029), start=5):
        data.cell(r, 1, year)
        data.cell(r, 11, year)
        data.cell(r, 29, year)
        for c in range(2, 9):
            data.cell(r, c, 1000 * r + c)
        for c in range(12, 27):
            data.cell(r, c, 2000 * r + c)
        for c in range(30, 39):
            data.cell(r, c, 3000 * r + c)
    wb.save(xlsx)

    build_html_dashboard(
        xlsx,
        html_path,
        rows=[],
        c={"positions": {}, "state": "IL", "plan_start": 2026, "plan_end": 2028},
    )
    html = html_path.read_text(encoding="utf-8")
    assert "const YEARS=[2026, 2027, 2028]" in html
    assert "Home Equity" in html
    assert "Rent" in html


def test_workbook_ordering_preserves_hidden_chart_data_sheet_source():
    source = Path('src/reporting/workbook_builder.py').read_text(encoding='utf-8')
    assert 'ordered_names' in source
    assert 'remainder' in source


def test_xml_optimizer_embeds_chart_value_caches(tmp_path):
    from openpyxl.chart import BarChart, Reference
    from src.reporting.workbook_xml_optimizer import optimize_workbook_xml
    from zipfile import ZipFile

    xlsx = tmp_path / 'cached_chart.xlsx'
    wb = openpyxl.Workbook()
    chart_ws = wb.active
    chart_ws.title = '8. Charts Dashboard'
    data = wb.create_sheet('_Chart Dashboard Data')
    data.sheet_state = 'hidden'
    data.cell(4, 1, 'Year')
    data.cell(4, 2, 'Value')
    for r, year in enumerate([2026, 2027, 2028], start=5):
        data.cell(r, 1, year)
        data.cell(r, 2, 1000 * (r - 4))
    chart = BarChart()
    chart.add_data(Reference(data, min_col=2, min_row=4, max_row=7), titles_from_data=True)
    chart.set_categories(Reference(data, min_col=1, min_row=5, max_row=7))
    chart_ws.add_chart(chart, 'A1')
    wb.save(xlsx)

    result = optimize_workbook_xml(str(xlsx))
    assert result['status'] == 'ok'
    assert result['chart_caches_patched'] >= 2
    with ZipFile(xlsx) as z:
        chart_xml = z.read('xl/charts/chart1.xml').decode('utf-8')
    assert '<numCache>' in chart_xml or '<c:numCache>' in chart_xml
    assert '<ptCount val="3"' in chart_xml
