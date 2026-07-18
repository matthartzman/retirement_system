from __future__ import annotations

import math
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from src.reporting.workbook_common import TEMPLATE_LAYOUT

ROOT = Path(__file__).resolve().parents[1]


def _min_wrapped_height(ws, row: int, col: int) -> float:
    """Same formula as minimize_row_heights(): the shortest height that still
    fits a wrapped string cell's text at its merged range's combined width."""
    cell = ws.cell(row, col)
    merge = next(
        (mr for mr in ws.merged_cells.ranges if mr.min_row == row and mr.min_col == col),
        None,
    )
    col_span = range(merge.min_col, merge.max_col + 1) if merge else range(col, col + 1)
    eff_width = max(
        sum((ws.column_dimensions[get_column_letter(c)].width or 8.43) for c in col_span),
        1.0,
    )
    lines = sum(max(1, math.ceil(len(line) / eff_width)) for line in str(cell.value).splitlines() or [''])
    font_size = float(cell.font.size) if (cell.font and cell.font.size) else 10.0
    return max(1, lines) * (font_size + 4.0)


def _zero_money(value) -> bool:
    try:
        return abs(float(value or 0)) < 0.50
    except Exception:
        return False


def _row_is_sole_column_a(ws, row: int) -> bool:
    """True when column A of ``row`` holds text and every other column in that
    row is empty -- so column A alone drives the row's needed height. That's the
    only case where measuring column A (as ``_min_wrapped_height`` does) equals
    the whole row's minimized height; in mixed rows a taller cell in another
    column legitimately sets the height."""
    a = ws.cell(row, 1).value
    if not (isinstance(a, str) and a.strip()):
        return False
    for col in range(2, (ws.max_column or 1) + 1):
        if ws.cell(row, col).value not in (None, ''):
            return False
    return True


def test_before_after_rebalancing_omits_zero_before_and_after_rows(built_workbook_path):
    wb = load_workbook(built_workbook_path, data_only=False, read_only=True)
    ws = wb['2B. Asset Allocation']
    start = next(r for r in range(1, ws.max_row + 1) if ws.cell(r, 1).value == 'BEFORE & AFTER REBALANCING')
    zero_rows = []
    in_table = False
    for row in range(start, ws.max_row + 1):
        a = ws.cell(row, 1).value
        b = ws.cell(row, 2).value
        e = ws.cell(row, 5).value
        if a == 'TAX-EFFICIENT REBALANCING SEQUENCE':
            break
        if a == 'Bucket' and b == 'Before $' and e == 'After $':
            in_table = True
            continue
        if not in_table:
            continue
        if a is None:
            in_table = False
            continue
        if isinstance(a, str) and ' —  Total:' in a:
            in_table = False
            continue
        if _zero_money(b) and _zero_money(e):
            zero_rows.append((row, a))
    wb.close()
    assert zero_rows == []


def test_asset_allocation_columns_are_compact_and_wrapped(built_workbook_path):
    wb = load_workbook(built_workbook_path, data_only=False)
    ws = wb['2B. Asset Allocation']
    pinned = TEMPLATE_LAYOUT.get('2B. Asset Allocation', {}).get('cols', {})
    if pinned:
        # The reference formatting workbook (template for column widths and
        # height.xlsx) pins exact widths for this sheet at generation time,
        # overriding the heuristic <=125 compactness cap below.
        for i in range(1, 11):
            letter = get_column_letter(i)
            if letter in pinned:
                assert ws.column_dimensions[letter].width == pinned[letter]
    else:
        width_total = sum(ws.column_dimensions[get_column_letter(i)].width or 8.43 for i in range(1, 11))
        assert width_total <= 125
    # Row heights are minimized to whatever the actual wrapped text needs at the
    # sheet's final (post-override) column widths -- not a fixed floor. Assert
    # this over every column-A-only wrapped row (the section descriptions and
    # rebalancing-sequence bullets, whose height column A alone drives) rather
    # than pinning magic row numbers, which drift onto empty spacer rows as
    # account/label content reshapes the layout.
    multi_row_a_merges = {
        mr.min_row
        for mr in ws.merged_cells.ranges
        if mr.min_col == 1 and mr.max_row > mr.min_row
    }
    sole_a_rows = [
        r
        for r in range(1, ws.max_row + 1)
        if r not in multi_row_a_merges
        and ws.cell(r, 1).alignment
        and ws.cell(r, 1).alignment.wrap_text
        and _row_is_sole_column_a(ws, r)
    ]
    assert sole_a_rows, 'expected wrapped column-A description/bullet rows'
    # At least one such row must wrap to more than one line, proving wrap +
    # height minimization actually engages rather than every row trivially
    # fitting on a single line.
    assert any(_min_wrapped_height(ws, r, 1) > 15.5 for r in sole_a_rows), \
        'expected at least one multi-line wrapped column-A row'
    for r in sole_a_rows:
        expected = _min_wrapped_height(ws, r, 1)
        actual = ws.row_dimensions[r].height or 0
        assert actual >= expected - 0.5, f'row {r} height {actual} clips its text (needs >= {expected})'
        assert actual <= expected + 0.5, f'row {r} height {actual} is not minimized (needs ~{expected})'
    wb.close()


def test_source_keeps_zero_filter_and_workbook_layout_pass_for_future_builds():
    summary_source = (ROOT / 'src' / 'reporting' / 'sheets_summary.py').read_text(encoding='utf-8')
    common_source = (ROOT / 'src' / 'reporting' / 'workbook_common.py').read_text(encoding='utf-8')
    builder_source = (ROOT / 'src' / 'reporting' / 'workbook_builder.py').read_text(encoding='utf-8')
    assert '_hide_zero_before_after_row' in summary_source
    assert 'if _hide_zero_before_after_row(before_val, after_val):' in summary_source
    assert 'if _hide_zero_before_after_row(bv, av):' in summary_source
    assert 'def optimize_workbook_layout' in common_source
    assert 'wrap_text=True' in common_source
    assert 'row_dimensions[row_idx].height' in common_source
    assert 'ordered_names' in builder_source
