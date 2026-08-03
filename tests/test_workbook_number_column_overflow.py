from __future__ import annotations

"""A numeric column too narrow for its formatted value renders as Excel's
"#####" overflow indicator. widen_overflowing_number_columns() is the
safety-net pass that catches this after every width-setting stage
(heuristic sizing, the reference template, user overrides from Settings ->
Workbook Formatting) has run -- any of those can shrink a column purely to
hit a target width or match a pinned template value without checking
whether the result still fits the actual data.
"""

from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

from src.reporting.workbook_common import (
    FMT_DOLLAR,
    FMT_INT,
    FMT_PCT,
    FMT_SCORE,
    _needed_number_width,
    widen_overflowing_number_columns,
)

ROOT = Path(__file__).resolve().parents[1]


def test_needed_number_width_matches_actual_display_length():
    cases = [
        (1234567, FMT_DOLLAR, "$1,234,567"),
        (-1234567, FMT_DOLLAR, "($1,234,567)"),
        (0, FMT_DOLLAR, "-"),  # FMT_DOLLAR's zero section is a literal "-"
        (0.256, FMT_PCT, "25.6%"),
        (-0.05, FMT_PCT, "-5.0%"),
        (12345678, FMT_INT, "12,345,678"),
        (87.3, FMT_SCORE, "87.3/100"),
        (2027, "0", "2027"),
    ]
    for value, fmt, expected_display in cases:
        assert _needed_number_width(value, fmt) == len(expected_display), (
            f"{value!r} under {fmt!r} should need width {len(expected_display)} "
            f"to show {expected_display!r}"
        )


def test_needed_number_width_ignores_general_format_and_non_numeric_values():
    assert _needed_number_width(1234567, "General") is None
    assert _needed_number_width(1234567, "") is None
    assert _needed_number_width("not a number", FMT_DOLLAR) is None


def test_widen_overflowing_number_columns_widens_a_too_narrow_dollar_column():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(1, 1, value="Label")
    ws.cell(2, 1, value=123456789)
    ws.cell(2, 1).number_format = FMT_DOLLAR
    ws.column_dimensions["A"].width = 6.0  # too narrow -- would show #####

    widen_overflowing_number_columns(wb)

    new_width = ws.column_dimensions["A"].width
    assert new_width > 6.0
    assert new_width >= _needed_number_width(123456789, FMT_DOLLAR)


def test_widen_overflowing_number_columns_leaves_already_wide_columns_alone():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(1, 1, value=1000)
    ws.cell(1, 1).number_format = FMT_INT
    ws.column_dimensions["A"].width = 25.0  # already plenty wide

    widen_overflowing_number_columns(wb)

    assert ws.column_dimensions["A"].width == 25.0


def test_widen_overflowing_number_columns_widens_the_widest_row_in_a_column():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.column_dimensions["A"].width = 6.0
    for row, value in enumerate([100, 987654321, 42], start=1):
        cell = ws.cell(row, 1, value=value)
        cell.number_format = FMT_INT

    widen_overflowing_number_columns(wb)

    assert ws.column_dimensions["A"].width >= _needed_number_width(987654321, FMT_INT)


def test_widen_overflowing_number_columns_handles_merged_numeric_cells():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    cell = ws.cell(1, 1, value=123456789)
    cell.number_format = FMT_DOLLAR
    for col in ("A", "B", "C"):
        ws.column_dimensions[col].width = 3.0  # combined 9.0, still too narrow

    widen_overflowing_number_columns(wb)

    total_width = sum(ws.column_dimensions[c].width for c in ("A", "B", "C"))
    assert total_width >= _needed_number_width(123456789, FMT_DOLLAR)


def test_column_overflow_pass_runs_before_row_height_minimization():
    # Widening a numeric column must happen before minimize_row_heights()
    # computes final row heights, since that pass sums final column widths
    # across a merged range to size wrapped text sharing those columns.
    src = (ROOT / "src/reporting/workbook_builder.py").read_text(encoding="utf-8")
    assert src.index("widen_overflowing_number_columns(wb") < src.index("minimize_row_heights(wb)")


def test_column_overflow_pass_runs_after_user_overrides_and_respects_them():
    # #251: apply_overrides() is documented to run "last so user edits
    # always win" -- widen_overflowing_number_columns must run after it in
    # the real build (so its widening reflects the final width) but must
    # NOT then re-widen a column the user explicitly overrode.
    src = (ROOT / "src/reporting/workbook_builder.py").read_text(encoding="utf-8")
    assert src.index("_apply_format_overrides(wb") < src.index("widen_overflowing_number_columns(wb")
    assert "protected_columns=" in src


def test_widen_overflowing_number_columns_leaves_a_protected_column_alone():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet"
    ws.cell(2, 1, value=123456789)
    ws.cell(2, 1).number_format = FMT_DOLLAR
    ws.column_dimensions["A"].width = 6.0  # user deliberately narrowed this

    widen_overflowing_number_columns(wb, protected_columns={"Sheet": {"A"}})

    assert ws.column_dimensions["A"].width == 6.0


def test_widen_overflowing_number_columns_still_widens_unprotected_columns_on_a_protected_sheet():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet"
    ws.cell(2, 1, value=123456789)
    ws.cell(2, 1).number_format = FMT_DOLLAR
    ws.cell(2, 2, value=123456789)
    ws.cell(2, 2).number_format = FMT_DOLLAR
    ws.column_dimensions["A"].width = 6.0
    ws.column_dimensions["B"].width = 6.0

    widen_overflowing_number_columns(wb, protected_columns={"Sheet": {"A"}})

    assert ws.column_dimensions["A"].width == 6.0
    assert ws.column_dimensions["B"].width >= _needed_number_width(123456789, FMT_DOLLAR)


def test_widen_overflowing_number_columns_only_protects_the_named_sheet():
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.cell(2, 1, value=123456789)
    ws1.cell(2, 1).number_format = FMT_DOLLAR
    ws1.column_dimensions["A"].width = 6.0
    ws2 = wb.create_sheet("Sheet2")
    ws2.cell(2, 1, value=123456789)
    ws2.cell(2, 1).number_format = FMT_DOLLAR
    ws2.column_dimensions["A"].width = 6.0

    widen_overflowing_number_columns(wb, protected_columns={"Sheet1": {"A"}})

    assert ws1.column_dimensions["A"].width == 6.0
    assert ws2.column_dimensions["A"].width >= _needed_number_width(123456789, FMT_DOLLAR)


def test_overridden_width_columns_resolves_via_sheet_renames(tmp_path):
    from src.reporting.workbook_format_config import merge_overrides, overridden_width_columns

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "3A. Executive Summary"
    merge_overrides({"3. Executive Summary": {"C": 20.0}}, input_dir=tmp_path)

    protected = overridden_width_columns(
        wb, input_dir=tmp_path, sheet_renames={"3. Executive Summary": "3A. Executive Summary"}
    )

    assert protected == {"3A. Executive Summary": {"C"}}


def test_overridden_width_columns_empty_when_no_overrides_saved(tmp_path):
    from src.reporting.workbook_format_config import overridden_width_columns

    wb = openpyxl.Workbook()
    assert overridden_width_columns(wb, input_dir=tmp_path) == {}
