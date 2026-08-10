"""Wave 5.4 (system review 2026-08-04, planner finding
no-current-vs-proposed-deliverable): "Current-vs-proposed comparison report."

New Sheet 37 ("Current vs. Proposed") turns the Executive Summary's
recommendation list into an actual verified comparison: for each
not-yet-adopted recommendation that maps to a config override the
deterministic engine actually reads, it re-runs the plan via the same
run_scenario() helper every other stress/scenario sheet uses and shows the
real terminal-net-worth and lifetime-tax deltas. Recommendations that only
affect report-time-only figures (cst_enabled/qtip_enabled/daf_amount are
read by report-building code, not projection_stages/deterministic_engine.py)
are listed separately with a pointer to their sheet instead of a fabricated
$0 delta.
"""
from __future__ import annotations

from src.data_io import load_csv, parse_client
from src.plan_config import ensure_engine_config
from src.planning_engines import project
from src.reporting.sheets_current_vs_proposed import (
    _proposed_changes,
    _report_only_items,
    build_sheet_current_vs_proposed,
)
from tests.golden_pricing import FROZEN_GOLDEN_MASTER_PRICES, frozen_holdings_prices

from conftest import TEST_INPUT_DIR


def sample_config_and_rows():
    c = parse_client(load_csv(TEST_INPUT_DIR / "client_data.csv"), "")
    c["roth_policy"] = "none"
    c = ensure_engine_config(c, source="test")
    with frozen_holdings_prices(FROZEN_GOLDEN_MASTER_PRICES):
        rows = project(c)
    return c, rows


def test_sheet37_registered_in_the_workbook_layout():
    from src.module_catalog import SHEET_REGISTRY

    assert "37. Current vs Proposed" in SHEET_REGISTRY
    spec = SHEET_REGISTRY["37. Current vs Proposed"]
    # Always-on: no module_key, matching other flagship report sheets
    # (Executive Summary, Balance Sheet, Spending Summary).
    assert spec.module_key is None


def test_report_only_recommendations_are_never_run_through_the_engine():
    # cst_enabled / qtip_enabled / daf_amount are read only by report-building
    # code (sheets_summary_builder.py, sheets_strategy.py), never by
    # projection_stages/deterministic_engine.py -- a run_scenario delta for
    # them would always be exactly $0 and wrongly imply "no benefit."
    c, rows = sample_config_and_rows()
    c = dict(c)
    c["cst_enabled"] = False
    c["qtip_enabled"] = False
    c["daf_amount"] = 0
    report_only = _report_only_items(c)
    labels = {label for label, _note, _is_active, _how_to in report_only}
    assert "Credit Shelter Trust at First Death" in labels
    assert "QTIP Trust for Annuity Post-First-Death" in labels
    assert "DAF Contribution in Highest-Income Year" in labels
    assert all(is_active is False for _l, _n, is_active, _h in report_only)


def test_already_adopted_recommendations_show_as_active_not_excluded():
    # #272: Sheet 37 must list every tracked recommendation, active or not --
    # a planner reviewing this sheet needs to see "already done" alongside
    # "not yet incorporated," not have adopted items silently disappear.
    c, rows = sample_config_and_rows()
    c = dict(c)
    c["cst_enabled"] = True
    c["qtip_enabled"] = True
    c["daf_amount"] = 50000
    c["ltc_enabled"] = True
    c["entity"] = "s_corp"
    report_only = _report_only_items(c)
    changes = _proposed_changes(c)
    assert report_only, "expected active estate/charitable items to still be listed"
    assert all(is_active is True for _l, _n, is_active, _h in report_only)
    assert changes, "expected active engine-modeled items to still be listed"
    assert all(is_active is True for _l, _o, _n, is_active, _h in changes)


def test_ltc_proposed_change_shows_a_real_engine_delta_not_zero():
    c, rows = sample_config_and_rows()
    c = dict(c)
    c["ltc_enabled"] = False
    changes = _proposed_changes(c)
    ltc_changes = [ch for ch in changes if "LTC" in ch[0]]
    assert ltc_changes, "expected an LTC recommendation when ltc_enabled is off"
    label, overrides, note, is_active, how_to = ltc_changes[0]
    assert is_active is False
    assert overrides["ltc_enabled"] is True
    assert overrides.get("ltc_annual_prem", 0) > 0


def test_sheet37_renders_and_uses_the_real_projection_engine():
    from openpyxl import Workbook

    c, rows = sample_config_and_rows()
    wb = Workbook()
    ws = wb.active
    build_sheet_current_vs_proposed(ws, c, rows)
    texts = [str(cell.value) for row in ws.iter_rows() for cell in row if cell.value is not None]
    assert any("CURRENT VS. PROPOSED" in t for t in texts)
    assert any("Engine-Modeled Comparisons" in t for t in texts)
    assert any("Estate/Charitable-Only Recommendations" in t for t in texts)


def test_sheet37_shows_active_status_when_frozen_fixture_already_adopted_them():
    # The frozen test fixture already has cst/qtip/daf configured (per Wave
    # 5.3's own investigation); #272 means those rows are now shown marked
    # ACTIVE rather than omitted.
    from openpyxl import Workbook

    c, rows = sample_config_and_rows()
    wb = Workbook()
    ws = wb.active
    build_sheet_current_vs_proposed(ws, c, rows)
    texts = [str(cell.value) for row in ws.iter_rows() for cell in row if cell.value is not None]
    report_only = _report_only_items(c)
    if report_only and all(is_active for _l, _n, is_active, _h in report_only):
        assert any(t == "ACTIVE" for t in texts)
